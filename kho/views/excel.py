# kho/views/excel.py
"""
Excel export/import views cho kho/products.
"""

from kho.utils import group_required
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.views.decorators.http import require_POST
import logging
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from core.sapo_client import get_sapo_client
from products.services.sapo_product_service import SapoProductService
from products.brand_settings import (
    is_brand_enabled,
    reload_settings
)
from products.services.metadata_helper import init_empty_metadata, update_variant_metadata
from products.services.dto import VariantMetadataDTO, BoxInfoDTO, PackedInfoDTO, NhanPhuInfoDTO

logger = logging.getLogger(__name__)


@group_required("WarehouseManager")
def export_products_excel(request: HttpRequest):
    """
    Export danh sách variants ra file Excel.
    Chỉ export brand hiện tại đang được lọc.
    """
    try:
        sapo_client = get_sapo_client()
        core_repo = sapo_client.core
        product_service = SapoProductService(sapo_client)
        
        # Lấy brand_id từ query parameters (mặc định 833608)
        DEFAULT_BRAND_ID = 833608
        brand_id = request.GET.get('brand_id', str(DEFAULT_BRAND_ID))
        try:
            brand_id = int(brand_id)
        except (ValueError, TypeError):
            brand_id = DEFAULT_BRAND_ID
        
        # Reload settings
        reload_settings()
        
        # Lấy variants theo brand_id từ Sapo API
        all_variants = []
        page = 1
        limit = 250
        expected_total = None
        
        while True:
            variants_response = core_repo.list_variants_raw(
                page=page,
                limit=limit,
                brand_ids=brand_id,
                composite=False,
                packsize=False
            )
            
            variants_data = variants_response.get("variants", [])
            
            metadata = variants_response.get("metadata", {})
            if page == 1:
                expected_total = metadata.get("total", 0)
            
            if not variants_data:
                break
            
            all_variants.extend(variants_data)
            
            if expected_total and expected_total > 0:
                if len(all_variants) >= expected_total:
                    break
            else:
                total_pages = metadata.get("total_pages")
                if total_pages:
                    if page >= total_pages:
                        break
                else:
                    if expected_total and expected_total > 0:
                        calculated_pages = (expected_total + limit - 1) // limit
                        if page >= calculated_pages:
                            break
            
            if len(variants_data) < limit:
                break
            
            page += 1
            
            if page > 100:
                break
        
        # Lấy product metadata cho từng variant
        product_map = {}
        product_ids = set(v.get("product_id") for v in all_variants if v.get("product_id"))
        
        product_ids_list = list(product_ids)
        batch_size = 50
        for i in range(0, len(product_ids_list), batch_size):
            batch_ids = product_ids_list[i:i+batch_size]
            for product_id in batch_ids:
                try:
                    product = product_service.get_product(product_id)
                    if product:
                        product_map[product_id] = product
                except Exception as e:
                    logger.warning(f"Failed to get product {product_id}: {e}")
                    continue
        
        # Parse variants và lấy metadata
        variants_data = []
        
        for variant_raw in all_variants:
            variant_id = variant_raw.get("id")
            product_id = variant_raw.get("product_id")
            
            product = product_map.get(product_id)
            if not product:
                brand = variant_raw.get("brand") or ""
            else:
                brand = product.brand or ""
            
            # Chỉ thêm variants nếu nhãn hiệu được bật
            if brand and not is_brand_enabled(brand):
                continue
            
            # Lấy metadata từ product nếu có
            variant_meta = None
            nhanphu_info = None
            if product:
                if product.gdp_metadata:
                    nhanphu_info = product.gdp_metadata.nhanphu_info
                
                for v in product.variants:
                    if v.id == variant_id:
                        variant_meta = v.gdp_metadata
                        break
            
            # Chuẩn bị dữ liệu cho Excel
            row_data = {
                "sku": variant_raw.get("sku", ""),
                "vari_id": variant_id,
                # Box info
                "full_box": variant_meta.box_info.full_box if variant_meta and variant_meta.box_info and variant_meta.box_info.full_box else "",
                "box_length_cm": variant_meta.box_info.length_cm if variant_meta and variant_meta.box_info and variant_meta.box_info.length_cm else "",
                "box_width_cm": variant_meta.box_info.width_cm if variant_meta and variant_meta.box_info and variant_meta.box_info.width_cm else "",
                "box_height_cm": variant_meta.box_info.height_cm if variant_meta and variant_meta.box_info and variant_meta.box_info.height_cm else "",
                # Packed info
                "packed_length_cm": variant_meta.packed_info.length_cm if variant_meta and variant_meta.packed_info and variant_meta.packed_info.length_cm else "",
                "packed_width_cm": variant_meta.packed_info.width_cm if variant_meta and variant_meta.packed_info and variant_meta.packed_info.width_cm else "",
                "packed_height_cm": variant_meta.packed_info.height_cm if variant_meta and variant_meta.packed_info and variant_meta.packed_info.height_cm else "",
                "packed_weight_with_box_g": variant_meta.packed_info.weight_with_box_g if variant_meta and variant_meta.packed_info and variant_meta.packed_info.weight_with_box_g else "",
                "packed_weight_without_box_g": variant_meta.packed_info.weight_without_box_g if variant_meta and variant_meta.packed_info and variant_meta.packed_info.weight_without_box_g else "",
                # TQ info
                "sku_tq": variant_meta.sku_tq if variant_meta and variant_meta.sku_tq else "",
                "name_tq": variant_meta.name_tq if variant_meta and variant_meta.name_tq else "",
                # Nhanphu info
                "nhanphu_vi_name": nhanphu_info.vi_name if nhanphu_info and nhanphu_info.vi_name else "",
                "nhanphu_en_name": nhanphu_info.en_name if nhanphu_info and nhanphu_info.en_name else "",
                "nhanphu_description": nhanphu_info.description if nhanphu_info and nhanphu_info.description else "",
                "nhanphu_material": nhanphu_info.material if nhanphu_info and nhanphu_info.material else "",
                # Update column
                "update": ""  # Cột update để user đánh dấu
            }
            variants_data.append(row_data)
        
        # Tạo Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Headers
        headers = [
            "sku", "vari_id",
            "full_box", "box_length_cm", "box_width_cm", "box_height_cm",
            "packed_length_cm", "packed_width_cm", "packed_height_cm",
            "packed_weight_with_box_g", "packed_weight_without_box_g",
            "sku_tq", "name_tq",
            "nhanphu_vi_name", "nhanphu_en_name", "nhanphu_description", "nhanphu_material",
            "update"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row_data in enumerate(variants_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="kho_products_export.xlsx"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting products Excel: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@group_required("WarehouseManager")
@require_POST
def import_products_excel(request: HttpRequest):
    """
    Import và cập nhật variants từ file Excel.
    Chỉ cập nhật những dòng có giá trị trong cột "update" (update != null).
    """
    try:
        if 'file' not in request.FILES:
            return JsonResponse({
                "status": "error",
                "message": "Không có file được upload"
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Kiểm tra file extension
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                "status": "error",
                "message": "File phải là định dạng Excel (.xlsx hoặc .xls)"
            }, status=400)
        
        # Đọc file Excel
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        
        # Đọc headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Tìm index của các cột
        header_map = {}
        for idx, header in enumerate(headers):
            if header:
                header_map[str(header).strip()] = idx
        
        # Kiểm tra các cột bắt buộc
        required_headers = ['vari_id', 'update']
        missing_headers = [h for h in required_headers if h not in header_map]
        if missing_headers:
            return JsonResponse({
                "status": "error",
                "message": f"Thiếu các cột bắt buộc: {', '.join(missing_headers)}"
            }, status=400)
        
        # Đọc dữ liệu
        sapo_client = get_sapo_client()
        product_service = SapoProductService(sapo_client)
        
        update_idx = header_map['update']
        vari_id_idx = header_map.get('vari_id')
        
        results = {
            "total_rows": 0,
            "processed": 0,
            "success": 0,
            "skipped": 0,
            "errors": 0,
            "error_details": []
        }
        
        # Process từng dòng
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            results["total_rows"] += 1
            
            # Kiểm tra cột update - chỉ quan tâm khi update != null
            update_cell = row[update_idx] if update_idx < len(row) else None
            update_value = update_cell.value if update_cell else None
            
            # Chỉ xử lý nếu có giá trị trong cột update (không null, không rỗng)
            if update_value is None or str(update_value).strip() == "":
                results["skipped"] += 1
                continue
            
            try:
                # Lấy vari_id
                vari_id = None
                if vari_id_idx is not None and vari_id_idx < len(row):
                    vari_id_value = row[vari_id_idx].value
                    if vari_id_value:
                        try:
                            vari_id = int(float(vari_id_value))
                        except:
                            pass
                
                if not vari_id:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Không tìm thấy variant_id")
                    continue
                
                # Lấy product_id từ variant
                variant_response = sapo_client.core.get_variant_raw(vari_id)
                variant = variant_response.get("variant", {})
                if not variant:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Variant {vari_id} không tồn tại")
                    continue
                
                product_id = variant.get("product_id")
                if not product_id:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Variant {vari_id} không có product_id")
                    continue
                
                # Lấy product hiện tại
                product = product_service.get_product(product_id)
                if not product:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Product {product_id} không tồn tại")
                    continue
                
                # Lấy metadata hiện tại
                current_metadata = product.gdp_metadata
                if not current_metadata:
                    variant_ids = [v.id for v in product.variants]
                    current_metadata = init_empty_metadata(product_id, variant_ids)
                
                # Tìm variant metadata hiện tại
                variant_meta = None
                for vm in current_metadata.variants:
                    if vm.id == vari_id:
                        variant_meta = vm
                        break
                
                if not variant_meta:
                    variant_meta = VariantMetadataDTO(id=vari_id)
                
                # Helper functions
                def get_cell_value(header_name, default=None):
                    if header_name in header_map:
                        idx = header_map[header_name]
                        if idx < len(row):
                            value = row[idx].value
                            return value if value is not None else default
                    return default
                
                def to_float_safe(val, default=None):
                    try:
                        if val is None or val == "":
                            return default
                        return float(val)
                    except:
                        return default
                
                def to_int_safe(val, default=None):
                    try:
                        if val is None or val == "":
                            return default
                        return int(float(val))
                    except:
                        return default
                
                # Update box_info
                full_box = to_int_safe(get_cell_value('full_box'))
                box_length = to_float_safe(get_cell_value('box_length_cm'))
                box_width = to_float_safe(get_cell_value('box_width_cm'))
                box_height = to_float_safe(get_cell_value('box_height_cm'))
                
                box_info = None
                if full_box is not None or box_length is not None or box_width is not None or box_height is not None:
                    box_info = BoxInfoDTO(
                        full_box=full_box if full_box is not None else (variant_meta.box_info.full_box if variant_meta.box_info else None),
                        length_cm=box_length if box_length is not None else (variant_meta.box_info.length_cm if variant_meta.box_info else None),
                        width_cm=box_width if box_width is not None else (variant_meta.box_info.width_cm if variant_meta.box_info else None),
                        height_cm=box_height if box_height is not None else (variant_meta.box_info.height_cm if variant_meta.box_info else None)
                    )
                elif variant_meta.box_info:
                    box_info = variant_meta.box_info
                
                # Update packed_info
                packed_length = to_float_safe(get_cell_value('packed_length_cm'))
                packed_width = to_float_safe(get_cell_value('packed_width_cm'))
                packed_height = to_float_safe(get_cell_value('packed_height_cm'))
                packed_weight_with_box = to_float_safe(get_cell_value('packed_weight_with_box_g'))
                packed_weight_without_box = to_float_safe(get_cell_value('packed_weight_without_box_g'))
                
                packed_info = None
                if packed_length is not None or packed_width is not None or packed_height is not None or packed_weight_with_box is not None or packed_weight_without_box is not None:
                    packed_info = PackedInfoDTO(
                        length_cm=packed_length if packed_length is not None else (variant_meta.packed_info.length_cm if variant_meta.packed_info else None),
                        width_cm=packed_width if packed_width is not None else (variant_meta.packed_info.width_cm if variant_meta.packed_info else None),
                        height_cm=packed_height if packed_height is not None else (variant_meta.packed_info.height_cm if variant_meta.packed_info else None),
                        weight_with_box_g=packed_weight_with_box if packed_weight_with_box is not None else (variant_meta.packed_info.weight_with_box_g if variant_meta.packed_info else None),
                        weight_without_box_g=packed_weight_without_box if packed_weight_without_box is not None else (variant_meta.packed_info.weight_without_box_g if variant_meta.packed_info else None)
                    )
                elif variant_meta.packed_info:
                    packed_info = variant_meta.packed_info
                
                # Update TQ info
                sku_tq = get_cell_value('sku_tq', '').strip() if get_cell_value('sku_tq') else None
                name_tq = get_cell_value('name_tq', '').strip() if get_cell_value('name_tq') else None
                
                # Update nhanphu_info (product level)
                nhanphu_vi_name = get_cell_value('nhanphu_vi_name', '').strip() if get_cell_value('nhanphu_vi_name') else None
                nhanphu_en_name = get_cell_value('nhanphu_en_name', '').strip() if get_cell_value('nhanphu_en_name') else None
                nhanphu_description = get_cell_value('nhanphu_description', '').strip() if get_cell_value('nhanphu_description') else None
                nhanphu_material = get_cell_value('nhanphu_material', '').strip() if get_cell_value('nhanphu_material') else None
                
                nhanphu_info = None
                if nhanphu_vi_name is not None or nhanphu_en_name is not None or nhanphu_description is not None or nhanphu_material is not None:
                    nhanphu_info = NhanPhuInfoDTO(
                        vi_name=nhanphu_vi_name if nhanphu_vi_name else (current_metadata.nhanphu_info.vi_name if current_metadata.nhanphu_info else None),
                        en_name=nhanphu_en_name if nhanphu_en_name else (current_metadata.nhanphu_info.en_name if current_metadata.nhanphu_info else None),
                        description=nhanphu_description if nhanphu_description else (current_metadata.nhanphu_info.description if current_metadata.nhanphu_info else None),
                        material=nhanphu_material if nhanphu_material else (current_metadata.nhanphu_info.material if current_metadata.nhanphu_info else None),
                        hdsd=current_metadata.nhanphu_info.hdsd if current_metadata.nhanphu_info else None
                    )
                elif current_metadata.nhanphu_info:
                    nhanphu_info = current_metadata.nhanphu_info
                
                # Tạo variant metadata mới
                new_variant_meta = VariantMetadataDTO(
                    id=vari_id,
                    price_tq=variant_meta.price_tq,
                    sku_tq=sku_tq if sku_tq else variant_meta.sku_tq,
                    name_tq=name_tq if name_tq else variant_meta.name_tq,
                    box_info=box_info,
                    packed_info=packed_info,
                    sku_model_xnk=variant_meta.sku_model_xnk,
                    web_variant_id=variant_meta.web_variant_id if variant_meta.web_variant_id else []
                )
                
                # Update variant metadata
                current_metadata = update_variant_metadata(
                    current_metadata,
                    vari_id,
                    new_variant_meta
                )
                
                # Update product level nhanphu_info nếu có
                if nhanphu_info:
                    current_metadata.nhanphu_info = nhanphu_info
                
                # Lưu vào Sapo
                success = product_service.update_product_metadata(
                    product_id,
                    current_metadata,
                    preserve_description=True
                )
                
                if success:
                    results["success"] += 1
                else:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Không thể lưu variant {vari_id}")
                
                results["processed"] += 1
                
            except Exception as e:
                results["errors"] += 1
                error_msg = f"Dòng {row_idx}: {str(e)}"
                results["error_details"].append(error_msg)
                logger.error(f"Error processing row {row_idx}: {e}", exc_info=True)
        
        return JsonResponse({
            "status": "success",
            "result": results
        })
        
    except Exception as e:
        logger.error(f"Error importing products Excel: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)

