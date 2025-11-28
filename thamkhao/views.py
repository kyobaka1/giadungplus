# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.template import loader
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from django.http import Http404,  HttpResponseRedirect
from django.views.decorators.csrf import csrf_exempt
import requests
import os
from rest_framework import viewsets
from rest_framework.response import Response
import xlrd
import openpyxl
from urllib.parse import urlencode
import xlwt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials
from oauth2client.service_account import ServiceAccountCredentials
from openai import OpenAI
import random
import xlsxwriter
from .apps import loginss, logintmdt, loginsp
from .functions import *
from PIL import Image
from urllib.parse import unquote
from collections import defaultdict
from lxml import html, etree
from .apps import loginss, logintmdt, loginsp
import json
import datetime
from calendar import monthrange
import calendar
import urllib.request
import itertools
import hashlib
import base64
import shutil
import time
import qrcode
from os import listdir
from django.db.models import Q
from os.path import isfile, join
from barcode import Code128
import math
import urllib.parse
from itertools import combinations
from django.core import serializers
from django.http import JsonResponse
import threading
from barcode.writer import ImageWriter
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from seleniumwire.utils import decode
from selenium.webdriver.support import expected_conditions as EC
start_time = time.time()
import xlrd
from django.shortcuts import render, redirect
from django.views.generic import ListView

MAIN_URL = "https://sisapsan.mysapogo.com/admin"
SERVER_ID = os.environ.get('SERVER_ID', None)

existing_shop_map = {
    "giadungplus_official": 10925,  #GIADUNGPLUS_OFFICIAL
    "lteng": 155174,  #LTENG
    "phaledo": 134366, #PHALEDO_SHOPEE_MALL
    "phaledo_store": 155687,
    "lteng_hcm": 155938,
}
connectID = ",".join(str(v) for v in existing_shop_map.values())

HOLIDAYS = {
    # '2025-09-02',  # ví dụ: định dạng 'YYYY-MM-DD'
    # '2025-10-01',
    '2025-09-01','2025-09-02','2025-9-9'
}

CUSTOMER_ID_JSON = {
    "MODELNHAPKHAU":[759930912],
    "NOTEKHO":[762249469],
    "VARIANTS":[759999534,792508285,792508285],
    "PRODUCTS": [760093681]

}


@csrf_exempt
def index(request):
    basename = "index"

    if request.GET['action'] == 'export-chat-history':
        LIST_CHAT = []
        for i in range(100):
            print(i)
            try:
                URL_2 = f"https://market-place.sapoapps.vn/search/conversation/filter?sortType=desc&connectionIds={connectID}&queryType=account&query=&page={int(i+1)}&limit=50"
                LIST_CONVER = js_get_url(URL_2)["conversations"]

                for conver in LIST_CONVER:
                   
                        URL_1 = f"https://market-place.sapoapps.vn/search/message/filter?&page=0&limit=20&conversationId={conver['id']}"
                        MES_ALL = js_get_url(URL_1)["messages"]
                        
                        count = 0
                        item_mes = {
                            'conversation_id': conver['id'],
                            'customer_name': conver['customer_name'],
                            'chat_log': []
                        }
                        for MES in MES_ALL:
                            if MES["send_from_customer"] == True:
                                count += 1
                                item_mes['chat_log'].append(MES["content"])

                        if count != 0:
                            LIST_CHAT.append(item_mes)
            except Exception as e:
                print(f"Error print: {e}")

        writejsonfile(LIST_CHAT,'logs/temp_mes.json')

    return render(request, 'index.html')

def realtime(request):
    options = webdriver.ChromeOptions()
    options.add_argument("user-data-dir=C://Users//Vuong//AppData//Local//Google//Chrome//User Data")
    options.add_argument('--profile-directory=Profile 1')
    options.add_argument('--allow-running-insecure-content')
    options.add_argument('--ignore-certificate-errors')
    xweb = webdriver.Chrome(executable_path='chromedriver.exe',options=options)
    xweb.get("https://banhang.shopee.vn/")
    time.sleep(10)
    for request in xweb.requests:
        if "banhang.shopee.vn/api/selleraccount/subaccount/get" in request.url:
            print("Update cookies!")
            header_dict = dict(request.headers)
            with open('logs/current-cookie.txt', 'w') as f:
                json.dump(header_dict, f)

    with open('logs/current-cookie.txt', 'r') as f:
        headers = json.load(f)
    loginsp.headers.update(headers)

    return render(request, 'realtime.html',{"message": "Send ZNS!"})


# PHẦN QUẢN TRỊ & NHẬP HÀNG

def thongtinthanhtoan(request):
    obj = {}

    # ---- helpers ----
    def parse_amount(order):
        if not order:
            return None
        cand = (
            order.get('total')
            or order.get('total_price')
            or order.get('current_total_price')
            or (order.get('total_price_set', {}).get('shop_money', {}).get('amount')
                if isinstance(order.get('total_price_set'), dict) else None)
            or order.get('grand_total')
            or order.get('total_with_discount')
            or order.get('subtotal_price')
        )
        if cand is None:
            return None
        try:
            return int(round(float(cand)))
        except Exception:
            return None

    def parse_sotien_param(val):
        s = str(val).strip()
        if not s:
            return None
        s = s.replace(',', '')
        try:
            return int(round(float(s)))
        except Exception:
            return None

    def parse_noidung_param(val):
        s = str(val).strip()
        return s if s else None

    # ---- lấy thông tin đơn (nếu có) ----
    order_amount = None
    order_code = None
    if 'donhang' in request.GET:
        code = request.GET['donhang']
        data = js_get_url(f"{MAIN_URL}/orders.json?query={code}")
        orders = data.get('orders', []) if isinstance(data, dict) else []
        order = orders[0] if orders else None
        if order:
            obj['order'] = order
            order_code = order.get('code', code)
            order_amount = parse_amount(order)

    # ---- ưu tiên sotien từ query, rồi tới từ đơn ----
    if 'sotien' in request.GET:
        amt_from_query = parse_sotien_param(request.GET['sotien'])
        if amt_from_query is not None:
            obj['sotien'] = amt_from_query
        elif order_amount is not None:
            obj['sotien'] = order_amount
    else:
        if order_amount is not None:
            obj['sotien'] = order_amount

    # ---- ưu tiên noidung từ query, rồi tới theo mã đơn, cuối cùng mặc định ----
    nd_from_query = parse_noidung_param(request.GET.get('noidung')) if 'noidung' in request.GET else None
    if nd_from_query is not None:
        obj['noidung'] = nd_from_query
    elif order_code:
        obj['noidung'] = f"P4Y {order_code}"
    else:
        obj['noidung'] = ""

    # ==== hằng số ngân hàng ====
    obj.update({
        "acc_mb": "0911169496888",
        "acc_tcb": "10000219923",
        "name_mb": "GIA DUNG PLUS",
        "name_tcb": "NGUYEN BICH TRAM",
        "bank_mb": "mbbank",          # hoặc '970422'
        "bank_tcb": "tpbank",    # hoặc '970407'
    })

    return render(request, 'qr-thanhtoan.html', obj)




def xnk_variant(request):
    excel_file_path = 'logs/log excel/ALL-VARIANTS.xls'
    def to_float_safe(val):
        try:
            if val is None:
                return 0.0
            s = str(val).strip()
            return float(s) if s else 0.0
        except ValueError:
            return 0.0

    # STEP 1 - Xử lý reload từ Excel
    if request.GET.get('action') == 'reload_from_file':
        wb = xlrd.open_workbook(excel_file_path)
        sheet = wb.sheet_by_index(0)

        required_headers = [
            "sku", "name_vn","nsx", "price_tq", "name_tq", "sku_tq",
            "sp_rong", "sp_dai", "sp_cao", "fullbox",
            "box_dai", "box_rong", "box_cao", "sku_nhapkhau", "notes","update"
        ]

        # Gộp tất cả notes active từ 2 nguồn
        all_notes = get_json_note(CUSTOMER_ID_JSON["VARIANTS"])
        headers = sheet.row_values(0)
        objects_list = []
        for i in range(1, sheet.nrows):
            row_values = sheet.row_values(i)
            obj = {header: value for header, value in zip(headers, row_values)}
            for key in required_headers:
                if key not in obj:
                    obj[key] = ""

            objects_list.append(obj)

        print(f"[+] Read done from excel:  {len(objects_list)} line!")
        for objx in objects_list:
            if objx["update"] == "" or objx["update"] == None:
                objx["update"] = 0

            if int(objx["update"]) == 1:
                vari_id = int(objx.get("vari_id", ""))
                print(f"- Process:  {vari_id}!")

                objx['vari_id'] = int(vari_id)
                objx['sp_dai'] = to_float_safe(objx['sp_dai'])
                objx['sp_rong'] = to_float_safe(objx['sp_rong'])
                objx['sp_cao'] = to_float_safe(objx['sp_cao'])
                objx['fullbox'] = int(objx['fullbox'])
                objx['box_dai'] = to_float_safe(objx['box_dai'])
                objx['box_rong'] = to_float_safe(objx['box_rong'])
                objx['box_cao'] = to_float_safe(objx['box_cao'])
                objx['sku_nhapkhau'] = str(objx['sku_nhapkhau']).strip().replace(" ","")

                matched_note = None
                for note in all_notes:
                    if str(note['vari_id']) == str(vari_id):
                        matched_note = note
                        break
                    if str(note['sku']) == str(objx["sku"]):
                        matched_note = note
                        break

                if matched_note:
                    print("Update!")
                    spapi_edit_note(matched_note['ticket_id'], matched_note["note_id"], objx)
                else:
                    print("New!")
                    spapi_new_note("792508285", objx)
            
    # STEP 2 - Export ngược ra Excel từ 2 nơi
    if request.GET.get('action') == 'export_to_excel':
        all_vari = get_all_variants("&product_types=normal&packsize=false&status=active")
        all_data = get_json_variants(CUSTOMER_ID_JSON["VARIANTS"])
        all_brand = get_list_brand("json")
        data_list = []
        if not all_data:
            return JsonResponse({"status": "error", "msg": "Không có dữ liệu hợp lệ"}, status=404)

        for vari in all_vari:
            #Lấy thêm thông tin variant:
            try:
                data = all_data[str(vari['id'])]
                data["nsx"] = all_brand[str(vari["brand_id"])]['name']
                data["sku"] = vari["sku"]
                data["vari_id"] = vari["id"]
                data["vn_name"] = vari["opt1"]
                
            except:
                data = {
                    "sku":vari["sku"], "name_vn":vari["opt1"],"nsx":all_brand[str(vari["brand_id"])]['name'], "price_tq":0, "name_tq":"", "sku_tq":"",
                    "sp_rong":0, "sp_dai":0, "sp_cao":0, "fullbox":0,
                    "box_dai":0, "box_rong":0, "box_cao":0, "sku_nhapkhau":0, "notes":0,"vari_id":vari["id"], "ticket_id":792508285
                }
            
            data_list.append(data)

        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        headers = list(data_list[0].keys())

        for col, header in enumerate(headers):
            ws.write(0, col, header)
        for row_idx, row_data in enumerate(data_list, start=1):
            for col_idx, header in enumerate(headers):
                ws.write(row_idx, col_idx, row_data.get(header, ""))

        wb.save(excel_file_path)

    all_vari = get_all_variants("&product_types=normal&packsize=false&status=active")
    all_data = get_json_variants(CUSTOMER_ID_JSON["VARIANTS"])
    all_brand = get_list_brand("json")

    for vari in all_vari:
        try:
            vari["data"] = all_data[str(vari['id'])]
            vari["nsx"] = all_brand[str(vari["brand_id"])]['name']
        except:
            vari["nsx"] = all_brand[str(vari["brand_id"])]['name']
            continue

    data_list_sorted = sorted(
        all_vari,
        key=lambda x: (x.get("nsx", ""), x.get("sku", ""))
    )

    return render(request, 'xnk_listproduct.html', {'data_list':data_list_sorted})

@csrf_exempt
def api_xnk_editvariant(request):

    if request.method != "POST":
        return JsonResponse({"status": "error", "msg": "Phải dùng phương thức POST"}, status=405)
    try:
        updates = json.loads(request.body)
    except Exception as e:
        return JsonResponse({"status": "error", "msg": f"Dữ liệu JSON không hợp lệ: {e}"}, status=400)

    result = spapi_edit_note(updates['ticket_id'], updates['note_id'], updates)
    if result.get("status") == "success":
        return JsonResponse({"status": "success","data": updates}, status=200)
    else:
        return JsonResponse({"status": "error", "msg": result.get("msg")}, status=500)


def kho_reportday(request):

    return render(request, 'kho_reportday.html')


@csrf_exempt
def api_kho_sumreport(request):
    # --- HỆ SỐ DANH MỤC THEO ĐỘ KHÓ ĐÓNG GÓI ---
    CATEGORY_MULTIPLIER = {
        "Thuỷ tinh": 1.3,
        "Tấm chắn dầu": 0.5,
        # Các nhóm còn lại mặc định = 1.0
    }

    def _cat_factor(cat_name: str) -> float:
        return CATEGORY_MULTIPLIER.get((cat_name or "").strip(), 1.0)

    def _adjusted_revenue_for_scope():
        """
        Tính doanh số đã hiệu chỉnh theo kho_filter (nếu có).
        - Nếu có kho_filter: dùng per_kho_categories[kho] + phần unknown mỗi kho.
        - Nếu không có kho_filter: dùng sum_data['categories'] + phần unknown chung.
        """
        if kho_filter:
            # Theo kho được chọn
            adjusted_sum = 0.0
            for scope in ["gele", "toky"]:
                kho_norm = scope2kho[scope]  # "HN" hoặc "HCM"
                if kho_norm not in kho_filter:
                    continue

                # Doanh số có danh mục của kho
                cat_rev_sum = 0.0
                for cat, cat_data in per_kho_categories[scope].items():
                    rev = float(cat_data.get("doanhso", 0))
                    adjusted_sum += rev * _cat_factor(cat)
                    cat_rev_sum += rev

                # Phần doanh số không gắn danh mục của kho (giữ nguyên hệ số 1.0)
                kho_total_rev = float(sum_data[scope].get("doanhso", 0))
                unknown_rev = max(0.0, kho_total_rev - cat_rev_sum)
                adjusted_sum += unknown_rev  # ×1.0
            return adjusted_sum
        else:
            # Tổng hai kho
            adjusted_sum = 0.0
            cat_rev_sum = 0.0
            for cat, cat_data in sum_data["categories"].items():
                rev = float(cat_data.get("doanhso", 0))
                adjusted_sum += rev * _cat_factor(cat)
                cat_rev_sum += rev

            total_rev_all = float(sum_data["sum"].get("doanhso", 0))
            unknown_rev = max(0.0, total_rev_all - cat_rev_sum)
            adjusted_sum += unknown_rev  # ×1.0
            return adjusted_sum

    def is_non_workday(dt: datetime.datetime):
        return dt.weekday() == 6 or dt.date().isoformat() in HOLIDAYS  # Chủ nhật = 6

    def next_business_day(dt: datetime.datetime):
        candidate = dt
        while is_non_workday(candidate):
            candidate += datetime.timedelta(days=1)
        return candidate

    def get_fast_deadline(ct: datetime.datetime):
        if 0 <= ct.hour < 18:
            # Trước 18h -> deadline 23:59:59 cùng ngày
            raw = ct.replace(hour=23, minute=59, second=59, microsecond=0)
        else:
            # Sau 18h -> deadline 12:00 trưa ngày kế tiếp
            raw = (ct + datetime.timedelta(days=1)).replace(hour=12, minute=0, second=0, microsecond=0)

        # Nếu deadline rơi vào Chủ nhật thì ép sang thứ 2 lúc 12:00
        if raw.weekday() == 6:  # Chủ nhật
            raw = (raw + datetime.timedelta(days=1)).replace(hour=12, minute=0, second=0, microsecond=0)

        return next_business_day(raw)


    def ensure_carrier_bucket(store, name):
        return store.setdefault(name, {
            'total': 0,
            'packed': 0,
            'handed_over': 0  # fulfilled
        })

    # cập nhật thống kê fast vs not fast (chỉ với đơn đã xuất kho)
    def add_fast_stats(scope, fast_flag, amount):
        bucket = fast_delivery[scope]
        if fast_flag:
            bucket['fast'] += 1
            bucket['fast_revenue'] += amount
        else:
            bucket['not_fast'] += 1
            bucket['not_fast_revenue'] += amount

    def _parse_ddmmyyyy(s):
        # all_chamcong dùng 'DD/MM/YYYY'
        return datetime.datetime.strptime(s, "%d/%m/%Y").date()

    def _norm_kho_label(lbl):
        u = (lbl or "").strip().upper()
        if u in ("SG", "SÀI GÒN", "SAIGON", "HO CHI MINH", "HCM"):
            return "HCM"
        if u in ("HN", "HÀ NỘI", "HA NOI"):
            return "HN"
        return u  # fallback
    # Kho theo scope
    scope2kho = {"gele": "HN", "toky": "HCM"}
    locid2kho = {241737: "HN", 548744: "HCM"}
    kho_filter = None


    all_chamcong = read_nhanvien_bangchamcong(1)

    start_date_str = request.GET.get("start_date")  # dạng '2025-07-01'
    end_date_str = request.GET.get("end_date")      # dạng '2025-07-31'
    location_ids_raw = request.GET.get("location_ids")  # dạng '241737,548744'


    start_dt = datetime.datetime.strptime(start_date_str, "%Y-%m-%d") - datetime.timedelta(hours=7)
    end_dt = datetime.datetime.strptime(end_date_str, "%Y-%m-%d") + datetime.timedelta(hours=16, minutes=59, seconds=59)
    window_start = start_dt - datetime.timedelta(days=3)
    window_end = end_dt + datetime.timedelta(days=1)

    created_on_min = window_start.strftime("%Y-%m-%dT%H:%M:%SZ")
    created_on_max = window_end.strftime("%Y-%m-%dT%H:%M:%SZ")

    # Chuẩn bị khoảng gói hàng cần lọc (local +7)
    packing_start_local = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
    packing_end_local = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    
    # --- 2. Prepare aggregation containers ---
    # Query params
    location_ids = [int(x) for x in location_ids_raw.split(",") if x.strip()]
    hourly_totals = defaultdict(lambda: defaultdict(int))
    sum_data = {
        'gele': {
            'total_time': 0.0, 'total_orders': 0,
            'total_time_xk': 0.0, 'total_orders_xk': 0,
            'sodonhang': 0, 'doanhso': 0, 'sosanpham': 0, 'products': {}, 'tile_sodonhang':0
        },
        'toky': {
            'total_time': 0.0, 'total_orders': 0,
            'total_time_xk': 0.0, 'total_orders_xk': 0,
            'sodonhang': 0, 'doanhso': 0, 'sosanpham': 0, 'products': {},'tile_sodonhang':0
        },
        'sum': {'sodonhang': 0, 'doanhso': 0, 'sosanpham': 0},
        'pending_deadlines' :{
                'before_today': {'sodonhang': 0, 'doanhso': 0},
                'before_18': {'sodonhang': 0, 'doanhso': 0},
                'after_18': {'sodonhang': 0, 'doanhso': 0}
            },
        'categories': {

        }
    }

    # Thống kê giao hàng nhanh
    fast_delivery = {
        'gele': {'fast': 0, 'not_fast': 0, 'fast_revenue': 0, 'not_fast_revenue': 0},
        'toky': {'fast': 0, 'not_fast': 0, 'fast_revenue': 0, 'not_fast_revenue': 0},
        'sum': {'fast': 0, 'not_fast': 0, 'fast_revenue': 0, 'not_fast_revenue': 0}
    }
    # trạng thái giao hàng theo deadline
    deadline_status = {
        'gele': {'not_shipped_on_time': 0, 'not_shipped_late': 0, 'shipped_on_time': 0, 'shipped_late': 0},
        'toky': {'not_shipped_on_time': 0, 'not_shipped_late': 0, 'shipped_on_time': 0, 'shipped_late': 0},
        'sum': {'not_shipped_on_time': 0, 'not_shipped_late': 0, 'shipped_on_time': 0, 'shipped_late': 0}
    }

    per_user = {'gele': {}, 'toky': {}}

    per_kho_categories = {'gele': {}, 'toky': {}}
    # breakdown theo đơn vị vận chuyển
    carrier_summary = {
        'gele': {},
        'toky': {},
        'sum': {}  # tổng cộng tất cả
    }
    # Load category mapping once (replicated from original)
    varis = {}


    for i in range(1, 4):
        resp = js_get_url(f"{MAIN_URL}/products.json?page={i}&limit=250")
        for vari in resp.get("products", []):
            varis[vari['id']] = vari.get('category')
    # --- 3. Pagination over orders ---
    for page in range(1, 100):
        params = {
            "created_on_min": created_on_min,
            "created_on_max": created_on_max,
            "status": "draft%2Cfinalized%2Ccompleted",
            "limit": 250,
            "page": page
        }
        orders_resp = js_get_url(
            f"{MAIN_URL}/orders.json?created_on_min={params['created_on_min']}"
            f"&created_on_max={params['created_on_max']}"
            f"&status={params['status']}&limit={params['limit']}&page={params['page']}"
        )

        orders = orders_resp.get('orders', [])
        if not orders:
            break

        for order in orders:
            # Filter by location_ids
            if order.get('location_id') not in location_ids:
                continue
            kho = 'gele' if order.get('location_id') == 241737 else 'toky'
            order = get_data_packing(order)
            if 'nguoi_goi' not in order:
                order['nguoi_goi'] = "NO-SCAN"
            if 'packing_status' not in order:
                order['packing_status'] = 0

            created_time = datetime.datetime.strptime(order['created_on'], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
            modified_on = datetime.datetime.strptime(order['modified_on'], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)

            if order.get('fulfillment_status') == "shipped":
                order['packing_status'] = 4

            # Parse time_packing
            if 'time_packing' in order:
                try:
                    if len(order['time_packing']) <= 16:
                        time_packing = datetime.datetime.strptime(order['time_packing'], "%H:%M %d-%m-%Y")
                    else:
                        time_packing = datetime.datetime.strptime(order['time_packing'], "%H:%M:%S %d-%m-%Y")
                except Exception:
                    time_packing = modified_on
            else:
                time_packing = modified_on

            # XỬ LÝ ĐƠN CHƯA GÓI (packing_status < 4) và CHƯA SHIPPED
            if order.get('packing_status', 0) < 4:
                # Xác định category deadline dựa trên created_time (đã +7)
                now_local = datetime.datetime.now() + datetime.timedelta(hours=7)  # nếu cần thời gian hiện tại
                today_local = datetime.datetime(now_local.year, now_local.month, now_local.day)
                cutoff_18 = today_local + datetime.timedelta(hours=18)

                if created_time.date() < today_local.date():
                    category = 'before_today'
                elif today_local <= created_time <= cutoff_18:
                    category = 'before_18'
                else:
                    category = 'after_18'

                amount = int(order.get('total', 0))
                sum_data['pending_deadlines'][category]['sodonhang'] += 1
                sum_data['pending_deadlines'][category]['doanhso'] += amount

            deadline = get_fast_deadline(created_time)
            if order.get('fulfillment_status') == "shipped" and order['fulfillments']:
                shipped_on = order['fulfillments'][-1].get('shipped_on')
                if shipped_on is None:
                    shipped_on = order['created_on']

                time_xuatkho = datetime.datetime.strptime(shipped_on, "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
                if shipped_on:
                    if time_xuatkho <= deadline:
                        status_key = 'shipped_on_time'
                    else:
                        status_key = 'shipped_late'
            else:
                # Chưa giao
                now_local = datetime.datetime.now() + datetime.timedelta(hours=7)
                if now_local > deadline:
                    status_key = 'not_shipped_late'
                else:
                    status_key = 'not_shipped_on_time'
            
            # Chỉ xét những đơn đã gói và time_packing nằm trong khoảng yêu cầu
            if order.get('packing_status', 0) < 4:
                continue
            if not (packing_start_local <= time_packing <= packing_end_local):
                continue

            # Tăng đếm theo kho và tổng
            deadline_status[kho][status_key] += 1
            deadline_status['sum'][status_key] += 1

            raw_carrier = order.get('dvvc', "SPX Express") or "SPX Express"
            hot_list = ["beDelivery", "GrabExpress", "Siêu Tốc - 4 Giờ", "AhaMove", "SPX Instant"]
            congkenh = ["GHN - Hàng Cồng Kềnh", "Hàng Cồng Kềnh", "NJV - Hàng Cồng Kềnh"]
            spx = ["Nhanh", "SPX Express"]
            if raw_carrier in hot_list:
                carrier = "Hoả Tốc"
            elif raw_carrier in congkenh:
                carrier = "Cồng Kềnh"
            elif raw_carrier in spx:
                carrier = "SPX Express"
            else:
                carrier = raw_carrier

            # Kho riêng
            c_bucket = ensure_carrier_bucket(carrier_summary[kho], carrier)
            c_bucket['total'] += 1
            if order.get('packing_status', 0) >= 4:
                c_bucket['packed'] += 1
            if order.get('fulfillment_status') == "shipped":
                c_bucket['handed_over'] += 1

            # Tổng chung
            c_bucket_sum = ensure_carrier_bucket(carrier_summary['sum'], carrier)
            c_bucket_sum['total'] += 1
            if order.get('packing_status', 0) >= 4:
                c_bucket_sum['packed'] += 1
            if order.get('fulfillment_status') == "shipped":
                c_bucket_sum['handed_over'] += 1

            # Skip invalid shipment structure
            if not order.get('fulfillments') or order['fulfillments'][-1].get('shipment') is None:
                continue

            # Cập nhật hourly
            hour = time_packing.hour
            hourly_totals[f"count_{kho}"][hour] += 1

            # Thời gian gói
            time_difference = (time_packing - created_time).total_seconds()
            sum_data[kho]['total_time'] += time_difference / (24 * 3600)
            sum_data[kho]['total_orders'] += 1
            
            # Breakdown theo người gói
            user_bucket = per_user[kho].setdefault(order['nguoi_goi'], {
                'cato': {},          # đếm số lượng theo danh mục (đã có)
                'cato_rev': {},      # NEW: doanh số theo danh mục
                'total_money': 0,
                'total_quantity': 0,
                'total_order': 0
            })

           
            # Tổng hợp sản phẩm
            order_total_qty = 0
            for line in order.get('order_line_items', []):
                product_id = line.get("product_id")
                qty = int(line.get('quantity', 0))
                amount = int(line.get('line_amount', 0))
                cat = varis.get(product_id) if product_id in varis else None
                sku = line.get('sku')
                if sku not in sum_data[kho]['products']:
                    sum_data[kho]['products'][sku] = {
                        'total': line.get('line_amount', 0),
                        'count': line.get('quantity', 0)
                    }
                else:
                    sum_data[kho]['products'][sku]['total'] += line.get('line_amount', 0)
                    sum_data[kho]['products'][sku]['count'] += line.get('quantity', 0)
                order_total_qty += line.get('quantity', 0)

                if product_id in varis:
                    cat = varis[product_id]
                    if cat:
                        user_bucket['cato'][cat] = user_bucket['cato'].get(cat, 0) + int(line.get('quantity', 0))
                    # NEW: cộng doanh số theo danh mục cho nhân sự
                    user_bucket['cato_rev'][cat] = user_bucket['cato_rev'].get(cat, 0) + int(line.get('line_amount', 0))

                if cat:
                    # Tổng hợp chung (cả hai kho)
                    cat_bucket = sum_data['categories'].setdefault(cat, {
                        'sosanpham': 0,
                        'doanhso': 0,
                        'order_count': 0  # Nếu muốn đếm số order liên quan
                    })
                    cat_bucket['sosanpham'] += qty
                    cat_bucket['doanhso'] += amount
                    # Để tránh đếm một order nhiều lần vào order_count, bạn có thể thêm logic dedupe nếu cần

                    # Nếu muốn breakdown theo kho riêng
                    kho_cat_bucket = per_kho_categories[kho].setdefault(cat, {
                        'sosanpham': 0,
                        'doanhso': 0,
                        'order_count': 0
                    })
                    kho_cat_bucket['sosanpham'] += qty
                    kho_cat_bucket['doanhso'] += amount

            order['total_quantity'] = order_total_qty
            user_bucket['total_money'] += int(order.get('total', 0))
            user_bucket['total_quantity'] += order.get('total_quantity', 0)
            user_bucket['total_order'] += 1
            
            # Xuất kho
            if order['fulfillments'][-1].get('composite_fulfillment_status') == "fulfilled":
                shipped_on = order['fulfillments'][-1].get('shipped_on')
                if shipped_on:
                    time_xuatkho = datetime.datetime.strptime(shipped_on, "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
                    time_difference_2 = (time_xuatkho - created_time).total_seconds()
                    sum_data[kho]['total_time_xk'] += time_difference_2 / (24 * 3600)
                    sum_data[kho]['total_orders_xk'] += 1

                    deadline = get_fast_deadline(created_time)

                    is_fast = time_xuatkho <= deadline

                    amount = int(order.get('total', 0))
                    add_fast_stats(kho, is_fast, amount)
                    add_fast_stats('sum', is_fast, amount)

    # --- 4. Summarize totals ---
    start_date_local = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_date_local = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    num_days = (end_date_local - start_date_local).days + 1
    if num_days <= 0:
        num_days = 1

    # Gộp người gói duy nhất từ cả 2 kho
    all_packers = set(list(per_user['gele'].keys()) + list(per_user['toky'].keys()))
    num_packers = len(all_packers)

    # Tổng đơn hàng từ cả 2 kho
    total_orders_all = 0
    for kho in ['gele', 'toky']:
        for nguoigoi, data in per_user[kho].items():
            sum_data['sum']['sodonhang'] += data['total_order']
            sum_data['sum']['doanhso'] += data['total_money']
            sum_data['sum']['sosanpham'] += data['total_quantity']

            sum_data[kho]['sodonhang'] += data['total_order']
            sum_data[kho]['doanhso'] += data['total_money']
            sum_data[kho]['sosanpham'] += data['total_quantity']

        for p in per_user[kho]:
            total_orders_all += per_user[kho][p]['total_order']
    
    sum_data['sum']['pcs_per_orders'] = round(float(sum_data['sum']['sosanpham']/sum_data['sum']['sodonhang']),2)

    avg_orders_per_packer_per_day = 0.0
    if num_packers > 0 and num_days > 0:
        avg_orders_per_packer_per_day = total_orders_all / num_packers / num_days
        avg_rev_per_packer_per_day = int(data['total_money']/ num_packers / num_days)

    efficiency = {
        'sum': {
            'packers_count': num_packers,
            'days': num_days,
            'total_orders': total_orders_all,
            'avg_orders': round(avg_orders_per_packer_per_day, 2),
            'avg_rev': avg_rev_per_packer_per_day
        }
    }
    # --- 5. Tổng hợp thêm: pending_deadlines.sum và overall packed vs unpacked ---
    # total pending (chưa gói) bằng tổng 3 category
    pending = sum_data['pending_deadlines']
    pending_sum_orders = (
        pending['before_today']['sodonhang']
        + pending['before_18']['sodonhang']
        + pending['after_18']['sodonhang']
    )
    pending_sum_revenue = (
        pending['before_today']['doanhso']
        + pending['before_18']['doanhso']
        + pending['after_18']['doanhso']
    )
    # gắn vào cấu trúc
    sum_data['pending_deadlines']['sum'] = {
        'sodonhang': pending_sum_orders,
        'doanhso': pending_sum_revenue
    }

    # packed / unpacked
    packed_orders = sum_data['sum']['sodonhang']
    unpacked_orders = sum_data['pending_deadlines']['sum']['sodonhang']
    packed_revenue = sum_data['sum']['doanhso']
    unpacked_revenue = sum_data['pending_deadlines']['sum']['doanhso']

    total_orders = packed_orders + unpacked_orders if (packed_orders + unpacked_orders) > 0 else 1
    packed_pct = round(packed_orders / total_orders * 100, 2)
    unpacked_pct = round(unpacked_orders / total_orders * 100, 2)

    pie_data = {
        'packed': packed_orders,
        'unpacked': unpacked_orders,
        'packed_pct': packed_pct,
        'unpacked_pct': unpacked_pct,
        'packed_revenue': packed_revenue,
        'unpacked_revenue': unpacked_revenue,
    }

    # --- tính % cho mỗi category ---
    total_doanhso_categories = sum(cat_data['doanhso'] for cat_data in sum_data['categories'].values()) or 1
    total_sosanpham_categories = sum(cat_data['sosanpham'] for cat_data in sum_data['categories'].values()) or 1

    # --- tính % cho mỗi kho ---
    if sum_data['gele']['sodonhang'] == 0:
        sum_data['gele']['sodonhang'] = 1
    if sum_data['toky']['sodonhang'] == 0:
        sum_data['toky']['sodonhang'] = 1

    sum_data['gele']['tile_sodonhang'] = int(float(sum_data['gele']['sodonhang']*100 / (sum_data['gele']['sodonhang']+sum_data['toky']['sodonhang'])))
    sum_data['toky']['tile_sodonhang'] = int(float(sum_data['toky']['sodonhang']*100 / (sum_data['gele']['sodonhang']+sum_data['toky']['sodonhang'])))
    
    # thêm phần % vào mỗi category
    for cat, cat_data in sum_data['categories'].items():
        cat_data['pct_doanhso'] = round(cat_data['doanhso'] / total_doanhso_categories * 100, 2)
        cat_data['pct_sosanpham'] = round(cat_data['sosanpham'] / total_sosanpham_categories * 100, 2)

    pie_cato = {
        'labels': [],
        'values': [],
        'percentages': []
    }
    for cat, cat_data in sorted(sum_data['categories'].items(), key=lambda x: x[1]['doanhso'], reverse=True):
        pie_cato['labels'].append(cat)
        pie_cato['values'].append(cat_data['doanhso'])
        pie_cato['percentages'].append(cat_data['pct_doanhso'])

    # Sort senders by revenue desc
    sorted_users = {
        'gele': dict(sorted(per_user['gele'].items(), key=lambda x: x[1]['total_money'], reverse=True)),
        'toky': dict(sorted(per_user['toky'].items(), key=lambda x: x[1]['total_money'], reverse=True))
    }

    # --- tính tỷ lệ % theo đơn vị vận chuyển trên tổng đơn ---
    total_all_orders_for_pct = total_orders_all if total_orders_all > 0 else 1
    for scope in ['gele', 'toky', 'sum']:
        for carrier, stats in carrier_summary[scope].items():
            stats['rate'] = round(stats['total'] / total_all_orders_for_pct * 100, 2)

    # tỷ lệ giao nhanh trên tổng đã xuất kho (chỉ tính fulfilled vì mới có shipped_on)
    def compute_fast_rate(scope):
        bucket = fast_delivery[scope]
        total_shipped = bucket['fast'] + bucket['not_fast'] if (bucket['fast'] + bucket['not_fast']) > 0 else 1
        bucket['fast_pct'] = round(bucket['fast'] / total_shipped * 100, 2)
        bucket['not_fast_pct'] = round(bucket['not_fast'] / total_shipped * 100, 2)

    compute_fast_rate('gele')
    compute_fast_rate('toky')
    compute_fast_rate('sum')

    # --- inventory onhand tổng / riêng từng kho ---
    def fetch_onhand(location_id):
        resp = js_get_url(
            f"{MAIN_URL}/reports/inventories/onhand.json?page=1&limit=1&location_ids={location_id}"
        )
        return resp.get("summary", {})

    gele_inv = fetch_onhand(241737)
    toky_inv = fetch_onhand(548744)

    gele_onhand = int(gele_inv.get("total_local_onhand", 0))
    gele_amount = int(gele_inv.get("total_local_amount", 0))
    toky_onhand = int(toky_inv.get("total_local_onhand", 0))
    toky_amount = int(toky_inv.get("total_local_amount", 0))

    sum_onhand = gele_onhand + toky_onhand
    sum_amount = (gele_amount or 0) + (toky_amount or 0)

    inventory_summary = {
        "gele": {"onhand": gele_onhand, "amount": gele_amount},
        "toky": {"onhand": toky_onhand, "amount": toky_amount},
        "sum": {"onhand": sum_onhand, "amount": sum_amount},
    }

    def compute_deadline_percentages(scope_counts):
        total = (
            scope_counts.get('not_shipped_late', 0)
            + scope_counts.get('shipped_on_time', 0)
            + scope_counts.get('shipped_late', 0)
        ) or 1  # tránh chia 0

        return {
            'not_shipped_late_pct': round(scope_counts.get('not_shipped_late', 0) / total * 100, 2),
            'shipped_on_time_pct': round(scope_counts.get('shipped_on_time', 0) / total * 100, 2),
            'shipped_late_pct': round(scope_counts.get('shipped_late', 0) / total * 100, 2),
        }

    deadline_summary = {
        'gele': {
            **deadline_status['gele'],
            **compute_deadline_percentages(deadline_status['gele'])
        },
        'toky': {
            **deadline_status['toky'],
            **compute_deadline_percentages(deadline_status['toky'])
        },
        'sum': {
            **deadline_status['sum'],
            **compute_deadline_percentages(deadline_status['sum'])
        }
    }

    # Gom giờ công theo (kho_norm, name)
    hours_by_staff = defaultdict(float)
    for row in all_chamcong:
        try:
            d = _parse_ddmmyyyy(row["date"])
        except Exception:
            continue
        if not (start_date_local <= d <= end_date_local):
            continue
        kho_norm = _norm_kho_label(row.get("kho"))
        name = (row.get("name") or "").strip()
        cong = float(row.get("cong") or 0.0)
        hours_by_staff[(kho_norm, name)] += cong

    # 2) Áp vào by_sender: key dạng "KHO: Tên", ví dụ "HN: Giang", "SG: Quý"
    def _split_sender_key(sender_key):
        # "HN: Giang" -> ("HN","Giang"); nếu không đúng format thì trả về (None, key)
        if ":" in sender_key:
            left, right = sender_key.split(":", 1)
            return _norm_kho_label(left), right.strip()
        return None, sender_key.strip()

    # Duyệt 2 scope, enrich từng nhân sự
    for scope in ["gele", "toky"]:
        enriched = {}
        for sender_key, data in sorted_users[scope].items():
            kho_label_in_key, staff_name = _split_sender_key(sender_key)
            kho_norm = kho_label_in_key or scope2kho.get(scope)

            # chỉ lấy giờ công đúng kho
            work_hours = 0.0
            if kho_norm == scope2kho[scope]:
                work_hours = float(hours_by_staff.get((kho_norm, staff_name), 0.0))

            total_money = float(data.get("total_money", 0.0))
            total_order = float(data.get("total_order", 0.0))

            # --- NEW: tính doanh số đã hiệu chỉnh theo danh mục cho từng nhân sự ---
            cato_rev = (data.get("cato_rev") or {})  # {'Thuỷ tinh': doanh_so, ...}
            cat_rev_sum = 0.0
            adjusted_user_rev = 0.0
            for cat, rev in cato_rev.items():
                rev = float(rev or 0.0)
                cat_rev_sum += rev
                adjusted_user_rev += rev * _cat_factor(cat)

            # phần doanh số không gắn danh mục (unknown) giữ nguyên hệ số 1.0
            unknown_rev_user = max(0.0, total_money - cat_rev_sum)
            adjusted_user_rev += unknown_rev_user

            # --- money_per_hour: dùng doanh số đã hiệu chỉnh ---
            if work_hours > 0:
                money_per_hour_adj = round(adjusted_user_rev / work_hours, 2)
                money_per_hour_raw = round(total_money / work_hours, 2)
                order_per_hour    = round(total_order / work_hours, 2)
            else:
                money_per_hour_adj = 0.0
                money_per_hour_raw = 0.0
                order_per_hour     = 0.0

            enriched[sender_key] = {
                **data,
                "work_hours": round(work_hours, 2),
                # Ghi đè money_per_hour = ĐÃ HIỆU CHỈNH
                "money_per_hour": int(money_per_hour_adj),
                # Xuất thêm để đối chiếu
                "money_per_hour_raw": int(money_per_hour_raw),
                "order_per_hour": round(order_per_hour, 1),
                # (tuỳ chọn) nếu muốn xem tổng doanh số đã hiệu chỉnh của nhân sự
                "total_money_adjusted": int(adjusted_user_rev),
            }
        sorted_users[scope] = enriched


    if location_ids_raw:
        locids = [int(x) for x in location_ids_raw.split(",") if x.strip()]
        kho_filter = set(locid2kho.get(lid) for lid in locids if lid in locid2kho)

    # --- TÍNH LẠI TỔNG GIỜ LÀM DỰA TRÊN BẢNG NHÂN SỰ (by_sender) ---
    # Chỉ cộng giờ của những người thật sự có gói đơn trong kỳ,
    # và tôn trọng kho_filter (HN/HCM).
    total_work_hours = 0.0
    if kho_filter:
        # Theo kho được chọn (VD: chỉ "HN" khi chọn location_id = 241737)
        for scope in ["gele", "toky"]:
            kho_norm = scope2kho[scope]  # "HN" hoặc "HCM"
            if kho_norm not in kho_filter:
                continue
            for sender_key, data in sorted_users[scope].items():
                total_work_hours += float(data.get("work_hours", 0.0))
    else:
        # Không lọc kho -> cộng cả hai kho
        for scope in ["gele", "toky"]:
            for sender_key, data in sorted_users[scope].items():
                total_work_hours += float(data.get("work_hours", 0.0))

    if total_work_hours > 0:
        adjusted_revenue = _adjusted_revenue_for_scope()
        rev_per_hour   = round(adjusted_revenue / total_work_hours, 2)
        order_per_hour = round(sum_data['sum']['sodonhang'] / total_work_hours, 2)
    else:
        rev_per_hour = 0.0
        order_per_hour = 0.0


    sum_data['sum']['hieusuat'] = {
        'total_work_hours': round(total_work_hours, 2),
        'doanhso_per_hour': int(rev_per_hour),     # << dùng giá trị ĐÃ hiệu chỉnh
        'order_per_hour': round(order_per_hour, 2)
    }
    # Build response payload
    response = {
        "input": {
            "start_date_str": start_date_str,
            "end_date_str": end_date_str,
            "location_ids": location_ids
        },
        "summary": sum_data,
        "by_sender": sorted_users,
        "hourly": {
            "categories": [str(h) for h in range(24)],
            "data_gele": [hourly_totals['count_gele'][h] for h in range(24)],
            "data_toky": [hourly_totals['count_toky'][h] for h in range(24)],
        },
        "efficiency": efficiency,
        "pie_data": pie_data,
        "pie_cato": pie_cato,
        "carrier_summary": carrier_summary,
        "fast_delivery": fast_delivery,
        "inventory_summary": inventory_summary,
        "deadline_summary": deadline_summary

    }

    return JsonResponse(response, safe=False)


@csrf_exempt
def api_kho_newnotereport(request):
    TICKET_CUSTOMER_IDS = CUSTOMER_ID_JSON["NOTEKHO"]
    if request.method != "POST":
        return JsonResponse({"status": "error", "msg": "Phải POST"}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"status": "error", "msg": "JSON không hợp lệ"}, status=400)

    location = data.get("location", "").strip()
    tags = data.get("tags", [])
    content_text = data.get("content", "").strip()
    timestamp = data.get("timestamp", "").strip()

    if not content_text:
        return JsonResponse({"status": "error", "msg": "Thiếu nội dung"}, status=400)
    if not location:
        return JsonResponse({"status": "error", "msg": "Thiếu location"}, status=400)
    if not isinstance(tags, list) or not tags:
        return JsonResponse({"status": "error", "msg": "Thiếu tags"}, status=400)

    # Sinh một ID phản hồi (ví dụ PH-<DDMMYYYY>-XXX)
    today = datetime.datetime.now().strftime("%d%m%Y")
    suffix = str(random.randint(0, 999)).zfill(3)
    ph_id = f"PH-{today}-{suffix}"

    # Xây payload note
    note_payload = {
        "ph_id": ph_id,
        "location": location,
        "tags": tags,
        "content": content_text,
        "timestamp": timestamp or datetime.datetime.utcnow().isoformat() + "Z"
    }

    # Gửi lên Sapo tạo note mới cho customer id 762249469
    url_note = f"{MAIN_URL}/customers/{TICKET_CUSTOMER_IDS}/notes.json"
    payload = {"content": json.dumps(note_payload, ensure_ascii=False)}

    resp = loginss.post(url_note, json=payload)

    if resp.status_code in (200, 201):
        return JsonResponse({"status": "ok", "note_id": ph_id})
    else:
        return JsonResponse({
            "status": "error",
            "msg": f"Lưu note thất bại ({resp.status_code})",
            "response_text": resp.text
        }, status=500)


@csrf_exempt
def api_kho_getnotereport(request):
    # kỳ vọng GET? cũng có thể hỗ trợ POST nếu cần, ở đây dùng GET
    start_date_str = request.GET.get("start_date")  # "2025-08-01"
    end_date_str = request.GET.get("end_date")      # "2025-08-05"

    if not start_date_str or not end_date_str:
        return JsonResponse({"status": "error", "msg": "Thiếu start_date hoặc end_date"}, status=400)

    try:
        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"status": "error", "msg": "Định dạng ngày không hợp lệ, dùng YYYY-MM-DD"}, status=400)

    # 1. Lấy tất cả note của customer phản hồi
    resp = js_get_url(f"{MAIN_URL}/customers/762249469.json")
    notes = resp.get("customer", {}).get("notes", [])
    filtered = []

    # 2. Duyệt và lọc
    for note in notes:
        if note.get("status") != "active":
            continue
        try:
            content = json.loads(note.get("content", "{}"))
        except json.JSONDecodeError:
            continue

        # Lấy timestamp trong content
        ts = content.get("timestamp")
        if not ts:
            continue
        # chuẩn hóa: loại bỏ Z hoặc offset nếu cần
        ts_clean = ts
        if ts_clean.endswith("Z"):
            ts_clean = ts_clean[:-1]
        elif "+" in ts_clean:
            ts_clean = ts_clean.split("+")[0]
        try:
            dt = datetime.datetime.fromisoformat(ts_clean)
        except ValueError:
            continue
        # convert sang timezone nếu muốn; ở đây coi timestamp là UTC
        local_date = (dt).date()

        if start_date <= local_date <= end_date:
            # gắn thêm note metadata
            entry = {
                "note_id": note.get("id"),
                "modified_on": note.get("modified_on"),
                "location": content.get("location"),
                "tags": content.get("tags"),
                "content": content.get("content"),
                "timestamp": content.get("timestamp"),
                "ph_id": content.get("ph_id")
            }
            filtered.append(entry)

    # 3. Sort theo timestamp mới nhất trước
    def sort_key(e):
        t = e.get("timestamp", "")
        try:
            tc = t
            if tc.endswith("Z"):
                tc = tc[:-1]
            elif "+" in tc:
                tc = tc.split("+")[0]
            return datetime.datetime.fromisoformat(tc)
        except:
            return datetime.datetime.min

    filtered.sort(key=sort_key, reverse=True)

    return JsonResponse(filtered, safe=False)


@csrf_exempt
def api_kho_getshopeereport(request):
    shops = ['giadungplus_official', 'phaledo']
    response = {}
    for shop in shops:
        connection_id = get_connection_id(shop, existing_shop_map)
        response[shop] = {}
        try:
            if doi_shop(connection_id,loginsp) == 1:
                URL_1 = "https://banhang.shopee.vn/api/sellergrowth/v2/ps/landing_page/?SPC_CDS=fa75a215-26fc-4960-8efc-1b1d87508841&SPC_CDS_VER=2"
                res = loginsp.get(URL_1)
                res = res.json()
                URL_2 = "https://banhang.shopee.vn/api/v2/shops/sellerCenter/ongoingPoints/?SPC_CDS=fa75a215-26fc-4960-8efc-1b1d87508841&SPC_CDS_VER=2"
                res2 = loginsp.get(URL_2)
                res2 = res2.json()

                response[shop]['shop_performance'] = res["metrics"]
                response[shop]['ongoing_points'] = res2["data"]

        except Exception as e:
            print(f"Error: {e}")

    return JsonResponse(response, safe=False)


def kho_product(request):
    excel_file_path = 'logs/log excel/ALL-VARIANTS.xls'

    all_vari = get_all_variants("&product_types=normal&packsize=false&status=active")
    all_data = get_json_variants(CUSTOMER_ID_JSON["VARIANTS"])
    all_brand = get_list_brand("json")

    for vari in all_vari:
        try:
            vari["data"] = all_data[str(vari['id'])]
            vari["nsx"] = all_brand[str(vari["brand_id"])]['name']
        except:
            vari["nsx"] = all_brand[str(vari["brand_id"])]['name']
            continue

    data_list_sorted = sorted(
        all_vari,
        key=lambda x: (x.get("nsx", ""), x.get("sku", ""))
    )

    return render(request, 'kho_product.html', {'data_list':data_list_sorted})

def kho_barcode(request):
    TICKET_CUSTOMER_IDS = ['760093681']
    excel_file_path = 'logs/log excel/ALL-PRODUCTS.xls'

    # STEP 1 - Xử lý reload từ Excel
    if request.GET.get('action') == 'reload_from_file':
        wb = xlrd.open_workbook(excel_file_path)
        sheet = wb.sheet_by_index(0)

        required_headers = ["sku", "sku_first","nsx", "brand", "vi_name", "en_name","descreption", "material"]

        # Gộp tất cả notes active từ 2 nguồn
        all_notes = []
        notes_by_ticket = {}  # Dùng để biết mỗi note đến từ ticket nào
        for ticket_id in TICKET_CUSTOMER_IDS:
            url_notes = f"{MAIN_URL}/customers/{ticket_id}.json"
            res = loginss.get(url_notes)
            customer = res.json().get("customer", {})
            active_notes = [n for n in customer.get("notes", []) if n.get("status") == "active"]
            for note in active_notes:
                note["_ticket_id"] = ticket_id  # Đánh dấu thuộc ticket nào
            all_notes.extend(active_notes)

        headers = sheet.row_values(0)
        objects_list = []
        for i in range(1, sheet.nrows):
            row_values = sheet.row_values(i)
            obj = {header: value for header, value in zip(headers, row_values)}
            for key in required_headers:
                if key not in obj:
                    obj[key] = ""
            objects_list.append(obj)

        for objx in objects_list:
            sku = str(objx.get("product_id", "")).strip()
            matched_note = None
            matched_ticket_id = None
            for note in all_notes:
                try:
                    content = json.loads(note.get("content", "{}"))
                    if str(content.get("product_id", "")).strip() == sku:
                        matched_note = note
                        matched_ticket_id = note["_ticket_id"]
                        break
                except:
                    continue

            if matched_note:
                spapi_edit_note(matched_ticket_id, matched_note["id"], objx)
            else:
                # Nếu không tìm thấy ở đâu, thì thêm vào ticket thứ hai
                spapi_new_note(TICKET_CUSTOMER_IDS[0], objx)

    # STEP 2 - Export ngược ra Excel từ 2 nơi
    if request.GET.get('action') == 'export_to_excel':

        data_list = []
        for ticket_id in TICKET_CUSTOMER_IDS:
            url_notes = f"{MAIN_URL}/customers/{ticket_id}.json"
            res = loginss.get(url_notes)
            if res.status_code != 200:
                continue

            customer = res.json().get("customer", {})
            active_notes = [n for n in customer.get("notes", []) if n.get("status") == "active"]
            for note in active_notes:
                try:
                    content = json.loads(note.get("content", "{}"))
                    data_list.append(content)
                except:
                    continue

        if not data_list:
            return JsonResponse({"status": "error", "msg": "Không có dữ liệu hợp lệ"}, status=404)

        for data_vari in data_list:
            #Lấy thêm thông tin sản phẩm.
            keys_to_remove = ["image_path", "name_vn", "brand","vi_name", "en_name", "descreption", "material"]
            for key in keys_to_remove:
                data_vari.pop(key, None)
            #Lấy thêm thông tin variant:
            try:
                this_vari = js_get_url(f"{MAIN_URL}/variants.json?query={data_vari['sku']}&limit=1")['variants'][0]
                data_vari["vari_id"] = this_vari["id"]
                data_vari["vn_name"] = this_vari["opt1"]
                #Lấy thêm thông tin product
                this_product = js_get_url(f"{MAIN_URL}/products/{this_vari['product_id']}.json")['product']
                data_vari["nsx"] = this_product['nsx']
            except:
                data_list.remove(data_vari)

        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        headers = list(data_list[0].keys())

        for col, header in enumerate(headers):
            ws.write(0, col, header)
        for row_idx, row_data in enumerate(data_list, start=1):
            for col_idx, header in enumerate(headers):
                ws.write(row_idx, col_idx, row_data.get(header, ""))

        wb.save(excel_file_path)
    
    data_list = []
    for ticket_id in TICKET_CUSTOMER_IDS:
        url_notes = f"{MAIN_URL}/customers/{ticket_id}.json"
        res = loginss.get(url_notes)
        if res.status_code != 200:
            continue

        customer = res.json().get("customer", {})
        active_notes = [n for n in customer.get("notes", []) if n.get("status") == "active"]
        for note in active_notes:
            try:
                content = json.loads(note.get("content", "{}"))
                content["product_id"] = int(float(content["product_id"]))
                data_list.append(content)
            except:
                continue

    return render(request, 'kho_barcode.html', {'data_list':data_list})

@csrf_exempt
def api_kho_editproduct(request):
    TICKET_CUSTOMER_IDS = ['760093681']

    if request.method != "POST":
        return JsonResponse({"status": "error", "msg": "Phải dùng phương thức POST"}, status=405)
    try:
        updates = json.loads(request.body)
    except Exception as e:
        return JsonResponse({"status": "error", "msg": f"Dữ liệu JSON không hợp lệ: {e}"}, status=400)

    sku = str(updates.get("product_id", "")).strip()
    if not sku:
        return JsonResponse({"status": "error", "msg": "Thiếu product_id trong dữ liệu gửi lên"}, status=400)

    # Duyệt qua từng ticket để tìm note
    matched_note = None
    matched_ticket_id = None

    for ticket_id in TICKET_CUSTOMER_IDS:
        url_notes = f"{MAIN_URL}/customers/{ticket_id}.json"
        res = loginss.get(url_notes)
        if res.status_code != 200:
            continue

        customer = res.json().get("customer", {})
        active_notes = [note for note in customer.get("notes", []) if note.get("status") == "active"]

        for note in active_notes:
            try:
                note_content = json.loads(note.get("content", "{}"))
                if str(int(float(note_content.get("product_id", "")))).strip() == sku:
                    matched_note = note
                    matched_ticket_id = ticket_id
                    break
            except json.JSONDecodeError:
                continue
        if matched_note:
            break  # Thoát khỏi vòng for ngoài nếu đã tìm thấy

    if matched_note and matched_ticket_id:
        result = spapi_edit_note(matched_ticket_id, matched_note["id"], updates)
        if result.get("status") == "success":
            return JsonResponse({"status": "success", "msg": "Cập nhật thành công"})
        else:
            return JsonResponse({"status": "error", "msg": result.get("msg")}, status=500)
    else:
        return JsonResponse({"status": "error", "msg": "Không tìm thấy SKU trong các notes hiện tại"}, status=404)


def xnk_code(request):
    TICKET_CUSTOMER_ID = '759930912'
    excel_file_path = 'logs/log excel/nhap-khau-model.xls'

    # STEP 1 - Xử lý reload từ Excel
    if request.GET.get('action') == 'reload_from_file':
        wb = xlrd.open_workbook(excel_file_path)
        sheet = wb.sheet_by_index(0)
        headers = sheet.row_values(0)

        objects_list = []
        for i in range(1, sheet.nrows):
            row_values = sheet.row_values(i)
            obj = {header: value for header, value in zip(headers, row_values)}
            objects_list.append(obj)

        # Lấy tất cả notes hiện tại
        url_notes = f"{MAIN_URL}/customers/{TICKET_CUSTOMER_ID}.json"
        res = loginss.get(url_notes)
        customer = res.json().get("customer", {})
        active_notes = [n for n in customer.get("notes", []) if n.get("status") == "active"]

        # So khớp theo SKU
        for objx in objects_list:
            sku = str(objx.get("sku", "")).strip()
            matched_note = None
            for note in active_notes:
                try:
                    content = json.loads(note.get("content", "{}"))
                    if str(content.get("sku", "")).strip() == sku:
                        matched_note = note
                        break
                except:
                    continue

            if matched_note:
                spapi_edit_note(TICKET_CUSTOMER_ID, matched_note["id"], objx)
            else:
                spapi_new_note(TICKET_CUSTOMER_ID, objx)

    # STEP 2 - Export ngược ra Excel
    if request.GET.get('action') == 'export_to_excel':
        url_notes = f"{MAIN_URL}/customers/{TICKET_CUSTOMER_ID}.json"
        res = loginss.get(url_notes)
        if res.status_code != 200:
            return JsonResponse({"status": "error", "msg": "Không lấy được notes"}, status=500)

        customer = res.json().get("customer", {})
        active_notes = [n for n in customer.get("notes", []) if n.get("status") == "active"]

        data_list = []
        for note in active_notes:
            try:
                content = json.loads(note.get("content", "{}"))
                data_list.append(content)
            except:
                continue

        if not data_list:
            return JsonResponse({"status": "error", "msg": "Không có dữ liệu hợp lệ"}, status=404)

        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        headers = list(data_list[0].keys())

        for col, header in enumerate(headers):
            ws.write(0, col, header)
        for row_idx, row_data in enumerate(data_list, start=1):
            for col_idx, header in enumerate(headers):
                ws.write(row_idx, col_idx, row_data.get(header, ""))

        wb.save(excel_file_path)

    # STEP 3 - Load dữ liệu render ra giao diện
    data_list = []
    url_notes = f"{MAIN_URL}/customers/{TICKET_CUSTOMER_ID}.json"
    res = loginss.get(url_notes)
    if res.status_code != 200:
        return JsonResponse({"status": "error", "msg": "Không lấy được notes"}, status=500)

    customer = res.json().get("customer", {})
    active_notes = [n for n in customer.get("notes", []) if n.get("status") == "active"]
    for note in active_notes:
        try:
            content = json.loads(note.get("content", "{}"))
            data_list.append(content)
        except:
            continue

    return render(request, 'listnkcode.html', {'data_list': data_list})

@csrf_exempt
def api_qt_editmodel(request):
    TICKET_CUSTOMER_ID = '759930912'

    if request.method != "POST":
        return JsonResponse({"status": "error", "msg": "Phải dùng phương thức POST"}, status=405)

    try:
        updates = json.loads(request.body)
    except Exception as e:
        return JsonResponse({"status": "error", "msg": f"Dữ liệu JSON không hợp lệ: {e}"}, status=400)

    sku = str(updates.get("sku", "")).strip()
    if not sku:
        return JsonResponse({"status": "error", "msg": "Thiếu SKU trong dữ liệu gửi lên"}, status=400)

    # Lấy danh sách notes active của khách
    url_notes = f"{MAIN_URL}/customers/{TICKET_CUSTOMER_ID}.json"
    res = loginss.get(url_notes)
    if res.status_code != 200:
        return JsonResponse({"status": "error", "msg": "Không lấy được danh sách notes"}, status=500)

    customer = res.json().get("customer", {})
    active_notes = [note for note in customer.get("notes", []) if note.get("status") == "active"]

    # Tìm note có SKU trùng
    matched_note = None
    for note in active_notes:
        try:
            note_content = json.loads(note.get("content", "{}"))
            if str(note_content.get("sku", "")).strip() == sku:
                matched_note = note
                break
        except json.JSONDecodeError:
            continue

    if matched_note:
        # Gọi lại hàm spapi_edit_note tái sử dụng
        result = spapi_edit_note(TICKET_CUSTOMER_ID, matched_note["id"], updates)
        if result.get("status") == "success":
            return JsonResponse({"status": "success", "msg": "Cập nhật thành công"})
        else:
            return JsonResponse({"status": "error", "msg": result.get("msg")}, status=500)
    else:
        return JsonResponse({"status": "error", "msg": "Không tìm thấy SKU trong các notes hiện tại"}, status=404)


def nhacungcap(request):
    action = request.GET['action']
    all_sup = json.loads(readfile("logs/all_sup.json"))
    if action == 'list':
        all_sup = json.loads(readfile("logs/all_sup.json"))
        for sup in all_sup["suppliers"]:
            if sup['description'] == None:
                sup['description'] = 'noname.jpg'
            elif '.jpg' not in sup['description'] and '.png' not in sup['description']:
                sup['description'] = 'noname.jpg'

            brand = js_get_url(f"{MAIN_URL}/brands.json?query={sup['code']}")["brands"][0] 
            sup["variants"] = js_get_url(f"{MAIN_URL}/variants.json?composite=false&packsize=false&limit=250&page=1&brand_ids={brand['id']}")["variants"]
            for vari in sup["variants"]:
                pass
            po = js_get_url(f"{MAIN_URL}/purchase_orders.json?page=1&limit=20&receive_statuses=received&supplier_ids={sup['id']}")
            if po['metadata']['total'] == 0:
                sup['last_time'] = "9999 years ago"
            else:
                po = po["purchase_orders"][0]
                time_nhap = datetime.datetime.strptime(po['receipts'][0]['processed_on'], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
                now = datetime.datetime.now()
                sup['last_time'] = f"{(now - time_nhap).days} days ago"
    if action == 'uoctinh':
        TICKET_CUSTOMER_IDS = CUSTOMER_ID_JSON["VARIANTS"]
        data_list = get_json_variants(TICKET_CUSTOMER_IDS)

        date_start = datetime.datetime.strptime(unquote(request.GET['date']), "%d/%m/%Y")
        all_sup = js_get_url(f"{MAIN_URL}/suppliers.json?statuses=active&page=1&limit=250")
        
        now = datetime.datetime.now()
        TIME_TO_START = (date_start - now).days
        SAISO = 1.1

        for sup in all_sup["suppliers"]:
            if sup['description'] == None:
                sup['description'] = 'noname.jpg'
            elif '.jpg' not in sup['description'] and '.png' not in sup['description']:
                sup['description'] = 'noname.jpg'

            sup["uoc_tinh_for_date"] = unquote(request.GET['date'])
            sup["list_hotsku"] = []
            sup["sum_khoi"] = 0
            sup["sum_hotsku"] = 0
            sup["sum_sanpham"] = 0

            sup['sum_qty_60_30'] = 0
            sup['sum_qty_30_0'] = 0
            sup['sum_qty_30_15'] = 0
            sup['sum_qty_15_0'] = 0

            TIME_TO_ARRIVE = 20
            if sup['group_name'] == "CONTAINER 30":
                TIME_TO_SALE = 30
            elif sup['group_name'] == "CONTAINER 60":
                TIME_TO_SALE = 60
            else:
                TIME_TO_SALE = 90

            brand = js_get_url(f"{MAIN_URL}/brands.json?query={sup['code']}")["brands"][-1]

            url_search = f"{MAIN_URL}/variants.json?composite=false&packsize=false&limit=250&page=1&brand_ids=" + str(brand["id"])
            all_variant = json.loads(loginss.get(url_search).text)
            if all_variant["metadata"]["total"] > 0:
                for variant in all_variant["variants"]:
                    now = datetime.datetime.now()
                    amonth = now - datetime.timedelta(days=60)
                    variant['start_date'] = amonth.strftime("%Y-%m-%d") + "T17:00:00Z"
                    variant["end_date"] = now.strftime("%Y-%m-%d") + "T16:59:59Z"
                    try:
                        variant['data'] = data_list[str(variant['id'])]

                    except:
                        break
                    if int(str(variant["data"]["fullbox"]).replace(".0","")) > 0:
                        variant["so_khoi"] = float((variant["data"]["box_cao"] * variant["data"]["box_dai"] * variant["data"]["box_rong"])/1000000/variant["data"]["fullbox"])
                    else:
                        variant["so_khoi"] = 0
                    

                    variant["available"] = int(variant["inventories"][0]["available"] + variant["inventories"][1]["available"])
                    variant["image"] = variant["images"][0]["full_path"]
                    variant['s_quantity'] = 0
                    variant["total_trans"] = 0
                    variant["trung_binh"] = 0
                    variant["brand"] = brand["name"]
                    # Tìm đơn thường có sản phẩm.
                    all_orders = js_get_url(f'{MAIN_URL}/orders.json?page=1&limit=1&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(str(variant['id']))+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                    if all_orders['metadata']['total'] > 0:
                        loop = math.ceil(all_orders['metadata']['total']/250)
                        for loop_x in range(loop):
                            all_orders = js_get_url(f'{MAIN_URL}/orders.json?page='+str(loop_x+1)+'&limit=250&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(str(variant['id']))+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                            for order in all_orders["orders"]:
                                if order['source_id'] != 7239422:
                                    for line in order['order_line_items']:
                                        if line['variant_id'] == variant["id"]:
                                            time_order = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ")
                                            day_from_now = abs((now - time_order).days)
                                            if day_from_now <= 60 and day_from_now >= 30:
                                                sup['sum_qty_60_30'] += int(line['quantity'])
                                            elif day_from_now <= 30 and day_from_now >= 0:
                                                variant['s_quantity'] += int(line['quantity'])
                                                sup['sum_qty_30_0'] += int(line['quantity'])
                                                if day_from_now >= 15:
                                                    sup['sum_qty_30_15'] += int(line['quantity'])
                                                else:
                                                    sup['sum_qty_15_0'] += int(line['quantity'])

                    # Tìm đơn có hệ số từ pack
                    all_vari_pack = js_get_url(f'{MAIN_URL}/variants.json?packsize=true&packsize_root_id='+str(variant['id']))
                    if all_vari_pack['metadata']['total'] > 0:
                        for pack_vari in all_vari_pack['variants']:
                            print("Xu ly packed vari: "+ pack_vari['sku'])
                            all_orders = js_get_url(f'{MAIN_URL}/orders.json?page=1&limit=1&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(pack_vari['id'])+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                            if all_orders['metadata']['total'] > 0:
                                loop = math.ceil(all_orders['metadata']['total']/250)
                                for loop_x in range(loop):
                                    all_orders = js_get_url(f'{MAIN_URL}/orders.json?page='+str(loop_x+1)+'&limit=250&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(pack_vari['id'])+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                                    for order in all_orders["orders"]:
                                        if order['source_id'] != 7239422:
                                            for line in order['order_line_items']:
                                                if line['variant_id'] == pack_vari['id']:
                                                    time_order = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ")
                                                    day_from_now = abs((now - time_order).days)
                                                    if day_from_now <= 60 and day_from_now >= 30:
                                                        sup['sum_qty_60_30'] += int(line['quantity']*pack_vari['packsize_quantity'])
                                                    elif day_from_now <= 30 and day_from_now >= 0:
                                                        variant['s_quantity'] += int(line['quantity']*pack_vari['packsize_quantity'])
                                                        sup['sum_qty_30_0'] += int(line['quantity']*pack_vari['packsize_quantity'])
                                                        if day_from_now >= 15:
                                                            sup['sum_qty_30_15'] += int(line['quantity']*pack_vari['packsize_quantity'])
                                                        else:
                                                            sup['sum_qty_15_0'] += int(line['quantity']*pack_vari['packsize_quantity'])
                                                    
                    # Tìm đơn có hệ số từ combo.
                    all_vari_combo = js_get_url(f'{MAIN_URL}/products.json?composite_item_variant_ids='+str(variant['id']))
                    if all_vari_pack['metadata']['total'] > 0:
                        for combo_vari in all_vari_combo['products']:
                            combo_vari = combo_vari["variants"][0]
                            for cmps in combo_vari['composite_items']:
                                if cmps['sub_variant_id'] == variant['id']:
                                    HESO = cmps['quantity']

                            all_orders = js_get_url(f'{MAIN_URL}/orders.json?page=1&limit=1&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(combo_vari['id'])+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                            if all_orders['metadata']['total'] > 0:
                                loop = math.ceil(all_orders['metadata']['total']/250)
                                for loop_x in range(loop):
                                    all_orders = js_get_url(f'{MAIN_URL}/orders.json?page='+str(loop_x+1)+'&limit=250&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(combo_vari['id'])+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                                    for order in all_orders["orders"]:
                                        if order['source_id'] != 7239422:
                                            for line in order['order_line_items']:
                                                if line['variant_id'] == combo_vari['id']:
                                                    time_order = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ")
                                                    day_from_now = abs((now - time_order).days)
                                                    if day_from_now <= 60 and day_from_now >= 30:
                                                        sup['sum_qty_60_30'] += int(line['quantity']*HESO)
                                                    elif day_from_now <= 30 and day_from_now >= 0:
                                                        variant['s_quantity'] += int(line['quantity']*HESO)
                                                        sup['sum_qty_30_0'] += int(line['quantity']*HESO)
                                                        if day_from_now >= 15:
                                                            sup['sum_qty_30_15'] += int(line['quantity']*HESO)
                                                        else:
                                                            sup['sum_qty_15_0'] += int(line['quantity']*HESO)

                    variant["trung_binh"] = float(variant["s_quantity"]/30)

                    TON_HANG_VE = variant["trung_binh"] * (TIME_TO_START + TIME_TO_ARRIVE) * SAISO
                    TON_HANG_CHAY = variant["trung_binh"] * 40 * SAISO
                    
                    if TON_HANG_CHAY > variant["available"]:
                        sup["sum_hotsku"] += 1
                        sup["list_hotsku"].append(variant)
                    if TON_HANG_VE >= variant["available"]:
                        TON_HANG_VE = 0  
                    else:
                        TON_HANG_VE = variant["available"] - TON_HANG_VE
                        
                    variant["goiynhap"] =  int(variant["trung_binh"] * TIME_TO_SALE * SAISO - TON_HANG_VE) - int(variant["inventories"][0]["incoming"])
                    if variant["goiynhap"] > 0:
                        sup["sum_khoi"] += float(variant["goiynhap"]*variant["so_khoi"])
                        sup["sum_sanpham"] += variant["goiynhap"]

        writejsonfile(all_sup,"logs/all_sup.json")
    #So sánh tốc độ bán 30 & 15
    for sup in all_sup["suppliers"]:
        sup["ss_30"] = 0
        sup["ss_15"] = 0

        if sup["sum_qty_30_15"] > 0: 
            sup["ss_15"] = float((sup["sum_qty_15_0"] - sup["sum_qty_30_15"])/sup["sum_qty_30_15"])*100
        if sup["sum_qty_60_30"] > 0: 
            sup["ss_30"] = float((sup["sum_qty_30_0"] - sup["sum_qty_60_30"])/sup["sum_qty_60_30"])*100

    all_sup["suppliers"] = sorted(all_sup["suppliers"], key=lambda x: x["sum_khoi"], reverse=True)
    obj = {
        'all_sup': all_sup["suppliers"],
        'action': action,           
    }

    return render(request, 'nhacungcap.html',obj)

def nhaphang(request):
    action = request.GET['action']
    if action == 'list':
        all_sup = js_get_url(f"{MAIN_URL}/suppliers.json?statuses=active&page=1&limit=250&ids=55410493")["suppliers"]
        for sup in all_sup:
            if sup['description'] == None:
                sup['description'] = 'noname.jpg'
            elif '.jpg' not in sup['description'] and '.png' not in sup['description']:
                sup['description'] = 'noname.jpg'

            brand = js_get_url(f"{MAIN_URL}/brands.json?query={sup['code']}")["brands"][0] 
            sup["variants"] = js_get_url(f"{MAIN_URL}/variants.json?composite=false&packsize=false&limit=250&page=1&brand_ids={brand['id']}")["variants"]
            for vari in sup["variants"]:
                pass

    obj = {
        'all_sup': all_sup[::-1],
        'action': action,           
    }
    return render(request, 'tongquannhaphang.html',obj)

def donnhaphang(request):
    action = request.GET['action']
    all_model = all_xnk_data()
    data_list = get_json_variants(TICKET_CUSTOMER_IDS = CUSTOMER_ID_JSON["VARIANTS"])

    if action == 'excel':
        obj = {
            'list_product': [],
            'list_id': []
        }
        po_id = request.GET['id']
        po = Purchase_Order.objects.get(id=po_id)
        print(f"Xu ly xuat excel: {po.madonnhap}")
        all_po_con = js_get_url(f"{MAIN_URL}/order_suppliers.json?page=1&limit=50&statuses=pending%2Cpartial%2Ccompleted&tags={po.madonnhap}")["order_suppliers"]
        
        for po_con in all_po_con:
            print(f"Xu ly PO: {po_con['code']}")
            for line in po_con["line_items"]:
                line['data'] = data_list[str(line['variant_id'])]
                if len(line['data']) < 10:
                    print(f"Vari {line['sku']} not found data!")
                else:
                    if line['data']['sku_nhapkhau'].strip() not in obj['list_id']:

                        count = 0
                        for xmodel in all_model:
                            if xmodel['sku'] == line['data']['sku_nhapkhau'].strip():
                                model = xmodel
                                count = 1
                                break
                        if count == 0:
                            print(f"{line['sku']} not have data!")
                        
                        else:
                            if model['sku'] == "":
                                print(f"{line['sku']} error with model!")
                            if model['en_name'] == "":
                                print(f"{line['sku']} error with model!")
                            if model['hs_code'] == "":
                                print(f"{line['sku']} error with model!")
                                
                            new_line = {'hs_code':model['hs_code'],'sku_nhapkhau':model['sku'],'vn_name':model['vn_name'],'en_name':model['en_name'],'cn_name':model['china_name'],'nsx_name':model['nsx_name'],'quantity':0,'unit':model['unit'],'price':model['usd_price'],'box':0,'total_price':0,'total_cpm':0,'material':model['material'],'total_weight':0}
                            new_line['quantity'] += line['quantity']
                            new_line['box'] += (line['quantity']/line['data']['fullbox'])
                            new_line['total_price'] += float(line['quantity'])*float(new_line['price'])
                            
                            new_line["total_cpm"] += (line['data']['box_cao']*line['data']['box_dai']*line['data']['box_rong']*line["quantity"])/1000000/line['data']['fullbox']

                            obj['list_id'].append(line['data']['sku_nhapkhau'].strip())
                            obj['list_product'].append(new_line)
                    else:
                        for new_line in obj['list_product']:
                            if line['data']['sku_nhapkhau'].strip() == new_line['sku_nhapkhau']:
                                if line['data']['fullbox'] == 0:
                                    line['data']['fullbox'] = 1
                                new_line['quantity'] += line['quantity']
                                new_line['box'] += (line['quantity']/line['data']['fullbox'])
                                new_line['total_price'] += float(line['quantity'])*float(new_line['price'])
                                new_line["total_cpm"] += (line['data']['box_cao']*line['data']['box_dai']*line['data']['box_rong']*line["quantity"])/1000000/line['data']['fullbox']
                               
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(f"assets/excel_po/ALL-{po.madonnhap}.xlsx")
        worksheet = workbook.add_worksheet()
        row = 0
        col = 0     
        worksheet.write(row, col, "MÃ HS")
        worksheet.write(row, col + 1, "TÊN TIẾNG VIỆT")
        worksheet.write(row, col + 2, "TÊN TIẾNG ANH")
        worksheet.write(row, col + 3, "TÊN TIẾNG TRUNG")
        worksheet.write(row, col + 4, "NSX")
        worksheet.write(row, col + 5, "QUANTITY")
        worksheet.write(row, col + 6, "UNIT")
        worksheet.write(row, col + 7, "BOX")
        worksheet.write(row, col + 8, "UNIT PRICE")
        worksheet.write(row, col + 9, "TOTAL PRICE")
        worksheet.write(row, col + 10, "TOTAL CAPACITY")
        worksheet.write(row, col + 11, "TOTAL WEIGHT")
        worksheet.write(row, col + 12, "MATERIAL")   
        worksheet.write(row, col + 13, "SKU-NHAPKHAU")  

        row = 1
        for item in obj['list_product']:
            col = 0
            worksheet.write(row, col, item["hs_code"])
            worksheet.write(row, col + 1, item["vn_name"])
            worksheet.write(row, col + 2, item["en_name"])
            worksheet.write(row, col + 3, item["cn_name"])
            worksheet.write(row, col + 4, item["nsx_name"])
            worksheet.write(row, col + 5, item["quantity"])
            worksheet.write(row, col + 6, item["unit"])
            worksheet.write(row, col + 7, item["box"])
            worksheet.write(row, col + 8, item["price"])
            worksheet.write(row, col + 9, item["total_price"])
            worksheet.write(row, col + 10, item["total_cpm"])
            worksheet.write(row, col + 11, (item["total_cpm"]*1000000/6000))
            worksheet.write(row, col + 12, item["material"])
            worksheet.write(row, col + 13, item["sku_nhapkhau"])
            row += 1              
        workbook.close()  
        obj['url_download'] = f"/static/excel_po/ALL-{po.madonnhap}.xlsx"
        return render(request, 're_download.html',obj)


    if action == 'new':
        if 'madonnhap' in request.GET:
            try:
                po = Purchase_Order.objects.get(madonnhap=str(request.GET['madonnhap']))
            except Purchase_Order.DoesNotExist:
                print("Im here 2!")
                po = Purchase_Order()
                po.madonnhap = request.GET['madonnhap']
                po.ship_to = request.GET['ship_to']
                po.port_start = request.GET['port_start']
                po.port_end = request.GET['port_end']
                po.ship_by = request.GET['ship_by']
                po.save()
         
        all_po = Purchase_Order.objects.all()
        obj = {
                'all_po': all_po[::-1],
                'urls': {
                    'action': action,
                }
        }
        return render(request, 'taodonnhaphang.html', obj)
    

    if action == 'xuatdoisoat':
        obj = {
        'list_product': [],
        'list_id': []
        }
        po_id = request.GET['id']
        po = Purchase_Order.objects.get(id=po_id)
        print(f"Xu ly xuat excel: {po.madonnhap}")
        all_po_con = js_get_url(f"{MAIN_URL}/order_suppliers.json?page=1&limit=50&statuses=pending%2Cpartial%2Ccompleted&tags={po.madonnhap}")["order_suppliers"]
        
        for po_con in all_po_con:
            print(f"Xu ly PO: {po_con['code']}")
            for line in po_con["line_items"]:
                line['po'] = po_con['code']
                line['data'] = data_list[str(line['variant_id'])]
                if len(line['data']) < 50:
                    print(f"Vari {line['sku']} not found data!")

                count = 0
                for xmodel in all_model:
                    if xmodel['sku'] == line['data']['sku_nhapkhau'].strip():
                        model = xmodel
                        count = 1
                        break

                if count == 0:
                    print(f"{line['sku']} not have data!")
                    
                else:
                    if model['sku'] == "":
                        print(f"{line['sku']} error with model!")
                    if model['en_name'] == "":
                        print(f"{line['sku']} error with model!")
                    if model['hs_code'] == "":
                        print(f"{line['sku']} error with model!")
                    
                    line['tax_value'] = model['usd_price']
                    line['tax_vat'] =  model['tax_vat']
                    line['tax_nk'] =  model['tax_nk']

                obj['list_product'].append(line)
        return render(request, 'xuatdoisoat.html', obj)
    
    if action == 'xuatinfokho':
        obj = {
        'poall': {},
        'po': []
        }
        po_id = request.GET['id']
        po = Purchase_Order.objects.get(id=po_id)
        print(f"[+] Xuất thông tin kho cho đơn hàng: {po.madonnhap}")
        all_po_con = js_get_url(f"{MAIN_URL}/order_suppliers.json?page=1&limit=50&statuses=pending%2Cpartial%2Ccompleted&tags={po.madonnhap}")["order_suppliers"]

        for po_con in all_po_con:
            po_con['madonnhap'] = po.madonnhap
            print(f"[+] PO con: {po_con['code']}")
            for line in po_con["line_items"]:
                line['po'] = po_con['code']
                line['data'] = data_list[str(line['variant_id'])]
                if line['data']['price_tq'] == 0:
                    print(line['data'])
                    print(f"Vari {line['sku']} not found data!")

                if po_con['location_id'] == 241737:
                    line['stock'] = line['data']['stock_hn']
                else:
                    line['stock'] = line['data']['stock_sg']

                    
            po_con["line_items"] = sorted(po_con["line_items"], key=lambda x: x['sku'])
            obj['po'].append(po_con)

        return render(request, 'xuatinfokho.html', obj)


    if action == 'packing_list':
        obj = {
            'list_product': [],
            'list_id': []
        }
        po_id = request.GET['id']
        po = Purchase_Order.objects.get(id=po_id)
        print(f"Xu ly xuat excel: {po.madonnhap}")
        all_po_con = js_get_url(f"{MAIN_URL}/order_suppliers.json?page=1&limit=50&statuses=pending%2Cpartial%2Ccompleted&tags={po.madonnhap}")["order_suppliers"]
        
        for po_con in all_po_con:
            print(f"Xu ly PO: {po_con['code']}")
            for line in po_con["line_items"]:
                line['data'] = data_list[str(line['variant_id'])]
                if len(line['data']) < 10:
                    print(f"Vari {line['sku']} not found data!")
                else:
                    count = 0
                    for xmodel in all_model:
                        if xmodel['sku'] == line['data']['sku_nhapkhau'].strip():
                            line['model'] = xmodel
                            count = 1
                            break
                    if count == 0:
                        print(f"{line['sku']} not have data!")

                line['box'] = int(line['quantity']/line['data']['fullbox'])
                line['cpm'] = float((line['data']['box_dai']*line['data']['box_rong']*line['data']['box_cao']/1000000)*line['box'])

                obj['list_product'].append(line)
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(f"assets/excel_po/ALL-{po.madonnhap}.xlsx")
        worksheet = workbook.add_worksheet()
        row = 0
        col = 0     
        worksheet.write(row, col, "MÃ HS")
        worksheet.write(row, col + 1, "TÊN TIẾNG VIỆT")
        worksheet.write(row, col + 2, "TÊN TIẾNG ANH")

        worksheet.write(row, col + 3, "SKU")
        worksheet.write(row, col + 4, "PHÂN LOẠI")

        worksheet.write(row, col + 5, "QUANTITY")

        worksheet.write(row, col + 6, "BOX")

        worksheet.write(row, col + 7, "CPM")

        row = 1
        for item in obj['list_product']:
            col = 0
            worksheet.write(row, col, item['model']["hs_code"])
            worksheet.write(row, col + 1, item['model']["vn_name"])
            worksheet.write(row, col + 2, item['model']["en_name"])

            worksheet.write(row, col + 3, item["sku"])
            worksheet.write(row, col + 4, item["variant_options"])

            worksheet.write(row, col + 5, item["quantity"])

            worksheet.write(row, col + 6, item["box"])
            worksheet.write(row, col + 7, item["cpm"])
            row += 1              
        workbook.close()  
        obj['url_download'] = f"/static/excel_po/ALL-{po.madonnhap}.xlsx"
        return render(request, 're_download.html',obj) 
    
    else:   
        if action == 'edit':
            if 'ship_by' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(ship_by=request.GET['ship_by'])
            if 'port_start' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(port_start=request.GET['port_start'])
            if 'port_end' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(port_end=request.GET['port_end'])
            if 'ship_to' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(ship_to=request.GET['ship_to'])
            if 'invoice' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(invoice=request.GET['invoice'])
            if 'invoice_date' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(invoice_date=request.GET['invoice_date'])
            if 'invoice_value' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(invoice_value=request.GET['invoice_value'])
            if 'vnd_ndt' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(vnd_ndt=request.GET['vnd_ndt'])
            if 'status' in request.GET:
                Purchase_Order.objects.filter(id=request.GET['id']).update(status=request.GET['status'])
        if action == 'edit-poc':
            poc_id = request.GET['vari_poc']
            Order_Supplier.objects.filter(id=poc_id).update(ship_fee=request.GET['ship_fee'])
        
        if action == 'view' or action == 'edit' or action == 'edit-poc':
            all_po = Purchase_Order.objects.exclude(status=3)
        elif action == 'doisoat':
            all_po = Purchase_Order.objects.filter(status=3)
        
        for po in all_po:
            po.total_cpm = 0
            po.total_quantity = 0
            po.total_box = 0
            po.total_cny = 0
            po.ship_tq = 0
            all_poc = js_get_url(f"{MAIN_URL}/order_suppliers.json?statuses=completed%2Cpartial%2Cpending&tags={po.madonnhap}")["order_suppliers"]
            
            for poc in all_poc:
                poc["total_quantity"] = 0
                poc["total_box"] = 0
                poc["total_cpm"] = 0
                poc["total_cny"] = 0
                poc["total_ship"] = 0
                try:
                    this_poc = Order_Supplier.objects.get(id=poc["id"])
                except Order_Supplier.DoesNotExist:
                    this_poc = Order_Supplier()
                    this_poc.id = poc["id"]
                    this_poc.code = poc["code"]
                    this_poc.save()
                else:
                    poc["total_ship"] = this_poc.ship_fee

                for item in poc["line_items"]:

                    item['data'] = data_list[str(item['variant_id'])]
                    if item['data']['fullbox'] == 0:
                        item['data']['fullbox'] = 1
                        
                    item["total_cpm"] = 1.05*(item['data']['box_cao']*item['data']['box_dai']*item['data']['box_rong']*item["quantity"])/1000000/item['data']['fullbox']
                    item["tq_variant"] = item['data']['name_tq']
                    
                    item["tq_price"] = float(item['data']['price_tq'])
                    if item["tq_price"] == 0:
                        print(f"{item['sku']} not have data price!")
                    if item['data']['box_cao'] == 0:
                        print(f"{item['sku']} not have data fullbox!")

                    item["tq_box"] = item['data']['fullbox']
                    poc["total_quantity"] += item["quantity"]

                    poc["total_cpm"] += item["total_cpm"]
                    poc["total_cny"] += item["quantity"]*item["tq_price"]
                    poc["total_box"] += item["quantity"]/item["tq_box"]

                    po.total_quantity += item["quantity"]
                    po.total_box += item["quantity"]/item["tq_box"]
                    po.total_cny += item["quantity"]*item["tq_price"]

                poc["total_money"] = poc["total_cny"] + poc["total_ship"]
                poc["total_vnd"] = poc["total_money"]*po.vnd_ndt
                #poc["ship_per_cpm"] = poc["total_ship"]/poc["total_cpm"]
                po.ship_tq += poc["total_ship"]
                po.total_cpm += poc["total_cpm"]
                po.save()

            po.total_vnd = po.total_cny*po.vnd_ndt
            po.poc = all_poc
        
        obj = {
                'all_po': all_po[::-1],
                'urls': {
                    'action': action,
                }
        }

        return render(request, 'donnhaphang.html', obj)

def editdonnhap(request):
    obj = {
        'all_vari': [],
        'sum': {},
        'dot_nhap': [],
        'real_po': [],
        'save_data': {},
        'url_download' : '',
        'time_to_sale': 60
    }
    data_list = get_json_variants(TICKET_CUSTOMER_IDS = CUSTOMER_ID_JSON["VARIANTS"])

    if request.method == 'GET':
        action = request.GET['action']
        if action == 'new':
            sup_code = request.GET['brand_id']

            obj['save_data']['brand_id'] = sup_code
            obj['dot_nhap'] = Purchase_Order.objects.filter(status__lt=3)
            obj['dot_nhap'] = json.loads(serializers.serialize('json', obj['dot_nhap']))

            sup = js_get_url(f"{MAIN_URL}/suppliers/{sup_code}.json")["supplier"]
            brand = js_get_url(f"{MAIN_URL}/brands.json?name={sup['code']}")["brands"][0] 

            all_products = []
            for i in range(3):
                products_x = js_get_url(f"{MAIN_URL}/products.json?page={int(i+1)}&status=active&limit=100&brand_id={brand['id']}")["products"]
                for pr in products_x:
                    all_products.append(pr)

            print(f"[+] Loading gợi ý nhập hàng cho brand: {brand['name']}")

            for x in range(1,3):
                url_search = f"{MAIN_URL}/variants.json?composite=false&status=active&packsize=false&limit=250&page={x}&brand_ids=" + str(brand["id"])
                all_variant = json.loads(loginss.get(url_search).text)
                if all_variant["metadata"]["total"] > 0:
                    for variant in all_variant["variants"]:
                        variant["tags"] = ""
                        for pr in all_products:
                            if str(variant["product_id"]) == str(pr["id"]):
                                variant["tags"] = pr["tags"]
                                break
                        variant["on_hand_hn"] = int(variant["inventories"][0]["on_hand"])
                        variant["on_hand_sg"] = int(variant["inventories"][1]["on_hand"])    
                        #Sắp về này cần là sắp về thật sự -> lên đơn nháp.
                        variant["sap_ve_hn"] = 0
                        variant["sap_ve_sg"] = 0

                        if "stop_import" not in variant["tags"]:
                            print(f"[-] Product is process {variant['sku']}")

                            variant['image'] = variant['images'][0]['full_path']
                            
                            in_po = js_get_url(f"{MAIN_URL}/reports/stock/stock_diff_log.json?location_ids=0%2C241737%2C548744&variant_id={variant['id']}&inventory_field=incoming&compare_type=not_equal_zero&limit=500&page=1")
                            if in_po["metadata"]["total"] > 0:
                                for po in in_po['document_codes']:
                                    if 'FUN' not in po['document_code'] and 'SRN' not in po['document_code'] and 'STN' not in po['document_code']:
                                        
                                        po_json = js_get_url(f"{MAIN_URL}/order_suppliers/{po['root_id']}.json")['order_supplier']
                                        try:
                                            this_po = Purchase_Order.objects.get(madonnhap=po_json["tags"][1])
                                        except Purchase_Order.DoesNotExist:
                                            pass
                                        else:
                                            if this_po.status > 1:
                                                if this_po.ship_to == "Cát Lái" or this_po.ship_to == "HCM" or this_po.ship_to == "Hồ Chí Minh" or this_po.ship_to == "Sài Gòn":   
                                                    variant["sap_ve_sg"] += int(po["inventory_field"])
                                                else:
                                                    variant["sap_ve_hn"] += int(po["inventory_field"])
                            

                            variant['data'] = data_list[str(variant['id'])]
                            if variant['data']['fullbox'] == 0:
                                variant['data']['fullbox'] = 1
                            variant["so_khoi_box"] = float((variant['data']['box_cao'] * variant['data']['box_dai'] * variant['data']['box_rong'])/1000000/variant['data']['fullbox'])
                            variant["tq_variant"] = variant['data']['name_tq']
                            variant["tq_price"] = variant['data']['price_tq']
                            variant["tq_box"] = variant['data']['fullbox']
                                
                            now = datetime.datetime.now()
                            amonth = now - datetime.timedelta(days=60)

                            TIME_TO_ARRIVE = 25
                            TIME_TO_SALE = 60
                            date_start = datetime.datetime.utcnow()
                            TIME_TO_START = (date_start - now).days
                            
                            variant['start_date'] = amonth.strftime("%Y-%m-%d") + "T17:00:00Z"
                            variant["end_date"] = now.strftime("%Y-%m-%d") + "T16:59:59Z"
                            variant["on_hand"] = int(variant["inventories"][0]["on_hand"] + variant["inventories"][1]["on_hand"])
                            variant["total_trans"] = 0
                            variant["display"] = 1

                            # Example usage
                            region_filter = {
                                "MIEN_NAM": "TP. Hồ Chí Minh, TP Hồ Chí Minh,Bình Phước,Bà Rịa - Vũng Tàu, Tây Ninh, Bà Rịa Vũng Tàu,Bà Rịa-Vũng Tàu, Bình Dương, Đồng Nai,Long An, Đồng Tháp, Tiền Giang, An Giang, Bến Tre, Vĩnh Long, Trà Vinh, Hậu Giang, Kiên Giang, Sóc Trăng, Bạc Liêu, Cà Mau, Cần Thơ,  Kon Tum, Gia Lai, Đắk Lắk, Đắk Nông và Lâm Đồng, Bình Định, Phú Yên, Khánh Hòa, Ninh Thuận và Bình Thuận",
                                "MIEN_BAC": "Hà Nội, Hải Phòng, Tỉnh Lào Cai, Yên Bái, Điện Biên, Hòa Bình, Lai Châu, Sơn La; Tỉnh Hà Giang, Cao Bằng, Bắc Kạn, Lạng Sơn, Tuyên Quang, Thái Nguyên, Phú Thọ, Bắc Giang, Quảng Ninh; Tỉnh Bắc Ninh, Hà Nam, Hải Dương, Hưng Yên, Nam Định, Ninh Bình, Thái Bình, Vĩnh Phúc, Thanh Hóa, Nghệ An, Hà Tĩnh, Quảng Bình, Quảng Trị, Thừa Thiên Huế, Đà Nẵng, Quảng Nam, Quảng Ngãi"
                            }

                            stats = []
                            
                            # Tìm đơn thường có sản phẩm.
                            all_orders = js_get_url(f'{MAIN_URL}/orders.json?page=1&limit=1&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(str(variant['id']))+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                            if all_orders['metadata']['total'] > 0:
                                loop = math.ceil(all_orders['metadata']['total']/250)
                                for loop_x in range(loop):
                                    all_orders = js_get_url(f'{MAIN_URL}/orders.json?page='+str(loop_x+1)+'&limit=250&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(str(variant['id']))+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                                    for order in all_orders["orders"]:
                                        if order['source_id'] != 7239422:
                                            stats += extract_order_stats(order, variant_id=variant['id'], heso=1, region_filter=region_filter)

                            # Tìm đơn có hệ số từ pack
                            all_vari_pack = js_get_url(f'{MAIN_URL}/variants.json?packsize=true&packsize_root_id='+str(variant['id']))
                            if all_vari_pack['metadata']['total'] > 0:
                                for pack_vari in all_vari_pack['variants']:
                                    print("[+] Sản phẩm có packed: "+ pack_vari['sku'])
                                    all_orders = js_get_url(f'{MAIN_URL}/orders.json?page=1&limit=1&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(pack_vari['id'])+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                                    if all_orders['metadata']['total'] > 0:
                                        loop = math.ceil(all_orders['metadata']['total']/250)
                                        for loop_x in range(loop):
                                            all_orders = js_get_url(f'{MAIN_URL}/orders.json?page='+str(loop_x+1)+'&limit=250&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(pack_vari['id'])+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                                            for order in all_orders["orders"]:
                                                if order['source_id'] != 7239422:
                                                    stats += extract_order_stats(order, variant_id=variant['id'], heso=1, region_filter=region_filter)
                                                   
                            # Tìm đơn có hệ số từ combo.
                            all_vari_combo = js_get_url(f'{MAIN_URL}/products.json?composite_item_variant_ids='+str(variant['id']))
                            if all_vari_pack['metadata']['total'] > 0:
                                for combo_vari in all_vari_combo['products']:
                                    print("[+] Sản phẩm có combo: "+ pack_vari['sku'])
                                    combo_vari = combo_vari["variants"][0]
                                    for cmps in combo_vari['composite_items']:
                                        if cmps['sub_variant_id'] == variant['id']:
                                            HESO = cmps['quantity']

                                    all_orders = js_get_url(f'{MAIN_URL}/orders.json?page=1&limit=1&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(combo_vari['id'])+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                                    if all_orders['metadata']['total'] > 0:
                                        loop = math.ceil(all_orders['metadata']['total']/250)
                                        for loop_x in range(loop):
                                            all_orders = js_get_url(f'{MAIN_URL}/orders.json?page='+str(loop_x+1)+'&limit=250&status=draft%2Cfinalized%2Ccompleted&variant_ids='+str(combo_vari['id'])+"&created_on_min="+str(variant['start_date'])+"&created_on_max="+str(variant['end_date']))
                                            for order in all_orders["orders"]:
                                                if order['source_id'] != 7239422:
                                                    stats += extract_order_stats(order, variant_id=variant['id'], heso=HESO, region_filter=region_filter)
                                                    
                            
                            available = {"MIEN_BAC": variant["on_hand_hn"], "MIEN_NAM": variant["on_hand_sg"]}
                            incoming = {"MIEN_BAC": variant["sap_ve_hn"], "MIEN_NAM": variant["sap_ve_sg"]}

                            variant["incoming"] = incoming
                            variant["available"] = available

                            variant["df_sugget"] = calc_suggest_import(stats, available_dict=available, incoming_dict=incoming, days_plan=TIME_TO_SALE).to_dict(orient="records")
                            variant["stock_history"] = calculate_onhand_60_days(variant["id"], available["MIEN_BAC"],available["MIEN_NAM"])
                            variant['nhap'] = 0
                            
                            obj['all_vari'].append(variant)

                    writejsonfile(obj, f"logs/save_po/{sup_code}.log")

        elif action == 'edit':
            if 'search' in request.GET:
                search_key = str(request.GET['search'])
            else:
                search_key = ''
            
            if 'time_to_sale' in request.GET:
                time_to_sale = int(request.GET['time_to_sale'])

            sup_code = request.GET['brand_id']
            obj = json.loads(readfile(f"logs/save_po/{sup_code}.log"))
            obj['save_data']['sum_cpm'] = 0
            obj['save_data']['brand_id'] = sup_code
            if 'vari' in request.GET:
                for vari in obj["all_vari"]:
                    if vari["id"] == int(request.GET["vari"]):
                        vari["nhap"] = int(request.GET["sothung"])
                        vari["box_nhap"] = float(vari["nhap"]/vari["tq_box"])
                        vari['gy_box'] = int(vari["box_nhap"])*vari["tq_box"]

                        if vari["nhap"] > 0:
                            vari['data'] = data_list[str(vari['id'])]
                            if vari['data']['fullbox'] == 0:
                                vari['data']['fullbox'] = 1

                            vari["so_khoi_box"] = float((vari['data']['box_cao'] * vari['data']['box_dai'] * vari['data']['box_rong'])/1000000/vari['data']['fullbox'])
                            
                            vari["tq_variant"] = vari['data']['name_tq']
                            vari["tq_price"] = vari['data']['price_tq']
                            vari["tq_box"] = vari['data']['fullbox']
                               
                    obj['save_data']['sum_cpm'] += vari["nhap"]*vari["so_khoi_box"]

            for vari in obj["all_vari"]:
                if search_key in vari['sku']:
                    vari['display'] = 1
                else:
                    vari['display'] = 0

            if 'time_to_sale' in request.GET:
                old_time_to_sale = obj['time_to_sale']
                for vari in obj["all_vari"]:
                    vari['df_sugget'][0]['suggest_import'] = int(vari['df_sugget'][0]['suggest_import']/old_time_to_sale*time_to_sale)
                    vari['df_sugget'][0]['suggest_mien_bac'] = int(vari['df_sugget'][0]['suggest_mien_bac']/old_time_to_sale*time_to_sale)
                    vari['df_sugget'][0]['suggest_mien_nam'] = int(vari['df_sugget'][0]['suggest_mien_nam']/old_time_to_sale*time_to_sale)


                obj['time_to_sale'] = time_to_sale
            obj['save_data']['search_key'] = search_key
            writejsonfile(obj, f"logs/save_po/{sup_code}.log")

        elif action == 'export':
            json_order = {
                "order_supplier": {
                    "location_id": 241737,
                    "assignee_id": 319911,
                    "supplier_id": 179824266,
                    "price_list_id": 717398,
                    "code": "TEST-001",
                    "due_on_min": "",
                    "due_on_max": "",
                    "line_items": [
                    ],
                    "tags": [
                        "CN",
                        "CN2"
                    ]
                }
            }

            sup_code = request.GET['brand_id']
            dotnhap = request.GET['dotnhap']

            try:
                this_po = Purchase_Order.objects.get(madonnhap=dotnhap)
            except Purchase_Order.DoesNotExist:
                pass
            else:
                if this_po.ship_to == "Cát Lái" or this_po.ship_to == "HCM" or this_po.ship_to == "Hồ Chí Minh" or this_po.ship_to == "Sài Gòn":   
                    json_order['order_supplier']["location_id"] = 548744
                else:
                    json_order['order_supplier']["location_id"] = 241737

            old_po = js_get_url(f"{MAIN_URL}/order_suppliers.json?page=1&statuses=pending&limit=1&tags="+ dotnhap + "&supplier_ids=" + str(sup_code))
            if old_po["metadata"]["total"] > 0:
                #Đã tồn tại thì chuyển sang chế độ edit.

                old_po = js_get_url(f"{MAIN_URL}/order_suppliers/{old_po['order_suppliers'][0]['id']}.json")['order_supplier']

                url_put = f"{MAIN_URL}/order_suppliers/{old_po['id']}.json"
                method = 'put'
                json_order['order_supplier']["line_items"] = []
                json_order['order_supplier']["supplier_id"] = old_po["supplier_id"]
                json_order['order_supplier']['code'] = old_po["code"]
                json_order['order_supplier']["id"] = old_po["id"]
                json_order['order_supplier']['tags'] = old_po["tags"]
                name_order = json_order['order_supplier']["code"]
                #print(f"Đã tồn tại & chuyển sang edit: {name_order}")
            else:
                #Nếu chưa tồn tại thì tạo mới
                url_post = f"{MAIN_URL}/order_suppliers.json"
                method = 'post'
                
                if dotnhap == "TN":
                    last_this = js_get_url(f"{MAIN_URL}/order_suppliers.json?page=1&limit=1&tags=TN")["order_suppliers"][0]["code"] 
                    json_order['order_supplier']["code"] = "TN-" + str(int(last_this.replace("TN-", "").replace("-2022","")) + 1 ) + "-2022"
                    json_order['order_supplier']["code"] =  "TN-2023-1"
                    json_order['order_supplier']["tags"] = ["TN"]
                else:
                    last_this = js_get_url(f"{MAIN_URL}/order_suppliers.json?page=1&limit=1")["order_suppliers"][0]["code"]
                    if "CN-2025-S" in last_this:
                        json_order['order_supplier']["code"] = "CN-2025-S" + str(int(last_this.replace("CN-2025-S", "")) + 1 )
                    elif "CN-2025-S" in last_this:
                        json_order['order_supplier']["code"] = "CN-2025-S" + str(int(last_this.replace("CN-2025-S", "")) + 1 )    
                    else:
                        json_order['order_supplier']["code"] = "CN-2025-S" + str(int(last_this.replace("OSN000", "")) + 1 )

                    json_order['order_supplier']["tags"] = ["CN", dotnhap]
                    #json_order['order_supplier']["code"] =  "CN-2025-S01"
                    print(json_order['order_supplier']["code"])

                
                json_order['order_supplier']["supplier_id"] = int(sup_code)
                name_order = json_order['order_supplier']["code"]

                #print(f"Chưa tồn tại nên tạo mới PO: {json_order['order_supplier']["code"]}")
            
            obj = json.loads(readfile(f"logs/save_po/{sup_code}.log"))

            for variant in obj["all_vari"]:
                variant['display'] = 1
                if variant["nhap"] > 0:
                    NEW_LINE = {
                        "price": int(float(variant["tq_price"])*3600),
                        "units": [
                            {
                                "variant_id": variant["id"],
                                "name": "chiếc"
                            }
                        ],
                        "tax": {
                            "tax_value": 0,
                            "tax_type": "exclude"
                        },
                        "quantity":  variant["nhap"],
                        "product_id": variant["product_id"],
                        "variant_id": variant["id"],
                        "excluded_tax_begin_amount": 0,
                        "excluded_tax_line_amount": 0
                    }
                    json_order['order_supplier']["line_items"].append(NEW_LINE)
            
            loginss.headers.update({'X-Sapo-Client': 'sapo-frontend-v3'})
            loginss.headers.update({'X-Sapo-ServiceId': 'sapo-frontend-v3'})
            if method == 'post':
                rs = loginss.post(url_post,json=json_order)
            else:
                rs = loginss.put(url_put,json=json_order)
            if len(rs.text) < 250:
                print(rs.text)

        elif action == 'load':
            po_id = request.GET['po_id']

            po = js_get_url(f"{MAIN_URL}/order_suppliers/{po_id}.json")['order_supplier']
            sup_code = po["supplier_id"]
            obj['save_data']['brand_id'] = sup_code
            obj = json.loads(readfile(f"logs/save_po/{sup_code}.log"))

            for variant in obj["all_vari"]:
                variant['display'] = 1
                flag = 0
                for line in po["line_items"]:
                    if variant["id"] == line["variant_id"]:
                        variant["nhap"] = line["quantity"]
                        variant["box_nhap"] = float(variant["nhap"]/variant["tq_box"])
                        flag = 1

                        if variant["nhap"] > 0:
                            variant['data'] = data_list[str(variant['id'])]
                            if variant['data']['fullbox'] == 0:
                                variant['data']['fullbox'] = 1

                            variant["so_khoi_box"] = float((variant['data']['box_cao'] * variant['data']['box_dai'] * variant['data']['box_rong'])/1000000/variant['data']['fullbox'])
                            
                            variant["tq_variant"] = variant['data']['name_tq']
                            variant["tq_price"] = variant['data']['price_tq']
                            variant["tq_box"] = variant['data']['fullbox']
         
                        break
                if flag == 0:
                    variant["nhap"] = 0
                    variant["box_nhap"] = 0

            writejsonfile(obj,f"logs/save_po/{sup_code}.log")
        
        elif action == 'excel':
            po_id = request.GET['po_id']

            po = js_get_url(f"{MAIN_URL}/order_suppliers/{po_id}.json")['order_supplier']
            sup_code = po["supplier_id"]
            obj['save_data']['brand_id'] = sup_code
            obj = json.loads(readfile(f"logs/save_po/{sup_code}.log"))
            for variant in obj["all_vari"]:
                variant["nhap"] = 0
                
            for variant in obj["all_vari"]:
                for line in po["line_items"]:
                    if variant["id"] == line["variant_id"]:
                        variant["nhap"] = line["quantity"]
                        break

            writejsonfile(obj,f"logs/save_po/{sup_code}.log")

            # Create a workbook and add a worksheet.
            workbook = xlsxwriter.Workbook(f"assets/excel_po/{po['supplier_data']['name']}-{po['code']}.xlsx")
            worksheet = workbook.add_worksheet()
            worksheet.set_default_row(60)
            worksheet.set_row(0,20)
            worksheet.set_column('A:A', 5)
            worksheet.set_column('B:B', 40)
            worksheet.set_column('C:C', 15)
            worksheet.set_column('D:D', 20)
            worksheet.set_column('E:E', 20)
            worksheet.set_column('I:I', 15)

            cell_text_wrap = workbook.add_format({'text_wrap': True,'align': 'center','valign': 'vcenter'})
            cell_align = workbook.add_format({'align': 'center','valign': 'vcenter'})
            cell_first = workbook.add_format({'bold': True,'bg_color': "#D8E4BC"})

            row = 0
            col = 0      
            worksheet.write(row, col, "#",cell_first)
            worksheet.write(row, col + 1, "Image",cell_first)
            worksheet.write(row, col + 2, "SKU",cell_first)
            worksheet.write(row, col + 3, "VN_Variant",cell_first)
            worksheet.write(row, col + 4, "名称",cell_first)
            worksheet.write(row, col + 5, "单价",cell_first)
            worksheet.write(row, col + 6, "数量",cell_first)
            worksheet.write(row, col + 7, "箱数",cell_first)
            worksheet.write(row, col + 8, "装箱数",cell_first)
            worksheet.write(row, col + 9, "外箱尺寸",cell_first)
            worksheet.write(row, col + 10, "总体积",cell_first)
            worksheet.write(row, col + 11, "单价",cell_first)
            row = 1
            count = 0
            # Iterate over the data and write it out row by row.
            obj["all_vari"] = sorted(obj["all_vari"], key=lambda x: x['sku'])

            for variant in obj["all_vari"]:
                print(f"Dang xu ly: {variant["sku"]}")

                if variant["nhap"] > 0:
                    variant['data'] = data_list[str(variant['id'])]
                    if variant['data']['fullbox'] == 0:
                        variant['data']['fullbox'] = 1

                    variant["so_khoi_box"] = float((variant['data']['box_cao'] * variant['data']['box_dai'] * variant['data']['box_rong'])/1000000/variant['data']['fullbox'])
                    variant["tq_variant"] = variant['data']['name_tq']
                    variant["tq_price"] = variant['data']['price_tq']
                    variant["tq_box"] = variant['data']['fullbox']
                    variant["box_size"] = str(variant['data']['box_cao']) + " * " + str(variant['data']['box_dai'])+ " * " + str(variant['data']['box_rong']) + " cm" 
                   
                    SO_KHOI = variant["so_khoi_box"]
                            
                    file_path = f"assets/saveimage/{variant['id']}.jpg"
                    variant["image"] = variant["images"][0]["full_path"]
                    r = requests.get(variant["image"], allow_redirects=True)
                    open(file_path, 'wb').write(r.content)
                        
                    worksheet.write(row, col, count,cell_align)

                    # Lấy kích thước của ô
                    cell_width = 75  # Giả sử chiều rộng của ô là 80
                    cell_height = 75  # Giả sử chiều cao của ô là 60

                    # Lấy kích thước của hình ảnh
                    with Image.open(file_path) as img:
                        image_width, image_height = img.size

                    # Tính tỉ lệ để hình ảnh vừa với ô
                    x_scale = float(cell_width) / float(image_width)
                    y_scale = float(cell_height) / float(image_height)

                    # Định dạng hình ảnh để tự động điều chỉnh kích thước
                    image_options = {
                        'x_offset': 5,
                        'y_offset': 5,
                        'x_scale': x_scale,
                        'y_scale': y_scale,
                    }
                    try:
                        tq_price = float(variant["tq_price"])
                    except ValueError:
                        tq_price = 0  # hoặc báo lỗi chi tiết


                    print(f"Gia TQ: {variant["tq_price"]}")

                    worksheet.insert_image(row, col + 1, "assets/saveimage/"+str(variant["id"])+".jpg",image_options)
                    worksheet.set_column(col + 1, col + 2, 12)
                    worksheet.write(row, col + 2, variant["sku"],cell_align)
                    worksheet.write(row, col + 3, variant["opt1"],cell_align)
                    worksheet.write(row, col + 4, variant["tq_variant"],cell_text_wrap)
                    worksheet.write(row, col + 5, tq_price,cell_align)
                    worksheet.write(row, col + 6, variant["nhap"],cell_align)
                    worksheet.write(row, col + 7, round(float(variant["nhap"]/variant["tq_box"]),1),cell_align)
                    worksheet.write(row, col + 8, str(variant["tq_box"])+" pcs/box",cell_align)
                    worksheet.write(row, col + 9, str(variant["box_size"]),cell_align)
                    worksheet.write(row, col + 10, round(float(variant["nhap"]*SO_KHOI*1.05),2),cell_align)
                    worksheet.write(row, col + 11,float(int(variant["nhap"])*tq_price),cell_align)
                    row += 1
                    count+=1
                    
            # Write a total using a formula.
            worksheet.write(row, 5, 'Total')
            worksheet.write(row, 6, '=SUM(G2:G'+ str(row) +')')
            worksheet.write(row, 7, '=SUM(H2:H'+ str(row) +')')
            worksheet.write(row, 9, '=SUM(J2:J'+ str(row) +')')
            worksheet.write(row, 10, '=SUM(K2:K'+ str(row) +')')

            workbook.close()
            obj['url_download'] = f"/static/excel_po/{po['supplier_data']['name']}-{po['code']}.xlsx"
            return render(request, 're_download.html',obj)
        
        elif action == 'nhanthung':
            all_model = all_xnk_data()
            po_id = request.GET['po_id']

            po = js_get_url(f"{MAIN_URL}/order_suppliers/{po_id}.json")['order_supplier']
            sup_code = po["supplier_id"]
            obj['save_data']['brand_id'] = sup_code
            obj = json.loads(readfile(f"logs/save_po/{sup_code}.log"))
            for variant in obj["all_vari"]:
                for line in po["line_items"]:
                    if variant["id"] == line["variant_id"]:
                        variant["nhap"] = line["quantity"]
                        break
            for variant in obj["all_vari"]:
                if variant["nhap"] > 0:
                    variant['data'] = data_list[str(variant['id'])]
                    if variant['data']['fullbox'] == 0:
                        variant['data']['fullbox'] = 1

                    variant["so_khoi_box"] = float((variant['data']['box_cao'] * variant['data']['box_dai'] * variant['data']['box_rong'])/1000000/variant['data']['fullbox'])
                    variant["tq_variant"] = variant['data']['name_tq']
                    variant["tq_price"] = variant['data']['price_tq']
                    variant["box_pcs"] = variant['data']['fullbox']
                    variant["box_info"] = str(variant['data']['box_cao']) + " * " + str(variant['data']['box_dai'])+ " * " + str(variant['data']['box_rong']) + " cm" 

                    count = 0
                    for xmodel in all_model:
                        if xmodel['sku'] == variant['data']['sku_nhapkhau'].strip():
                            model = xmodel
                            count = 1
                            break
                    if count == 0:
                        print(f"{line['sku']} not have data!")
                    else:
                        variant["en_name"] = model["en_name"]
                        variant["nsx_name"] = model["nsx_name"]
                        variant["nsx_address"] = model["nsx_address"]
                        variant["hs_code"] = model["hs_code"]
                        variant["vn_name"] = model["vn_name"]
                        variant["material"] = model["material"]

                    variant["po_code"] = po["code"]
                    variant["po_day"] = datetime.datetime.strptime(po['created_on'], "%Y-%m-%dT%H:%M:%SZ").strftime("%Y-%m-%d")

            writejsonfile(obj,f"logs/save_po/{sup_code}.log")
            obj["all_vari"] = sorted(obj["all_vari"], key=lambda x: x['sku'])
            return render(request, 'nhanthungxnk.html',obj)
        
        elif action == 'nhanthungfix':
            print("Nhan thung fix")
            return render(request, 'nhanthungfix.html')

        elif action == 'view':
            po_id = request.GET['po_id']
            po = js_get_url(f"{MAIN_URL}/order_suppliers/{po_id}.json")['order_supplier']
            sup_code = po["supplier_id"]
            obj['save_data']['brand_id'] = sup_code
            obj = json.loads(readfile(f"logs/save_po/{sup_code}.log"))
            obj['sum'] = {'sum_nhap':0, 'sum_box':0,'sum_money':0,'sum_khoi':0}
            for variant in obj["all_vari"]:
                flag = 0
                for line in po["line_items"]:
                    if variant["id"] == line["variant_id"]:
                        variant["nhap"] = line["quantity"]
                        variant["box_nhap"] = float(variant["nhap"]/variant["tq_box"])
                        flag = 1

                        if int(variant["nhap"]) > 0:
                            variant["image"] = variant["images"][0]["full_path"]
                            r = requests.get(variant["image"], allow_redirects=True)
                            name = 'assets/saveimage/' + str(variant['id']) + ".jpg"
                            if os.path.exists(name):
                                pass
                            else:
                                print('Save image: '+ name)
                                open(name, 'wb').write(r.content)
                            
                            variant['data'] = get_data_variant(variant['id'])
                            
                            variant["so_khoi_box"] = float((variant['data']['box_cao'] * variant['data']['box_dai'] * variant['data']['box_rong'])/1000000/variant['data']['fullbox'])
                            variant["tq_variant"] = variant['data']['name_tq']
                            variant["tq_price"] = variant['data']['price_tq']
                            variant["box_pcs"] = variant['data']['fullbox']
                        break

                variant["sum_khoi"] = round(float(variant["nhap"]*variant["so_khoi_box"]),2)
                variant["sum_tien"] = float(variant["nhap"]*variant["tq_price"])
                
                if flag == 0:
                    variant["nhap"] = 0
                    variant["box_nhap"] = 0
                    
                if variant["nhap"] > 0:
                    obj['sum']['sum_nhap'] += variant["nhap"]
                    obj['sum']['sum_box'] += variant["box_nhap"]
                    obj['sum']['sum_khoi'] += variant["sum_khoi"]
                    obj['sum']['sum_money'] += variant["sum_tien"]

            obj["all_vari"] = sorted(obj["all_vari"], key=lambda x: x['sku'])
            writejsonfile(obj, f"logs/save_po/{sup_code}.log")
            return render(request, 'viewdonnhap.html',obj)

    obj["real_po"] = js_get_url(f"{MAIN_URL}/order_suppliers.json?page=1&limit=20&supplier_ids={obj['save_data']['brand_id']}&statuses=pending")["order_suppliers"]
    
    return render(request, 'edit-test.html',obj)

def updategiavon(request):
    obj = {
        'save_po' : {},
        'all_line': [],
        'thongbao': ''
    }
    #Nhận file lên
    if request.method == 'POST':
        # Đọc file lên.
        action = request.POST['action']
        upload_file = request.FILES['document']
        os.remove("logs/log excel/update_giavon.xlsx")
        fs = FileSystemStorage()
        fs.save('logs/log excel/update_giavon.xlsx',upload_file)
        loc = ("logs/log excel/update_giavon.xlsx")
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(0)     
        sheet.cell_value(0, 0)
        # Thông tin chung
        tokhaihq = ''
        tygiacny = 0
        tygiausd = 0

        for i in range(sheet.nrows):
            array_pr = sheet.row_values(i)
            if i == 1:
                tokhaihq = str(array_pr[6])
                tygiacny = int(array_pr[1])
            if i == 11:
                tygiausd = int(array_pr[1])

            # Update từng dòng 1.
            if i >= 29:
                #Nếu PO chưa có dữ liệu
                if str(array_pr[0]) not in obj['save_po']:
                    this_po = js_get_url(f"https://sisapsan.mysapogo.com/admin/order_suppliers.json?query={str(array_pr[0])}&page=1&limit=20")["order_suppliers"][0]
                    info_po = js_get_url(f"https://sisapsan.mysapogo.com/admin/purchase_orders.json?order_supplier_id={this_po['id']}")["purchase_orders"][0]
                    this_po['receipts'] = info_po['receipts'][0]['code']
                    obj['save_po'][str(array_pr[0])] = this_po
               
                line_info = {'vid': int(array_pr[16]),'pid':0,'p': str(array_pr[0]),'s' : str(array_pr[1]),'gt' : float(array_pr[2]),'s1' : int(array_pr[3]),'s2' : int(array_pr[4]),'tg' : float(array_pr[5]),'tv' : float(array_pr[6]),'tn' : float(array_pr[7]),'pu': 0,'np': int(array_pr[9]),'nq': int(array_pr[14]),'date': str(array_pr[15]),'op': 0,'oq': 0,'li': obj['save_po'][str(array_pr[0])]['location_id'],'rc': obj['save_po'][str(array_pr[0])]['receipts'],}
                # Tìm variant đó đã nè.
                vari = js_get_url(f"https://sisapsan.mysapogo.com/admin/variants/{line_info['vid']}.json")["variant"]
                line_info['pid'] = vari["product_id"]
                # Tìm old quantity
                for i in range(1,100):
                    traces = js_get_url(f"https://sisapsan.mysapogo.com/admin/reports/inventories/variants/{vari['id']}.json?page={i}&limit=250&location_ids={line_info['li']}")["variant_inventories"]
                    if len(traces) == 0:
                        break
                    for trace in traces:
                        if trace['trans_object_code'] != None:
                            if trace['trans_object_code'] == line_info['rc']:
                                line_info['oq'] = int(trace['onhand']) - int(trace['onhand_adj'])
                if line_info['oq'] < 0:
                    line_info['oq'] = 0
                # Tìm old price
                line_info['op'] = line_info['np']
                line_info['pu'] = int((line_info['oq']*line_info['op']+line_info['nq']*line_info['np'])/(line_info['nq']+line_info['oq']))

                obj['all_line'].append(line_info)

        #Ghi file này lên Sapo
        if 'update_for' in request.POST and len(request.POST['update_for']) > 0:
            TRAVE = js_get_url(f"https://sisapsan.mysapogo.com/admin/price_adjustments.json?query={request.POST['update_for']}")

            if len(TRAVE["price_adjustments"]) > 0:
                TRAVE = TRAVE["price_adjustments"][0]
            else:
                #Nếu chưa có thì tạo
                TRAVE = {"location_id":this_po["location_id"],"code":request.POST['update_for'],"tags":[],"note":"","line_items":[]}
                xTRAVE = loginss.post(f"{MAIN_URL}/price_adjustments.json", json=TRAVE)
                TRAVE = json.loads(xTRAVE.text)["price_adjustment"]

        PD_JSON ={"price_adjustment":{"code":TRAVE["code"],"note":"","line_items":[]}}
        
        for line in obj['all_line']:
            line_string = json.dumps(line).replace(" ", "", 2000)
            LINE_JSON = {"product_id":line['pid'],"variant_id":line['vid'],"note":"X","price":int(line['pu']),"product_type":"normal"}
            LINE_JSON['note'] = line_string
            PD_JSON["price_adjustment"]["line_items"].append(LINE_JSON)
       
        rs = loginss.put(f"https://sisapsan.mysapogo.com/admin/price_adjustments/{TRAVE['id']}.json", json=PD_JSON)
        if len(rs.text) < 200:
            obj['thongbao'] = f"Tác vụ đã thất bại. Lỗi: {rs.text}"
        else:
            obj['thongbao'] = f"Tác vụ thành công! Cảm ơn bạn đã cập nhật giá vốn cho sản phẩm."

    return render(request, 'updategiavon.html',obj)

def ptgiavon(request):
    obj = {
        'shop': {},
        'sum': {'now':{'sumds':0,'sum_gv':0, 'tile':0},'pre':{'sumds':0,'sum_gv':0, 'tile':0}},
        'toky': {'now':{'sumds':0,'sum_gv':0, 'tile':0},'pre':{'sumds':0,'sum_gv':0, 'tile':0}},
        'gele': {'now':{'sumds':0,'sum_gv':0, 'tile':0},'pre':{'sumds':0,'sum_gv':0, 'tile':0}},
        'orders': []
    }
    LGV = get_list_giavon()

    today = datetime.datetime.today()
    end_day = today.replace(hour=23, minute=59, second=59, microsecond=999999)

    if 'time' in request.GET:
        time = request.GET['time']
        if time == 'today':
            ss_day = today
        elif time == 'yes':
            ss_day = today - datetime.timedelta(days=1)
            end_day = ss_day
        else:
            ss_day = datetime.datetime.strptime(unquote(request.GET['time']), "%d/%m/%Y")
            end_day = datetime.datetime.strptime(unquote(request.GET['end_time']), "%d/%m/%Y")
        
        print(f'[+] Report gói hàng: Ngày {ss_day} đến {end_day}!')

        delta_days = (end_day - ss_day).days
        start_cungky = ss_day - datetime.timedelta(days=delta_days) - datetime.timedelta(days=1)
        start_cungky = start_cungky.replace(hour=0, minute=0, second=0, microsecond=0)
        end_cungky = ss_day - datetime.timedelta(days=1)
        end_cungky = end_cungky.replace(hour=23, minute=59, second=59, microsecond=999999)


        for i in range(1,500):
            URL = f"{MAIN_URL}/orders.json?status=finalized%2Ccompleted%2Cdraft&created_on_max={str(end_day.strftime('%Y-%m-%d'))}T16%3A59%3A59Z&created_on_min={str((ss_day - datetime.timedelta(days=1)).strftime('%Y-%m-%d'))}T17%3A00%3A00Z&limit=250&page={i}"
            orders = js_get_url(URL)["orders"]
            if len(orders) == 0:
                break
            for order in orders:
                if order['source_id'] != 7239422:
                    giavonx = 0
                    date_start = datetime.datetime.strptime('2023-12-18', "%Y-%m-%d")
                    # Xử lý date
                    created_on = datetime.datetime.strptime(order['created_on'], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(7)
                    if created_on < date_start:
                        break
                    #KHO TÔ KÝ
                    # Tách làm 2 kho gele và toky
                    if order['location_id'] == 241737:
                        kho = 'gele'
                    else:
                        kho = 'toky'
                    
                    if len(order['tags']) >= 2 and "_" in order['tags'][1]:
                        SHOP = order['tags'][1].split("_")[1]
                    else:
                        SHOP = "SAPO"

                    if SHOP not in obj['shop']:
                        obj['shop'][SHOP] = {'now':{'sumds':0,'sum_gv':0, 'tile':0},'pre':{'sumds':0,'sum_gv':0, 'tile':0}}

                    obj[kho]['now']['sumds'] += int(order['total'])
                    obj['shop'][SHOP]['now']['sumds'] += int(order['total'])
                    obj['sum']['now']['sumds'] += int(order['total'])
                        
                    # Tìm giá vốn và cộng vào
                    for line in order["order_line_items"]:
                      
                        giavon = find_giavon(line['variant_id'], created_on, order['location_id'], LGV,(line['line_amount']/line["quantity"] - line['distributed_discount_amount']/line["quantity"]))
                        
                        obj[kho]['now']['sum_gv'] += giavon * line['quantity']
                        obj['shop'][SHOP]['now']['sum_gv'] += giavon * line['quantity']
                        obj['sum']['now']['sum_gv'] += giavon * line['quantity']
                        giavonx += giavon * line['quantity']

                    if order['total'] == 0:
                        order['error'] = 0
                        obj["orders"].append(order)
                    else:
                        tile = (giavonx/order['total'])*100
                        if tile >= 75:
                            order['error'] = 1
                            obj["orders"].append(order)
                        if tile <= 40:
                            order['error'] = 2
                            obj["orders"].append(order)
                        order['tile'] = tile
                        order['giavonx'] = giavonx

        obj['sum']['now']['tile'] = (obj['sum']['now']['sum_gv'] / obj['sum']['now']['sumds'])*100
        obj['gele']['now']['tile'] = (obj['gele']['now']['sum_gv'] / obj['gele']['now']['sumds'])*100
        obj['toky']['now']['tile'] = (obj['toky']['now']['sum_gv'] / obj['toky']['now']['sumds'])*100

        for shop, data in obj['shop'].items():
            sumds_shop_now = data['now']['sumds']
            sumgv_shop_now = data['now']['sum_gv']
            if sumds_shop_now != 0:
                obj['shop'][shop]['now']['tile'] = (sumgv_shop_now / sumds_shop_now) * 100
            else:
                obj['shop'][shop]['now']['tile'] = 0

        convert_int_in_obj(obj)

    return render(request, 'ptgiavon.html', obj)

def ptsanpham(request):
    LGV = get_list_giavon()

    today = datetime.datetime.today()
    end_day = today.replace(hour=23, minute=59, second=59, microsecond=999999)

    if 'time' in request.GET:
        time = request.GET['time']
        if time == 'week':
            ss_day = today  - datetime.timedelta(days=7)
            end_day = today
        elif time == 'month':
            ss_day = today
            end_day = ss_day  - datetime.timedelta(days=30)
        else:
            ss_day = datetime.datetime.strptime(unquote(request.GET['time']), "%d/%m/%Y")
            end_day = datetime.datetime.strptime(unquote(request.GET['end_time']), "%d/%m/%Y")
        
        print(f'[+] Report sản phẩm: Ngày {ss_day} đến {end_day}!')

        delta_days = (end_day - ss_day).days
        start_cungky = ss_day - datetime.timedelta(days=delta_days) - datetime.timedelta(days=1)
        start_cungky = start_cungky.replace(hour=0, minute=0, second=0, microsecond=0)
        end_cungky = ss_day - datetime.timedelta(days=1)
        end_cungky = end_cungky.replace(hour=23, minute=59, second=59, microsecond=999999)
    else:
        end_day = today
        ss_day = end_day  - datetime.timedelta(days=30)       

    if 'brand_id' in request.GET:
        brand_id = request.GET['brand_id']
    else:
        brand_id = ''
    
    if 'suggest' in request.GET:
        suggest = request.GET['suggest']
    else:
        suggest = 0.45

    all_product = []
    brand = js_get_url(f"{MAIN_URL}/brands/{brand_id}.json")['brands']['name']
    for i in range(1,3):
        all_vari = js_get_url(f"{MAIN_URL}/variants.json?limit=250&packsize=false&product_type=normal&page={i}&status=active&brand_ids={brand_id}")
        if len(all_vari["variants"]) > 0:
            for vari in all_vari["variants"]:
                print(f"Dang xu ly: {vari['sku']}")
                vari["total_quantity"] = 0
                vari["total_price"] = 0
                vari["total_giavon"] = 0

                vari['data'] = get_data_variant(vari['id'])
                vari['image'] = vari['images'][0]['full_path']
                # Tìm giá vốn
                vari['giavon_sg'] = find_giavon(vari['id'], today,548744, LGV,0)
                vari['giavon_hn'] = find_giavon(vari['id'], today,241737, LGV,0)

                #Tồn kho
                vari['kho_sg'] = int(vari['inventories'][1]['on_hand'])
                vari['kho_hn'] = int(vari['inventories'][0]['on_hand'])
                vari['kho_sum'] = int(vari['kho_sg']+vari['kho_hn'])

                if (vari['kho_sg']+vari['kho_hn']) > 0:
                    vari['giavon_now'] = int((vari['giavon_sg']*vari['kho_sg']+vari['giavon_hn']*vari['kho_hn'])/(vari['kho_sg']+vari['kho_hn']))
                else:
                    vari['giavon_now'] = int(vari['giavon_sg'])

                daily_sales_current = {}
                daily_sales_previous = {}

                #for i in range(1,1):
                all_orders = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=250&status=draft%2Cfinalized%2Ccompleted&variant_ids={vari['id']}&created_on_max={str(end_day.strftime('%Y-%m-%d'))}T16%3A59%3A59Z&created_on_min={str((ss_day - datetime.timedelta(days=1)).strftime('%Y-%m-%d'))}T17%3A00%3A00Z")
                
                for order in all_orders["orders"]:
                    if order['source_id'] != 7239422:                    
                        time_order = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ")
                        day_from_now = abs((today - time_order).days)

                        for line in order['order_line_items']:
                            if line['variant_id'] == vari["id"]:
                                vari["total_quantity"] += int(line['quantity'])
                                vari["total_price"] += int((line['line_amount']-line['distributed_discount_amount']))
                                giavon = find_giavon(line['variant_id'], time_order, order['location_id'], LGV,(line['line_amount']/line["quantity"] - line['distributed_discount_amount']/line["quantity"]))
                                
                                vari["total_giavon"] += giavon*int(line['quantity'])
                    if vari["total_quantity"] > 0:
                        vari['tb_price'] = int(vari["total_price"]/ vari["total_quantity"])
                        if vari["total_price"] == 0:
                            vari["total_price"] = 1
                        vari["tb_giavon"] = int(vari["total_giavon"]/vari["total_quantity"])
                        vari["ratio_giavon"] = (vari["total_giavon"]/vari["total_price"])*100

                        vari['duban'] = vari['kho_sum']/(vari["total_quantity"]/30)

                        vari['sg_gia'] = int(vari["tb_giavon"]/suggest)


                all_product.append(vari)
        else:
            break

    all_product.sort(key=lambda x: x['product_id'])
    all_brand = json.loads(loginss.get(f"{MAIN_URL}/brands.json?page=1&limit=250").text)["brands"]
    all_brand  = loc_brand(all_brand)
    obj = {
        'products': all_product,
        'brands': all_brand,
        'urls': {
            'action': 'action',
            'brand_id': request.GET['brand_id'] if 'brand_id' in request.GET else '',
            'search': request.GET['search'] if 'search' in request.GET else '',
            }
        }
    return render(request, 'ptsanpham.html', obj)

def flash_sale(request):
    obj = {
        'save': {'connect_id':'',},
        'products': []

    }

    connect_id = ''
    if 'connect_id' in request.GET:
        #Input vào tên của khách hàng.
        connect_id = request.GET['connect_id']
        obj["save"]["connect_id"] = connect_id
        request.session['connect_id'] = connect_id
    if 'connect_id' in request.session:
        obj['save']['connect_id'] = request.session['connect_id']
        connect_id = request.session['connect_id']
    else:
        connect_id = 134366
    shop_name = existing_shop_map[int(connect_id)]

    #Cần thanh lý

    ALL_OR_THANH_LY = js_get_url(f"https://sisapsan.mysapogo.com/admin/products.json?tags=THANH+L%C3%9D&page=1&limit=100")["products"]
    ALL_ID_THANH_LY = []
    for xorder in ALL_OR_THANH_LY:
        ALL_ID_THANH_LY.append(xorder["id"])

    # Lấy tất cả sản phẩm
    for i in range(1,20):
        URL = f"https://market-place.sapoapps.vn/products/v2/filter?page={i}&tenantId=1262&mappingStatus=0&syncStatus=0&connectionIds={connectID}"
        text = js_get_url(URL)
        if "products" in text:
            products = text["products"]
            for product in products:
                product["flag_tl"] = 0
                for vari in product["variants"]:
                    if vari["sapo_product_id"] in ALL_ID_THANH_LY:
                        product["flag_tl"] = 1
                    
                obj['products'].append(product)
        else:
            break

    # Nếu cần xuất file
    if 'xuatfile' in request.GET:
        product_id = request.GET['product_id']
        product_ids = product_id.split(",")

        all_product = []

        loc = (f"logs/flash_sale/{shop_name}.xlsx")
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(0)     
        sheet.cell_value(0, 0)
        for i in range(sheet.nrows):
            if i > 0:
                array_pr = sheet.row_values(i)
                data_form = {'name': array_pr[0],'item_id': array_pr[1],  'variation_id': array_pr[2], 'price':array_pr[6],'flash_sale_price':array_pr[7] }
                all_product.append(data_form)
        
        xuatfile = []    
        for ids in product_ids:
            for product in all_product:
                if ids == str(product['item_id']):
                    xuatfile.append(product)

        workbook = xlsxwriter.Workbook(f'logs/flash_sale/flash_sale_total.xlsx')
        worksheet = workbook.add_worksheet()
        # Header cho các cột
        headers = ['item_id','variation_id', 'flash_sale_price']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        # Viết dữ liệu từ all_product vào file Excel
        for row, product in enumerate(xuatfile, start=1):
            for col, key in enumerate(headers):
                worksheet.write(row, col, product.get(key, ''))

        # Header cho các cột
        headers = ['Mã sản phẩm','Mã phân loại hàng', 'Giá đã giảm']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)
            
        workbook.close()


    if 'update' in request.GET:
        # Tạo workbook và worksheet
        workbook = xlsxwriter.Workbook(f'logs/flash_sale/{shop_name}-new.xlsx')
        worksheet = workbook.add_worksheet()
        # Header cho các cột
        headers = ['name','item_id', 'variation_id','sapo_variant_id', 'sapo_variant_sku','vari_name', 'price','flash_sale_price']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        all_vari = []
        for product in obj['products']:
            for vari in product["variants"]:
                match = re.search(r"Phân loại\s*(.*)", vari["name"])
                if match:
                    vari["vari_name"] = match.group(1)
                else:
                    vari["vari_name"] = ""

                all_vari.append({
                    'name': vari['name'],
                    'vari_name': vari['vari_name'],
                    'item_id': vari['item_id'],
                    'variation_id':vari['variation_id'],
                    'sapo_variant_id':vari['sapo_variant_id'],
                    'sapo_variant_sku':vari['sapo_variant_sku'],
                    'price':vari['price'],
                    'flash_sale_price':0

                })

        all_vari.sort(key=lambda x: x['item_id'])
        # Viết dữ liệu từ all_product vào file Excel
        for row, product in enumerate(all_vari, start=1):
            for col, key in enumerate(headers):
                worksheet.write(row, col, product.get(key, ''))
        
        # Đóng workbook
        workbook.close()


    return render(request, 'flashsale.html', obj)
# PHẦN KẾ TOÁN

def ketoan_doisoat(request):
    obj = {
        'all_notify': []
    }
  
    ngay_thang_nam = datetime.datetime.now()
    ngay = ngay_thang_nam.day
    thang = ngay_thang_nam.month
    nam = ngay_thang_nam.year

    if request.method == "GET" and 'autodoisoat' in request.GET:
        print("Doi soat tu dong")
        all_json_success = []
        DS_NAME = "PDS_AUTOTIKI_0902"
        GHTK_JSON = {
            "delivery_collation": {
                "delivery_service_provider_id": 145496,
                "location_id": 241737,
                "code": DS_NAME,
                "created_on": "",
                "line_items": [],
                "account_id": 319911,
                "status": "collating"
            }
        }
        for i in [1,2,3]:
            all_json = js_get_url(f"{MAIN_URL}/shipments.json?limit=250&location_ids=241737&composite_fulfillment_statuses=received%2Cfulfilled_cancelling%2Cfulfilled_cancelled&delivery_service_provider_ids=145496&collation_statuses=unresolved&page={i}")
            for fun_json in all_json["fulfillments"]:
                line_item = {
                    "tenant_id": 236571,
                    "account_id": 828281,
                    "fulfillment_id": fun_json["id"],
                    "code_shipment": fun_json['shipment']['tracking_code'],
                    "code_order": "",
                    "system_cod_amount": fun_json["shipment"]["cod_amount"],
                    "system_ship_fee": fun_json["shipment"]["freight_amount"],
                    "service_provider_cod_amount": fun_json["shipment"]["cod_amount"],
                    "service_provider_ship_fee": fun_json["shipment"]["freight_amount"],
                    "external_cod_amount": fun_json["shipment"]["cod_amount"],
                    "freight_payer": "shop",
                    "code": fun_json['shipment']['tracking_code'],
                    "order_id": fun_json['order_id'],
                    "created_on": fun_json['created_on'],
                    "fulfillment_status": "received"
                }
                GHTK_JSON["delivery_collation"]["line_items"].append(line_item)
                all_json_success.append(fun_json["order_id"])


        if len(GHTK_JSON["delivery_collation"]["line_items"]) > 0:
            url_search = f"{MAIN_URL}/delivery_collations.json"
            loginss.headers.update({'X-Sapo-LocationId':'241737'})
            headers = {'Content-type': 'application/json;charset=UTF-8', 'Accept': 'application/json, text/plain, */*', 'Accept-Encoding': 'gzip, deflate, br'}
            rs = loginss.post(url_search,json=GHTK_JSON, headers=headers)
            print(rs.text)
            if len(rs.text) > 300: 
                print("TẠO THÀNH CÔNG PHIẾU ĐỐI SOÁT: "+ DS_NAME)
                return render(request, 'ketoan_doisoat.html', {'error':200, 'notify': DS_NAME})
            else:
                return render(request, 'ketoan_doisoat.html', {'error':444, 'notify': rs.text})
        else:
            return render(request, 'ketoan_doisoat.html', {'error':444, 'notify': 'Không có đơn hàng nào hợp lệ.'})
    
    # ĐỐI SOÁT GIAO HÀNG TIẾT KIỆM
    if request.method == "POST" and 'ghtk' in request.POST["doitac"]:
        upload_file = request.FILES['document']
        os.remove("logs/input_doisoat.xlsx")
        fs = FileSystemStorage()
        fs.save('logs/input_doisoat.xlsx',upload_file)
        loc = ("logs/input_doisoat.xlsx")
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(0)     
        sheet.cell_value(0, 0)
        all_json = js_get_url(f"{MAIN_URL}/delivery_collations.json?delivery_service_provider_ids=69025")
        DS_NAME = "GHTK_" + str(ngay) +"_"+str(thang)+"_" + str(nam) + "_" +str(random.randint(100, 999))
        GHTK_JSON = {
                "delivery_collation": {
                    "delivery_service_provider_id": 69025,
                    "location_id": 241737,
                    "code": DS_NAME,
                    "created_on": "",
                    "line_items": [],
                    "account_id": 319911,
                    "status": "collating"
                }
        }

        COUNT = 1
        for i in range(sheet.nrows):
            if i > 0:
                array_pr = sheet.row_values(i)
                if len(array_pr[1]) > 15 and len(array_pr[1]) < 45:
                    
                    if abs(int(array_pr[9])) > 0:
                        MA_DON_HANG = str(array_pr[1])
                        TONG_PHI = int(array_pr[9]) + int(array_pr[10]) + int(array_pr[11]) + int(array_pr[12]) + int(array_pr[13]) + int(array_pr[14]) + int(array_pr[15]) + int(array_pr[16]) + int(array_pr[17]) + int(array_pr[18]) + int(array_pr[19]) + int(array_pr[20])+ int(array_pr[21])+ int(array_pr[22])+ int(array_pr[23])+ int(array_pr[24]) + int(array_pr[25])
                        COD = int(array_pr[8])
                        fun_json = js_get_url(f"{MAIN_URL}/fulfillments.json?composite_fulfillment_statuses=received%2cfulfilled_cancelled&query={MA_DON_HANG}")
                        if fun_json["metadata"]["total"]  > 0:
                            fun_json = fun_json["fulfillments"][0]
                            if fun_json["shipment"]["collation_status"] == "unresolved":
                                line_item = {
                                    "tenant_id": 236571,
                                    "account_id": 828281,
                                    "fulfillment_id": fun_json["id"],
                                    "code_shipment": fun_json['shipment']['tracking_code'],
                                    "code_order": "",
                                    "system_cod_amount": fun_json["shipment"]["cod_amount"],
                                    "system_ship_fee": fun_json["shipment"]["freight_amount"],
                                    "service_provider_cod_amount": COD,
                                    "service_provider_ship_fee": abs(TONG_PHI),
                                    "external_cod_amount": COD,
                                    "freight_payer": "shop",
                                    "code": fun_json['shipment']['tracking_code'],
                                    "order_id": fun_json['order_id'],
                                    "created_on": fun_json['created_on'],
                                    "fulfillment_status": "received"
                                }
                                GHTK_JSON["delivery_collation"]["line_items"].append(line_item) 
                               
                                obj['all_notify'].append({'error':0, 'tb': f"{COUNT}# {MA_DON_HANG} được thêm vào phiếu đối soát!"})
                            else:
                                obj['all_notify'].append({'error':1, 'tb': f"{COUNT}# {MA_DON_HANG} đã có trong phiếu đối soát khác rồi!"})
                        else:
                            obj['all_notify'].append({'error':1, 'tb': f"{COUNT}# {MA_DON_HANG} không tồn tại trên Sapo!"})
                                
                    else:
                        obj['all_notify'].append({'error':1, 'tb': f"{COUNT}# {MA_DON_HANG} không có phí giao hàng, đơn hàng lỗi!"})
                COUNT += 1
        if len(GHTK_JSON["delivery_collation"]["line_items"]) > 0:
            url_search = "{MAIN_URL}/delivery_collations.json"
            loginss.headers.update({'X-Sapo-LocationId':'241737'})
            headers = {'Content-type': 'application/json;charset=UTF-8', 'Accept': 'application/json, text/plain, */*', 'Accept-Encoding': 'gzip, deflate, br'}
            rs = loginss.post(url_search,json=GHTK_JSON, headers=headers)
            if len(rs.text) > 400:
                obj['all_notify'].append({'error':3, 'tb': f"Đã tạo thành công phiếu đối soát: {DS_NAME}"})
            else:
                obj['all_notify'].append({'error':2, 'tb': f"Tạo phiếu đối soát: {DS_NAME} thất bại!"})
                obj['all_notify'].append({'error':2, 'tb': f"Lỗi: {rs.text}"})
                
        else:
            obj['all_notify'].append({'error':2, 'tb': f"Không có đơn hàng nào hợp lệ!"})
        
        return render(request, 'ketoan_doisoat.html', obj)

    # ĐỐI SOÁT SHOPEE
    elif request.method == "POST" and 'shopee' in request.POST["doitac"]:
        upload_file = request.FILES['document']
        FILE_UPLOAD = upload_file.name

        os.remove("logs/input_doisoat.xlsx")
        fs = FileSystemStorage()
        fs.save('logs/input_doisoat.xlsx',upload_file)
        loc = ("logs/input_doisoat.xlsx")
        wb = xlrd.open_workbook(loc)
        sheet = wb.sheet_by_index(0)     
        sheet.cell_value(0, 0)

        url_search = js_get_url(f"{MAIN_URL}/delivery_collations.json?delivery_service_provider_ids=128394")
        DS_NAME = "SHOPEE_" + str(ngay) +"_"+str(thang)+"_" + str(nam) + "_" +str(random.randint(100, 999))
        DS_NAME_TOKY = "SHOPEE_" + str(ngay) +"_"+str(thang)+"_" + str(nam) + "_" +str(random.randint(100, 999))
        
        SHOPEE_JSON_GELE = {
            "delivery_collation": {
                "delivery_service_provider_id": 128394,
                "location_id": 241737,
                "code": DS_NAME,
                "created_on": "",
                "line_items": [],
                "account_id": 319911,
                "status": "collating",
                "note": FILE_UPLOAD
            }
        }
        SHOPEE_JSON_TOKY = {
            "delivery_collation": {
                "delivery_service_provider_id": 128394,
                "location_id": 548744,
                "code": DS_NAME_TOKY,
                "created_on": "",
                "line_items": [],
                "account_id": 319911,
                "status": "collating",
                "note": FILE_UPLOAD
            }
        }

        
        columns = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
        # Chuyển đổi dữ liệu thành JSON
        data_json = []
        for row_idx in range(1, sheet.nrows):  # Bắt đầu từ hàng 1, bỏ qua tiêu đề
            row_data = {columns[col_idx]: sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)}
            data_json.append(row_data)

        # In kết quả
        COUNT = 0
        for item in data_json:
            COUNT += 1
            MA_DON_HANG = item["Mã đơn hàng"]
            sum_vc = item["Phí vận chuyển Người mua trả"] + item ["Phí vận chuyển được trợ giá từ Shopee"] + item["Phí vận chuyển thực tế"]
            sum_trahang = item["Phí trả hàng"] + item["Giảm trừ phí dịch vụ Shipping Fee Saver (Trả Hàng)"]
            sum_shopee = item["Phí hoa hồng Tiếp thị liên kết"] + item["Phí cố định"] + item["Phí Dịch Vụ"] + item["Phí thanh toán"]

            sum_fee = abs(int(sum_vc+sum_trahang+sum_shopee))

            sum_cod = int(item['Giá gốc']) +int(item['Số tiền bạn trợ giá cho sản phẩm'])+int(item['Số tiền hoàn lại'])+int(item['Sản phẩm được trợ giá từ Shopee']) + int(item['Mã giảm giá'])+int(item['Người Bán Hoàn Xu'])

            fun_json = js_get_url(f"{MAIN_URL}/fulfillments.json?composite_fulfillment_statuses=received%2cfulfilled_cancelled&query={MA_DON_HANG}")
            
            if fun_json["metadata"]["total"]  > 0:
                fun_json = fun_json["fulfillments"][0]
                if fun_json["shipment"]["collation_status"] == "unresolved":
                    line_item = {
                        "tenant_id": fun_json["tenant_id"],
                        "account_id": 828281,
                        "fulfillment_id": fun_json["id"],
                        "code_shipment": fun_json['shipment']['tracking_code'],
                        "code_order": "",
                        "system_cod_amount": fun_json["shipment"]["cod_amount"],
                        "system_ship_fee": fun_json["shipment"]["freight_amount"],
                        "service_provider_cod_amount": sum_cod,
                        "service_provider_ship_fee": sum_fee,
                        "external_cod_amount": sum_cod,
                        "freight_payer": "shop",
                        "code": fun_json['shipment']['tracking_code'],
                        "order_id": fun_json['order_id'],
                        "created_on": fun_json['created_on'],
                        "fulfillment_status": "received"
                    }
                    if fun_json['stock_location_id'] == 241737:
                        SHOPEE_JSON_GELE["delivery_collation"]["line_items"].append(line_item)
                    else:
                        SHOPEE_JSON_TOKY["delivery_collation"]["line_items"].append(line_item)

                    obj['all_notify'].append({'error':0, 'tb': f"{COUNT}# {MA_DON_HANG} được thêm vào phiếu đối soát!"})
                else:
                    obj['all_notify'].append({'error':1, 'tb': f"{COUNT}# {MA_DON_HANG} đã có trong phiếu đối soát khác rồi!"})

        if len(SHOPEE_JSON_GELE["delivery_collation"]["line_items"]) > 0:
            url_search = f"{MAIN_URL}/delivery_collations.json"
            loginss.headers.update({'X-Sapo-LocationId':'241737'})
            headers = {'Content-type': 'application/json;charset=UTF-8', 'Accept': 'application/json, text/plain, */*', 'Accept-Encoding': 'gzip, deflate, br'}
            rs = loginss.post(url_search,json=SHOPEE_JSON_GELE, headers=headers)
            if len(rs.text) > 400:
                obj['all_notify'].append({'error':3, 'tb': f"Đã tạo thành công phiếu đối soát: {DS_NAME}"})
            else:
                obj['all_notify'].append({'error':2, 'tb': f"Tạo phiếu đối soát: {DS_NAME} thất bại!"})
                obj['all_notify'].append({'error':2, 'tb': f"Lỗi: {rs.text}"})

        if len(SHOPEE_JSON_TOKY["delivery_collation"]["line_items"]) > 0:
            url_search = f"{MAIN_URL}/delivery_collations.json"
            loginss.headers.update({'X-Sapo-LocationId':'548744'})
            headers = {'Content-type': 'application/json;charset=UTF-8', 'Accept': 'application/json, text/plain, */*', 'Accept-Encoding': 'gzip, deflate, br'}
            rs = loginss.post(url_search,json=SHOPEE_JSON_TOKY, headers=headers)
            if len(rs.text) > 400:
                obj['all_notify'].append({'error':3, 'tb': f"Đã tạo thành công phiếu đối soát: {DS_NAME_TOKY}"})
            else:
                obj['all_notify'].append({'error':2, 'tb': f"Tạo phiếu đối soát: {DS_NAME_TOKY} thất bại!"})
                obj['all_notify'].append({'error':2, 'tb': f"Lỗi: {rs.text}"})

    return render(request, 'ketoan_doisoat.html', obj)

def ketoan_1mdonhang(request):
    obj = {
        "orders": [],
        "products": []
    }
    source_order = {}
    all_order = js_get_url(f"{MAIN_URL}/order_sources.json?page=1&limit=250")["order_sources"]
    for source in all_order:
        source_order[source["id"]] = source["name"] 

    all_order = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=1&created_on_max=2023-05-22T16%3A59%3A59Z&created_on_min=2022-12-31T17%3A00%3A00Z")
    ROUND = math.ceil(all_order["metadata"]["total"]/250)
    for i in range(ROUND):
        orders = js_get_url(f"{MAIN_URL}/orders.json?page={str(i+1)}&limit=250&created_on_max=2023-05-22T16%3A59%3A59Z&created_on_min=2022-12-31T17%3A00%3A00Z")["orders"]
        for order in orders:
            order["link_code"] = order['code']
            order["source_name"] = source_order[order["source_id"]]
            MIEN_NAM =  "TP. Hồ Chí Minh, TP Hồ Chí Minh,Bình Phước,Bà Rịa - Vũng Tàu, Tây Ninh, Bà Rịa Vũng Tàu,Bà Rịa-Vũng Tàu, Bình Dương, Đồng Nai,Long An, Đồng Tháp, Tiền Giang, An Giang, Bến Tre, Vĩnh Long, Trà Vinh, Hậu Giang, Kiên Giang, Sóc Trăng, Bạc Liêu, Cà Mau, Cần Thơ,  Kon Tum, Gia Lai, Đắk Lắk, Đắk Nông và Lâm Đồng, Bình Định, Phú Yên, Khánh Hòa, Ninh Thuận và Bình Thuận"
            MIEN_TRUNG = "Thanh Hóa, Nghệ An, Hà Tĩnh, Quảng Bình, Quảng Trị, Thừa Thiên Huế, Đà Nẵng, Quảng Nam, Quảng Ngãi"
            MIEN_BAC = "Hà Nội, Hải Phòng, Tỉnh Lào Cai, Yên Bái, Điện Biên, Hòa Bình, Lai Châu, Sơn La; Tỉnh Hà Giang, Cao Bằng, Bắc Kạn, Lạng Sơn, Tuyên Quang, Thái Nguyên, Phú Thọ, Bắc Giang, Quảng Ninh; Tỉnh Bắc Ninh, Hà Nam, Hải Dương, Hưng Yên, Nam Định, Ninh Bình, Thái Bình, Vĩnh Phúc."

            # Mã liên kết -> liên kết với bên đối soát vận đơn
            if order['channel'] != None and order['reference_number'] != None:
                #Shopee
                order["link_code"] = order['reference_number']
            else:
                if(len(order['fulfillments']) > 0):
                    if order['fulfillments'][-1]["shipment"] != None:
                        order["link_code"] = order["fulfillments"][-1]["shipment"]["tracking_code"]
                else:
                    order["link_code"] = order['code']
            
            # Shop bán hàng
            if len(order['tags']) >= 2:
                order['shop'] = order['tags'][1]
            # Xử lý địa chỉ
            if order['shipping_address'] != None :
                if order['shipping_address']['city'] == None:
                    order['shipping_address']['city'] = "Hà Nội"

                if order['shipping_address']['city'] in MIEN_NAM:
                    VUNG_MIEN = "Miền Nam"
                elif order['shipping_address']['city'] in MIEN_TRUNG:
                    VUNG_MIEN = "Miền Trung"
                elif order['shipping_address']['city'] in MIEN_BAC:
                    VUNG_MIEN = "Miền Bắc"
                order['city'] = order['shipping_address']['city']
                order['district'] = order['shipping_address']['district']
                order['ward'] = order['shipping_address']['ward']
            else:
                VUNG_MIEN = "Miền Bắc"
                order['city'] = "Hà Nội"
                order['district'] = "Quận Hà Đông"
                order['ward'] = " "
            
            order["billing_address"] = None
            order["shipping_address"] = None
            #Xử lý t.t khách hàng
            order['customer_code'] = order['customer_data']['code']
            order['customer_data'] = None

            #Khách hàng đã quay lại hay chưa ?
            order['real_items'] = []
            for item in order['order_line_items']:
                if item['product_type'] == "normal" and item['is_packsize'] == False:
                    flag = 0
                    for line in order['real_items']:
                        if line['id'] == item['variant_id']:
                            line['quantity'] += int(item['quantity'])
                            flag = 1
                            break
                    if flag == 0:
                        order['real_items'].append({
                            'id': item['variant_id'],
                            'sku': item['sku'][3:],
                            'variant_options': item['variant_options'],  
                            'quantity': int(item['quantity']),
                            'unit': item['unit']
                        })
                elif item['product_type'] == "normal" and item['is_packsize'] == True:
                    flag = 0
                    for line in order['real_items']:
                        if line['id'] == item['pack_size_root_id']:
                            line['quantity'] += int(item['quantity']*item['pack_size_quantity'])
                            flag = 1
                            break
                    if flag == 0:
                        vari = js_get_url(f"{MAIN_URL}/variants/{item['pack_size_root_id']}.json")['variant']
                        order['real_items'].append({
                            'id': vari['id'],
                            'sku': vari['sku'][3:],
                            'variant_options': vari['opt1'],
                            'unit': 'chiếc',
                            'quantity': int(item['quantity']*item['pack_size_quantity'])
                        })

                if item['product_type'] == "composite":
                    for xitem in item['composite_item_domains']:
                        flag = 0
                        for line in order['real_items']:
                            if line['id'] == xitem['variant_id']:
                                line['quantity'] += int(xitem['quantity'])
                                flag = 1
                                break
                        if flag == 0:
                            vari = js_get_url(f"{MAIN_URL}/variants/{xitem['variant_id']}.json")['variant']
                            order['real_items'].append({
                                'id': xitem['variant_id'],
                                'sku': vari['sku'][3:],
                                'unit': 'chiếc',
                                'variant_options': vari['opt1'],  
                                'quantity': int(xitem['quantity'])
                            })
                if item['product_type'] == None:
                    order['real_items'].append({
                        'id': 0,
                        'sku': "SERVICE",
                        'unit': item['unit'],
                        'variant_options': item['note'],  
                        'quantity': int(item['quantity'])
                    })   
                    
            order['total_quantity'] = 0
            for line in order['real_items']:
                line['order'] = order['code']
                order['total_quantity'] += line['quantity']
                obj['products'].append(line)

            # Xử lý sản phẩm trong đơn
            obj["orders"].append(order)

    writejsonfile(obj['orders'],'logs/power-bi/orders/orders.json')
    writejsonfile(obj['products'],'logs/power-bi/product_orders/product_orders.json')
    
    return render(request, 'ketoan_1mdonhang.html', obj)

# PHẦN KHO HÀNG

def kho_sos(request):
    # TÍNH TỔNG DOANH THU & TỔNG GIÁ VỐN TRÊN SAPO TEST
    LGV = {}
    # ĐẦU TIÊN -> PHẢI LẤY GIÁ VỐN CÁI ĐÃ
    ALL_PA = js_get_url(f"https://sisapsan.mysapogo.com/admin/price_adjustments.json?query=SUPFINAL&page=1&limit=250")["price_adjustments"]
    for PA in ALL_PA:
        for line in PA["line_items"]:
            line['data'] = json.loads(line['note'])

            if line['variant_id'] not in LGV:
                LGV[line['variant_id']] = []
                LGV[line['variant_id']].append(line['data'])
            else:
                LGV[line['variant_id']].append(line['data'])

    for key in LGV.keys():
        LGV[key] = sorted(LGV[key], key=lambda x: datetime.datetime.strptime(x['date'], "%d/%m/%Y"))

    #VÀO TỪNG ĐƠN ĐỂ XỬ LÝ
    obj = {'sumds':0, 'sum_gv': 0, 'tile': 0}
    for i in range(1,500):
        orders = js_get_url(f"{MAIN_URL}/orders.json?status=completed%2Cfinalized&limit=250&page={i}&location_ids=548744")["orders"]
        if len(orders) == 0:
            break
        for order in orders:
            giavon = 0
            giavonx = 0

            date_start = datetime.datetime.strptime('2023-12-18', "%Y-%m-%d")
            obj['sumds'] += int(order['total'])

            # Xử lý date
            created_on = datetime.datetime.strptime(order['created_on'], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(7)
            
            if created_on < date_start:
                break
            # Tìm giá vốn và cộng vào
            for line in order["order_line_items"]:
                
                #variant_id, nếu có
                if line['variant_id'] in LGV:
                    # Tìm giá vốn phù hợp
                    for giavon_item in LGV[line['variant_id']]:
                        giavon_date = datetime.datetime.strptime(giavon_item['date'], "%d/%m/%Y")
                        if created_on < giavon_date:
                            giavon = int(giavon_item['pu'])
                            break
                        else:
                            giavon = int(giavon_item['pu'])
                    
                    obj['sum_gv'] += int(giavon*line['quantity'])
                    giavonx += int(giavon*line['quantity'])
                #nếu không có:
                else:
                    #print(f"{line['sku']} chua co gia von trong don hang {order['code']} trong kho hang {order['location_id']}!")
                    obj['sum_gv'] += int(line['line_amount']*0.6)
                    giavonx += int(line['line_amount']*0.6)


            if order['total'] == 0:
                print(f"Don ko thu duoc gi: {order['code']}")
            else:
                tile = (giavonx/order['total'])*100
                if tile < 40 and order['total'] > 300000:
                    print(f"Don hang sai sai: {order['code']} - gia von chiem {tile} - cu the: {giavonx} / {order['total']}")



    obj['tile'] = (obj['sum_gv'] / obj['sumds'])*100
    return render(request, 'copyanh.html', obj)

def kho_start(request):
    if request.method == 'GET':
        obj = {
            "boloc": {'channel': '', 'dvvc': '', 'kho': '', 'shop':''},
            "report": {
                'toky': {'dagoi':0,'chogoi':0,'ds_dagoi':0},
                'gele': {'dagoi':0,'chogoi':0,'ds_dagoi':0}
            },
            "channel": {
                "Shopee": {"name":"Shopee","total":0},
                "Lazada": {"name":"Lazada","total":0},
                "Tiki": {"name":"Tiki","total":0},
                "Tiktok": {"name":"Tiktok","total":0},
                "Sapo": {"name":"Sapo Orders","total":0}
            },
            "list_dvvc":[],
            "list_orders": [],
            'list_shop': [],
            'shop_orders_count': {},
            'error_lk' : 0
        }
        dvvc_map = {474723:'ViettelPost',474733:'GHSV',467880: 'ViettelPost',67310: 'Sapo Express',69025: 'GHTK',204246: 'BEST Express',70835: 'Tự Ship / Hà Nội',190219: 'Hola Ship',373257: 'Shopee Xpress',373707: 'J&T Express' }
        list_id = []
        list_order = []
        """
        list_id_tmdt = []
        orders = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?channelType=1,2,6,4&syncStatus=2,3&connectionIds=10925,124264,134366,1551744&page=1&limit=250&channelOrderStatus=READY_TO_SHIP,RETRY_SHIP,IN_CANCEL&sortBy=ISSUED_AT&orderBy=desc")
        get_cookie_tmdt(orders)
        for x_order in orders["orders"]:
            list_id_tmdt.append(x_order['id'])
        
        URL_SYNC = f"https://market-place.sapoapps.vn/v2/orders/sync?ids={','.join(str(e) for e in list_id_tmdt)}&accountId=518175"
        loginss.put(URL_SYNC)
        """ 

        channel = ''
        if "channel" in request.GET:
            channel = request.GET["channel"]
            obj['boloc']['channel'] = channel
        shop = ''
        if "shop" in request.GET:
            shop = request.GET["shop"]
            obj['boloc']['shop'] = shop

        dvvc = ''
        if "dvvc" in request.GET:
            dvvc = request.GET["dvvc"]
            obj['boloc']['dvvc'] = dvvc
        kho = ''
        if "kho" in request.GET:
            kho = request.GET["kho"]
            obj['boloc']['kho'] = kho
            request.session['kho'] = kho

        if 'kho' in request.session:
            obj['boloc']['kho'] = request.session['kho']
            kho = request.session['kho']

        if kho == 'geleximco':
            location = '241737'
        elif kho == 'toky':
            location = '548744'
        else:
            location = '241737,548744'


        # Lấy số lượng đơn hàng đồng bộ thất bại
        orders = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=20&connectionIds={connectID}&sortBy=ISSUED_AT&orderBy=desc&fromDate=1703437200&hasMapping=false")
        
        if 'metadata' in orders:
            if orders['metadata'] == None:
                obj["error_lk"] = 0
            else:
                obj["error_lk"] = len(orders["orders"])

            for order in orders["orders"]:
                URL = f"https://market-place.sapoapps.vn/v2/orders/sync?ids={order['id']}&accountId=319911"
                logintmdt.put(URL)

        else:
            obj["error_lk"] = 0

        # Lấy tất cả những đơn hàng ở trạng thái (yêu cầu gói & đã gói)
        for i in range(15):
            orders = js_get_url(f"{MAIN_URL}/orders.json?status=finalized&limit=250&page={int(i+1)}&packed_status=processing%2Cpacked&composite_fulfillment_status=wait_to_pack%2Cpacked_processing%2Cpacked")
            for order in orders['orders']:
                order['dvvc'] = ''
                list_order.append(order)
                list_id.append(order['id'])
            if len(orders['orders']) == 0:
                break
              
        # Lấy tất cả những đơn hàng TMĐT còn lại
        for i in range(200):
            orders = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?page={int(i+1)}&limit=200&connectionIds={connectID}&channelOrderStatus=READY_TO_SHIP,RETRY_SHIP,PROCESSED&sortBy=ISSUED_AT&orderBy=desc")
            if i == 0:
                #get_cookie_tmdt(orders)
                pass
            if "orders" in orders:
                for x_order in orders["orders"]:
                    #Nếu đơn hàng chưa có trong list_id (tức là chưa có ở trên)
                    if x_order['sapo_order_id'] not in list_id:
                        if x_order['shipping_carrier_name'] != None:
                            order_js = js_get_url(f"{MAIN_URL}/orders/{x_order['sapo_order_id']}.json")
                            if "error" not in order_js and order_js != {}:
                                order_js = order_js['order']
                                order_js['dvvc'] = x_order['shipping_carrier_name']
                                order_js['status_sp'] = x_order['channel_order_status']
                                list_order.append(order_js)
                    else:
                        for order in list_order:
                            if order['id'] == x_order['sapo_order_id']:
                                if x_order['shipping_carrier_name'] != None:
                                    order['dvvc'] = x_order['shipping_carrier_name']
                                    order['status_sp'] = x_order['channel_order_status']
                                    break
            else:
                break
        
        for order in list_order:
            order['packing_status'] = 0
            #Xử lý data gói hàng đang lưu trên Sapo.
            order = get_data_packing(order)

            if order['packing_status'] == 0:
                
                order['mall'] = 0
                if order["process_status_id"] == None:
                    order["process_status_id"] = 0

                for item in order['order_line_items']:
                    if item['product_type'] == "normal" and item['is_packsize'] == True:
                        item['quantity'] = int(item['quantity']*item['pack_size_quantity'])

                if str(order['location_id']) in location:
                    
                    time_order = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
                    now = datetime.datetime.now()

                    order["order_date"] = time_order.strftime("%d-%m (%H:%M)")
                    # Assume ngay_hien_tai là ngày hiện tại (dạng YYYY-MM-DD)
                    ngay_hien_tai = now.strftime("%Y-%m-%d")
                    # Assume ngay_hom_qua là ngày hôm qua (dạng YYYY-MM-DD)
                    ngay_hom_qua = (now - datetime.timedelta(days=1)).strftime("%Y-%m-%d")

                    # Assume thoi_diem_het_han la 21:00:00
                    thoi_diem_het_han = datetime.datetime(now.year, now.month, now.day, 21, 0, 0)

                    # Khởi tạo giá trị mặc định
                    order["time_to_send"] = 0
                    order["sos"] = 0
                    order["delay"] = 0
                    order["time_to_huy"] = 0

                    if time_order.strftime("%Y-%m-%d") == ngay_hien_tai:
                        # Đơn hàng được phát sinh trong ngày hôm nay
                        if time_order.hour < 14:
                            order["time_to_send"]  = int((thoi_diem_het_han - now).total_seconds() / 3600)
                            order["sos"] = 1
                            order["delay"] = 0
                        else:
                            order["time_to_send"] = int((thoi_diem_het_han + datetime.timedelta(days=1) - now).total_seconds() / 3600)
                            order["sos"] = 0
                            order["delay"] = 0
                    elif time_order.strftime("%Y-%m-%d") == ngay_hom_qua:
                        # Đơn hàng được phát sinh trong ngày hôm qua
                        order["time_to_send"]  = int((thoi_diem_het_han - now).total_seconds() / 3600)
                        order["sos"] = 1
                        order["delay"] = 1 if time_order.hour < 12 else 0
                    else:
                        order["delay"] = 2
                        order["sos"] = 1
                        # Đơn hàng được phát sinh trước ngày hôm qua
                        order["time_to_send"] = int((thoi_diem_het_han - now).total_seconds() / 3600)
                        order["time_to_huy"] = int((now + datetime.timedelta(days=3)).strftime("%Y%m%d%H%M%S"))


                    if len(order['tags']) >= 2:
                        for tags in order['tags']:
                            if 'offical' in tags.lower() or 'lteng' in tags.lower() or 'giadungplus' in tags.lower() or 'phaledo' in tags.lower():
                                order['shop'] = tags.split("_")[1]
                            else:
                                order['shop'] = "Gia Dụng Plus"
                        
                        if order['shop'] in ['Phaledo Offcial'] or order['shop'] in ['phaledo']:
                            order['mall'] = 1
                        else:
                            order['mall'] = 0

                    if len(order['fulfillments']) > 0 and order['fulfillments'][-1]["shipment"] is not None:
                        order['chuanbihang'] = 1
                    else:
                        order['chuanbihang'] = 0
                    # Đơn TMĐT hoặc quay lại -> giao ngoài
                    if order['source_id'] in [1880152,6510687,1880149,1880150] and order['channel'] == None:
                        if order['fulfillments'][-1]['status'] != 'cancelled':
                            if order['fulfillments'][-1]['shipment'] != None:
                                order['dvvc'] = dvvc_map[order['fulfillments'][-1]['shipment']['delivery_service_provider_id']]
                            obj['channel']['Sapo']['total'] += 1
                        obj['list_orders'].append(order)
                    # Đơn Sapo
                    elif order["source_id"] in [5483992,4864539,4893087,4339735,1880148,1880147,1880146]:
                        order['sos'] = 0
                        obj['channel']['Sapo']['total'] += 1
                        if order['fulfillments'][-1]['shipment'] != None:
                            if 'SPXVN' in order['fulfillments'][-1]['shipment']['tracking_code']:
                                order['dvvc'] = "SPX Express"
                            elif order['fulfillments'][-1]['shipment']['tracking_code'][:2] == "G8":
                                order['dvvc'] = "Giao Hàng Nhanh"
                            elif order['fulfillments'][-1]['shipment']['tracking_code'][:2] == "82":
                                order['dvvc'] = "J&T Express"

                        obj['list_orders'].append(order)  
                    # Đơn Shopee Real
                    elif order["source_id"] in [1880152] and order['channel'] != None and (channel == 'shopee' or channel == ''):
                        obj['channel']['Shopee']['total'] += 1
                        if order['dvvc'] == dvvc or dvvc == '':
                            if order['shop'] == shop or shop == '':
                                obj['list_orders'].append(order)

                            # Đếm số lượng đơn của từng cửa hàng
                            obj['shop_orders_count'][str(order['shop'])] = obj['shop_orders_count'].get(str(order['shop']), 0) + 1

                    # Đơn Tiktok
                    elif order["source_id"] in [6510687]:
                        obj['channel']['Tiktok']['total'] += 1
                        if channel == 'tiktok' or channel == '':
                            if order['dvvc'] not in obj['list_dvvc']:
                                obj['list_dvvc'].append(order['dvvc'])
                            if order['dvvc'] == dvvc or dvvc == '':
                                obj['list_orders'].append(order)
                    #Đơn Lazada
                    elif order["source_id"] in [1880149]:
                        obj['channel']['Lazada']['total'] += 1
                        obj['list_orders'].append(order)
                    # Đơn Tiki
                    elif order["source_id"] in [1880150]:
                        obj['channel']['Tiki']['total'] += 1
                        obj['list_orders'].append(order)
                    for item in order["order_line_items"]:
                        if item['variant_id'] != None:
                            file_path = f"assets/saveimage/{item['variant_id']}.jpg"
                            if os.path.isfile(file_path):
                                pass
                            else:
                                print(f"{MAIN_URL}/variants/{item['variant_id']}.json")
                                variant = js_get_url(f"{MAIN_URL}/variants/{item['variant_id']}.json")['variant']
                                variant["image"] = variant["images"][0]["full_path"]
                                r = requests.get(variant["image"], allow_redirects=True)
                                open(file_path, 'wb').write(r.content)
        
        obj['list_orders'].reverse()
        obj['list_orders'].sort(key=lambda x: (-int(x['process_status_id']), x['id']))

        return render(request, 'kho_start.html',obj)
    
    elif request.method == 'POST':
        obj = {

        }
        action = request.POST['action']
        orders_id = request.POST['orders_id'].split(",")
        if action == 'pick-up':
            for order in orders_id:
                continue
            return redirect('kho_pickup')

def kho_hoatoc(request):
    obj = {
        "boloc": {'channel': '', 'dvvc': '', 'kho': '', 'shop':'','auto': 'ON'},
        "list_orders": [],
    }


    kho = ''
    if "kho" in request.GET:
        kho = request.GET["kho"]
        obj['boloc']['kho'] = kho
        request.session['kho'] = kho

    if 'kho' in request.session:
        obj['boloc']['kho'] = request.session['kho']
        kho = request.session['kho']

    if kho == 'geleximco':
        location = '241737'
        with open('logs/auto-gele.txt', 'r') as file:
            auto_status = file.read().strip()

        if 'auto' in request.GET:
            auto_status = request.GET['auto']
            if auto_status.upper() in ['ON', 'OFF']:
                with open('logs/auto-gele.txt', 'w') as file:
                    file.write(auto_status.upper())
            obj['boloc']['auto'] = auto_status

    else:
        location = '548744'
        with open('logs/auto-toky.txt', 'r') as file:
            auto_status = file.read().strip()

        if 'auto' in request.GET:
            auto_status = request.GET['auto']
            if auto_status.upper() in ['ON', 'OFF']:
                with open('logs/auto-toky.txt', 'w') as file:
                    file.write(auto_status.upper())
            obj['boloc']['auto'] = auto_status

    list_order = []

    orders = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?connectionIds={connectID}&page=1&limit=50&channelOrderStatus=READY_TO_SHIP,RETRY_SHIP,PROCESSED,UNPAID&shippingCarrierIds=134097,108346,17426,60176,35696,47741,4095,14895,176002,4329&sortBy=ISSUED_AT&orderBy=desc")
    if "orders" in orders:
        for x_order in orders["orders"]:
            if x_order['shipping_carrier_name'] != None:
                order_js = js_get_url(f"{MAIN_URL}/orders/{x_order['sapo_order_id']}.json")
                if "error" not in order_js and order_js != {}:
                    order_js = order_js['order']
                    order_js['dvvc'] = x_order['shipping_carrier_name']
                    order_js['status_sp'] = x_order['channel_order_status']
                    list_order.append(order_js)
        
    for order in list_order:
        order['packing_status'] = 0
        order = get_data_packing(order)
        order['mall'] = 0
        if order["process_status_id"] == None:
            order["process_status_id"] = 0

        for item in order['order_line_items']:
            if item['product_type'] == "normal" and item['is_packsize'] == True:
                item['quantity'] = int(item['quantity']*item['pack_size_quantity'])

        if str(order['location_id']) in location:
            
            time_order = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
            now = datetime.datetime.now()

            order["order_date"] = time_order.strftime("%H:%M - %d-%m")
            
            if len(order['tags']) >= 2:
                for tags in order['tags']:
                    if 'offical' in tags.lower() or 'lteng' in tags.lower() or 'giadungplus' in tags.lower() or 'phaledo' in tags.lower():
                        order['shop'] = tags.split("_")[1]
                    else:
                        order['shop'] = "Gia Dụng Plus"
                
            if order['shop'] in ['phaledo']:
                order['mall'] = 1
            else:
                order['mall'] = 0
                    
            if len(order['fulfillments']) > 0 and order['fulfillments'][-1]["shipment"] is not None:
                order['chuanbihang'] = 1
            else:
                order['chuanbihang'] = 0

            for item in order["order_line_items"]:
                    if item['variant_id'] != None:
                        file_path = f"assets/saveimage/{item['variant_id']}.jpg"
                        if os.path.isfile(file_path):
                            pass
                        else:
                            print(f"{MAIN_URL}/variants/{item['variant_id']}.json")
                            variant = js_get_url(f"{MAIN_URL}/variants/{item['variant_id']}.json")['variant']
                            variant["image"] = variant["images"][0]["full_path"]
                            r = requests.get(variant["image"], allow_redirects=True)
                            open(file_path, 'wb').write(r.content)
        
            obj['list_orders'].append(order)

    obj['list_orders'].reverse()
    obj['list_orders'].sort(key=lambda x: (-int(x['process_status_id']), x['id']))

    return render(request, 'kho_hoatoc.html', obj)

def kho_qlpickup(request):
    obj = {
        'list_orders': [],
        'error': 3,
        'notice': "",
        'boloc':{'phieu':0}
    }
    
    phieu = 1
    if "phieu" in request.GET:
        phieu = int(request.GET["phieu"])
        obj['boloc']['phieu'] = phieu
        request.session['phieu'] = phieu
    
    obj['boloc']['phieu'] = request.session['phieu']
    phieu = request.session['phieu']

    HOME_PARAM = os.environ.get('DJANGO_HOME', None)
    if HOME_PARAM == "HN":
        URL_PICKUP = 'logs/pickup/gele.txt'
    elif HOME_PARAM == "HCM":
        URL_PICKUP = 'logs/pickup/toky.txt'
    else:
        URL_PICKUP = 'logs/pickup/test.txt'

    with open(URL_PICKUP, 'r') as file:
        temp_pickup = file.read().strip()

    all_pickup = json.loads(temp_pickup)
    temp_pickup = all_pickup['all'][str(phieu)]

    #Delete
    if "delete" in request.GET:
        delete_id = int(request.GET['delete'])
        if delete_id in temp_pickup:
            temp_pickup.remove(delete_id)

    #Thêm hàng loạt từ kho_start
    if "orders_id" in request.GET:
        order_ids = request.GET['orders_id'].split(",")
        for order_id in order_ids:
            order_search = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")['order']
            order_search = get_data_packing(order_search)

            # Kiểm tra trạng thái (đã giao/ đã hoàn thành / đã huỷ/ khách yêu cầu huỷ)
            if order_search['status'] in ['cancelled']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã huỷ!'
            elif order_search['status'] in ['complete']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã hoàn thành!'

            if order_search['fulfillments'][-1]['composite_fulfillment_status'] in ['retry_delivery','fulfilled','received','packed_cancelled_client','fulfilled_cancelling','fulfilled_cancelled']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã giao đi rồi!'

            # Kiểm tra xem đơn đang yêu cầu huỷ hay không.
            orders = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?channelType=1,2,6,4&connectionIds={connectID}&page=1&limit=250&channelOrderStatus=IN_CANCEL&sortBy=ISSUED_AT&orderBy=desc")
            if len(orders) > 100:
                for x_order in orders["orders"]:
                    if order_search['id'] == x_order['sapo_order_id']:
                        obj['error'] = 1
                        obj['notice'] = 'Khách hàng đang yêu cầu xin huỷ đơn này!'
                        break
            if order_search['packing_status'] <= 3:
                if obj['error'] == 0:
                    temp_pickup.append(order_search["id"])


    # Thêm đơn hàng vào list nhặt hàng.
    if 'search_add' in request.GET:
        obj['error'] = 0
        search_add = (request.GET['search_add']).replace(" ","").upper()
        SEARCH_FLAG = 0
        # Tìm đơn hàng từ mã vận đơn được gửi lên.
        if search_add != "":
            order_search = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=100&query={search_add}")["orders"]
            if len(order_search) > 0:
                order_search = order_search[0]
                SEARCH_FLAG = 1
            else:
                MADONHANG = search_order_shopee(search_add)
                if len(MADONHANG) > 1:
                    order_search = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=100&query={MADONHANG}")["orders"]
                    order_search = order_search[0]
                    SEARCH_FLAG = 1
                else:
                    obj['error'] = 1
                    obj['notice'] = 'Không tìm thấy đơn hàng hợp lệ!'
                    
        if SEARCH_FLAG == 1:
            if order_search["id"] not in temp_pickup:
                order_search = get_data_packing(order_search)

                # Kiểm tra trạng thái (đã giao/ đã hoàn thành / đã huỷ/ khách yêu cầu huỷ)
                if order_search['status'] in ['cancelled']:
                    obj['error'] = 1
                    obj['notice'] = 'Đơn hàng đã huỷ!'
                elif order_search['status'] in ['complete']:
                    obj['error'] = 1
                    obj['notice'] = 'Đơn hàng đã hoàn thành!'

                elif order_search['packing_status'] == 4:
                    obj['error'] = 1
                    obj['notice'] = 'Đơn hàng đã được gói!'
                
                elif order_search['packing_status'] > 4:
                    obj['error'] = 1
                    obj['notice'] = 'Đơn hàng đã được phân hàng!'

                if order_search['fulfillments'][-1]['composite_fulfillment_status'] in ['retry_delivery','fulfilled','received','packed_cancelled_client','fulfilled_cancelling','fulfilled_cancelled']:
                    obj['error'] = 1
                    obj['notice'] = 'Đơn hàng đã giao đi rồi!'

                # Kiểm tra xem đơn đang yêu cầu huỷ hay không.
                orders = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?channelType=1,2,6,4&connectionIds={connectID}&page=1&limit=250&channelOrderStatus=IN_CANCEL&sortBy=ISSUED_AT&orderBy=desc")
                if len(orders) > 100:
                    for x_order in orders["orders"]:
                        if order_search['id'] == x_order['sapo_order_id']:
                            obj['error'] = 1
                            obj['notice'] = 'Khách hàng đang yêu cầu xin huỷ đơn này!'
                            break

                if order_search['packing_status'] <= 3:
                    if obj['error'] == 0:
                        temp_pickup.append(order_search["id"])


    print(temp_pickup)
    temp_pickup = list(set(temp_pickup))
    #Phân tích lại và xử lý pick_up order để loại bỏ những đơn: Đang y/cầu huỷ, đã huỷ, đã gói hàng, đã bàn giao cho bên vận chuyển. Chỉ nhặt những đơn hàng hợp lệ.
    for order_id in temp_pickup:
        this_order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")['order']
        this_order = get_data_packing(this_order)
        ERROR = 0
        # Kiểm tra trạng thái (đã giao/ đã hoàn thành / đã huỷ/ khách yêu cầu huỷ)
        if this_order['status'] in ['cancelled']:
            ERROR = 1
        elif this_order['status'] in ['complete']:
            ERROR = 1
        if this_order['fulfillments'][-1]['composite_fulfillment_status'] in ['retry_delivery','fulfilled','received','packed_cancelled_client','fulfilled_cancelling','fulfilled_cancelled']:
            ERROR = 1
        if this_order['packing_status'] > 3:
            ERROR = 1
        
        time_order = datetime.datetime.strptime(this_order["created_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
        now = datetime.datetime.now()

        this_order["order_date"] = time_order.strftime("%d-%m (%H:%M)")
        # Assume ngay_hien_tai là ngày hiện tại (dạng YYYY-MM-DD)

        if len(this_order['tags']) >= 2:
            this_order['shop'] = this_order['tags'][1].split("_")[1]
            
            if this_order['shop'] in ['Phaledo Offcial']:
                this_order['mall'] = 1
            else:
                this_order['mall'] = 0

        if len(this_order['fulfillments']) > 0 and this_order['fulfillments'][-1]["shipment"] is not None:
            this_order['chuanbihang'] = 1
        else:
            this_order['chuanbihang'] = 0

        if ERROR == 1:
            temp_pickup.remove(this_order['id'])

        else:
            obj["list_orders"].append(this_order)

    with open(URL_PICKUP, 'w') as file:
        all_save = all_pickup
        all_save["all"][str(phieu)] = temp_pickup
        file.write(json.dumps(all_save))

    if "orders_id" in request.GET:
        return render(request, 're_print.html',{'order_ids':request.GET['orders_id']})
    else:
        return render(request, 'kho_qlpickup.html',obj)

def kho_pickup(request):
    obj = {
        'list_orders': [],
        'list_product': [],
        "boloc":{"phieu":1}
    }

    phieu = 1
    if "phieu" in request.GET:
        phieu = int(request.GET["phieu"])
        obj['boloc']['phieu'] = phieu
        request.session['phieu'] = phieu
    
    obj['boloc']['phieu'] = request.session['phieu']
    phieu = request.session['phieu']

    HOME_PARAM = os.environ.get('DJANGO_HOME', None)
    if HOME_PARAM == "HN":
        URL_PICKUP = 'logs/pickup/gele.txt'
        LOCALTION_ID = 241737
    elif HOME_PARAM == "HCM":
        URL_PICKUP = 'logs/pickup/toky.txt'
        LOCALTION_ID = 548744
    else:
        URL_PICKUP = 'logs/pickup/test.txt'
        LOCALTION_ID = 241737

    with open(URL_PICKUP, 'r') as file:
        temp_pickup = file.read().strip()

    all_pickup = json.loads(temp_pickup)
    temp_pickup = all_pickup['all'][str(phieu)]

    #Phân tích lại và xử lý pick_up order để loại bỏ những đơn: Đang y/cầu huỷ, đã huỷ, đã gói hàng, đã bàn giao cho bên vận chuyển. Chỉ nhặt những đơn hàng hợp lệ.
    for order_id in temp_pickup:
        this_order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")['order']
        this_order = get_data_packing(this_order)
        ERROR = 0
        # Kiểm tra trạng thái (đã giao/ đã hoàn thành / đã huỷ/ khách yêu cầu huỷ)
        if this_order['status'] in ['cancelled']:
            ERROR = 1
        elif this_order['status'] in ['complete']:
            ERROR = 1
        if this_order['fulfillments'][-1]['composite_fulfillment_status'] in ['retry_delivery','fulfilled','received','packed_cancelled_client','fulfilled_cancelling','fulfilled_cancelled']:
            ERROR = 1
        if this_order['packing_status'] > 3:
            ERROR = 1

        if ERROR == 1:
            temp_pickup.remove(this_order['id'])
        else:
            obj["list_orders"].append(this_order)
    
    with open(URL_PICKUP, 'w') as file:
        all_save = all_pickup
        all_save["all"][str(phieu)] = temp_pickup
        file.write(json.dumps(all_save))

    # XỬ LÝ LIST SẢN PHẨM CẦN NHẶT
    for this_order in obj["list_orders"]:
        
        for item in this_order['order_line_items']:
            if item['product_type'] == "normal" and item['is_packsize'] == False:
                flag = 0
                for line in obj['list_product']:
                    if line['id'] == item['variant_id']:
                        line['quantity'] += int(item['quantity'])
                        flag = 1
                        break
                if flag == 0:
                    obj['list_product'].append({
                        'id': item['variant_id'],
                        'sku': item['sku'][3:],
                        'barcode': str(item['barcode']),
                        'variant_options': item['variant_options'],  
                        'quantity': int(item['quantity'])
                    })
            elif item['product_type'] == "normal" and item['is_packsize'] == True:
                flag = 0
                for line in obj['list_product']:
                    if line['id'] == item['variant_id']:
                        line['quantity'] += int(item['quantity']*item['pack_size_quantity'])
                        flag = 1
                        break
                if flag == 0:
                    obj['list_product'].append({ 
                        'id': item['variant_id'],
                        'sku': item['sku'][3:],
                        'barcode': str(item['barcode']),
                        'variant_options': item['variant_options'],  
                        'quantity': int(item['quantity']*item['pack_size_quantity'])
                    })

            if item['product_type'] == "composite":
                for xitem in item['composite_item_domains']:
                    flag = 0
                    for line in obj['list_product']:
                        if line['id'] == xitem['variant_id']:
                            line['quantity'] += int(xitem['quantity'])
                            flag = 1
                            break
                    if flag == 0:
                        vari = js_get_url(f"{MAIN_URL}/variants/{xitem['variant_id']}.json")['variant']
                        obj['list_product'].append({
                            'id': xitem['variant_id'],
                            'sku': vari['sku'][3:],
                            'barcode': str(vari['barcode']),
                            'variant_options': "chiếc",  
                            'quantity': int(xitem['quantity'])
                        })
    for vari in obj['list_product']:
        this_vari = js_get_url(f"{MAIN_URL}/variants/{vari['id']}.json")['variant']

        for iven in this_vari['inventories']:
            
            if iven['location_id'] == LOCALTION_ID:
                vari['tonkho'] = int(iven['on_hand'])
                if iven['bin_location'] is None:
                    vari['kho'] = ""
                    vari['pl1'] = "Loading"
                    vari['pl2'] = ""

                elif "/" in iven['bin_location']:
                    vari['kho'] = iven['bin_location'].split("/")[0]
                    vari['pl1'] = iven['bin_location'].split("/")[1]
                    vari['pl2'] = iven['bin_location'].split("/")[2]
                else:
                    vari['kho'] = ""
                    vari['pl1'] = "Loading"
                    vari['pl2'] = ""

    obj['list_product'].sort(key=lambda x: x['pl2'])

    return render(request, 'kho_pickup.html',obj)

def kho_scanphanhang(request):
    obj = {
        'nguoi_goi':'',
        'nhanvien_data': {},
        'pwd': '',
        'list_orders': [],
        'error': 0,
        'notice': "",
        'boloc':{'phieu':1}
    }
    phieu = 1
    if "phieu" in request.GET:
        phieu = int(request.GET["phieu"])
        obj['boloc']['phieu'] = phieu
        request.session['phieu'] = phieu
    
    obj['boloc']['phieu'] = request.session['phieu']
    phieu = request.session['phieu']

    #check_login_sapo()
    # Đọc và xử lý dữ liệu từ file toky.txt
    with open('logs/nhanvien/toky.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for line in lines:
            name, password = line.strip().split('/')
            obj['nhanvien_data'][f'TOKY: {name}'] = password

    # Đọc và xử lý dữ liệu từ file geleximco.txt
    with open('logs/nhanvien/geleximco.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for line in lines:
            name, password = line.strip().split('/')
            obj['nhanvien_data'][f'GELE: {name}'] = password
    if 'nguoi_goi' in request.GET:
        nguoi_goi = request.GET['nguoi_goi']
        request.session['nguoi_goi'] = nguoi_goi
        obj['nguoi_goi'] = nguoi_goi

    if 'nguoi_goi' in request.session:
        obj['nguoi_goi'] = request.session['nguoi_goi']

    if 'pwd' in request.session:
        obj['pwd'] = request.session['pwd']

    HOME_PARAM = os.environ.get('DJANGO_HOME', None)
    if HOME_PARAM == "HN":
        URL_PICKUP = 'logs/pickup/gele.txt'
    elif HOME_PARAM == "HCM":
        URL_PICKUP = 'logs/pickup/toky.txt'
    else:
        URL_PICKUP = 'logs/pickup/test.txt'

    with open(URL_PICKUP, 'r') as file:
        temp_pickup = file.read().strip()

    all_pickup = json.loads(temp_pickup)
    temp_pickup = all_pickup['all'][str(phieu)]
    
    temp_pickup = list(set(temp_pickup))
    #Phân tích lại và xử lý pick_up order để loại bỏ những đơn: Đang y/cầu huỷ, đã huỷ, đã gói hàng, đã bàn giao cho bên vận chuyển. Chỉ nhặt những đơn hàng hợp lệ.
    for order_id in temp_pickup:
        this_order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")['order']
        this_order = get_data_packing(this_order)
        ERROR = 0
        # Kiểm tra trạng thái (đã giao/ đã hoàn thành / đã huỷ/ khách yêu cầu huỷ)
        if this_order['status'] in ['cancelled']:
            ERROR = 1
        elif this_order['status'] in ['complete']:
            ERROR = 1
        if this_order['fulfillments'][-1]['composite_fulfillment_status'] in ['retry_delivery','fulfilled','received','packed_cancelled_client','fulfilled_cancelling','fulfilled_cancelled']:
            ERROR = 1
        if this_order['packing_status'] > 3:
            ERROR = 1
        
        time_order = datetime.datetime.strptime(this_order["created_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
        now = datetime.datetime.now()

        this_order["order_date"] = time_order.strftime("%d-%m (%H:%M)")
        # Assume ngay_hien_tai là ngày hiện tại (dạng YYYY-MM-DD)

        if len(this_order['tags']) >= 2:
            this_order['shop'] = this_order['tags'][1].split("_")[1]
            
            if this_order['shop'] in ['Phaledo Offcial']:
                this_order['mall'] = 1
            else:
                this_order['mall'] = 0

        if len(this_order['fulfillments']) > 0 and this_order['fulfillments'][-1]["shipment"] is not None:
            this_order['chuanbihang'] = 1
        else:
            this_order['chuanbihang'] = 0

        if ERROR == 1:
            temp_pickup.remove(this_order['id'])
        else:
            obj["list_orders"].append(this_order)

    with open(URL_PICKUP, 'w') as file:
        all_save = all_pickup
        all_save["all"][str(phieu)] = temp_pickup
        file.write(json.dumps(all_save))

    return render(request, 'kho_scanphanhang.html', obj)

def kho_phanhang(request):
    #check_login_sapo()
    obj = {
        'order': {},
        'error': 0,
        'notice': '',
        'print_status': 'ON',
        'nhanvien_data': {}
    }
    nguoi_goi = "Quý"
    flag = 0
    pwd = ""
    if 'nguoi_goi' in request.GET:
        nguoi_goi = request.GET['nguoi_goi']
        request.session['nguoi_goi'] = nguoi_goi

    if 'pwd' in request.GET:
        pwd = request.GET['pwd']
        request.session['pwd'] = pwd

    nguoi_goi = request.session['nguoi_goi']
    pwd = request.session['pwd']
     
    # Đọc và xử lý dữ liệu từ file toky.txt
    with open('logs/nhanvien/toky.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for line in lines:
            name, password = line.strip().split('/')
            obj['nhanvien_data'][f'TOKY: {name}'] = password

    # Đọc và xử lý dữ liệu từ file geleximco.txt
    with open('logs/nhanvien/geleximco.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for line in lines:
            name, password = line.strip().split('/')
            obj['nhanvien_data'][f'GELE: {name}'] = password

    flag = 0
    for name, password in obj['nhanvien_data'].items():
        if nguoi_goi == name and pwd == password:
            flag = 1
            break            

    if flag == 1:
        if 'mavandon' in request.GET:
            mavandon = request.GET['mavandon'].replace(" ", "")
            save_mavandon = mavandon
            order_id = js_get_url(f"{MAIN_URL}/fulfillments.json?page=1&limit=1&query={mavandon}")['fulfillments']
            if len(order_id) > 0:
                order_id = order_id[0]['order_id']
            else:
                mavandon = search_order_shopee(mavandon)
                if len(mavandon) > 1:
                    order_id = js_get_url(f"{MAIN_URL}/fulfillments.json?page=1&limit=1&query={mavandon}")['fulfillments']
                    order_id = order_id[0]['order_id']
                    SEARCH_FLAG = 1
                    
                else:
                    obj['error'] = 1
                    obj['notice'] = f"Không tìm được vận đơn liên quan!"
                    return render(request, 'kho_packing.html',obj)
            
        if 'order_id' in request.GET:
            order_id = request.GET['order_id']

        if 'done' in request.GET:
            order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")["order"]
            #Xử lý data gói hàng đang lưu trên Sapo.
            order = get_data_packing(order)
            if "split" not in order:
                order["split"] = 1

            if order["split"] > 1 and 'dontach' in request.GET:
                dontach = int(request.GET['dontach'])
                if dontach == 1:
                    update_data_one(order["code"],{'nguoi_phan':nguoi_goi,'time_packing': datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})
                else:
                    update_data_one(order["code"],{f"nguoi_phan{dontach}":nguoi_goi,f"time_packing{dontach}": datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})

                order = get_data_packing(order)
                pack_goi = 0
                for i in range(order["split"]):

                    if i == 0:
                        if 'nguoi_phan' in order:
                            pack_goi += 1
                    else:
                        j = i+1
                        if f"nguoi_phan{j}" in order:
                            pack_goi += 1
                print(f"So don tach da goi: {pack_goi}")
                if pack_goi == order["split"]:
                    update_data_one(order["code"],{f"packing_status":5})
            else:
                update_data_one(order["code"],{f"packing_status":5,'nguoi_chia':nguoi_goi,'time_chia': datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})

            #send_zns_xacnhan(243976, order, 2)
            return redirect('kho_scanphanhang')
        else:
            order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")['order']
            order = get_data_packing(order)

            if 'split' not in order:
                order['split'] = 1

            if len(order['tags']) >= 2:
                order['shop'] = order['tags'][1].split("_")[-1:][0].replace("Offcial","")

            if order["packing_status"] == 4:
                obj['error'] = 1
                obj['notice'] = f"Đơn hàng đã được {order["nguoi_goi"]} gói vào lúc {order["time_packing"]}!"

            if order["packing_status"] == 5:
                obj['error'] = 1
                obj['notice'] = f"Đơn hàng đã được {order["nguoi_chia"]} phân hàng vào lúc {order["time_chia"]}!"

            # Kiểm tra trạng thái (đã giao/ đã hoàn thành / đã huỷ/ khách yêu cầu huỷ)
            if order['status'] in ['cancelled']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã huỷ!'
            elif order['status'] in ['complete']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã hoàn thành!'

            if order['fulfillments'][-1]['composite_fulfillment_status'] in ['retry_delivery','fulfilled','received','packed_cancelled_client','fulfilled_cancelling','fulfilled_cancelled']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã giao đi rồi!'

            # Kiểm tra xem đơn đang yêu cầu huỷ hay không.
            orders = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?channelType=1,2,6,4&connectionIds={connectID}&page=1&limit=250&channelOrderStatus=IN_CANCEL&sortBy=ISSUED_AT&orderBy=desc")
            if len(orders) > 100:
                for x_order in orders["orders"]:
                    if order['id'] == x_order['sapo_order_id']:
                        obj['error'] = 1
                        obj['notice'] = 'Khách hàng đang yêu cầu xin huỷ đơn này!'
                        break
            # Kiểm tra xem đơn hàng đã có MVĐ hay chưa ?
            if order['fulfillments'][-1]['shipment'] == None:
                obj['error'] = 4
            
            # Xử lý kèm keo
            list_kemkeo = []
            all_kemkeo = js_get_url(f"{MAIN_URL}/products.json?tags=add_keo_dan&page=1&litmit=250")["products"]
            for pr in all_kemkeo:
                for vari in pr['variants']:
                    list_kemkeo.append(vari['id'])
            order['kemkeo'] = 0

            if obj['error'] == 0:
                order["REAL_TACH"] = 0
                order['real_items'] = []
                for item in order['order_line_items']:
                    item['pr_name'] = ''
                    # Sản phẩm thường
                    if item['product_name'] != None:
                        if '/' in item['product_name']:
                            item['pr_name'] = item['product_name'].split('/')[0].upper()

                    if item['variant_id'] in list_kemkeo:
                        order['kemkeo'] = 1

                    if item['product_type'] == "normal" and item['is_packsize'] == False:
                        flag = 0
                        for line in order['real_items']:
                            if line['id'] == item['variant_id']:
                                line['quantity'] += int(item['quantity'])
                                flag = 1
                                break
                        if flag == 0:
                            order['real_items'].append({
                                'id': item['variant_id'],
                                'sku': item['sku'][3:],
                                'variant_options': item['variant_options'],  
                                'quantity': int(item['quantity']),
                                'unit': item['unit'],
                                'old_id': 0,
                                'barcode': item['barcode'],
                                'pr_name':item['pr_name'],
                                'print':0,
                            })
                    elif item['product_type'] == "normal" and item['is_packsize'] == True:
                        flag = 0
                        for line in order['real_items']:
                            if line['id'] == item['pack_size_root_id']:
                                line['quantity'] += int(item['quantity']*item['pack_size_quantity'])
                                flag = 1
                                break
                        if flag == 0:
                            vari = js_get_url(f"{MAIN_URL}/variants/{item['pack_size_root_id']}.json")['variant']
                            order['real_items'].append({
                                'id': vari['id'],
                                'sku': vari['sku'][3:],
                                'barcode': vari['barcode'],
                                'variant_options': vari['opt1'],
                                'old_id': item['variant_id'],
                                'unit': 'chiếc (tách từ combo)',
                                'pr_name':item['pr_name'],
                                'quantity': int(item['quantity']*item['pack_size_quantity']),
                                'print':0
                            })

                    if item['product_type'] == "composite":
                        for xitem in item['composite_item_domains']:
                            flag = 0
                            for line in order['real_items']:
                                if line['id'] == xitem['variant_id']:
                                    line['quantity'] += int(xitem['quantity'])
                                    flag = 1
                                    break
                            if flag == 0:
                                vari = js_get_url(f"{MAIN_URL}/variants/{xitem['variant_id']}.json")['variant']
                                order['real_items'].append({
                                    'id': xitem['variant_id'],
                                    'sku': vari['sku'][3:],
                                    'barcode': vari['barcode'],
                                    'old_id': item['variant_id'],
                                    'pr_name':item['pr_name'],
                                    'unit': 'chiếc',
                                    'variant_options': vari['opt1'],  
                                    'quantity': int(xitem['quantity']),
                                    'print':0
                                })
                    if item['product_type'] == None:
                        order['real_items'].append({
                            'id': 0,
                            'sku': "SERVICE",
                            'unit': item['unit'],
                            'old_id': 0,
                            'variant_options': item['note'],  
                            'pr_name':item['pr_name'],
                            'quantity': int(item['quantity']),
                            'print':0
                        })   
                        
                order['total_quantity'] = 0
                for line in order['real_items']:
                    order['total_quantity'] += line['quantity']

                if order['note'] != None:
                    if len(order['note']) > 5:
                        file_path = 'assets/notify/have_note.mp3'
                        """
                        pygame.mixer.init()
                        pygame.mixer.music.load(file_path)
                        pygame.mixer.music.play()
                        """

                if order["split"] > 1 and obj["error"] == 0:
                    print("XỬ LÝ ĐƠN TÁCH KHI GÓI HÀNG!")
                    order['d_tach_values'] = list(range(1, order['split'] + 1))

                    tmdt_order = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=1&connectionIds={connectID}&query={order['reference_number']}")['orders'][0]
                    logintmdt.put(f"https://market-place.sapoapps.vn/v2/orders/sync?ids={tmdt_order['id']}&accountId=3199")

                    shop_name = existing_shop_map.get(tmdt_order['connection_id'], "")

                    if doi_shop(tmdt_order['connection_id'],loginsp) == 1:
                        if shop_name == 'giadungplus_official':
                            shop_id = 241961702
                        elif shop_name == 'phaledo':
                            shop_id = 1009027554


                    #Thông tin kiện hàng
                    URL = f"https://banhang.shopee.vn/api/v3/order/batch_get_packages_multi_shop?SPC_CDS=f8cfde35-a66a-4a25-8b35-c7ea97c759aa&SPC_CDS_VER=2"
                    RS = json.loads(loginsp.post(URL,json={"orders":[{"order_id":order["shopee_id"],"shop_id":shop_id,"region_id":"VN"}]}).text)
                    

                    URL = f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=20&connectionIds={connectID}&query={order['reference_number']}&sortBy=ISSUED_AT&orderBy=desc"
                    ORDER_TMDT = js_get_url(URL)["orders"][0]
                    for x in order['real_items']:
                        x["item_id_shopee"] = 0

                    for pr in ORDER_TMDT["products"]:
                        for x in order['real_items']:
                            if int(pr["sapo_variant_id"]) == int(x["id"]) or int(pr["sapo_variant_id"]) == int(x["old_id"]):
                                x["item_id_shopee"] = pr["variation_id"]

                    LIST_PACK =RS['data']['list'][0]["package_list"]
                    D_TACH = 0
                    #Nếu nhập vào là mã vận đơn.
                    if not save_mavandon.startswith('SON'):
                        for PACKEDD in LIST_PACK:
                            D_TACH +=1
                            #Nếu nó là đơn tách này thì mới xử lý.
                            if save_mavandon == PACKEDD['third_party_tn']:
                                order["REAL_TACH"] = D_TACH
                                flag = 0
                                for s_item in PACKEDD["items"]:
                                    for x in order['real_items']:
                                        if str(x["item_id_shopee"]) == str(s_item["model_id"]):
                                            x["dontach"] = D_TACH
                                            flag = 1

                                if flag == 0:
                                    for x in order['real_items']:
                                        if x["item_id_shopee"] == 0:
                                            x["dontach"] = D_TACH
                    else:
                        D_TACH +=1
                        flag = 0
                        for s_item in PACKEDD["items"]:
                            for x in order['real_items']:
                                if str(x["item_id_shopee"]) == str(s_item["model_id"]):
                                    x["dontach"] = D_TACH
                                    flag = 1

                        if flag == 0:
                            for x in order['real_items']:
                                if x["item_id_shopee"] == 0:
                                    x["dontach"] = D_TACH

            obj['order'] = order
            return render(request, 'kho_phanhang.html',obj)
    else:
        obj['error'] = 1
        obj['notice'] = "Mật khẩu không chính xác!"
        return render(request, 'kho_phanhang.html',obj) 
    

    return render(request, 'kho_scanphanhang.html', obj)

def kho_goibihuy(request):

    # Những đơn hàng đã gói (packing status = 4, nhưng đơn ở trạng thái Hủy -> phải xác nhận là đã nhận lại hàng thì mới biến mất.)
    obj = {
            "boloc": {'channel': '', 'dvvc': '','kho': ''},
            "channel": {
                "Shopee": {"name":"Shopee","total":0},
                "Lazada": {"name":"Lazada","total":0},
                "Tiki": {"name":"Tiki","total":0},
                "Tiktok": {"name":"Tiktok","total":0},
                "Sapo": {"name":"Sapo Orders","total":0}
            },
            "list_dvvc":[],
            "list_orders": []
        }
    dvvc_map = {474723:'ViettelPost',474733:'GHSV',467880: 'ViettelPost',67310: 'Sapo Express',69025: 'GHTK',204246: 'BEST Express',70835: 'Tự Ship / Hà Nội',190219: 'Hola Ship',373257: 'Shopee Xpress',373707: 'J&T Express' }
    list_id = []
    list_order = []

    channel = ''
    if "channel" in request.GET:
        channel = request.GET["channel"]
        obj['boloc']['channel'] = channel
    dvvc = ''
    if "dvvc" in request.GET:
        dvvc = request.GET["dvvc"]
        obj['boloc']['dvvc'] = dvvc
    
    if 'done' in request.GET:
        order_id = request.GET["order_id"]
        location_id = request.GET["location_id"]
        update_status(int(order_id),17238,location_id)

    kho = ''
    if "kho" in request.GET:
        kho = request.GET["kho"]
        obj['boloc']['kho'] = kho
        request.session['kho'] = kho
    

    if 'kho' in request.session:
        obj['boloc']['kho'] = request.session['kho']
        kho = request.session['kho']
    if kho == 'geleximco':
        location = '241737'
    elif kho == 'toky':
        location = '548744'
    else:
        location = '241737,548744'
    
    if 'search' in request.GET:
        orders = js_get_url(f"{MAIN_URL}/orders.json?query={request.GET['search']}")
        for order in orders['orders']:
            order['dvvc'] = ''
            list_order.append(order)            
    else:
        for i in range(1,5):
            # Lấy tất cả những đơn hàng ở trạng thái (yêu cầu gói & đã gói)
            orders = js_get_url(f"{MAIN_URL}/orders.json?location_ids="+str(location)+"&status=cancelled&limit=250&page="+str(i))
            for order in orders['orders']:
                order['dvvc'] = ''
                list_order.append(order)

    for order in list_order:

        order = get_data_packing(order)

        for item in order['order_line_items']:
            if item['product_type'] == "normal" and item['is_packsize'] == True:
                item['quantity'] = int(item['quantity']*item['pack_size_quantity'])

        if len(order['tags']) >= 2:
            order['shop'] = order['tags'][1].split("_")[-1:][0]
        if order['packing_status'] >= 4:
            # Đơn TMĐT hoặc quay lại -> giao ngoài
            if order['source_id'] in [1880152,6510687,1880149,1880150] and order['channel'] == None:
                obj['channel']['Sapo']['total'] += 1
                if channel == 'sapo' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
            # Đơn Sapo
            elif order["source_id"] in [5483992,4864539,4893087,4339735,1880148,1880147,1880146]:
                obj['channel']['Sapo']['total'] += 1
                if channel == 'sapo' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)          
                          
            # Đơn Shopee real
            elif order["source_id"] in [1880152] and order['channel'] != None:
                obj['channel']['Shopee']['total'] += 1
                if channel == 'shopee' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                       
            # Đơn Tiktok
            elif order["source_id"] in [6510687]:
                obj['channel']['Tiktok']['total'] += 1
                if channel == 'tiktok' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                    
                         
            #Đơn Lazada
            elif order["source_id"] in [1880149]:
                obj['channel']['Lazada']['total'] += 1
                if channel == 'lazada' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                   
            # Đơn Tiki
            elif order["source_id"] in [1880150]:
                obj['channel']['Tiki']['total'] += 1
                if channel == 'tiki' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                    
            
            for item in order["order_line_items"]:
                if item['variant_id'] != None:
                    file_path = f"assets/saveimage/{item['variant_id']}.jpg"
                    if os.path.isfile(file_path):
                        pass
                    else:
                        print(f"{MAIN_URL}/variants/{item['variant_id']}.json")
                        variant = js_get_url(f"{MAIN_URL}/variants/{item['variant_id']}.json")['variant']
                        variant["image"] = variant["images"][0]["full_path"]
                        r = requests.get(variant["image"], allow_redirects=True)
                        open(file_path, 'wb').write(r.content)


    return render(request, 'kho_goibihuy.html', obj)

def kho_donbansot(request):

    # Những đơn hàng đã gói,.trạng thái OK (chưa huỷ).
    obj = {
            "boloc": {'channel': '', 'dvvc': '','kho': ''},
            "channel": {
                "Shopee": {"name":"Shopee","total":0},
                "Lazada": {"name":"Lazada","total":0},
                "Tiki": {"name":"Tiki","total":0},
                "Tiktok": {"name":"Tiktok","total":0},
                "Sapo": {"name":"Sapo Orders","total":0}
            },
            "list_dvvc":[],
            "list_orders": []
        }
    dvvc_map = {474733:'GHSV',467880: 'ViettelPost',67310: 'Sapo Express',69025: 'GHTK',204246: 'BEST Express',70835: 'Tự Ship / Hà Nội',190219: 'Hola Ship',373257: 'Shopee Xpress',373707: 'J&T Express' }
    list_id = []
    list_order = []

    channel = ''
    if "channel" in request.GET:
        channel = request.GET["channel"]
        obj['boloc']['channel'] = channel
    dvvc = ''
    if "dvvc" in request.GET:
        dvvc = request.GET["dvvc"]
        obj['boloc']['dvvc'] = dvvc
    
    if 'done' in request.GET:
        order_id = request.GET["order_id"]
        location_id = request.GET["location_id"]
        update_status(int(order_id),16685,location_id)

    kho = ''
    if "kho" in request.GET:
        kho = request.GET["kho"]
        obj['boloc']['kho'] = kho
        request.session['kho'] = kho
    

    if 'kho' in request.session:
        obj['boloc']['kho'] = request.session['kho']
        kho = request.session['kho']
    if kho == 'geleximco':
        location = '241737'
    elif kho == 'toky':
        location = '548744'
    else:
        location = '241737,548744'
    
    if 'search' in request.GET:
        orders = js_get_url(f"{MAIN_URL}/orders.json?query={request.GET['search']}")
        for order in orders['orders']:
            order['dvvc'] = ''
            list_order.append(order)            
    else:
        for i in range(1,8):
            # Lấy tất cả những đơn hàng ở trạng thái (yêu cầu gói & đã gói)
            orders = js_get_url(f"{MAIN_URL}/orders.json?location_ids="+str(location)+"&composite_fulfillment_status=wait_to_pack%2Cpacked_processing%2Cpacked&status=draft%2Cfinalized&limit=250&page="+str(i))
            for order in orders['orders']:
                order['dvvc'] = ''
                list_order.append(order)

    for order in list_order:
        order = get_data_packing(order)
        
        for item in order['order_line_items']:
            if item['product_type'] == "normal" and item['is_packsize'] == True:
                item['quantity'] = int(item['quantity']*item['pack_size_quantity'])

        if len(order['tags']) >= 2:
            order['shop'] = order['tags'][1].split("_")[-1:][0]
          
        if order['packing_status'] == 4:
            # Đơn TMĐT hoặc quay lại -> giao ngoài
            if order['source_id'] in [1880152,6510687,1880149,1880150] and order['channel'] == None:
                obj['channel']['Sapo']['total'] += 1
                if channel == 'sapo' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
            # Đơn Sapo
            elif order["source_id"] in [5483992,4864539,4893087,4339735,1880148,1880147,1880146]:
                obj['channel']['Sapo']['total'] += 1
                if channel == 'sapo' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)          
                          
            # Đơn Shopee real
            elif order["source_id"] in [1880152] and order['channel'] != None:
                obj['channel']['Shopee']['total'] += 1
                if channel == 'shopee' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                       
            # Đơn Tiktok
            elif order["source_id"] in [6510687]:
                obj['channel']['Tiktok']['total'] += 1
                if channel == 'tiktok' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                    
                         
            #Đơn Lazada
            elif order["source_id"] in [1880149]:
                obj['channel']['Lazada']['total'] += 1
                if channel == 'lazada' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                   
            # Đơn Tiki
            elif order["source_id"] in [1880150]:
                obj['channel']['Tiki']['total'] += 1
                if channel == 'tiki' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                    
            
            for item in order["order_line_items"]:
                if item['variant_id'] != None:
                    file_path = f"assets/saveimage/{item['variant_id']}.jpg"
                    if os.path.isfile(file_path):
                        pass
                    else:
                        print(f"{MAIN_URL}/variants/{item['variant_id']}.json")
                        variant = js_get_url(f"{MAIN_URL}/variants/{item['variant_id']}.json")['variant']
                        variant["image"] = variant["images"][0]["full_path"]
                        r = requests.get(variant["image"], allow_redirects=True)
                        open(file_path, 'wb').write(r.content)


    return render(request, 'kho_donbansot.html', obj)

def kho_setupmayin(request):

    if 'mayin' in request.GET:
        mayin = request.GET['mayin']
        
        print("Set up printer: "+mayin)
        writefile(mayin, 'logs/mayin.log')

    return render(request, 'kho_setupmayin.html',{'mayin': mayin})

def kho_thudoitra(request):
    obj = {
        "save": {}
    }

    if 'username' in request.GET:
        #Input vào tên của khách hàng.
        username = request.GET['username']
        obj["save"]["username"] = username
    
    if 'shop_name' in request.GET:
        #Input vào tên của khách hàng.
        shop_name = request.GET['shop_name']
        obj["save"]["shop_name"] = shop_name

    if 'sanphamthieu' in request.GET:
        #Input vào tên của khách hàng.
        sanphamthieu = request.GET['sanphamthieu']
        obj["save"]["sanphamthieu"] = sanphamthieu

    if 'phuonganxuly' in request.GET:
        #Input vào tên của khách hàng.
        phuonganxuly = request.GET['phuonganxuly']
        obj["save"]["phuonganxuly"] = phuonganxuly

    if 'print' in request.GET:
        return render(request, 'kho_thudoitra_html.html', obj)
    return render(request, 'kho_thudoitra.html', obj)

def kho_scanpacking(request):
    obj = {
        'nguoi_goi':'',
        'nhanvien_data': {},
        'pwd': ''
    }

    obj['nhanvien_data'] = read_nhanvien_pwd(1)
    obj['nguoi_goi'] = request.session.get('nguoi_goi', '')
    obj['pwd']        = request.session.get('pwd', '')

    return render(request, 'kho_scanpacking.html', obj)

def kho_donhang(request):
    ##check_login_sapo()
    if request.method == 'GET':
        obj = {
            "boloc": {'channel': '', 'dvvc': ''},
            "channel": {
                "Shopee": {"name":"Shopee","total":0},
                "Lazada": {"name":"Lazada","total":0},
                "Tiki": {"name":"Tiki","total":0},
                "Tiktok": {"name":"Tiktok","total":0},
                "Sapo": {"name":"Sapo Orders","total":0}
            },
            "list_dvvc":[],
            "list_orders": []
        }
        dvvc_map = {474733:'GHSV',467880: 'ViettelPost',67310: 'Sapo Express',69025: 'GHTK',204246: 'BEST Express',70835: 'Tự Ship / Hà Nội',190219: 'Hola Ship',373257: 'Shopee Xpress',373707: 'J&T Express' }
        list_id = []
        list_order = []

        if 'search' in request.GET:
            orders = js_get_url(f"{MAIN_URL}/orders.json?query={request.GET['search']}")
            for order in orders['orders']:
                order['dvvc'] = ''
                list_order.append(order)            
        else:
            for i in range(1,1):
                # Lấy tất cả những đơn hàng ở trạng thái (yêu cầu gói & đã gói)
                orders = js_get_url(f"{MAIN_URL}/orders.json?status=finalized%2Ccompleted&limit=250&page="+str(i))
                for order in orders['orders']:
                    order['dvvc'] = ''
                    list_order.append(order)

        channel = ''
        if "channel" in request.GET:
            channel = request.GET["channel"]
            obj['boloc']['channel'] = channel
        dvvc = ''
        if "dvvc" in request.GET:
            dvvc = request.GET["dvvc"]
            obj['boloc']['dvvc'] = dvvc

        for order in list_order:
            for item in order['order_line_items']:
                if item['product_type'] == "normal" and item['is_packsize'] == True:
                    item['quantity'] = int(item['quantity']*item['pack_size_quantity'])
            order["created_on"] = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ")
            
            #Xử lý data
            order = get_data_packing(order)

            if order['fulfillment_status'] == 'shipped':
                order['ship_on'] = datetime.datetime.strptime(order["fulfillments"][-1]["shipped_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)

            # Lấy thông tin từ Shopee
            tmdt_order = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=1&connectionIds={connectID}&query={order['reference_number']}")['orders'][0]

            directory_path = 'logs/print-cover/'

            shop_name = existing_shop_map.get(tmdt_order['connection_id'], "")
            link_to_cover = directory_path + "/" + shop_name + "/"
            if doi_shop(tmdt_order['connection_id'],loginsp) == 1:
                if shop_name == 'giadungplus_official':
                    shop_id = 241961702
                elif shop_name == 'phaledo':
                    shop_id = 1009027554

                try:
                    URL = f"https://banhang.shopee.vn/api/v3/order/get_order_list_search_bar_hint?keyword={order['reference_number']}&&category=1&order_list_tab=100&SPC_CDS=a2c0b37e-fa4d-420d-a821-dc94b4265519&SPC_CDS_VER=2"
                    load_data = loginsp.get(URL).text
                    #Sẽ trả về order_id, nếu không có order_id tức là request thất bại.

                    SHOPEE_ID =json.loads(load_data)['data']['order_sn_result']['list'][0]['order_id']
                    print("Shopee ID:" + str(SHOPEE_ID))
                    URL = f"https://banhang.shopee.vn/api/v3/logistics/get_logistics_tracking_history?SPC_CDS=74a0f00a-140f-49c7-87b8-98cf0a4b0397&SPC_CDS_VER=2&order_id={SHOPEE_ID}"

                    load_data = json.loads(loginsp.get(URL).text)
                    order['history_gv'] = load_data['data']['list'][0]

                    for data in order['history_gv']['tracking_info']:
                        data['date'] = datetime.datetime.fromtimestamp(int(data['ctime']))
                except Exception as e:
                    print(f"Error print: {e}")

            if len(order['tags']) >= 2:
                order['shop'] = order['tags'][1].split("_")[-1:][0]
                    
            # Đơn TMĐT hoặc quay lại -> giao ngoài
            if order['source_id'] in [1880152,6510687,1880149,1880150] and order['channel'] == None:
                obj['channel']['Sapo']['total'] += 1
                if channel == 'sapo' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
            # Đơn Sapo
            elif order["source_id"] in [5483992,4864539,4893087,4339735,1880148,1880147,1880146]:
                obj['channel']['Sapo']['total'] += 1
                if channel == 'sapo' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)          
                          
            # Đơn Shopee real
            elif order["source_id"] in [1880152] and order['channel'] != None:
                obj['channel']['Shopee']['total'] += 1
                if channel == 'shopee' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                       
            # Đơn Tiktok
            elif order["source_id"] in [6510687]:
                obj['channel']['Tiktok']['total'] += 1
                if channel == 'tiktok' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                            
            #Đơn Lazada
            elif order["source_id"] in [1880149]:
                obj['channel']['Lazada']['total'] += 1
                if channel == 'lazada' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                   
            # Đơn Tiki
            elif order["source_id"] in [1880150]:
                obj['channel']['Tiki']['total'] += 1
                if channel == 'tiki' or channel == '':
                    if order['dvvc'] not in obj['list_dvvc']:
                        obj['list_dvvc'].append(order['dvvc'])
                    if order['dvvc'] == dvvc or dvvc == '':
                        obj['list_orders'].append(order)
                    
            
            for item in order["order_line_items"]:
                if item['variant_id'] != None:
                    file_path = f"assets/saveimage/{item['variant_id']}.jpg"
                    if os.path.isfile(file_path):
                        pass
                    else:
                        print(f"{MAIN_URL}/variants/{item['variant_id']}.json")
                        variant = js_get_url(f"{MAIN_URL}/variants/{item['variant_id']}.json")['variant']
                        variant["image"] = variant["images"][0]["full_path"]
                        r = requests.get(variant["image"], allow_redirects=True)
                        open(file_path, 'wb').write(r.content)

        return render(request, 'kho_donhang.html',obj)

def kho_packing(request):
    #check_login_sapo()
    obj = {
        'order': {},
        'error': 0,
        'notice': '',
        'print_status': 'ON',
        'nhanvien_data': {}
    }
    nguoi_goi = "Quý"
    flag = 0
    pwd = ""
    if 'nguoi_goi' in request.GET:
        nguoi_goi = request.GET['nguoi_goi']
        request.session['nguoi_goi'] = nguoi_goi

    if 'pwd' in request.GET:
        pwd = request.GET['pwd']
        request.session['pwd'] = pwd

    nguoi_goi = request.session['nguoi_goi']
    pwd = request.session['pwd']

    obj['pwd'] = pwd
    obj['nguoi_goi'] = nguoi_goi
     
    obj['nhanvien_data'] = read_nhanvien_pwd(1)

    flag = 0
    for nhanvien in obj['nhanvien_data']:
        f_nv = f"{nhanvien['kho']}: {nhanvien['name']}"
        if nguoi_goi == f_nv:
            if pwd == nhanvien['pwd']:
                flag = 1

    if flag == 1:
        if 'mavandon' in request.GET:
            mavandon = request.GET['mavandon'].replace(" ", "")
            save_mavandon = mavandon
            order_id = js_get_url(f"{MAIN_URL}/fulfillments.json?page=1&limit=1&query={mavandon}")['fulfillments']
            if len(order_id) > 0:
                order_id = order_id[0]['order_id']
            else:
                mavandon = search_order_shopee(mavandon)
                if len(mavandon) > 2:
                    order_id = js_get_url(f"{MAIN_URL}/fulfillments.json?page=1&limit=1&query={mavandon}")['fulfillments']
                    order_id = order_id[0]['order_id']
                    SEARCH_FLAG = 1
                else:
                    obj['error'] = 1
                    obj['notice'] = f"Không tìm được vận đơn liên quan!"
                    return render(request, 'kho_packing.html',obj)
            
        if 'order_id' in request.GET:
            order_id = request.GET['order_id']

        if 'done' in request.GET:
            order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")["order"]
            #Xử lý data gói hàng đang lưu trên Sapo.
            order["split"] = 1
            order = get_data_packing(order)

            if order["split"] > 1 and 'dontach' in request.GET:
                dontach = int(request.GET['dontach'])
                if dontach == 1:
                    update_data_one(order["code"],{'nguoi_goi':nguoi_goi,'time_packing': datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})
                else:
                    update_data_one(order["code"],{f"nguoi_goi{dontach}":nguoi_goi,f"time_packing{dontach}": datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})

                order = get_data_packing(order)
                pack_goi = 0
                for i in range(order["split"]):

                    if i == 0:
                        if 'nguoi_goi' in order:
                            pack_goi += 1
                    else:
                        j = i+1
                        if f"nguoi_goi{j}" in order:
                            pack_goi += 1
                print(f"So don tach da goi: {pack_goi}")
                if pack_goi == order["split"]:
                    update_data_one(order["code"],{f"packing_status":4})
            else:
                update_data_one(order["code"],{f"packing_status":4,'nguoi_goi':nguoi_goi,'time_packing': datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})

            #send_zns_xacnhan(243976, order, 2)
            return redirect('kho_scanpacking')
        else:
            order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")['order']
            order["split"] = 1
            order = get_data_packing(order)

            if order["packing_status"] == 4:
                obj['error'] = 1
                obj['notice'] = f"Đơn hàng đã được {order["nguoi_goi"]} gói vào lúc {order["time_packing"]}!"

            # Kiểm tra trạng thái (đã giao/ đã hoàn thành / đã huỷ/ khách yêu cầu huỷ)
            if order['status'] in ['cancelled']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã huỷ!'
            elif order['status'] in ['complete']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã hoàn thành!'

            if order['fulfillments'][-1]['composite_fulfillment_status'] in ['retry_delivery','fulfilled','received','packed_cancelled_client','fulfilled_cancelling','fulfilled_cancelled']:
                obj['error'] = 1
                obj['notice'] = 'Đơn hàng đã giao đi rồi!'

            # Kiểm tra xem đơn đang yêu cầu huỷ hay không.
            orders = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?channelType=1,2,6,4&connectionIds={connectID}&page=1&limit=250&channelOrderStatus=IN_CANCEL&sortBy=ISSUED_AT&orderBy=desc")
            if len(orders) > 100:
                for x_order in orders["orders"]:
                    if order['id'] == x_order['sapo_order_id']:
                        obj['error'] = 1
                        obj['notice'] = 'Khách hàng đang yêu cầu xin huỷ đơn này!'
                        break
            # Kiểm tra xem đơn hàng đã có MVĐ hay chưa ?
            if order['fulfillments'][-1]['shipment'] == None:
                obj['error'] = 4
            

            XULY = 0
            # Xử lý về vấn đề là có duyệt nữa hay ko. Ko soi -> biến scan=true, scan=on . Ngược lại thì ko có scan, hoặc scan=false.
            if 'scan' in request.GET:
                scan = request.GET['scan']
                print(scan)
                if scan == 'true' or scan == 'on':
                    XULY = 0
                else:
                    XULY = 1
            else:
                XULY = 1

            if XULY == 1:
                # Xử lý kèm keo
                list_kemkeo = []
                all_kemkeo = js_get_url(f"{MAIN_URL}/products.json?tags=add_keo_dan&page=1&litmit=250")["products"]
                for pr in all_kemkeo:
                    for vari in pr['variants']:
                        list_kemkeo.append(vari['id'])
                order['kemkeo'] = 0
                order["REAL_TACH"] = 0

                if obj['error'] == 0:
                    
                    order['real_items'] = []
                    for item in order['order_line_items']:
                        if item['variant_id'] in list_kemkeo:
                            order['kemkeo'] = 1

                        if item['product_type'] == "normal" and item['is_packsize'] == False:
                            flag = 0
                            for line in order['real_items']:
                                if line['id'] == item['variant_id']:
                                    line['quantity'] += int(item['quantity'])
                                    flag = 1
                                    break
                            if flag == 0:
                                order['real_items'].append({
                                    'id': item['variant_id'],
                                    'sku': item['sku'][3:],
                                    'variant_options': item['variant_options'],  
                                    'quantity': int(item['quantity']),
                                    'unit': item['unit'],
                                    'old_id': 0,
                                    'barcode': item['barcode'],
                                    'print':0
                                })
                        elif item['product_type'] == "normal" and item['is_packsize'] == True:
                            flag = 0
                            for line in order['real_items']:
                                if line['id'] == item['pack_size_root_id']:
                                    line['quantity'] += int(item['quantity']*item['pack_size_quantity'])
                                    flag = 1
                                    break
                            if flag == 0:
                                vari = js_get_url(f"{MAIN_URL}/variants/{item['pack_size_root_id']}.json")['variant']
                                order['real_items'].append({
                                    'id': vari['id'],
                                    'sku': vari['sku'][3:],
                                    'barcode': vari['barcode'],
                                    'variant_options': vari['opt1'],
                                    'old_id': item['variant_id'],
                                    'unit': 'chiếc (tách từ combo)',
                                    'quantity': int(item['quantity']*item['pack_size_quantity']),
                                    'print':0
                                })

                        if item['product_type'] == "composite":
                            for xitem in item['composite_item_domains']:
                                flag = 0
                                for line in order['real_items']:
                                    if line['id'] == xitem['variant_id']:
                                        line['quantity'] += int(xitem['quantity'])
                                        flag = 1
                                        break
                                if flag == 0:
                                    vari = js_get_url(f"{MAIN_URL}/variants/{xitem['variant_id']}.json")['variant']
                                    order['real_items'].append({
                                        'id': xitem['variant_id'],
                                        'sku': vari['sku'][3:],
                                        'barcode': vari['barcode'],
                                        'old_id': item['variant_id'],
                                        'unit': 'chiếc',
                                        'variant_options': vari['opt1'],  
                                        'quantity': int(xitem['quantity']),
                                        'print':0
                                    })
                        if item['product_type'] == None:
                            order['real_items'].append({
                                'id': 0,
                                'sku': "SERVICE",
                                'unit': item['unit'],
                                'old_id': 0,
                                'variant_options': item['note'],  
                                'quantity': int(item['quantity']),
                                'print':0
                            })   
                            
                    order['total_quantity'] = 0
                    for line in order['real_items']:
                        order['total_quantity'] += line['quantity']

                    if order['note'] != None:
                        if len(order['note']) > 5:
                            file_path = 'assets/notify/have_note.mp3'
                            """
                            pygame.mixer.init()
                            pygame.mixer.music.load(file_path)
                            pygame.mixer.music.play()
                            """

                if order["split"] > 1 and obj["error"] == 0:
                    print("XỬ LÝ ĐƠN TÁCH KHI GÓI HÀNG!")
                    order['d_tach_values'] = list(range(1, order['split'] + 1))

                    tmdt_order = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=1&connectionIds={connectID}&query={order['reference_number']}")['orders'][0]
                    logintmdt.put(f"https://market-place.sapoapps.vn/v2/orders/sync?ids={tmdt_order['id']}&accountId=3199")

                    shop_name = existing_shop_map.get(tmdt_order['connection_id'], "")

                    if doi_shop(tmdt_order['connection_id'],loginsp) == 1:
                        if shop_name == 'giadungplus_official':
                            shop_id = 241961702
                        elif shop_name == 'phaledo':
                            shop_id = 1009027554


                    #Thông tin kiện hàng
                    URL = f"https://banhang.shopee.vn/api/v3/order/batch_get_packages_multi_shop?SPC_CDS=f8cfde35-a66a-4a25-8b35-c7ea97c759aa&SPC_CDS_VER=2"
                    RS = json.loads(loginsp.post(URL,json={"orders":[{"order_id":order["shopee_id"],"shop_id":shop_id,"region_id":"VN"}]}).text)
                    

                    URL = f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=20&connectionIds={connectID}&query={order['reference_number']}&sortBy=ISSUED_AT&orderBy=desc"
                    ORDER_TMDT = js_get_url(URL)["orders"][0]
                    for x in order['real_items']:
                        x["item_id_shopee"] = 0

                    for pr in ORDER_TMDT["products"]:
                        for x in order['real_items']:
                            if int(pr["sapo_variant_id"]) == int(x["id"]) or int(pr["sapo_variant_id"]) == int(x["old_id"]):
                                x["item_id_shopee"] = pr["variation_id"]

                    LIST_PACK =RS['data']['list'][0]["package_list"]
                    D_TACH = 0
                    #Nếu nhập vào là mã vận đơn.
                    if not save_mavandon.startswith('SON'):
                        for PACKEDD in LIST_PACK:
                            D_TACH +=1
                            #Nếu nó là đơn tách này thì mới xử lý.
                            if save_mavandon == PACKEDD['third_party_tn']:
                                order["REAL_TACH"] = D_TACH
                                flag = 0
                                for s_item in PACKEDD["items"]:
                                    for x in order['real_items']:
                                        if str(x["item_id_shopee"]) == str(s_item["model_id"]):
                                            x["dontach"] = D_TACH
                                            flag = 1

                                if flag == 0:
                                    for x in order['real_items']:
                                        if x["item_id_shopee"] == 0:
                                            x["dontach"] = D_TACH
                    else:
                        D_TACH +=1
                        flag = 0
                        for s_item in PACKEDD["items"]:
                            for x in order['real_items']:
                                if str(x["item_id_shopee"]) == str(s_item["model_id"]):
                                    x["dontach"] = D_TACH
                                    flag = 1

                        if flag == 0:
                            for x in order['real_items']:
                                if x["item_id_shopee"] == 0:
                                    x["dontach"] = D_TACH

                obj['order'] = order
                return render(request, 'kho_packing.html',obj)
    
            else:
                if obj['error'] == 0:
                    obj['thongbao'] = "Scan gói hàng thành công!"
                    order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")["order"]
                    #Xử lý data gói hàng đang lưu trên Sapo.
                    order["split"] = 1
                    order = get_data_packing(order)

                    if order["split"] > 1 and 'dontach' in request.GET:
                        dontach = int(request.GET['dontach'])
                        if dontach == 1:
                            update_data_one(order["code"],{'nguoi_goi':nguoi_goi,'time_packing': datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})
                        else:
                            update_data_one(order["code"],{f"nguoi_goi{dontach}":nguoi_goi,f"time_packing{dontach}": datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})

                        order = get_data_packing(order)
                        pack_goi = 0
                        for i in range(order["split"]):

                            if i == 0:
                                if 'nguoi_goi' in order:
                                    pack_goi += 1
                            else:
                                j = i+1
                                if f"nguoi_goi{j}" in order:
                                    pack_goi += 1
                        print(f"So don tach da goi: {pack_goi}")
                        if pack_goi == order["split"]:
                            update_data_one(order["code"],{f"packing_status":4})
                    else:
                        update_data_one(order["code"],{f"packing_status":4,'nguoi_goi':nguoi_goi,'time_packing': datetime.datetime.now().strftime("%H:%M %d-%m-%Y")})

                obj['order'] = order
                return render(request, 'kho_scanpacking.html',obj)

    else:
        obj['error'] = 1
        obj['notice'] = "Mật khẩu không chính xác!"
        return render(request, 'kho_packing.html',obj) 
      
@csrf_exempt
def api_kho_packinginfo(request):
    try:
        payload = json.loads(request.body)
        nguoi_goi = payload.get('nguoi_goi', '').strip()
        pwd       = payload.get('pwd', '').strip()
        mavandon  = payload.get('mavandon', '').strip()
        save_mavandon = mavandon
        request.session['nguoi_goi'] = nguoi_goi
        request.session['pwd']       = pwd
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 1, 'notice': 'Dữ liệu đầu vào không hợp lệ.'}, status=400)


    # 1. Xác thực người gói
    nhanvien_data = read_nhanvien_pwd(1)
    valid = any(f"{nv['kho']}: {nv['name']}" == nguoi_goi and nv['pwd'] == pwd
                for nv in nhanvien_data)
    if not valid:
        return JsonResponse({'error': 1, 'notice': 'Mật khẩu hoặc người gói không chính xác.'}, status=403)

    # 2. Tìm order_id từ mã vận đơn
    resp = js_get_url(f"{MAIN_URL}/fulfillments.json?page=1&limit=1&query={mavandon}")
    ful = resp.get('fulfillments', [])
    if ful:
        order_id = ful[0]['order_id']
    else:
        # Fallback tìm qua Shopee
        print("[+] Fallback: Tìm order thông qua Shopee API")
        fallback = search_order_shopee(mavandon)
        if len(fallback) > 2:
            resp2 = js_get_url(f"{MAIN_URL}/fulfillments.json?page=1&limit=1&query={fallback}")
            ful2 = resp2.get('fulfillments', [])
            if ful2:
                order_id = ful2[0]['order_id']
            else:
                return JsonResponse({'error': 1, 'notice': f'Không tìm thấy vận đơn với keyword: {mavandon}'})
        else:
            return JsonResponse({'error': 1, 'notice': f'Không tìm thấy vận đơn với keyword: {mavandon}'})

    # 3. Lấy chi tiết đơn và xử lý packing
    order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json").get('order', {})
    order['split'] = 1
    order = get_data_packing(order)

    # --- KIỂM TRA TRẠNG THÁI TRƯỚC KHI TRẢ VỀ DATA ---
    status = order.get('status')
    if status == 'cancelled':
        return JsonResponse({'error': 1, 'notice': f'Đơn {order['code']} đã huỷ!'})
    elif status == 'complete':
        return JsonResponse({'error': 1, 'notice': f'Đơn {order['code']}  đã hoàn thành!'})

    final_ful = order.get('fulfillments', [])[-1].get('composite_fulfillment_status', '')
    if final_ful in [
        'retry_delivery','fulfilled','received',
        'packed_cancelled_client','fulfilled_cancelling','fulfilled_cancelled'
    ]:
        return JsonResponse({'error': 1, 'notice': f'Đơn {order['code']} đã giao đi rồi!'})

    if order.get("packing_status") == 4:
        return JsonResponse({
            'error': 1,
            'notice': f'Đơn hàng đã được {order.get("nguoi_goi", "")} gói vào lúc {order.get("time_packing", "")}!'
        })
    cancel_resp = js_get_url(
        "https://market-place.sapoapps.vn/v2/orders"
        "?channelType=1,2,6,4"
        f"&connectionIds={connectID}"
        "&page=1&limit=250"
        "&channelOrderStatus=IN_CANCEL"
        "&sortBy=ISSUED_AT&orderBy=desc"
    )
    if any(o.get('sapo_order_id') == order.get('id') for o in cancel_resp.get('orders', [])):
        return JsonResponse({'error': 1, 'notice': 'Khách hàng đang yêu cầu xin huỷ đơn này!'})
    
    # --- END KIỂM TRA TRẠNG THÁI ---
    # --- 1. Lấy danh sách variant phải dán keo ---
    resp = js_get_url(f"{MAIN_URL}/products.json?tags=add_keo_dan&page=1&limit=250")
    kemkeo_ids = [
        v['id']
        for pr in resp.get('products', [])
        for v in pr.get('variants', [])
    ]
    # Đánh dấu đơn có keo hay không
    order['kemkeo'] = int(any(
        item.get('variant_id') in kemkeo_ids
        for item in order.get('order_line_items', [])
    ))

    order['REAL_TACH'] = 0
    # --- 2. Xử lý real_items chỉ khi không có lỗi trước đó ---

    grouped = {}  # sẽ gom nhóm theo variant_id (hoặc id root)

    for item in order.get('order_line_items', []):
        ptype = item.get('product_type')
        qty_raw = int(item.get('quantity', 0))

        # 2.1. NORMAL không tách combo
        if ptype == 'normal' and not item.get('is_packsize', False):
            key = item['variant_id']
            qty = qty_raw
            name = item.get('variant_options', '')
            unit = item.get('unit', '')
            barcode = item.get('barcode', '')

            # gom nhóm
            if key in grouped:
                grouped[key]['quantity'] += qty
            else:
                grouped[key] = {
                    'id': key,
                    'sku': item.get('sku', '')[3:],
                    'variant_options': name,
                    'quantity': qty,
                    'unit': unit,
                    'old_id': 0,
                    'barcode': barcode,
                    'print': 0
                }

        # 2.2. NORMAL tách combo (packsize)
        elif ptype == 'normal' and item.get('is_packsize', False):
            root_id = item['pack_size_root_id']
            factor = int(item.get('pack_size_quantity', 1))
            qty = qty_raw * factor
            # lấy info của root variant
            vari = js_get_url(f"{MAIN_URL}/variants/{root_id}.json").get('variant', {})
            name = vari.get('opt1', '')
            unit = 'chiếc (từ combo)'
            barcode = vari.get('barcode', '')

            if root_id in grouped:
                grouped[root_id]['quantity'] += qty
            else:
                grouped[root_id] = {
                    'id': root_id,
                    'sku': vari.get('sku','')[3:],
                    'variant_options': name,
                    'quantity': qty,
                    'unit': unit,
                    'old_id': item['variant_id'],
                    'barcode': barcode,
                    'print': 0
                }

        # 2.3. COMPOSITE (đã ghép sẵn)
        elif ptype == 'composite':
            for ci in item.get('composite_item_domains', []):
                cid = ci['variant_id']
                qty = int(ci.get('quantity', 0))
                vari = js_get_url(f"{MAIN_URL}/variants/{cid}.json").get('variant', {})
                name = vari.get('opt1', '')
                unit = 'chiếc (từ combo)'
                barcode = vari.get('barcode','')

                if cid in grouped:
                    grouped[cid]['quantity'] += qty
                else:
                    grouped[cid] = {
                        'id': cid,
                        'sku': vari.get('sku','')[3:],
                        'variant_options': name,
                        'quantity': qty,
                        'unit': unit,
                        'old_id': item['variant_id'],
                        'barcode': barcode,
                        'print': 0
                    }

        # 2.4. Dịch vụ / Không xác định
        else:
            key = f"service_{len(grouped)}"
            qty = qty_raw
            name = item.get('note', '')
            unit = item.get('unit', '')
            grouped[key] = {
                'id': 0,
                'sku': 'SERVICE',
                'variant_options': name,
                'quantity': qty,
                'unit': unit,
                'old_id': 0,
                'barcode': '',
                'print': 0
            }

    order['real_items'] = list(grouped.values())

    try:
        if order["split"] > 1:
            print("[+] PACED SPLIT CHECK!")
            order['d_tach_values'] = list(range(1, order['split'] + 1))

            tmdt_order = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=1&connectionIds={connectID}&query={order['reference_number']}")['orders'][0]
            logintmdt.put(f"https://market-place.sapoapps.vn/v2/orders/sync?ids={tmdt_order['id']}&accountId=3199")

            shop_name = existing_shop_map.get(tmdt_order['connection_id'], "")

            if doi_shop(tmdt_order['connection_id'],loginsp) == 1:
                if shop_name == 'giadungplus_official':
                    shop_id = 241961702
                elif shop_name == 'phaledo':
                    shop_id = 1009027554

            #Thông tin kiện hàng
            URL = f"https://banhang.shopee.vn/api/v3/order/batch_get_packages_multi_shop?SPC_CDS=f8cfde35-a66a-4a25-8b35-c7ea97c759aa&SPC_CDS_VER=2"
            RS = json.loads(loginsp.post(URL,json={"orders":[{"order_id":order["shopee_id"],"shop_id":shop_id,"region_id":"VN"}]}).text)    

            URL = f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=20&connectionIds={connectID}&query={order['reference_number']}&sortBy=ISSUED_AT&orderBy=desc"
            ORDER_TMDT = js_get_url(URL)["orders"][0]
            for x in order['real_items']:
                x["item_id_shopee"] = 0

            for pr in ORDER_TMDT["products"]:
                for x in order['real_items']:
                    if int(pr["sapo_variant_id"]) == int(x["id"]) or int(pr["sapo_variant_id"]) == int(x["old_id"]):
                        x["item_id_shopee"] = pr["variation_id"]

            LIST_PACK =RS['data']['list'][0]["package_list"]
            D_TACH = 0
            #Nếu nhập vào là mã vận đơn.
            if not save_mavandon.startswith('SON'):
                for PACKEDD in LIST_PACK:
                    D_TACH +=1
                    #Nếu nó là đơn tách này thì mới xử lý.
                    if save_mavandon == PACKEDD['third_party_tn']:
                        order["REAL_TACH"] = D_TACH
                        flag = 0
                        for s_item in PACKEDD["items"]:
                            for x in order['real_items']:
                                if str(x["item_id_shopee"]) == str(s_item["model_id"]):
                                    x["dontach"] = D_TACH
                                    flag = 1

                        if flag == 0:
                            for x in order['real_items']:
                                if x["item_id_shopee"] == 0:
                                    x["dontach"] = D_TACH

    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 1, 'notice': 'Dữ liệu đầu vào không hợp lệ.'}, status=400)
    

    return JsonResponse({
        'error': 0,
        'order': order
    }, status=200)
         
@csrf_exempt
def api_kho_packingcomplete(request):
    try:
        payload   = json.loads(request.body)
        nguoi_goi = payload.get('nguoi_goi', '').strip()
        order_id  = payload.get('order_id', '').replace(' ', '')
        dontach   = payload.get('dontach')  # may be None
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 1, 'notice': 'Dữ liệu không hợp lệ.'}, status=400)

    # 3. Lấy và xử lý đơn
    order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json").get('order', {})
    order['split'] = 1
    order = get_data_packing(order)

     # 4. Cập nhật trạng thái gói
    now_str = datetime.datetime.now().strftime("%H:%M %d-%m-%Y")
    if order.get('split', 1) > 1 and dontach is not None:
        d = int(dontach)
        if d == 1:
            update_data_one(order['code'], {
                'nguoi_goi':  nguoi_goi,
                'time_packing': now_str
            })
        else:
            update_data_one(order['code'], {
                f'nguoi_goi{d}':   nguoi_goi,
                f'time_packing{d}': now_str
            })
        pack_goi = 0
        for i in range(order['split']):
            if i == 0:
                if 'nguoi_goi' in order:
                    pack_goi += 1
            else:
                if f'nguoi_goi{i+1}' in order:
                    pack_goi += 1
        if pack_goi == order['split']:
            update_data_one(order['code'], {'packing_status': 4})
    else:
        update_data_one(order['code'], {
            'packing_status': 4,
            'nguoi_goi':      nguoi_goi,
            'time_packing':   now_str
        })

    return JsonResponse({'error': 0, 'order': order})

def kho_indonsapo(request):

    json_order = js_get_url(f"{MAIN_URL}/orders/{request.GET['order_id']}.json")['order']
    
    json_order["tracking_number"] = json_order["fulfillments"][-1]["shipment"]["tracking_code"]
    json_order["route_number"] = json_order["fulfillments"][-1]["shipment"]["shipping_address"]["city"]
    ean = Code128(json_order["tracking_number"], writer=ImageWriter())
    filename = ean.save("assets/barcode-son/bar-"+str(json_order["code"]),options={"write_text": False,"module_width":1, "module_height":30, "font_size": 12, "text_distance": -3, "quiet_zone": 1})
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=20,
        border=1,
    )

    qr.add_data(json_order["tracking_number"])
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save("assets/barcode-qr/qr-"+str(json_order["code"])+".png")

    json_order["shipping_name"] = json_order["fulfillments"][-1]["shipment"]["shipping_address"]["full_name"]
    json_order["shipping_phone"] = json_order["fulfillments"][-1]["shipment"]["shipping_address"]["phone_number"][:-3] +"***"
    if json_order["source_id"] == 1880152:
        json_order["shipping_address"] = json_order["fulfillments"][-1]["shipment"]["shipping_address"]["address1"]
    else:
        json_order["shipping_address"] = json_order["fulfillments"][-1]["shipment"]["shipping_address"]["address1"] + " - " + str(json_order["fulfillments"][-1]["shipment"]["shipping_address"]["ward"]) + " - " + json_order["fulfillments"][-1]["shipment"]["shipping_address"]["district"] + " - " + json_order["fulfillments"][-1]["shipment"]["shipping_address"]["city"]
    json_order["cod"] = format(int(json_order["fulfillments"][-1]["shipment"]["cod_amount"]), ',d')
    json_order["camera_on"] = json_order["fulfillments"][-1]["packed_on"].replace("T","  #").replace("Z", "")
    json_order["weight"] = "%.2f" % float(json_order["fulfillments"][-1]["shipment"]["weight"]/1000)
    
    json_order['real_items'] = []
    for item in json_order['order_line_items']:
        if item['product_type'] == "normal" and item['is_packsize'] == False:
            flag = 0
            for line in json_order['real_items']:
                if line['id'] == item['variant_id']:
                    line['quantity'] += int(item['quantity'])
                    flag = 1
                    break
            if flag == 0:
                json_order['real_items'].append({
                    'id': item['variant_id'],
                    'sku': item['sku'][3:],
                    'variant_options': item['variant_options'],  
                    'quantity': int(item['quantity']),
                    'unit': item['unit']
                })
        elif item['product_type'] == "normal" and item['is_packsize'] == True:
            flag = 0
            for line in json_order['real_items']:
                if line['id'] == item['pack_size_root_id']:
                    line['quantity'] += int(item['quantity']*item['pack_size_quantity'])
                    flag = 1
                    break
            if flag == 0:
                vari = js_get_url(f"{MAIN_URL}/variants/{item['pack_size_root_id']}.json")['variant']
                json_order['real_items'].append({
                    'id': vari['id'],
                    'sku': vari['sku'][3:],
                    'variant_options': vari['opt1'],
                    'unit': 'chiếc',
                    'quantity': int(item['quantity']*item['pack_size_quantity'])
                })

        if item['product_type'] == "composite":
            for xitem in item['composite_item_domains']:
                flag = 0
                for line in json_order['real_items']:
                    if line['id'] == xitem['variant_id']:
                        line['quantity'] += int(xitem['quantity'])
                        flag = 1
                        break
                if flag == 0:
                    print()
                    vari = js_get_url(f"{MAIN_URL}/variants/{xitem['variant_id']}.json")['variant']
                    json_order['real_items'].append({
                        'id': xitem['variant_id'],
                        'sku': vari['sku'][3:],
                        'unit': 'chiếc',
                        'variant_options': vari['opt1'],  
                        'quantity': int(xitem['quantity'])
                    })
        if item['product_type'] == None:
            json_order['real_items'].append({
                'id': 0,
                'sku': "SERVICE",
                'unit': item['unit'],
                'variant_options': item['note'],  
                'quantity': int(item['quantity'])
            })    

    json_order['total_quantity'] = 0
    for line in json_order['real_items']:
        json_order['total_quantity'] += line['quantity']
    
    return render(request, 'kho_phieugiaohangsapo.html',{'order':json_order})

def kho_printnow(request):
    order_ids = request.GET['orders_id'].split(",")
    prints='yes'

    if 'prints' in request.GET:
        prints = request.GET['prints']

    for order_id in order_ids:
        order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")['order']
        order['real_items'] = []
        for item in order['order_line_items']:
            item['pr_name'] = ''
            
            # Sản phẩm thường
            if item['product_name'] != None:
                if '/' in item['product_name']:
                    item['pr_name'] = item['product_name'].split('/')[0]

            if item['unit'] == None:
                item['unit'] = "cái"
            if item['product_type'] == "normal" and item['is_packsize'] == False:
                flag = 0
                for line in order['real_items']:
                    if line['id'] == item['variant_id']:
                        line['quantity'] += int(item['quantity'])
                        flag = 1
                        break
                if flag == 0:
                    order['real_items'].append({
                        'old_id': 0,
                        'id': item['variant_id'],
                        'sku': item['sku'][3:],
                        'variant_options': item['variant_options'],  
                        'quantity': int(item['quantity']),
                        'unit': item['unit'],
                        'pr_name':item['pr_name'],
                        'print':0
                    })

            # Sản phẩm thường & packed nhiều.
            elif item['product_type'] == "normal" and item['is_packsize'] == True:
                flag = 0
                for line in order['real_items']:
                    if line['id'] == item['pack_size_root_id']:
                        line['quantity'] += int(item['quantity']*item['pack_size_quantity'])
                        flag = 1
                        break
                if flag == 0:
                    vari = js_get_url(f"{MAIN_URL}/variants/{item['pack_size_root_id']}.json")['variant']
                    order['real_items'].append({
                        'old_id': item['variant_id'],
                        'id': vari['id'],
                        'sku': vari['sku'][3:],
                        'variant_options': vari['opt1'],
                        'old_id': item['variant_id'],
                        'unit': "cái",
                        'pr_name':item['pr_name'],
                        'quantity': int(item['quantity']*item['pack_size_quantity']),
                        'print':0
                    })
                    
            if item['product_type'] == "composite":
                for xitem in item['composite_item_domains']:
                    flag = 0
                    for line in order['real_items']:
                        if line['id'] == xitem['variant_id']:
                            line['quantity'] += int(xitem['quantity'])
                            flag = 1
                            break
                    if flag == 0:
                        vari = js_get_url(f"{MAIN_URL}/variants/{xitem['variant_id']}.json")['variant']
                        order['real_items'].append({
                            'id': xitem['variant_id'],
                            'sku': vari['sku'][3:],
                            'barcode': vari['barcode'],
                            'old_id': item['variant_id'],
                            'unit': "cái",
                            'pr_name':item['pr_name'],
                            'variant_options': vari['opt1'],  
                            'quantity': int(xitem['quantity']),
                            'print':0
                        })

        order['total_quantity'] = 0
        for line in order['real_items']:
            if line['sku'] != 'KEO':
                order['total_quantity'] += line['quantity']

        order['real_items'] = sorted(
            order['real_items'],
            key=lambda item: item['sku'].split('-')[0]
        )

        # Hàm lấy phần số từ SKU trước dấu '-'
        def get_sku_number(sku):
            try:
                # Tách phần trước dấu '-' và chuyển thành số (nếu có thể)
                sku_number = sku.split('-')[0]
                # Kiểm tra xem phần số có phải là số hay không
                if sku_number.isdigit():
                    return int(sku_number)
                else:
                    # Nếu không phải số, trả về một giá trị mặc định cho các SKU không phải số
                    return float('inf')  # Đặt các SKU không phải số ở cuối
            except Exception as e:
                # Nếu có lỗi trong quá trình tách hoặc chuyển đổi, trả về giá trị mặc định
                return float('inf')

        # Sắp xếp order['real_items'] theo SKU (sắp xếp theo phần số trong SKU trước dấu '-')
        order['real_items'] = sorted(order['real_items'], key=lambda item: get_sku_number(item['sku']))

        order = get_data_packing(order)

        if order['fulfillment_status'] == 'shipped':
            order['ship_on'] = datetime.datetime.strptime(order["fulfillments"][-1]["shipped_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)

        # In đơn Shopee.
        if order['source_id'] == 1880152 and order['channel'] != None and order['reference_number'] != None and order['account_id'] == 319911:
            if print_shopee(order,prints) == 1:
                if order['packing_status'] < 3:
                    order['packing_status'] = 3
                    time.sleep(1)
                    x = update_data_one(order["code"],{'packing_status':order['packing_status']})
                else:
                    print("[e] Double print error!")
            else:
                print("LOI KHI IN DON SHOPEE - LIEN HE ADMIN NGOC VUONG!")

        # Đơn Lazada Tiki các kiểu
        elif order['fulfillments'][-1]['shipment'] != None:
            if order['fulfillments'][-1]['shipment']['delivery_service_provider_id'] in [122586,248139,145496,128395]:
                print("No print Lazada, Tiki, Tiktok")
            else:
                print("Sapo!")
                if order['fulfillments'][-1]['shipment']["tracking_code"][:3] == "FUN":
                    pass
                else:
                    print_sapo(order['id'])
                    x = update_data_one(order["code"],{'packing_status':3})
     
    return redirect('kho_start')

def kho_checkprint(request):
    if 'mavandon' in request.GET:
        mavandon = request.GET['mavandon'].upper()
        this_order = js_get_url(f"https://market-place.sapoapps.vn/v2/orders?page=1&limit=20&connectionIds={connectID}&query={mavandon}&sortBy=ISSUED_AT&orderBy=desc")
        
        if "orders" in this_order:
            this_order = this_order["orders"][0]

            x = update_data_one(this_order["sapo_order_code"],{'packing_status':3, 'dvvc': this_order["shipping_carrier_name"]})
 
            if x == 0:
                logintmdt.put(f"https://market-place.sapoapps.vn/v2/orders/sync?ids={this_order['id']}&accountId=3199")
                time.sleep(2)
                x = update_data_one(this_order["sapo_order_code"],{'packing_status':3, 'dvvc': this_order["shipping_carrier_name"]})
 
            if x == 0:
                logintmdt.put(f"https://market-place.sapoapps.vn/v2/orders/sync?ids={this_order['id']}&accountId=3199")
                time.sleep(2)
                x = update_data_one(this_order["sapo_order_code"],{'packing_status':3, 'dvvc': this_order["shipping_carrier_name"]})
 
    return render(request, 'kho_checkprint.html')

def kho_nhansanpham(request):

    ARRAY_INFO = []
    vari_list = []
    soluong = int(request.GET["soluong"])
    size = str(request.GET["size"])
    vari_id = int(float(request.GET["id"]))
    variant = json.loads(loginss.get(f"{MAIN_URL}/variants.json?ids="+str(vari_id)).text)["variants"][0]
    product = json.loads(loginss.get(f"{MAIN_URL}/products/{variant['product_id']}.json").text)["product"]

    brand = js_get_url(f"{MAIN_URL}/brands/{variant['brand_id']}.json")['brands']['name']
    NSX = js_get_url(f"{MAIN_URL}/suppliers.json?query={brand}")["suppliers"][0]
    
    nsx = NSX['addresses'][0]['full_name']
    diachi = NSX['addresses'][0]['address1']
    
    variant["brand_name"] = nsx              
    variant["nsx_name"] = nsx
    variant["skugon"] = variant["sku"][3:]
    variant["nsx_diachi"] = diachi 
    
    TICKET_CUSTOMER_IDS = ['760093681']
    for ticket_id in TICKET_CUSTOMER_IDS:
        url_notes = f"{MAIN_URL}/customers/{ticket_id}.json"
        res = loginss.get(url_notes)
        if res.status_code != 200:
            continue

        customer = res.json().get("customer", {})
        active_notes = [n for n in customer.get("notes", []) if n.get("status") == "active"]
        for note in active_notes:
            try:
                content = json.loads(note.get("content", "{}"))
                if int(float(variant['product_id'])) == int(float(content['product_id'])):
                    variant['vi_name'] = content['vi_name']
                    variant['en_name'] = content['en_name']
                    variant['descreption'] = content['descreption']
                    variant['material'] = content['material']
                    variant['brand'] = content['brand']

            except:
                continue

    for i in range(soluong):
        vari_list.append(variant)

    return render(request, 'kho_nhansanpham.html',{'vari_list': vari_list, 'size': size})

def kho_quanlysos(request):
    obj = {
            "boloc": {'channel': '', 'dvvc': '', 'kho':''},
            "channel": {
                "Shopee": {"name":"Shopee","total":0},
                "Lazada": {"name":"Lazada","total":0},
                "Tiki": {"name":"Tiki","total":0},
                "Tiktok": {"name":"Tiktok","total":0},
                "Sapo": {"name":"Sapo Orders","total":0}
            },
            "list_dvvc":[],
            "list_orders": []
    }
    dvvc_map = {474723:'ViettelPost',474733:'GHSV',467880: 'ViettelPost',67310: 'Sapo Express',69025: 'GHTK',204246: 'BEST Express',70835: 'Tự Ship / Hà Nội',190219: 'Hola Ship',373257: 'Shopee Xpress',373707: 'J&T Express' }
    list_id = []
    list_order = []
    # Lấy tất cả những đơn hàng ở trạng thái (yêu cầu gói & đã gói)
    orders = js_get_url(f"{MAIN_URL}/orders.json?limit=250&status=finalized%2Cdraft&fulfillment_status=unshipped&page=1")
    for order in orders['orders']:
            order['dvvc'] = ''
            list_order.append(order)
            list_id.append(order['id'])
    page = 2
    total_orders = orders['metadata']['total']
    # Vòng lặp để xử lý các trang dựa trên số lượng đơn hàng
    while total_orders > (page - 1) * 250:
        # Lấy dữ liệu cho các trang tiếp theo
        orders = js_get_url(f"{MAIN_URL}/orders.json?limit=250&status=finalized%2Cdraft&fulfillment_status=unshipped&page={page}")
        for order in orders['orders']:
            order['dvvc'] = ''
            list_order.append(order)
            list_id.append(order['id'])
        # Tăng số trang và cập nhật số lượng đơn hàng
        page += 1    

    channel = ''
    if "channel" in request.GET:
        channel = request.GET["channel"]
        obj['boloc']['channel'] = channel
    dvvc = ''
    if "dvvc" in request.GET:
        dvvc = request.GET["dvvc"]
        obj['boloc']['dvvc'] = dvvc
    kho = ''
    if "kho" in request.GET:
        kho = request.GET["kho"]
        obj['boloc']['kho'] = kho
        request.session['kho'] = kho

    if 'kho' in request.session:
        obj['boloc']['kho'] = request.session['kho']
        kho = request.session['kho']

    if kho == 'geleximco':
        location = '241737'
    elif kho == 'toky':
        location = '548744'
    else:
        location = '241737,548744'

    for order in list_order:

        flag = 0
        if len(order['fulfillments']) == 0:
            flag = 1
        elif len(order['fulfillments']) > 0:
            if order['fulfillments'][-1]['composite_fulfillment_status'] != 'fulfilled_cancelled':
                flag = 1
        if str(order['location_id']) in location:
            if order['source_id'] == 1880152 and flag==1:
                time_order = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
                order["order_date"] = time_order.strftime("%d-%m (%H:%M)")
                    
                if len(order['fulfillments']) > 0 and order['fulfillments'][-1]["shipment"] is not None:
                    order['chuanbihang'] = 1
                else:
                    order['chuanbihang'] = 0

                if len(order['tags']) >= 2:
                    if "_" in order['tags'][1]:
                        order['shop'] = order['tags'][1].split("_")[1]
                    else:
                        order['shop'] = "Gia Dụng Plus"
                        
                order = get_data_packing(order)

                if 'shipdate' in order:
                    if "/" in order['shipdate']:
                        ship_by_datetime = datetime.datetime.now()
                    else:
                        ship_by_datetime = datetime.datetime.fromtimestamp(order['shipdate'])

                    today = datetime.datetime.now().date()
            
                    if ship_by_datetime.date() <= today:
                        order['hangui'] = "Hôm nay"
                    else:
                        order['hangui'] = "Ngày mai"
                else:
                    now = datetime.datetime.now()
                    midnight_today = datetime.datetime(now.year, now.month, now.day, 14, 0, 0)
                    # Tính toán biến hangui
                    if time_order < midnight_today:
                        order['hangui'] = "Hôm nay"
                    else:
                        order['hangui'] = "Ngày mai"

                order['sos'] = 1
                if order['sos'] >= 0:
                    if order["source_id"] in [1880152] and order['channel'] != None:
                        obj['channel']['Shopee']['total'] += 1
                        if channel == 'shopee' or channel == '':
                                if order['dvvc'] not in obj['list_dvvc']:
                                    obj['list_dvvc'].append(order['dvvc'])
                                if order['dvvc'] == dvvc or dvvc == '':
                                    obj['list_orders'].append(order)

    obj['list_orders'].reverse()
    obj['list_orders'].sort(key=lambda x: (-x['sos'], x['id']))

    return render(request, 'kho_quanlysos.html',obj)

def kho_bangiao(request):
    obj = {

    }
    print("Start!")
    # Using readline()
    file1 = open('mvd.txt', 'r')
    count = 0
    
    while True:
        count += 1
    
        # Get next line from file
        line = file1.readline()
    
        # if line is empty
        # end of file is reached
        if not line:
            break
        order = js_get_url(f"{MAIN_URL}/orders.json?query={line.strip()}")["orders"][0]
        print(order["reference_number"])

    file1.close()

    return render(request, 'kho_bangiao.html',obj)

# PHẦN BÁN HÀNG
def kd_sanpham(request):
    obj = {
        "orders": [],
        "products": [],
        "qa": []
    }

    # ==== BƯỚC 1: XÁC THỰC GOOGLE SHEET ====
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
    creds = Credentials.from_service_account_file("logs/app-project-sapo-plus-eb0edec5c0dc.json", scopes=SCOPES)
    client = gspread.authorize(creds)

    # ==== BƯỚC 2: MỞ FILE GOOGLE SHEET ====
    sheet_url = "https://docs.google.com/spreadsheets/d/1RzS18QjF-sg-yGRz_SpVIhFOK9gz1x9mIkkAClbd1mk/edit?usp=sharing"
    worksheet = client.open_by_url(sheet_url).worksheet("ALL")  # Hoặc tên sheet cụ thể
    # ==== BƯỚC 4: CHUYỂN DỮ LIỆU VÀO obj["qa"] ====

    rows = worksheet.get_all_values()

    for i, row in enumerate(rows[2:], start=2):  # bỏ dòng tiêu đề
        try:
            # Bỏ qua dòng rỗng
            if not any(row):  
                continue
            question = str(row[0]).strip()
            answer1 = str(row[1]).strip()
            answer2 = str(row[2]).strip() if len(row) > 2 else ""
            tags_raw = str(row[3]).strip() if len(row) > 3 else ""
            related_products = str(row[4]).strip() if len(row) > 4 else ""

            # Bỏ dòng nếu thiếu câu hỏi hoặc câu trả lời chính
            if not question or not answer1:
                continue

            tags = [tag.strip() for tag in tags_raw.split(",") if tag.strip()]
            
            obj["qa"].append({
                "question": question,
                "answer": answer1,
                "answer_followup": answer2 if answer2 else None,
                "tags": tags,
                "related_products": related_products
            })

        except Exception as e:
            print(f"[!] Lỗi tại dòng {i}: {e}")
            continue

    all_value = []
    all_products = js_get_url(f"{MAIN_URL}/products.json?product_types=normal&page=1&limit=250")["products"]
    for pr in all_products:
        all_value.append(pr)
    all_products = js_get_url(f"{MAIN_URL}/products.json?product_types=normal&page=2&limit=250")["products"]
    for pr in all_products:
        all_value.append(pr)     

    for pr in all_value:
        pr['stock'] = 0
        pr['can_sale'] = 0
        pr['dang_ve'] = 0
        pr['list_qa'] = []
        for vari in pr["variants"]:
            if vari['packsize'] == False:
                try:
                    info_vari = Product.objects.get(id=vari['id'])
                except Product.DoesNotExist:
                    pass
                else:
                    vari['sp_height'] = info_vari.sp_height
                    vari['sp_width'] = info_vari.sp_width
                    vari['sp_length'] = info_vari.sp_length
                    vari['sp_weight'] = info_vari.sp_weight

                pr['stock'] += int(vari['inventories'][0]['on_hand'] + vari['inventories'][1]['on_hand'])
                pr['can_sale'] += int(vari['inventories'][0]['available'] + vari['inventories'][1]['available'])
                pr['dang_ve'] += int(vari['inventories'][0]['incoming'] + vari['inventories'][1]['incoming'])
            else:
                vari['error'] = 1       

    # Lấy ngày hôm nay (UTC, dạng ISO)
    today = datetime.datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    start_date = today - datetime.timedelta(days=30)
    end_date_str = today.isoformat() + "Z"
    start_date_str = start_date.isoformat() + "Z"
    all_sales = []
    base_url = "https://sisapsan.mysapogo.com/admin/reports/sales/by_variant.json"
    for i in range(4):
        params = {
            "location_ids": "",
            "start_date": start_date_str,
            "end_date": end_date_str,
            "limit": 250,
            "page": int(i+1),
            "criteria": "appeared_rate"
        }

        report_url = f"{base_url}?{urlencode(params)}"
        report_sales = js_get_url(report_url).get('items')

        if len(report_sales) > 0:
            all_sales.extend(report_sales)
        else:
            break
    for pr in all_value:
        for vari in pr["variants"]:
            vari['luotban30ngay'] = 0
            vari['ds30ngay'] = 0
            vari['don30ngay'] = 0
            for sales in all_sales:
                if str(sales['variant_id']) == str(vari['id']):
                    vari['luotban30ngay'] = int(sales['quantity'])
                    vari['ds30ngay'] = int(sales['amount_for_sorting'])
                    vari['don30ngay'] = int(sales['order_for_sorting'])
                    break
    for pr in all_value:
        pr['luotban30ngay'] = 0
        pr['ds30ngay'] = 0
        pr['don30ngay'] = 0
        for vari in pr["variants"]:
            pr['luotban30ngay'] += int(vari['luotban30ngay'])
            pr['ds30ngay'] += int(vari['ds30ngay'])
            pr['don30ngay'] += int(vari['don30ngay'])

    sorted_products = sorted(all_value,
        key=lambda x: (x["status"] != "active", -x["luotban30ngay"] if x["status"] == "active" else 0)
    )

    obj["products"] = sorted_products
    return render(request, 'kd_sanpham.html', obj)

@csrf_exempt
def kd_api_getproductinfo(request): 
    product_id = str(request.GET['id'])
    base_url = f"https://sisapsan.mysapogo.com/admin/products/{product_id}.json"
    product = js_get_url(base_url).get('product')

    product['data'] = parse_product_description(product['description'])

    # ==== LẤY Q&A
    product["qa"] = []
    list_qa = []
    # ==== BƯỚC 1: XÁC THỰC GOOGLE SHEET ====
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
    creds = Credentials.from_service_account_file("logs/app-project-sapo-plus-eb0edec5c0dc.json", scopes=SCOPES)
    client = gspread.authorize(creds)

    # ==== BƯỚC 2: MỞ FILE GOOGLE SHEET ====
    sheet_url = "https://docs.google.com/spreadsheets/d/1RzS18QjF-sg-yGRz_SpVIhFOK9gz1x9mIkkAClbd1mk/edit?usp=sharing"
    worksheet = client.open_by_url(sheet_url).worksheet("ALL")  # Hoặc tên sheet cụ thể
    # ==== BƯỚC 4: CHUYỂN DỮ LIỆU VÀO obj["qa"] ====

    rows = worksheet.get_all_values()

    for i, row in enumerate(rows[2:], start=2):  # bỏ dòng tiêu đề
        try:
            # Bỏ qua dòng rỗng
            if not any(row):  
                continue
            question = str(row[0]).strip()
            status = str(row[1]).strip()
            insight = str(row[2]).strip()
            tuduy = str(row[3]).strip()
            answer1 = str(row[4]).strip()
            answer2 = str(row[5]).strip() if len(row) > 2 else ""
            answer3 = str(row[6]).strip() if len(row) > 2 else ""
            tags_raw = str(row[7]).strip() if len(row) > 3 else ""
            products = str(row[8]).strip() if len(row) > 4 else ""

            # Bỏ dòng nếu thiếu câu hỏi hoặc câu trả lời chính
            if not question or not answer1:
                continue

            if not status or status != "Đang áp dụng":
                continue

            tags = [tag.strip() for tag in tags_raw.split(",") if tag.strip()]
            list_qa.append({
                "question": question,
                "status": status,
                "insight": insight if insight else None,
                "tuduy": tuduy if tuduy else None,
                "answer": answer1 if answer1 else None,
                "answer2": answer2 if answer2 else None,
                "answer3": answer3 if answer3 else None,
                "tags": tags,
                "products": products
            })

        except Exception as e:
            print(f"[!] Lỗi tại dòng {i}: {e}")
            continue

    for qa in list_qa:
        for v in product.get("variants", []):
            sku = v.get("sku") or ""
            parts = sku.split("-")
            if len(parts) >= 2:
                short_code = parts[1]   # VD: TG-0201-DEN -> 0201
                # Nếu short_code có trong list products của QA
                if short_code in qa.get("products", []):
                    product["qa"].append(qa)
                    break   # đã match 1 variant thì add xong, thoát vòng for variant

    return JsonResponse(product)

def kd_qa(request):
    obj = {
        "qa": []
    }
    # ==== BƯỚC 1: XÁC THỰC GOOGLE SHEET ====
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
    creds = Credentials.from_service_account_file("logs/app-project-sapo-plus-eb0edec5c0dc.json", scopes=SCOPES)
    client = gspread.authorize(creds)

    # ==== BƯỚC 2: MỞ FILE GOOGLE SHEET ====
    sheet_url = "https://docs.google.com/spreadsheets/d/1RzS18QjF-sg-yGRz_SpVIhFOK9gz1x9mIkkAClbd1mk/edit?usp=sharing"
    worksheet = client.open_by_url(sheet_url).worksheet("ALL")  # Hoặc tên sheet cụ thể
    # ==== BƯỚC 4: CHUYỂN DỮ LIỆU VÀO obj["qa"] ====

    rows = worksheet.get_all_values()

    for i, row in enumerate(rows[2:], start=2):  # bỏ dòng tiêu đề
        try:
            # Bỏ qua dòng rỗng
            if not any(row):  
                continue
            question = str(row[0]).strip()
            status = str(row[1]).strip()
            insight = str(row[2]).strip()
            tuduy = str(row[3]).strip()
            answer1 = str(row[4]).strip()
            answer2 = str(row[5]).strip() if len(row) > 2 else ""
            answer3 = str(row[6]).strip() if len(row) > 2 else ""
            tags_raw = str(row[7]).strip() if len(row) > 3 else ""
            products = str(row[8]).strip() if len(row) > 4 else ""

            # Bỏ dòng nếu thiếu câu hỏi hoặc câu trả lời chính
            if not question or not answer1:
                continue

            if not status or status != "Đang áp dụng":
                continue

            tags = [tag.strip() for tag in tags_raw.split(",") if tag.strip()]
            obj["qa"].append({
                "question": question,
                "status": status,
                "insight": insight if insight else None,
                "tuduy": tuduy if tuduy else None,
                "answer": answer1 if answer1 else None,
                "answer2": answer2 if answer2 else None,
                "answer3": answer3 if answer3 else None,
                "tags": tags,
                "products": products
            })

        except Exception as e:
            print(f"[!] Lỗi tại dòng {i}: {e}")
            continue

    return render(request, 'kd_qa.html', obj)

def kd_addqa(request):
    if request.method == "GET":
        if 'xoa' in request.GET:
            product_id = request.GET['product_id']
            stt = request.GET['stt']

            product = js_get_url(f"{MAIN_URL}/products/{product_id}.json")
            new_descreption = ""
            
            new_descreption = product['product']['description'].split('*#SPLITCAUHOI*#')[0]
            cauhoia = product['product']['description'].split('*#SPLITCAUHOI*#')[1:]

            count = 1
            for cauhoi in cauhoia:
                if int(stt) != int(count):
                    new_descreption += "<p>*#SPLITCAUHOI*#</p>"+cauhoi
                count += 1  

            product['product']['description'] = new_descreption
            #print(pr['product']['description'])
            rp = loginss.put(f"{MAIN_URL}/products/{product_id}.json",json=product)
            print(rp.text)

    if request.method == "POST":
        product_id = request.POST['product_id']
        cauhoi = request.POST['cauhoi']
        ycuakhach = request.POST['ycuakhach']
        cautraloi = request.POST['cautraloi']
        
        product = js_get_url(f"{MAIN_URL}/products/{product_id}.json")
        product['product']['description'] += f"<p>*#SPLITCAUHOI*#</p><p>{cauhoi}***{ycuakhach}***{cautraloi}</p>"
        loginss.put(f"{MAIN_URL}/products/{product_id}.json",json=product)

    obj = {
        "orders": [],
        "products": []
    }
    all = []
    all_products = js_get_url(f"{MAIN_URL}/products.json?product_types=normal&page=1&limit=250")["products"]
    for pr in all_products:
        all.append(pr)
    all_products = js_get_url(f"{MAIN_URL}/products.json?product_types=normal&page=2&limit=250")["products"]
    for pr in all_products:
        all.append(pr)     

    for pr in all:
        pr['stock'] = 0
        pr['can_sale'] = 0
        pr['dang_ve'] = 0
        pr['list_qa'] = []
        for vari in pr["variants"]:
            if vari['packsize'] == False:
                try:
                    info_vari = Product.objects.get(id=vari['id'])
                except Product.DoesNotExist:
                    pass
                else:
                    vari['sp_height'] = info_vari.sp_height
                    vari['sp_width'] = info_vari.sp_width
                    vari['sp_length'] = info_vari.sp_length
                    vari['sp_weight'] = info_vari.sp_weight

                pr['stock'] += int(vari['inventories'][0]['on_hand'] + vari['inventories'][1]['on_hand'])
                pr['can_sale'] += int(vari['inventories'][0]['available'] + vari['inventories'][1]['available'])
                pr['dang_ve'] += int(vari['inventories'][0]['incoming'] + vari['inventories'][1]['incoming'])
            else:
                vari['error'] = 1       

        if pr['description'] != None and 'SPLITCAUHOI' in pr['description']:
            pr['thongtin'] = pr['description'].split('<p>*#SPLITCAUHOI*#</p>')[0]
            pr['cauhoi'] = pr['description'].split('<p>*#SPLITCAUHOI*#</p>')[1:]
        
            for cauhoi in pr['cauhoi']:
                cauhoix = {'hoi':'','y':'','rep':''}
                if len(cauhoi.split('***')) >= 2:
                    cauhoix['hoi'] = cauhoi.split('***')[0].replace("<p>","")
                    cauhoix['y'] = cauhoi.split('***')[1]
                    cauhoix['rep'] = cauhoi.split('***')[2].replace("</p>","")
                    pr['list_qa'].append(cauhoix)

    obj["products"] = all
    return render(request, 'kd_sanpham.html', obj)

def kd_donhang(request):
    obj = {
        "orders": [],
        "products": []
    }
    all_brand = json.loads(loginss.get(f"{MAIN_URL}/brands.json?page=1&limit=250").text)["brands"]
    
    for brand in all_brand:
        all_vari = js_get_url(f"{MAIN_URL}/variants.json?limit=250&packsize=false&product_type=normal&page=1&status=active&brand_ids={brand['id']}")
        for vari in all_vari["variants"]:
            print(vari["sku"])
            vari['image'] = vari["images"][0]["full_path"]
            for price in vari["variant_prices"]:
                if price["name"] == "Giá sỉ":
                    vari["giasi"] = int(price["value"])
                elif price["name"] == "Giá bán lẻ":
                    vari["giale"] = int(price["value"])

            vari["stock"] = vari["inventories"][0]["on_hand"]            
            obj["products"].append(vari)

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(f"assets/excel_po/ALL-PRODUCT-X.xlsx")
    worksheet = workbook.add_worksheet()
    worksheet.set_default_row(60)
    worksheet.set_row(0,20)
    worksheet.set_column('A:A', 40)
    worksheet.set_column('B:B', 40)
    worksheet.set_column('C:C', 15)
    worksheet.set_column('D:D', 20)
    worksheet.set_column('E:E', 20)
    worksheet.set_column('I:I', 15)

    cell_text_wrap = workbook.add_format({'text_wrap': True,'align': 'center','valign': 'vcenter'})
    cell_align = workbook.add_format({'align': 'center','valign': 'vcenter'})
    cell_first = workbook.add_format({'bold': True,'bg_color': "#D8E4BC"})

    row = 0
    col = 0      
    worksheet.write(row, col, "SKU",cell_first)
    worksheet.write(row, col + 1, "ẢNH",cell_first)
    worksheet.write(row, col + 2, "TÊN SẢN PHẨM",cell_first)
    worksheet.write(row, col + 3, "PHÂN LOẠI",cell_first)
    worksheet.write(row, col + 4, "GIÁ SỈ",cell_first)
    worksheet.write(row, col + 5, "GIÁ LẺ THAM KHẢO",cell_first)
    worksheet.write(row, col + 6, "TỒN KHO",cell_first)
    row = 1
    # Iterate over the data and write it out row by row.
    for variant in obj["products"]:
        
        file_path = f"assets/saveimage/{variant['id']}.jpg"
        if os.path.isfile(file_path):
            pass
        else:
            variant["image"] = variant["images"][0]["full_path"]
            r = requests.get(variant["image"], allow_redirects=True)
            open(file_path, 'wb').write(r.content)
            
        worksheet.write(row, col, variant["sku"],cell_align)
        # Lấy kích thước của ô
        cell_width = 75  # Giả sử chiều rộng của ô là 80
        cell_height = 75  # Giả sử chiều cao của ô là 60

        # Lấy kích thước của hình ảnh
        with Image.open(file_path) as img:
            image_width, image_height = img.size

        # Tính tỉ lệ để hình ảnh vừa với ô
        x_scale = float(cell_width) / float(image_width)
        y_scale = float(cell_height) / float(image_height)

        # Định dạng hình ảnh để tự động điều chỉnh kích thước
        image_options = {
            'x_offset': 5,
            'y_offset': 5,
            'x_scale': x_scale,
            'y_scale': y_scale,
        }
        worksheet.insert_image(row, col + 1, "assets/saveimage/"+str(variant["id"])+".jpg",image_options)
        worksheet.set_column(col + 1, col + 2, 12)
        worksheet.write(row, col + 2, variant["name"],cell_align)
        worksheet.write(row, col + 3, variant["opt1"],cell_align)
        worksheet.write(row, col + 4, variant["giasi"],cell_align)
        worksheet.write(row, col + 5, variant["giale"],cell_align)
        worksheet.write(row, col + 6, variant["stock"],cell_align)

        row += 1

    workbook.close()
    obj['url_download'] = f"/static/excel_po/ALL-PRODUCT-X.xlsx"
    return render(request, 're_download.html',obj)

def kd_vandon(request):
    obj = {
        "orders": [],
    }    
    process = get_list_process()
    
    if 'update_data' in request.GET:
        pass

    if 'update_status' in request.GET:
        order_id = request.GET['order_id']
        status = request.GET['update_status']
        order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")['order']

        status_data = {}
        if order['fulfillments'][0]['shipment']['note'] != None:
            if len(order['fulfillments'][0]['shipment']['note']) > 100:
                status_data = json.loads(order['fulfillments'][0]['shipment']['note'])

        status_data['status_vandon'] = status
        update_data(order["fulfillments"][0]['id'],status_data)

    if 'action' in request.GET:
        action = request.GET['xem']
    else:
        action = 'hoan'

    if action == 'hoan':
        orders = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=250&composite_fulfillment_status=fulfilled_cancelling%2Cfulfilled_cancelled")["orders"]
        for order in orders:
            if order['process_status_id'] != None:
                if str(order['process_status_id']) in process:
                    order['process'] = process[str(order['process_status_id'])]
                else:
                    order['process'] = 'Khác'
            else:
                order['process'] = 'Chưa xử lý'
            if order['fulfillments'][0]['shipment']['note'] != None:
                if len(order['fulfillments'][0]['shipment']['note']) > 100:
                    status_data = json.loads(order['fulfillments'][0]['shipment']['note'])
                    if 'status_vandon' in status_data:
                        order['process'] = status_data['status_vandon']

            order['status_vc'] = order['fulfillments'][-1]['composite_fulfillment_status']
            time_order = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
            
            order["order_date"] = time_order.strftime("%d-%m")

            obj['orders'].append(order)
    


    return render(request, 'kd_giaovan.html', obj)

# ---------- Main view ----------

def kd_repdanhgia(request):
    obj = {"save": {}, "comment": {}}

    # --- 1. Lấy thông tin shop ---
    shop_name = get_shop_name(request)
    obj["save"]["shop_name"] = shop_name
    connection_id = get_connection_id(shop_name, existing_shop_map)
    chat_id_connect = 0

    # --- 2. Lấy thông tin chat ---
    ALL_CONNECT = js_get_url("https://market-place.sapoapps.vn/go/messenger/connection-chats?key=MzE5OTEx")
    chat_id_connect = find_chat_id_connect(connection_id, ALL_CONNECT)

    # --- 3. Lấy thông tin khách hàng ---
    username = request.GET.get('username', None)
    if username:
        obj["save"]["username"] = username
    COVER_ID = find_cover_id(username, chat_id_connect)
    obj["save"]["COVER_ID"] = COVER_ID

    # --- 4. Handler (người xử lý đánh giá) ---
    handler = get_handler(request)
    if handler:
        obj["save"]["handler"] = handler

    # --- 5. Thông tin đơn hàng và đánh giá ---
    order_sn = request.GET.get('order_sn', None)
    comment_id = request.GET.get('comment_id', None)
    rate_star = int(request.GET.get('rate_star', 0))
    obj["save"].update({"order_sn": order_sn, "comment_id": comment_id, "rate_star": rate_star})

    # --- 6. Lấy dữ liệu đánh giá của đơn hàng từ Shopee ---
    if username and doi_shop(connection_id, loginsp) == 1:
        URL = f"https://banhang.shopee.vn/api/v3/settings/search_shop_rating_comments_new/?SPC_CDS=4352421f-1de6-47ae-b399-d0b69ec45e70&SPC_CDS_VER=2&search_request={order_sn}&rating_star=5,4,3,2,1&page_number=1&page_size=20&cursor=0&from_page_number=1&language=vi"
        all_data = json.loads(loginsp.get(URL).text)["data"]["list"]
        obj['orders'] = all_data
        order_comment_rep, comment = get_order_comment_rep(all_data, comment_id)
        obj['comment'] = comment
        order = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=250&query={order_sn}")["orders"][0]
        obj["comment"]["order_info"] = order

    # --- 7. Sinh gợi ý đánh giá ---
    if 'male' in request.GET:
        male = request.GET['male']
        name = request.GET['name']
        obj["save"].update({"male": male, "name": name.title()})
        excel_file_path = 'logs/log excel/GOIYDANHGIA.xlsx'
        listgoiy, all_tags = load_goi_y_danh_gia(excel_file_path, shop_name, rate_star, male, name)
        obj['goiy'] = listgoiy
        obj['tags'] = all_tags

    # --- 8. Xử lý phản hồi đánh giá ---
    if 'repcomment' in request.GET and request.GET['repcomment'] != '':
        repcomment = request.GET['repcomment']
        now = datetime.datetime.now()
        log_dir = "logs/comment"
        log_filename = f"{now.month}_{now.year}-{SERVER_ID}.jsonl"
        log_path = os.path.join(log_dir, log_filename)

        log_entry = {
            "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
            "handler": handler,
            "comment_id": comment_id,
            "shop_name": shop_name,
            "rate_star": rate_star,
            "order_sn": order_sn,
            "username": username,
            "repcomment": repcomment
        }

        log_rep_comment(log_path, log_entry)
        obj['repcomment_done'] = 1

        if rate_star >= 4:
            # Gửi phản hồi lên Shopee
            URL = "https://banhang.shopee.vn/api/v3/settings/reply_shop_rating/?SPC_CDS=128340d3-0357-49e4-8ba5-5c8a09dd052b&SPC_CDS_VER=2"
            POST_JSON = {"order_id": int(obj['comment']['order_id']), "comment_id": int(comment_id), "comment": repcomment}
            rs = loginsp.post(URL, json=POST_JSON)
            # Gửi tin nhắn nếu chưa từng rep
            MESS_REP = [
                f"Hơi muộn nhưng shop vừa đọc được đánh giá của {male} {name} rồi ạ.\nCảm ơn {male} rất nhiều vì đã giành thời gian để đánh giá & ủng hộ shop!",
                repcomment.replace("\n", "\\n").replace('"', '*'),
                "Nếu mình có đặt đơn hàng tiếp theo, hãy nhắc shop, để shop gửi 1 phần QUÀ TẶNG ngẫu nhiên cho mình thay lời cảm ơn ạ!"
            ]
            if order_comment_rep == 0 and COVER_ID != 0:
                send_messages_to_customer(COVER_ID, MESS_REP, logintmdt)
            elif order_comment_rep == 0 and COVER_ID == 0:
                log_entry['send_flag'] = 0
                with open('logs/comment/message/log.jsonl', "a", encoding="utf-8") as f:
                    f.write(json.dumps(log_entry, ensure_ascii=False) + "\n")
        else:
            # Ghi vào Google Sheet cho đánh giá xấu
            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            creds = ServiceAccountCredentials.from_json_keyfile_name("logs/app-project-sapo-plus-eb0edec5c0dc.json", scope)
            client = gspread.authorize(creds)
            sheet = client.open("DANHGIAXAU-DUYET").sheet1
            sheet.append_row([
                now.strftime("%Y-%m-%d %H:%M:%S"),  # timestamp
                handler,
                comment_id,
                order_sn,
                str(rate_star) + "★",
                comment.get('comment', ''),
                "Giải thích vấn đề của khách gặp phải",
                repcomment,
                0  # duyet
            ])

        obj['repcomment_done'] = 1

    return render(request, 'kd_repdanhgia.html', obj)

def kd_repall(request):
    # 1. Nếu có auto_ai trong GET → render ngay
    if request.GET.get("auto_ai"):
        return render(request, "kd_repauto.html")
 
    return render(request, 'kd_repall.html')

def kd_tenkhach(request):
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import json, time, random, threading
    import requests
    from openai import OpenAI

    API_KEY = "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"
    with open('logs/openai/prompt-customer-name.txt', 'r', encoding='utf-8') as f:
        prompt_content = f.read()

    # --- Tham số tinh chỉnh nhanh ---
    MAX_CHAT_WORKERS = 8          # số luồng gọi ChatGPT song song
    MAX_PUT_WORKERS  = 8          # số luồng PUT song song
    BATCH_SIZE       = 50         # cỡ batch gửi lên ChatGPT
    RETRY_MAX        = 4          # số lần retry tối đa cho mỗi call
    BACKOFF_BASE     = 0.8        # backoff cơ sở (giây)

    print_lock = threading.Lock()  # để print gọn gàng khi đa luồng

    def log_safe(msg):
        with print_lock:
            print(msg, flush=True)

    def retry_backoff_sleep(attempt):
        # ví dụ: 0.8s, 1.6s, 3.2s, 6.4s + jitter
        time.sleep((BACKOFF_BASE * (2 ** attempt)) + random.uniform(0, 0.3))

    def call_chatgpt_batch(batch, prompt_content, api_key):
        """
        Gọi ChatGPT cho 1 batch, có retry + backoff.
        Trả về list[dict]: [{id, short_name, sex}, ...]
        """
        batch_str = json.dumps(batch, ensure_ascii=False, indent=2)
        full_prompt = f"Dữ liệu đầu vào:\n{batch_str}\n"

        for attempt in range(RETRY_MAX):
            try:
                client = OpenAI(api_key=api_key)  # tạo client trong luồng để thread-safe
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": prompt_content},
                        {"role": "user",   "content": full_prompt}
                    ],
                    temperature=0.7,
                    timeout=120  # đề phòng treo
                )
                result = response.choices[0].message.content

                # Thử parse JSON cứng
                try:
                    parsed = json.loads(result)
                    if isinstance(parsed, list):
                        return parsed
                    # nếu không phải list, cố tìm mảng JSON trong chuỗi
                except Exception:
                    pass

                # Fallback: tìm khối JSON đầu tiên trong text
                start = result.find('[')
                end   = result.rfind(']')
                if start != -1 and end != -1 and end > start:
                    parsed = json.loads(result[start:end+1])
                    return parsed

                # Nếu đến đây vẫn không parse được:
                log_safe(f"[ChatBatch] Parse lỗi, trả về rỗng. Batch size={len(batch)}")
                return []

            except Exception as e:
                log_safe(f"[ChatBatch] Lỗi attempt {attempt+1}/{RETRY_MAX}: {e}")
                if attempt < RETRY_MAX - 1:
                    retry_backoff_sleep(attempt)
                else:
                    return []

    def split_batches(lst, batch_size=BATCH_SIZE):
        for i in range(0, len(lst), batch_size):
            yield lst[i:i + batch_size]

    # =======================
    # Trong view: thay vào chỗ "=== Thêm code ở đây ==="
    # =======================
    if request.GET.get("tool_ai"):
        save_all_custo = {}
        all_custo = []
        print("[+] Bắt đầu dùng AI để xác định tên và giới tính.")
        for i in range(50):
            URL_SEARCH = f"{MAIN_URL}/customers.json?page={int(i+1)}&limit=100"
            ALL_CUSTOMERS = js_get_url(URL_SEARCH)['customers']
            for custo in ALL_CUSTOMERS:
                if custo['tax_number'] is None and "****" not in custo['name']:
                    save_all_custo[str(custo['id'])] = custo
                    all_custo.append({ 'id': custo['id'], 'name': custo['name'] })
    
        # === Thêm code ở đây === (đa luồng) ===
        client = OpenAI(api_key=API_KEY)  # KHÔNG dùng client này cho đa luồng, chỉ để check nhanh nếu muốn
        final_results = []

        # Chia batch
        batches = list(split_batches(all_custo, batch_size=BATCH_SIZE))

        # 1) GỌI CHATGPT SONG SONG
        log_safe(f"==> Bắt đầu gọi ChatGPT: {len(batches)} batch, {len(all_custo)} khách chưa có tax_number")
        with ThreadPoolExecutor(max_workers=MAX_CHAT_WORKERS) as executor:
            future_to_idx = {
                executor.submit(call_chatgpt_batch, batch, prompt_content, API_KEY): idx
                for idx, batch in enumerate(batches, start=1)
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    batch_result = future.result() or []
                    final_results.extend(batch_result)
                    log_safe(f"[ChatBatch DONE] {idx}/{len(batches)} (+{len(batch_result)} khách)")
                except Exception as e:
                    log_safe(f"[ChatBatch FAIL] {idx}/{len(batches)}: {e}")

        log_safe(f"==> Tổng hợp được {len(final_results)} khách từ ChatGPT")

        # 2) CHUẨN BỊ PAYLOADS
        put_tasks = []
        for custo in final_results:
            try:
                info_custo = save_all_custo[str(custo['id'])]
            except KeyError:
                # bảo vệ khi ChatGPT trả id không có trong trang hiện tại
                continue

            payload = {
                "customer": {
                    "id": int(custo['id']),
                    "name": info_custo["name"],
                    "code": info_custo["code"],
                    "customer_group_id": info_custo["customer_group_id"],
                    "sex": custo.get('sex'),
                    "tax_number": "1",
                    "website": custo.get('short_name'),
                    "status": info_custo.get("status"),
                    "assignee_id": info_custo.get("assignee_id"),
                    "tags": info_custo.get("tags", []),
                    "apply_incentives": info_custo.get("apply_incentives", "group"),
                }
            }
            url = f"{MAIN_URL}/customers/{custo['id']}.json"
            put_tasks.append((url, payload))

        # 3) GỬI PUT SONG SONG (dùng loginss đã đăng nhập sẵn)
        success_cnt = 0
        log_safe(f"==> Bắt đầu PUT: {len(put_tasks)} bản ghi")
        with ThreadPoolExecutor(max_workers=MAX_PUT_WORKERS) as executor:
            future_to_url = {
                executor.submit(put_customer, loginss, url, payload): url
                for (url, payload) in put_tasks
            }
            for future in as_completed(future_to_url):
                url = future_to_url[future]
                ok = False
                try:
                    ok = future.result()
                except Exception as e:
                    log_safe(f"[PUT EXC] {url}: {e}")
                if ok:
                    success_cnt += 1

        log_safe(f"==> PUT thành công {success_cnt}/{len(put_tasks)}")
        # === Hết block đa luồng ===

    return render(request, 'kd_tenkhach.html')


def kd_updatereview(request):
    SISAPO_BASE = "https://sisapsan.mysapogo.com/admin"  # domain Sapo của anh

    def iso_utc_from_epoch(sec: int) -> str:
        try:
            return datetime.datetime.utcfromtimestamp(int(sec)).replace(
                tzinfo=datetime.timezone.utc
            ).isoformat()
        except Exception:
            return ""

    def get_order_by_sn(order_sn: str) -> dict | None:
        """Tìm đơn theo order_sn qua API Sapo của anh."""
        if not order_sn:
            return None
        url = f"{MAIN_URL}/orders.json?query={order_sn}"
        data = js_get_url(url) or {}
        orders = data.get("orders") or []
        return orders[0] if orders else None

    def note_already_exists(customer_id: int, comment_id: int) -> bool:
        try:
            url = f"{SISAPO_BASE}/customers/{customer_id}/notes.json?limit=50"
            res = loginss.get(url, timeout=20)
            res.raise_for_status()
            j = res.json() or {}
            notes = j.get("notes") or []

            # Tìm substring bền vững cho JSON: "comment_id":<id>
            mark = f"{int(comment_id)}"
            for n in notes:
                content = (n.get("content") or "").strip()
                if mark in content:
                    return True
        
        except Exception as e:
            print(f"Không kiểm tra duplicate note được {e}")
        return False

    def to_timestamp_vn(date_str: str, end_of_day: bool = False) -> int:
        TZ_VN = datetime.timezone(datetime.timedelta(hours=7))
        """
        Chuyển 'DD-MM-YYYY' → timestamp theo giờ Việt Nam (UTC+7).
        - end_of_day=False: 00:00:00 (đầu ngày)
        - end_of_day=True : 23:59:59 (cuối ngày)  -> dùng cho khoảng thời gian inclusive
        """
        dt = datetime.datetime.strptime(date_str, '%d-%m-%Y').replace(tzinfo=TZ_VN)
        if end_of_day:
            dt = dt.replace(hour=23, minute=59, second=59, microsecond=0)
        else:
            dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
        return int(dt.timestamp())


    # ---- Tham số ----
    shop_name   = request.GET.get("shop_name", "giadungplus_official").strip()
    rating_star = request.GET.get("rating_star", "1,2,3,4,5").strip()
    rate_type   = request.GET.get("rate_type", "new").strip().lower()

    # Ví dụ cố định theo yêu cầu
    time_start = to_timestamp_vn('15-08-2025', end_of_day=False)      # 1755968400
    time_end_old   = to_timestamp_vn('27-08-2025', end_of_day=True)       # 1756227599


    # ---- Đổi shop theo shop_name ----
    connection_id = get_connection_id(shop_name, existing_shop_map)
    doi_shop(connection_id, loginsp)

    # ---- Base URL chung (không có page_number/page_size) ----
    base_url = (
        "https://banhang.shopee.vn/api/v3/settings/search_shop_rating_comments_new/"
        "?SPC_CDS=15c18032-c3ae-45ea-9393-85c234ac4a32"
        "&SPC_CDS_VER=2"
        f"&rating_star={rating_star}"
        "&language=vi"
        f"&time_start={time_start}&time_end={time_end_old}"
    )

    # ---- Tự động xác định total & total_pages ----
    page_size = int(request.GET.get("page_size", 50))   # cho phép override; mặc định 50
    probe_url = f"{base_url}&page_number=1&page_size={page_size}&cursor=0&from_page_number=1"

    total = 0
    try:
        # loginsp thường là requests.Session đã đăng nhập/đổi shop
        resp = loginsp.get(probe_url, timeout=30)
        resp.raise_for_status()
        j = resp.json()
        total = int(((j.get("data") or {}).get("page_info") or {}).get("total") or 0)
    except Exception as e:
        # fallback an toàn: nếu không lấy được total, coi như 1 trang để crawler tự xoay sở
        total = 0
    print(f"[+] TỔNG TRONG THỜI GIAN NÀY CÓ {total} ĐÁNH GIÁ")    
   
    # Nếu API trả null page_size, ta vẫn dùng page_size mình gửi (50)
    total_pages = max(1, math.ceil(total / page_size)) if total else 1
    # Ví dụ total = 795, page_size = 50 -> total_pages = 16

    # ---- Crawl như cũ, nhưng dùng total_pages đã tính ----
    all_comment = crawl_shopee_ratings(
        base_url,
        max_pages=total_pages,
        page_size=page_size,
        delay=0.1
        # Không cần rate_type ở crawler nữa vì đã khóa time_start/time_end ở URL
    )

    from concurrent.futures import ThreadPoolExecutor, as_completed

    def process_one_comment(comment):
        osn = (comment.get('order_sn') or '').strip()
        if not osn:
            return ('skip:no_osn', None)

        order = get_order_by_sn(osn)
        if not order:
            return (f'skip:no_order:{osn}', None)

        customer_id = (order.get('customer_id')
                       or (order.get('customer') or {}).get('id')
                       or (order.get('customer_data') or {}).get('id'))
        if not customer_id:
            return (f'skip:no_customer:{osn}', None)

        cmid = int(comment.get('comment_id', 0))
        if cmid and note_already_exists(int(customer_id), cmid):
            # Dù trùng note, vẫn thêm tag COMMENT_SHOPEE nếu thiếu
            put_tuple = build_customer_update_from_order(order)
            return (f'skip:dup:{cmid}', put_tuple)

        payload = build_note_payload(comment, order)

        oli = order.get('order_line_items') or []
        if len(oli) == 1:
            payload['variant_id'] = oli[0].get('variant_id', 0)
        else:
            url = ("https://market-place.sapoapps.vn/v2/orders?page=1&limit=1"
                   f"&connectionIds={connectID}"
                   f"&query={payload['order_sn']}&sortBy=ISSUED_AT&orderBy=desc")
            o = js_get_url(url)
            if isinstance(o, dict) and (o.get('orders') or []):
                for v in (o['orders'][0].get('products') or []):
                    if v.get('item_id') == payload.get('product_id'):
                        payload['variant_id'] = v.get('sapo_variant_id', 0); break
            payload.setdefault('variant_id', 0)

        content = pack_note(payload, shorten_values=True)
        spapi_new_note(int(customer_id), content)   # POST note

        put_tuple = build_customer_update_from_order(order)  # (cid, url, payload) or None
        return (f'ok:{cmid}', put_tuple)



    COMMENT_TAG = "COMMENT_SHOPEE"

    def _dedup_keep_order(seq):
        seen = set(); out = []
        for x in (seq or []):
            if x is None: 
                continue
            if x not in seen:
                seen.add(x); out.append(x)
        return out

    def _strip_none(d: dict) -> dict:
        """Loại bỏ các key có value None để không vô tình overwrite thành null trên Sapo."""
        return {k: v for k, v in d.items() if v is not None}

    def build_customer_update_from_order(order: dict):
        """
        Lấy trực tiếp từ order['customer_data'] để dựng payload đầy đủ.
        Không gọi thêm API. Chỉ thêm COMMENT_SHOPEE vào tags.
        """
        cd = (order.get("customer_data") or {})
        cid = cd.get("id") or order.get("customer_id")
        if not cid:
            return None
        cid = int(cid)

        # Tags từ order => thêm COMMENT_SHOPEE
        tags = list(cd.get("tags") or [])
        if COMMENT_TAG not in tags:
            tags.append(COMMENT_TAG)
        tags = _dedup_keep_order(tags)

        # Dựng payload theo schema anh yêu cầu, ưu tiên GIỮ NGUYÊN giá trị hiện có trong order
        cust_payload = {
            "id":                cid,
            "name":              cd.get("name"),
            "code":              cd.get("code"),
            "customer_group_id": cd.get("customer_group_id"),
            "sex":               cd.get("sex"),
            "tax_number":        cd.get("tax_number") if cd.get("tax_number") not in (None, "") else "1",
            "website":           cd.get("website"),
            "status":            cd.get("status") or "active",
            "assignee_id":       cd.get("assignee_id"),
            "tags":              tags,
            "apply_incentives":  cd.get("apply_incentives") or "group",
        }

        # Tránh gửi None → không overwrite rỗng
        cust_payload = _strip_none(cust_payload)

        url = f"{MAIN_URL}/customers/{cid}.json"
        payload = {"customer": cust_payload}
        return cid, url, payload

    results = []
    customers_for_put = {}  # {customer_id: (url, payload)}

    print(f"[+] THU THẬP DỮ LIỆU VỀ ĐƯỢC: {len(all_comment)}")

    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = [ex.submit(process_one_comment, cmt) for cmt in all_comment]
        for fu in as_completed(futs):
            try:
                r, put_tuple = fu.result()
            except Exception as e:
                r, put_tuple = (f'err:{e.__class__.__name__}', None)
            results.append(r)
            if put_tuple:
                cid, url, payload = put_tuple
                customers_for_put[cid] = (url, payload)  # de-dup theo customer_id

    def do_put(task):
        url, payload = task
        try:
            res = loginss.put(url, json=payload, timeout=25)
            code = res.status_code
            return {"url": url, "status_code": code, "ok": (200 <= code < 300), "error": None}
        except Exception as e:
            return {"url": url, "status_code": None, "ok": False, "error": str(e)}

    put_results = []
    if customers_for_put:
        with ThreadPoolExecutor(max_workers=8) as ex:
            futs = [ex.submit(do_put, task) for task in customers_for_put.values()]
            for fu in as_completed(futs):
                put_results.append(fu.result())

    ok_count = sum(1 for r in put_results if r.get("ok"))
    fail_count = len(put_results) - ok_count
    print(f"[PUT CUSTOMER TAGS] OK={ok_count}  FAIL={fail_count}")

    return render(request, 'kd_tenkhach.html')

@csrf_exempt
def kd_repauto(request):
    API_KEY = "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"
    shop_name = get_shop_name(request)
    connection_id = get_connection_id(shop_name, existing_shop_map)

    log_path = "assets/openai/new-comment.json"
    log_comment = "logs/openai/log_comment.json"
    log_name = "logs/openai/log-name.json"

    if request.GET.get("update") == "1":
        cmt_id = request.GET.get("cmt_id")
        try:
            cmt_id = int(cmt_id)
        except ValueError:
            return JsonResponse({"error": "Invalid comment_id"}, status=400)

        if not os.path.exists(log_path):
            return JsonResponse({"error": "Log file not found"}, status=404)

        try:
            body = json.loads(request.body)
            new_reply = body.get("new_reply", "").strip()

            with open(log_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            comment_found = False
            for item in data:
                if item.get("comment_id") == cmt_id:
                    item["repcomment_suggest"] = new_reply
                    item["status"] = "updated_by_user"
                    item["suggested_at"] = datetime.datetime.utcnow().isoformat()
                    comment_found = True
                    break

            if not comment_found:
                return JsonResponse({"error": "Comment ID not found"}, status=404)

            with open(log_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return JsonResponse({"success": True, "comment_id": cmt_id})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


    # ----------------------------
    if request.GET.get("make_name") == "ok":
        if request.method == "GET":
            SOLUONG   = int(request.GET.get("soluong", 100))
            MAXPAGES  = math.ceil(SOLUONG / 50)

            # ---- Tham số mới/hiện có ----
            shop_name   = request.GET.get("shop_name", "giadungplus_official").strip()
            rating_star = request.GET.get("rating_star", "4,5").strip()
            rate_type   = request.GET.get("rate_type", "new").strip().lower()

            # Ví dụ mặc định 1593536400 như URL anh gửi
            time_start  = int(request.GET.get("time_start", 1593536400))

            # time_end (chỉ dùng cho "old"): = now - 30 ngày (epoch giây, UTC)
            time_end_old = int((datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=30)).timestamp())

            # ---- Reset log cũ ----
            if os.path.exists(log_name):
                os.remove(log_name)

            # ---- Đổi shop theo shop_name ----
            connection_id = get_connection_id(shop_name, existing_shop_map)
            doi_shop(connection_id, loginsp)

            # ---- Base URL chung ----
            base_url = (
                "https://banhang.shopee.vn/api/v3/settings/search_shop_rating_comments_new/"
                "?SPC_CDS=15c18032-c3ae-45ea-9393-85c234ac4a32"
                "&SPC_CDS_VER=2"
                f"&rating_star={rating_star}"
                "&reply_status=1"
                "&language=vi"
            )

            # ---- GHIM KHUNG THỜI GIAN KHI LẤY CŨ (old) ----
            if rate_type == "old":
                base_url += f"&time_start={time_start}&time_end={time_end_old}"
            # (new thì không set, để mặc định backend Shopee trả mới→cũ)

            # ---- Crawl như cũ ----
            all_comment = crawl_shopee_ratings(
                base_url,
                max_pages=MAXPAGES,
                page_size=50,
                delay=0.1
                # Không cần rate_type ở crawler nữa, vì đã chốt khung thời gian ở URL
            )

            # ---- Chuẩn hoá dữ liệu cho OpenAI như cũ ----
            all_name = []

            for comment in all_comment:
                # LẤY THÔNG TIN TRÊN SAPO VỀ. ĐƠN HÀNG >> KHÁCH HÀNG.
                URL_ORDERS = f"{MAIN_URL}/orders.json?query={comment["order_sn"]}"
                ORDER = js_get_url(URL_ORDERS)['orders'][0]

                if ORDER["customer_data"]["sex"] == "male":
                    gender = "Anh"
                else:
                    gender = "Chị"
                #LƯU VÀO ITEM.
                item = {
                    "comment_id": comment["comment_id"],
                    "shopee_order_id": comment["order_id"],
                    "sapo_user_id": ORDER['customer_data']["id"],
                    "order_sn":   comment["order_sn"],
                    "full_name":  ORDER["customer_data"]["name"].replace("+ ",""),
                    "gender":  gender,
                    "short_name":  ORDER["customer_data"]["website"],
                    "user_name":  comment["user_name"],
                    "avatar":     "https://cf.shopee.vn/file/" + comment["user_portrait"]
                }

                if ORDER['customer_data']['sale_order'] != None:
                    comment['old_order'] = ORDER['customer_data']['sale_order']["order_purchases"]
                else:
                    comment['old_order'] = 0

                if not any(it["shopee_order_id"] == comment["order_id"] for it in all_name):
                    all_name.append(item)

            with open(log_name, 'w', encoding='utf-8') as log_file:
                json.dump(all_name, log_file, ensure_ascii=False, indent=2)

            with open(log_comment, 'w', encoding='utf-8') as log_file:
                json.dump(all_comment, log_file, ensure_ascii=False, indent=2)

            return JsonResponse(all_name, safe=False)


        if request.method == "POST":
            def _strip_none(d: dict) -> dict:
                """Loại bỏ các key có value None để không vô tình overwrite thành null trên Sapo."""
                return {k: v for k, v in d.items() if v is not None}

            try:
                body = json.loads(request.body)
            except Exception:
                return JsonResponse({'success': False, 'msg': 'Body không đúng JSON.'})

            comment_id = str(body.get('comment_id'))
            gender = body.get('gender')
            short_name = body.get('short_name')

            if not comment_id or (gender is None and short_name is None):
                return JsonResponse({'success': False, 'msg': 'Thiếu tham số.'})

            # Đọc dữ liệu
            if not os.path.exists(log_name):
                return JsonResponse({'success': False, 'msg': 'File không tồn tại.'})
            try:
                with open(log_name, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            except Exception:
                return JsonResponse({'success': False, 'msg': 'File lỗi hoặc không phải JSON.'})

            # Cập nhật thông tin
            found = False
            for item in data:
                if int(item["comment_id"]) == int(comment_id):
                    if gender is not None:
                        item['gender'] = gender
                    if short_name is not None:
                        item['short_name'] = short_name
                    found = True

                    if gender == "Anh":
                        sex = "male"
                    else:
                        sex = "female"
                    
                    order = js_get_url(f"{MAIN_URL}/customers/{item['sapo_user_id']}.json")

                    cd = (order.get("customer") or {})

                    # Dựng payload theo schema anh yêu cầu, ưu tiên GIỮ NGUYÊN giá trị hiện có trong order
                    cust_payload = {
                        "id":                cd.get("id"),
                        "name":              cd.get("name"),
                        "code":              cd.get("code"),
                        "customer_group_id": cd.get("customer_group_id"),
                        "sex":               sex,
                        "tax_number":        "1",
                        "website":           item['short_name'],
                        "status":            cd.get("status") or "active",
                        "assignee_id":       cd.get("assignee_id"),
                        "apply_incentives":  cd.get("apply_incentives") or "group",
                    }

                    # Tránh gửi None → không overwrite rỗng
                    cust_payload = _strip_none(cust_payload)
                    url = f"{MAIN_URL}/customers/{item['sapo_user_id']}.json"
                    payload = {"customer": cust_payload}
                    put_customer(loginss,url,payload)
                    break

            if not found:
                return JsonResponse({'success': False, 'msg': 'Không tìm thấy comment_id.'})


            # Ghi lại file
            try:
                with open(log_name, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception:
                return JsonResponse({'success': False, 'msg': 'Lỗi ghi file.'})

            return JsonResponse({'success': True, 'msg': 'Cập nhật thành công.'})

    # ----------------------------
    # Tạo file và cho tải về để upload lên chat gpt plus -> Tiết kiệm chi phí hơn #
    # ----------------------------
    if request.GET.get("makerep") == "ok":
        
        with open(log_name, 'r', encoding='utf-8') as f:
            all_name = json.load(f)
        with open(log_comment, 'r', encoding='utf-8') as f:
            all_comment = json.load(f)

        xuat_file_all = []
        for comment in all_comment:
            count = 0
            for name in all_name:
                if int(comment["order_id"]) == int(name["shopee_order_id"]):
                    comment["gender"] = name["gender"]
                    comment["short_name"] = name["short_name"]
                    count +=1
                    break

            if count == 0:
                print(comment["comment_id"])
                print("Not find!")

            item = {
                "comment_id": comment["comment_id"],
                "shopee_order_id": comment["order_id"],
                "order_sn": comment["order_sn"],
                "product_name": comment["product_name"],
                "rating_star": comment["rating_star"],
                "gender": comment["gender"],
                "short_name": comment["short_name"],
                "review_of_customer": comment["comment"],
                "old_order_buy": comment["old_order"]
            }

            xuat_file_all.append(item)

        with open(log_path, 'w', encoding='utf-8') as log_file:
            json.dump(xuat_file_all, log_file, ensure_ascii=False, indent=2)

        return JsonResponse({'success': True, 'msg': 'Cập nhật thành công.'})
    
    # ----------------------------
    # 3. Gửi đánh giá lên Shopee and chat với khách
    # ----------------------------
    if request.GET.get("send_shopee") == "ok":

        if os.path.exists(log_path):
            with open(log_path, "r", encoding="utf-8") as f:
                comment_list = json.load(f)

        # --- 2. Lấy thông tin chat ---
        ALL_CONNECT = js_get_url("https://market-place.sapoapps.vn/go/messenger/connection-chats?key=MzE5OTEx")
        chat_id_connect = find_chat_id_connect(connection_id, ALL_CONNECT)

        if doi_shop(connection_id,loginsp) == 1:
            for comment in comment_list:
                # --- Lấy dữ liệu từ Shopee ---
                URL = f"https://banhang.shopee.vn/api/v3/settings/search_shop_rating_comments_new/?SPC_CDS=4352421f-1de6-47ae-b399-d0b69ec45e70&SPC_CDS_VER=2&search_request={order_sn}&rating_star=5,4,3,2,1&page_number=1&page_size=20&cursor=0&from_page_number=1&language=vi"
                all_data = json.loads(loginsp.get(URL).text)["data"]["list"]
                order_comment_rep, commentx = get_order_comment_rep(all_data, comment_id)

                # --- Gửi đánh giá lên Shopee ---
                URL = "https://banhang.shopee.vn/api/v3/settings/reply_shop_rating/?SPC_CDS=128340d3-0357-49e4-8ba5-5c8a09dd052b&SPC_CDS_VER=2"
                POST_JSON = {"order_id":int(commentx['order_id']),"comment_id":int(comment["comment_id"]),"comment":comment["repcomment_suggest"]}
                rs = loginsp.post(URL,json=POST_JSON)
                
                now = datetime.datetime.now()
                log_dir = "logs/comment"
                log_filename = f"{now.month}_{now.year}-{SERVER_ID}.jsonl"
                log_path = os.path.join(log_dir, log_filename)

                log_entry = {
                    "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
                    "handler": handler,
                    "comment_id": comment_id,
                    "shop_name": shop_name,
                    "rate_star": rate_star,
                    "order_sn": order_sn,
                    "username": username,
                    "repcomment": repcomment
                }

                log_rep_comment(log_path, log_entry)
                obj['repcomment_done'] = 1

                if rate_star >= 4:
                    # Gửi phản hồi lên Shopee
                    URL = "https://banhang.shopee.vn/api/v3/settings/reply_shop_rating/?SPC_CDS=128340d3-0357-49e4-8ba5-5c8a09dd052b&SPC_CDS_VER=2"
                    POST_JSON = {"order_id": int(obj['comment']['order_id']), "comment_id": int(comment_id), "comment": repcomment}
                    rs = loginsp.post(URL, json=POST_JSON)
                    # Gửi tin nhắn nếu chưa từng rep
                    MESS_REP = [
                        f"Hơi muộn nhưng shop vừa đọc được đánh giá của {male} {name} rồi ạ.\nCảm ơn {male} rất nhiều vì đã giành thời gian để đánh giá & ủng hộ shop!",
                        repcomment.replace("\n", "\\n").replace('"', '*'),
                        "Nếu mình có đặt đơn hàng tiếp theo, hãy nhắc shop, để shop gửi 1 phần QUÀ TẶNG ngẫu nhiên cho mình thay lời cảm ơn ạ!"
                    ]
                    if order_comment_rep == 0 and COVER_ID != 0:
                        send_messages_to_customer(COVER_ID, MESS_REP, logintmdt)
                    elif order_comment_rep == 0 and COVER_ID == 0:
                        log_entry['send_flag'] = 0
                        with open('logs/comment/message/log.jsonl', "a", encoding="utf-8") as f:
                            f.write(json.dumps(log_entry, ensure_ascii=False) + "\n")
    
    # ----------------------------
    # 4. Tạo lịch sử đánh giá để trainning cho AI
    # ----------------------------
    if request.GET.get("make_history") == "ok":
        
        connection_id = get_connection_id('giadungplus_official', existing_shop_map)
        doi_shop(connection_id,loginsp)
        
        base_url = (
        "https://banhang.shopee.vn/api/v3/settings/search_shop_rating_comments_new/"
        "?SPC_CDS=15c18032-c3ae-45ea-9393-85c234ac4a32"
        "&SPC_CDS_VER=2"
        "&rating_star=5,4,3,2,1"
        "&reply_status=2"
        "&language=vi"
        )

        all_data = crawl_shopee_ratings(base_url, max_pages=25, page_size=50)
        print(f"Đã crawl tổng cộng {len(all_data)} đánh giá.")
        # Lưu ra file
        all_data_final = []
        for xitem in all_data:
            item = {
                "user_name_of_customer": xitem['user_name'],
                "product_name_buy": xitem['product_name'],
                "comment_of_customer": xitem['comment'],
                "reply_of_shop": xitem['reply'],
                "images": xitem['images'],
                "rating_star": xitem['rating_star']
            }
            all_data_final.append(item)

        with open("logs/openai/history_rep/shopee_ratings.json", "w", encoding="utf-8") as f:
            json.dump(all_data_final, f, ensure_ascii=False, indent=2)
    
    # ----------------------------
    # 5. Tạo file json để gửi cho AI
    # ----------------------------
    if request.GET.get("make_file_500") == "ok":
        
        connection_id = get_connection_id('giadungplus_official', existing_shop_map)
        doi_shop(connection_id,loginsp)
        
        base_url = (
        "https://banhang.shopee.vn/api/v3/settings/search_shop_rating_comments_new/"
        "?SPC_CDS=15c18032-c3ae-45ea-9393-85c234ac4a32"
        "&SPC_CDS_VER=2"
        "&rating_star=5,4"
        "&reply_status=1"
        "&language=vi"
        )

        all_comment = crawl_shopee_ratings(base_url, max_pages=2, page_size=50)
        for comment in all_comment:
            order = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=250&query={comment['order_sn']}")["orders"][0]
            comment["brand"] = order["tags"]
            comment["name_all"] =order["customer_data"]["name"]
            comment["buy_history"] = order["customer_data"]["sale_order"]
            comment["image_of_customer"] = f"https://cf.shopee.vn/file/{comment['user_portrait']}"
            
        # Lưu vào log file
        with open(f"logs/openai/history_rep/file_500_{str(random.randint(1, 5000))}.json", "w", encoding="utf-8") as f:
            json.dump(all_comment, f, ensure_ascii=False, indent=2)


    # 3. Mặc định: trả về nội dung file log nếu có
    if os.path.exists(log_path):
        with open(log_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return JsonResponse(data, safe=False)

    return JsonResponse([], safe=False)

def kd_repnguoimua(request):
    #check_login_sapo()
    obj = {
        "order": {},
    }    
    if 'code' in request.GET:
        code = request.GET['code']
        order = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=100&query={code}")['orders'][0]
        order["mix_name"] = []
        order["full_name"] = order["billing_address"]["full_name"]
        words = order["billing_address"]["full_name"].split()
        combinations = []
        for r in range(1, len(words) + 1):
            combinations.extend(itertools.combinations(words, r))    
        for combination in combinations:
            text = ' '.join(combination)
            order["mix_name"].append(text)
        obj['order'] = order

    if 'order_id' in request.GET:
        order_id = request.GET['order_id']
        male = request.GET['male']
        name = request.GET['name']
        order = js_get_url(f"{MAIN_URL}/orders/{order_id}.json")["order"]

        thanks_1 = f"𝐆𝐢𝐚 𝐃ụ𝐧𝐠 𝐏𝐥𝐮𝐬 gửi lời cảm ơn đến {str(male).capitalize()} {name.capitalize()} đã từng chọn & mua hàng tại shop."
        thanks_2 = f"{(male).capitalize()} {name} đã từng trải nghiệm về hủ gia vị/ bình dầu của chúng em thì cũng biết nó xịn & chất lượng như thế nào rồi ạ."
        thanks_3 = f"Gia Dụng Plus vừa ra thương hiệu mới chuyên về hủ gia vị & bình dầu ăn cao cấp có tên là: ᴘʜᴀʟᴇᴅᴏ (Shopee Mall)"
        thanks_4 = f"Để tri ân cũng như chúng mừng ra mắt thương hiệu, chúng em gửi tới mình ưu đãi【giảm 20% tổng đơn hàng】so với giá đang bán trên Gia Dụng Plus."
        thanks_5 = f"【Đặc biệt】: Khi mua hàng trên Phaledo - {str(male)} {name} sẽ được bảo hành 6 tháng bể vỡ (bể là đổi) miễn phí để mình thêm yên tâm sử dụng ạ."
        
        obj['goiy'] = [thanks_1,thanks_2,thanks_3,thanks_4,thanks_5]
        obj['order'] = order
    return render(request, 'kd_repnguoimua.html', obj)

def kd_connect(request):
    obj = {
        "orders": [],
    }    
    
    return render(request, 'kd_giaovan.html', obj)

def kd_giale(request):
    
    #check_login_sapo()
    list_source = js_get_url(f"{MAIN_URL}/order_sources.json?query=&page=1&limit=50")["order_sources"]
          
    if request.method == "GET":
        
        if 'order_id' in request.GET:
            order_id = request.GET["order_id"]
            url_search = f"{MAIN_URL}/orders.json?ids="+ str(order_id)
            order_json = json.loads(loginss.get(url_search).text)["orders"][0]
            for source in list_source:
                    if source['id'] == order_json["source_id"]:  
                        order_json['source_order'] = source["name"]
                                
            order_json['nhan_vien'] = "Ngọc Vương"

            order_json['ngay_len_don'] = (
                datetime.datetime.strptime(order_json['modified_on'], "%Y-%m-%dT%H:%M:%SZ")  # parse
                + datetime.timedelta(hours=7)                                # cộng +7h
            ).strftime("%d-%m-%Y")

            if order_json['delivery_fee'] == None:
                order_json['delivery_fee'] = {}
                order_json['delivery_fee']['fee'] = 0

            url_search = f"{MAIN_URL}/orders/order_status.json?filter_type=customer_info&customer_ids="+str(order_json['customer_id'])
            dashboard_order = json.loads(loginss.get(url_search).text)["dashboard_order"]['statuses'][0]

            url_search = f"{MAIN_URL}/customers/"+str(order_json['customer_id'])+'.json'
            customer = loginss.get(url_search).text
            if len(customer) > 0:
                customer = json.loads(customer)["customer"]
            # 1. Avatar
            avatar = "/static/customer.png"
            if customer.get("social_customers"):
                avatar_url = customer["social_customers"][0].get("avatar_url")
                if avatar_url:
                    avatar = avatar_url

            # 2. Rank
            sale_order = customer.get("sale_order") or {}
            total_sales = int(sale_order.get("total_sales") or 0)
            order_purchases = int(sale_order.get("order_purchases") or 0)

            rank_customer = 0
            if total_sales >= 10_000_000:
                rank_customer = 3
            elif total_sales >= 3_000_000:
                rank_customer = 2
            elif total_sales >= 1_000_000:
                rank_customer = 1

            if order_purchases >= 15:
                rank_customer = 3
            elif order_purchases > 6 and rank_customer < 2:
                rank_customer = 2
            elif order_purchases > 3 and rank_customer < 1:
                rank_customer = 1

            customer['rank_customer'] = rank_customer
            customer['avatar'] = avatar
            

            if order_json['source_order'] == 'Shopee':
                order_json['dia_chi'] = order_json['shipping_address']['address1']
                if 'Phường' not in order_json['dia_chi']:
                    order_json['dia_chi'] += ', ' + order_json['shipping_address']['ward'] + ', ' +order_json['shipping_address']['district'] +', ' + order_json['shipping_address']['city']
            else:
                order_json['dia_chi'] = order_json['shipping_address']['address1'] + ', ' + order_json['shipping_address']['ward'] + ', ' +order_json['shipping_address']['district'] +', ' + order_json['shipping_address']['city']

            order_json['total_product_price'] = 0

            for line in order_json["order_line_items"]:
                
                if line["variant_id"] != None:
                    line['product_name'] = line['product_name'].split("/")[0]
                    url_search = f"{MAIN_URL}/variants.json?ids="+str(line["variant_id"])
                    vari = json.loads(loginss.get(url_search).text)["variants"][0]
                    line["image"] = vari["images"][0]["full_path"]
                    r = requests.get(line["image"], allow_redirects=True)
                    name = 'assets/saveimage/' + str(vari['id']) + ".jpg"
                    if os.path.exists(name):
                        pass
                    else:
                        print('Save image: '+ name)
                        open(name, 'wb').write(r.content)

                else:
                    line["sku"] = "SERVICE"
                    line["product_name"] = "DỊCH VỤ CỘNG THÊM"
                    
                    line["variant_options"] = line["note"]
                    line["variant_id"] = "love-cus"

                line["quantity"] = int(line["quantity"])
                line["line_amount"] = int(line["line_amount"])
                line["don_gia"] = int(line["price"] - line["discount_value"])
                line["gia_ao"] = int(line["don_gia"] * 1.35)
                order_json['total_product_price'] += int(line["line_amount"]) - int(line["tax_amount"])

                fulfillments = order_json.get("fulfillments", [])
                if fulfillments:
                    if fulfillments[-1].get("shipment", {}) != None:
                        order_json['phiship'] = fulfillments[-1].get("shipment", {}).get("freight_amount") or 0
                    else:
                        order_json['phiship'] = 0
                else:
                    order_json['phiship'] = 0

                if 'TIENSHIP' in order_json['note'] and order_json['phiship'] == 0:
                    order_json['phiship'] = int(order_json['note'].split("TIENSHIP:")[1])
                    print(f"[+] Đơn hàng có phí ship là: {order_json['phiship']}")

                prepayment = order_json.get("prepayments", [])
                if prepayment:  # có ít nhất 1 phần tử
                    order_json['dathanhtoan'] = prepayment[-1].get("paid_amount") or 0
                else:
                    order_json['dathanhtoan'] = 0

                order_json['yeucauthanhtoan'] = order_json['total'] - order_json['dathanhtoan'] 

                if 'TIENSHIP' in order_json['note'] and order_json['phiship'] == 0:
                    order_json['phiship'] = int(order_json['note'].split("TIENSHIP:")[1])


            if 'giale' in request.GET:
                return render(request, 'kd_giale.html', {'order': order_json, 'customer': customer})
            elif 'bangbaogia' in request.GET:
                return render(request, 'kd_bangbaogia.html', {'order': order_json, 'customer': customer})
            else:
                return render(request, 'kd_giasi.html', {'order': order_json, 'customer': customer})

        else:
            if 'source' in request.GET:
                source = str(request.GET["source"])
            else:
                source = 'all'
            if source == 'all':
                source = 'source_id=6510687%2C5483992%2C4917102%2C4893087%2C4864539%2C4557803%2C4394223%2C4339892%2C4339891%2C4339836%2C4339736%2C4339735%2C2172553%2C1880153%2C1880152%2C1880151%2C1880150%2C1880149%2C1880148%2C1880147%2C1880146'
            elif source == 'social':
                source = 'source_id=5483992%2C1880147%2C1880148%2C4339735%2C4339736%2C1880146'
            elif source == 'tmdt':
                source = 'source_id=1880150%2C1880149%2C1880152'
            url_search = f"{MAIN_URL}/orders.json?status=draft%2Cfinalized%2Ccompleted&limit=250&"+ source
            packing_queue = json.loads(loginss.get(url_search).text)["orders"]
            
            doanh_so = 0

            for order in packing_queue:
                for source in list_source:
                    if source['id'] == order["source_id"]:  
                        order['source_order'] = source["name"]
                
                order['nhan_vien'] = "Ngọc Vương"

                time_create = datetime.datetime.strptime(order["created_on"], "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=7)
                start = datetime.datetime.now().replace(hour=0,minute=0,second=0,microsecond=0)
                end = datetime.datetime.now().replace(hour=23,minute=59,second=59,microsecond=59)

            return render(request, 'kd_showdon.html', {'packing_queue': packing_queue})
    else:

        return render(request, 'indexkd.html')

def kd_realtime(request):
    list_nhan_vien = get_list_nhan_vien()
    list_gia_von = get_list_gia_goc()
    brand_all = get_list_brand()
    list_time = {'00':0,'01':0,'02':0,'03':0,'04':0,'05':0,'06':0,'07':0,'08':0,'09':0,'10':0,'11':0,'12':0,'13':0,'14':0,'15':0,'16':0,'17':0,'18':0,'19':0,'20':0,'21':0,'22':0,'23':0}
    date_time_today = {'start':datetime.datetime.combine(datetime.datetime.today(), datetime.time(0, 0, 0, 1)).strftime("%Y-%m-%dT%H:%M:%SZ"), 'end':datetime.datetime.combine(datetime.datetime.today(), datetime.time(23, 59, 59, 999999)).strftime("%Y-%m-%dT%H:%M:%SZ")}

    array_sanpham_ds = {}
    array_donhang = []
    array_sanpham_sl = {}
    brand_all = []
    source_all = []
    list_time = {'00':0,'01':0,'02':0,'03':0,'04':0,'05':0,'06':0,'07':0,'08':0,'09':0,'10':0,'11':0,'12':0,'13':0,'14':0,'15':0,'16':0,'17':0,'18':0,'19':0,'20':0,'21':0,'22':0,'23':0}
    

    if 'realtime' in request.GET:
        gan_date = datetime.datetime.now()
        xa_date = datetime.datetime.now()

    else:
        url_query = f"{MAIN_URL}/brands.json"
        json_order_list = json.loads(loginss.get(url_query).text)
        for brand in json_order_list["brands"]:
            if brand["id"] != 1150688 and brand["id"] != 1150686 and brand["id"] != 1150685 and brand["id"] != 1150687 and brand["id"] != 1150684 and brand["id"] != 1150683 and brand["id"] != 1150682 and brand["id"] != 1150681 and brand["id"] != 1150680 and brand["id"] != 1150675: 
                brand["doanh_thu"] = 0
                brand_all.append(brand)

        url_query = f"{MAIN_URL}/order_sources.json?limit=250&page=1"
        json_order_list = json.loads(loginss.get(url_query).text)
        for brand in json_order_list["order_sources"]:
            source_all.append(brand)
        if 'new' not in request.GET:
            return render(request, 'timereport.html', {
                "brands": brand_all, "source_all":source_all
            })

        if 'date' in request.GET:
            date = str(request.GET["date"])
            xa_date = datetime.datetime(int(date.split(" - ")[0].split("/")[2]), int(date.split(" - ")[0].split("/")[0]), int(date.split(" - ")[0].split("/")[1]))
            gan_date = datetime.datetime(int(date.split(" - ")[1].split("/")[2]), int(date.split(" - ")[1].split("/")[0]), int(date.split(" - ")[1].split("/")[1]))

        if 'brand' in request.GET:
            brand_input = str(request.GET["brand"])
            if "," in brand_input:
                brand_input = brand_input[:-1]

        if 'source' in request.GET:
            source = str(request.GET["source"])
            if "," in source:
                source = source[:-1]

    report_all = {
        "all_orders": [],
        "cong_no": 0,
        "max_order":0,
        "khach":{"cu":0,"moi":0,"c_cu":0,"c_moi":0,"ti_le_quay_lai":0},
        "kho": {"ton_kho":0, "tong_san_pham":0, "tong_goi":0, "tong_ds_goi":0,"tong_goi_yes":0, "tong_ds_goi_yes":0 },
        "sum_all": {
            "don_hang":0,"doanh_so_ao" : 0, "doanh_so_thuc":0, "da_huy":0, "san_pham": 0, "gia_von":0,"c_doanh_so_ao" : 0, "c_doanh_so_thuc":0, "c_da_huy":0,
            "san_pham":0,"don_hang_yes":0,"doanh_so_thuc_yes":0, "ss_don_hang":0,"ss_doanh_so_thuc":0,"ti_le_gia_von":0,
            "tra_hang":0,"c_tra_hang":0,"ds_trung_binh_don":0,"sp_trung_binh_don":0
        },
        "facebook":{
            "ti_le_doanh_thu":0,"doanh_so": 0, "don_hang":0, "trung_binh_don":0, "tin_nhan":0, "tuong_tac":0, "ti_le_chuyen_doi":0,'ads':[],
            "spend_ads":0,"ti_le_ads":0, "tl_comback":0, "dt_comback":0, "pt_comback":0
        },
        "shopee":{
            "ti_le_doanh_thu":0,"doanh_so": 0,'gui_ngoai':0,'c_gui_ngoai':0,'cp_gui_ngoai':0,'tl_gui_ngoai':0,'gui_shopee':0, "don_hang":0, "trung_binh_don":0, 
            "official":0,"vimora":0,"mall":0,"do_su":0, 'dashboard':{},'ti_le_tro_ship':0,'tro_ship_tb':0,'don_duoc_tro_ship':1,
            'tl_extra':0,'phi_san':0,'phi_dv':0,'phi_tt':0,'tong_ship':0,'ship_shopee_tra':0,'ship_khach_tra':0,'mgg_shop':0,'mgg_shopee':0,'tro_gia':0,'slove':0,'ti_le_phi_san':0,'tb_ship_khach_tra':0,
            'm1':0,'m2':0,'m3':0,'m4':0,'m5':0,'tm1':0,'tm2':0,'tm3':0,'tm4':0,'tm5':0,'ads':{},
            "tl_comback":0, "dt_comback":0, "pt_comback":0
        },
        "lazada":{
            "lazada":0,"tiki":0,"ti_le_doanh_thu":0,"doanh_so": 0, "don_hang":0, "trung_binh_don":0, "tl_comback":0, "dt_comback":0, "pt_comback":0
        },
        "zalo":{
            "ti_le_doanh_thu":0,"doanh_so": 0, "don_hang":0, "trung_binh_don":0, "tl_comback":0, "dt_comback":0, "pt_comback":0
        },
        "daily":{
            "ti_le_doanh_thu":0,"doanh_so": 0, "don_hang":0, "trung_binh_don":0
        },
        "ship":{'ds_ship':0,"pt_ship":0, "tl_ship":0, "cp_ship":0,"thu_khach":0}
    }
    CONG_NO = js_get_url(f'{MAIN_URL}/delivery_service_providers/search.json?limit=250&page=1&statuses=active&query=&location_ids=241737')['delivery_service_providers']
    for no in CONG_NO:
        report_all['cong_no'] += int(no['debt'])

    report_all = real_time_report(report_all,gan_date,xa_date, array_sanpham_ds, array_sanpham_sl, list_nhan_vien,list_gia_von,array_donhang,list_time)
    list_nhan_vien = sorted(list_nhan_vien, key=lambda k: k['doanh_so'], reverse=True)

    url_query = f"{MAIN_URL}/reports/inventories/onhand.json"
    json_order_list = json.loads(loginss.get(url_query).text)

    report_all["kho"]["ton_kho"] += int(json_order_list["summary"]["total_global_amount"])
    report_all["kho"]["tong_san_pham"] += int(json_order_list["summary"]["total_global_onhand"])

    result_vari = []
    url_query = f"{MAIN_URL}/variants.json"
    json_order_list = json.loads(loginss.get(url_query).text)
    if json_order_list["metadata"]["total"] > 0:    
        ROUND = math.ceil(json_order_list["metadata"]["total"]/250)
        for i in range(ROUND):
            url_search = url_query + "?limit=250&page=" + str(i+1)
            json_order_list = json.loads(loginss.get(url_search).text)
            list_vari = json_order_list["variants"]
            # Iterate over the data and write it out row by row.
            for vari in list_vari:
                if str(vari["id"]) in array_sanpham_ds:
                    if "FB_"+str(vari["id"]) not in array_sanpham_sl:
                        array_sanpham_sl["FB_"+str(vari["id"])] = 0
                    if "SI_"+str(vari["id"]) not in array_sanpham_sl:
                        array_sanpham_sl["SI_"+str(vari["id"])] = 0
                    if "LZ_"+str(vari["id"]) not in array_sanpham_sl:
                        array_sanpham_sl["LZ_"+str(vari["id"])] = 0
                    if "SP_"+str(vari["id"]) not in array_sanpham_sl:
                        array_sanpham_sl["SP_"+str(vari["id"])] = 0

                    result_vari.append({"id":vari["id"],"sku":vari["sku"],"image":vari["images"][0]["full_path"],
                        "rs_ds":int(array_sanpham_ds[str(vari["id"])]),"rs_sl":int(array_sanpham_sl[str(vari["id"])]),
                        "rs_sp":int(array_sanpham_sl["SP_"+str(vari["id"])]),"rs_si":int(array_sanpham_sl["SI_"+str(vari["id"])]),
                        "rs_fb":int(array_sanpham_sl["FB_"+str(vari["id"])]),"rs_lz":int(array_sanpham_sl["LZ_"+str(vari["id"])]),
                        "rs_tonkho": int(vari["inventories"][0]["on_hand"] + vari["inventories"][1]["on_hand"])
                        })

                    for brand in brand_all:
                        if vari["brand_id"] == brand["id"]:
                            brand["doanh_thu"] += int(array_sanpham_ds[str(vari["id"])])
                            break

    brand_all = sorted(brand_all, key=lambda k: k['doanh_thu'], reverse=True)
    brand_all = brand_all[:10]

    for x in list_time.values():
        if report_all['max_order'] < int(x):
            report_all['max_order'] = x*1.05


    for ads in report_all['facebook']['ads']:
        report_all['facebook']['spend_ads'] += int(ads['fb_amount_spend'])
    if report_all['facebook']['doanh_so']+report_all['zalo']['doanh_so'] > 0:
        report_all['facebook']['ti_le_ads'] = "%.2f" % float(( report_all['facebook']['spend_ads']*100/(report_all['facebook']['doanh_so']+report_all['zalo']['doanh_so'])))
        report_all['facebook']['spend_ads'] = human_format(int(report_all['facebook']['spend_ads']))

    report_all["khach"]["ti_le_quay_lai"] = "%.2f" % float(report_all["khach"]["c_cu"]*100 / (report_all["khach"]["c_cu"] + report_all["khach"]["c_moi"]))
    report_all["sum_all"]["ds_trung_binh_don"] = "%.0f" % float((report_all["sum_all"]["doanh_so_thuc"]/report_all["sum_all"]["c_doanh_so_thuc"]))
    report_all["sum_all"]["sp_trung_binh_don"] = "%.2f" % float((report_all["sum_all"]["san_pham"]/report_all["sum_all"]["c_doanh_so_thuc"]))

    url = "https://banhang.shopee.vn/api/report/miscellaneous/upload_router_info?SPC_CDS=b66f9fb0-3b25-4509-86a4-f3263fe6aee5&SPC_CDS_VER=2"
    payload = {"route_pattern":"/^(?:\\/(?=$))?$/i","region":"vn","is_cn":False,"nav_type":1}
    loginsp.post(url,json=payload)

    url = "https://banhang.shopee.vn/api/mydata/homepage/key-metrics?SPC_CDS=744159b6-aba6-4abd-9298-99a0a0392a84&SPC_CDS_VER=2&order_type=confirmed"
    rs = loginsp.get(url).text
    report_all['shopee']['dashboard'] = json.loads(rs)['data']
    report_all['shopee']['dashboard']['conversion_rate'] = "%.2f" % float(report_all['shopee']['dashboard']['conversion_rate']*100)
    report_all['shopee']['dashboard']['uv_pct_diff'] = "%.2f" % float(report_all['shopee']['dashboard']['uv_pct_diff']*100)
    report_all['shopee']['dashboard']['pv_pct_diff'] = "%.2f" % float(report_all['shopee']['dashboard']['pv_pct_diff']*100)
    report_all['shopee']['dashboard']['conversion_rate_pct_diff'] = "%.2f" % float(report_all['shopee']['dashboard']['conversion_rate_pct_diff']*100)
   
    if 'realtime' in request.GET:
        START_TIME = datetime.datetime.now().replace(hour=00, minute=00, second=00)
        END_TIME = datetime.datetime.now().replace(hour=23, minute=59, second=59)
        START_TIME = int(round(START_TIME.timestamp()))
        END_TIME = int(round(END_TIME.timestamp()))

        url = "https://banhang.shopee.vn/api/report/miscellaneous/upload_router_info?SPC_CDS=b66f9fb0-3b25-4509-86a4-f3263fe6aee5&SPC_CDS_VER=2"
        payload = {"route_pattern":"/^\\/portal\\/marketing\\/pas\\/assembly(?:\\/(?=$))?$/i","region":"vn","is_cn":False,"nav_type":2}
        loginsp.post(url,json=payload)
        url = f"https://banhang.shopee.vn/api/marketing/v3/pas/report/homepage_report_by_time/?start_time={START_TIME}&end_time={END_TIME}&agg_interval=96&campaign_type=cpc_homepage&SPC_CDS=9311b433-7d79-463c-b7ec-5e609dc100dd&SPC_CDS_VER=2"
        rs = loginsp.get(url).text
        report_all['shopee']['ads'] = json.loads(rs)['data']['homepage_aggregate']
        report_all['shopee']['ads']['tl_ads'] = "%.2f" % float((report_all['shopee']['ads']['cost']*100 /report_all['shopee']['ads']['order_gmv']))
        report_all['shopee']['ads']['tl_sum_ads'] = "%.2f" % float((report_all['shopee']['ads']['cost']*100 /report_all['shopee']['official']))
    
        for key in report_all['shopee']['ads'].keys():
            if  type(report_all['shopee']['ads'][key]) != str and report_all['shopee']['ads'][key] > 1000:
                report_all['shopee']['ads'][key] = human_format(int(report_all['shopee']['ads'][key]))

    report_all['shopee']['slove'] = "%.2f" % float((report_all['shopee']['tro_gia'] + report_all['shopee']['ship_shopee_tra'])/report_all['shopee']['phi_san'])
    report_all['shopee']['ti_le_phi_san'] = "%.2f" % float((report_all['shopee']['phi_san']*100 /report_all['shopee']['doanh_so']))
    report_all['shopee']['tb_ship_khach_tra'] = int((report_all['shopee']['ship_khach_tra'] /report_all['shopee']['don_hang']))
    report_all['shopee']['ti_le_tro_ship'] = "%.2f" % float((report_all['shopee']['ship_shopee_tra']*100 /report_all['shopee']['tong_ship']))
    report_all['shopee']['tl_extra'] = "%.2f" % float((report_all['shopee']['phi_dv']*100 /report_all['shopee']['gui_shopee']))

    if report_all['facebook']['doanh_so'] > 0:
        report_all['facebook']['tl_comback'] = "%.2f" % float((report_all['facebook']['tl_comback']*100 /report_all['facebook']['don_hang']))
        report_all['facebook']['pt_comback'] = "%.2f" % float((report_all['facebook']['dt_comback']*100 /report_all['facebook']['doanh_so']))
    if report_all['shopee']['doanh_so'] > 0:
        report_all['shopee']['tl_comback'] = "%.2f" % float((report_all['shopee']['tl_comback']*100 /report_all['shopee']['don_hang']))
        report_all['shopee']['pt_comback'] = "%.2f" % float((report_all['shopee']['dt_comback']*100 /report_all['shopee']['doanh_so']))

    if report_all['zalo']['doanh_so'] > 0:
        report_all['zalo']['tl_comback'] = "%.2f" % float((report_all['zalo']['tl_comback']*100 /report_all['zalo']['don_hang']))
        report_all['zalo']['pt_comback'] = "%.2f" % float((report_all['zalo']['dt_comback']*100 /report_all['zalo']['doanh_so']))

    if report_all['lazada']['doanh_so'] > 0:
        report_all['lazada']['tl_comback'] = "%.2f" % float((report_all['lazada']['tl_comback']*100 /report_all['lazada']['don_hang']))
        report_all['lazada']['pt_comback'] = "%.2f" % float((report_all['lazada']['dt_comback']*100 /report_all['lazada']['doanh_so']))

    if report_all['ship']['ds_ship'] > 0:
        report_all['ship']['pt_ship'] = report_all['ship']['cp_ship'] - report_all['ship']['thu_khach']
        report_all['ship']['tl_ship'] = "%.2f" % float((report_all['ship']['pt_ship']*100 /report_all['ship']['ds_ship']))

    if report_all['shopee']['gui_ngoai'] != 0:
        report_all['shopee']['tt_gui_ngoai'] = "%.2f" % float((report_all['shopee']['cp_gui_ngoai']*100 /report_all['shopee']['gui_ngoai']))

    report_all['shopee']['tm1'] = "%.2f" % float((report_all['shopee']['m1']*100 /report_all['shopee']['don_hang']))
    report_all['shopee']['tm2'] = "%.2f" % float((report_all['shopee']['m2']*100 /report_all['shopee']['don_hang']))
    report_all['shopee']['tm3'] = "%.2f" % float((report_all['shopee']['m3']*100 /report_all['shopee']['don_hang']))
    report_all['shopee']['tm4'] = "%.2f" % float((report_all['shopee']['m4']*100 /report_all['shopee']['don_hang']))
    report_all['shopee']['tm5'] = "%.2f" % float((report_all['shopee']['m5']*100 /report_all['shopee']['don_hang']))

    report_all['shopee']['tro_ship_tb'] = human_format(report_all['shopee']['ship_shopee_tra'] /report_all['shopee']['don_duoc_tro_ship'])
    report_all['shopee']['phi_san'] = human_format(report_all['shopee']['phi_san'])
    report_all['shopee']['tong_ship'] = human_format(report_all['shopee']['tong_ship'])
    report_all['shopee']['ship_khach_tra'] = human_format(report_all['shopee']['ship_khach_tra'])
    report_all['shopee']['ship_shopee_tra'] = human_format(report_all['shopee']['ship_shopee_tra'])
    report_all['shopee']['gui_ngoai'] = human_format(report_all['shopee']['gui_ngoai'])
    report_all['shopee']['cp_gui_ngoai'] = human_format(int(report_all['shopee']['cp_gui_ngoai']))
    report_all['shopee']['tro_gia'] = human_format(int(report_all['shopee']['tro_gia']))
    report_all['shopee']['phi_dv'] = human_format(int(report_all['shopee']['phi_dv']))

    return render(request, 'index.html',{'report': report_all, "list_nhan_vien": list_nhan_vien, "result_vari":result_vari, "brands": brand_all,'list_donhang':array_donhang,'CONG_NO':CONG_NO,'list_time':list_time})

def kd_chatshopee(request):
    print("Update Chat Shopee")
    logintmdt.get(f"https://market-place.sapoapps.vn/go/messenger/conversations/17615/crawl")
    time.sleep(15)        
    print("Update Chat Shopee")
    for i in range(0,20):
        rs = logintmdt.get(f"https://market-place.sapoapps.vn/search/conversation/filter?sortType=desc&connectionIds={connectID}&replied=false&queryType=account&replied=false&page={i+1}&limit=20")
        rs = json.loads(rs.text)
        for chat in rs["conversations"]:
            print(f"Update chat: {chat['customer_name']}")
            logintmdt.get(f"https://market-place.sapoapps.vn/go/messenger/conversations/{chat['connection_chat_id']}/crawl?conversationId={chat['id']}")
            time.sleep(15)

    return render(request, 'copyanh.html')    

def mkt_huongcontent(request):
    if request.method == 'GET':
        return render(request, 'mkt_huongcontent.html')


def mkt_listproduct(request):
    obj = {
        "orders": [],
        "products": [],
        "qa": []
    }

    # ==== BƯỚC 1: XÁC THỰC GOOGLE SHEET ====
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
    creds = Credentials.from_service_account_file("logs/app-project-sapo-plus-eb0edec5c0dc.json", scopes=SCOPES)
    client = gspread.authorize(creds)

    # ==== BƯỚC 2: MỞ FILE GOOGLE SHEET ====
    sheet_url = "https://docs.google.com/spreadsheets/d/1RzS18QjF-sg-yGRz_SpVIhFOK9gz1x9mIkkAClbd1mk/edit?usp=sharing"
    worksheet = client.open_by_url(sheet_url).worksheet("ALL")  # Hoặc tên sheet cụ thể
    # ==== BƯỚC 4: CHUYỂN DỮ LIỆU VÀO obj["qa"] ====

    rows = worksheet.get_all_values()

    for i, row in enumerate(rows[2:], start=2):  # bỏ dòng tiêu đề
        try:
            # Bỏ qua dòng rỗng
            if not any(row):  
                continue
            question = str(row[0]).strip()
            answer1 = str(row[1]).strip()
            answer2 = str(row[2]).strip() if len(row) > 2 else ""
            tags_raw = str(row[3]).strip() if len(row) > 3 else ""
            related_products = str(row[4]).strip() if len(row) > 4 else ""

            # Bỏ dòng nếu thiếu câu hỏi hoặc câu trả lời chính
            if not question or not answer1:
                continue

            tags = [tag.strip() for tag in tags_raw.split(",") if tag.strip()]
            
            obj["qa"].append({
                "question": question,
                "answer": answer1,
                "answer_followup": answer2 if answer2 else None,
                "tags": tags,
                "related_products": related_products
            })

        except Exception as e:
            print(f"[!] Lỗi tại dòng {i}: {e}")
            continue

    all_value = []
    all_products = js_get_url(f"{MAIN_URL}/products.json?product_types=normal&page=1&limit=250")["products"]
    for pr in all_products:
        all_value.append(pr)
    all_products = js_get_url(f"{MAIN_URL}/products.json?product_types=normal&page=2&limit=250")["products"]
    for pr in all_products:
        all_value.append(pr)     

    for pr in all_value:
        pr['stock'] = 0
        pr['can_sale'] = 0
        pr['dang_ve'] = 0
        pr['list_qa'] = []
        pr['shop_post'] = {'giadungplus':0}
        # duyệt qua các tags
        for tag in pr['tags'].split(","):
            if 'shop:' in tag:
                shop_name = tag.split(":")[1].strip()
                if shop_name:
                    pr['shop_post'][shop_name] = 0
                   

        for vari in pr["variants"]:
            if vari['packsize'] == False:
                try:
                    info_vari = Product.objects.get(id=vari['id'])
                except Product.DoesNotExist:
                    pass
                else:
                    vari['sp_height'] = info_vari.sp_height
                    vari['sp_width'] = info_vari.sp_width
                    vari['sp_length'] = info_vari.sp_length
                    vari['sp_weight'] = info_vari.sp_weight

                pr['stock'] += int(vari['inventories'][0]['on_hand'] + vari['inventories'][1]['on_hand'])
                pr['can_sale'] += int(vari['inventories'][0]['available'] + vari['inventories'][1]['available'])
                pr['dang_ve'] += int(vari['inventories'][0]['incoming'] + vari['inventories'][1]['incoming'])
            else:
                vari['error'] = 1       

    all_value2 = []
    all_products = js_get_url(f"{MAIN_URL}/products.json?product_types=composite&page=1&limit=250")["products"]
    for pr in all_products:
        all_value2.append(pr)

    for pr in all_value2:
        for vari in pr["variants"]:
            for xpr in all_value:
                count = 0
                for composite_items in vari['composite_items']:
                    for xvari in xpr['variants']:
                        if count == 0:
                            if xvari['id'] == composite_items['sub_variant_id']:
                                xpr['variants'].append(vari)
                                count+=1
                                break


    # Lấy ngày hôm nay (UTC, dạng ISO)
    today = datetime.datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    start_date = today - datetime.timedelta(days=30)
    end_date_str = today.isoformat() + "Z"
    start_date_str = start_date.isoformat() + "Z"
    all_sales = []
    base_url = "https://sisapsan.mysapogo.com/admin/reports/sales/by_variant.json"
    for i in range(4):
        params = {
            "location_ids": "",
            "start_date": start_date_str,
            "end_date": end_date_str,
            "limit": 250,
            "page": int(i+1),
            "criteria": "appeared_rate"
        }

        report_url = f"{base_url}?{urlencode(params)}"
        report_sales = js_get_url(report_url).get('items')

        if len(report_sales) > 0:
            all_sales.extend(report_sales)
        else:
            break

    TMDT_PRODUCT = []
    for i in range(20):
        URL_SEARCH = f"https://market-place.sapoapps.vn/products/v2/filter?page={int(i+1)}&limit=250&tenantId=1262&connectionIds={connectID}"
        xproduct = js_get_url(URL_SEARCH).get("products")
        if xproduct == None:
            break
        else:
            for pr in xproduct:
                TMDT_PRODUCT.append(pr)

    for pr in all_value:
        for vari in pr["variants"]:
            vari['luotban30ngay'] = 0
            vari['ds30ngay'] = 0
            vari['don30ngay'] = 0
            for sales in all_sales:
                if str(sales['variant_id']) == str(vari['id']):
                    vari['luotban30ngay'] = int(sales['quantity'])
                    vari['ds30ngay'] = int(sales['amount_for_sorting'])
                    vari['don30ngay'] = int(sales['order_for_sorting'])
                    break

    for pr in all_value:
        pr['post_count'] = 0
        pr['luotban30ngay'] = 0
        pr['ds30ngay'] = 0
        pr['don30ngay'] = 0
        for vari in pr["variants"]:
            pr['luotban30ngay'] += int(vari['luotban30ngay'])
            pr['ds30ngay'] += int(vari['ds30ngay'])
            pr['don30ngay'] += int(vari['don30ngay'])
            vari['post_status'] = 0
            vari['post_link'] = []
            for xpr in TMDT_PRODUCT:
                for xvari in xpr["variants"]:
                    if xvari['sapo_variant_id'] != None:
                        if int(xvari['sapo_variant_id']) == int(vari["id"]):
                            xvari['shop_name'] = get_connection_name(xpr['connection_id'], existing_shop_map)

                            for key in pr["shop_post"].keys():
                                if key in xvari['shop_name']:  # ví dụ 'giadungplus' in 'giadungplus_official'
                                    pr["shop_post"][key] = 1

                            xvari['image_pr'] = xpr['image']
                            vari['post_link'].append(xvari)
                            vari['post_status'] = 1

            if vari['post_status'] == 1:
                pr['post_count'] += 1

            if vari['packsize'] == True:
                pr['post_count'] += 1

            if vari['sellable'] == False:
                pr['post_count'] += 1

        if pr['post_count'] == 0:
            pr['post_status'] = 0
        elif pr['post_count'] < len(pr["variants"]):
            pr['post_status'] = 1
        else:
            pr['post_status'] = 2

    sorted_products = sorted(all_value,
        key=lambda x: (x["status"] != "active", -x["luotban30ngay"] if x["status"] == "active" else 0)
    )

    obj["products"] = sorted_products


    return render(request, 'mkt_listproduct.html', obj)

        

def mkt_copyanh_tmall(request):
    if request.method == 'GET':
        return render(request, 'mkt_copyanhtmall.html')

    minput = request.POST.get('data', '')
    soup = BeautifulSoup(minput, 'html.parser')
    image_tags = soup.find_all('img')

    result = []
    seen = set()

    for tag in image_tags:
        src = tag.get('src')
        if not src:
            continue

        # Bỏ qua icon SVG
        if src.lower().endswith('.svg'):
            continue

        # Chuẩn hóa URL: loại bỏ các hậu tố như _q50.jpg_.webp thành .jpg
        jpg_url = re.sub(r"\.jpg.*$", ".jpg", src, flags=re.IGNORECASE)

        # Loại bỏ trùng lặp
        if jpg_url in seen:
            continue

        seen.add(jpg_url)
        result.append(jpg_url)

    return render(request, 'mkt_copyanhtmall.html', {'data': result})


def show_listsanpham(request):
    obj = {'all':[]}

    doi_shop(10925,loginsp)
    now = datetime.datetime.today()
    for i in range(10):
        URL = f"https://banhang.shopee.vn/api/v3/mpsku/list/v2/get_product_list?SPC_CDS=fa1bf47d-d561-41e3-b518-208e2bdc458a&SPC_CDS_VER=2&page_number={int(i+1)}&page_size=30&list_type=all&need_ads=true"
        all_pr = json.loads(loginsp.get(URL).text)
        if 'data' in all_pr:
            if 'products' in all_pr['data']:
                for pr in all_pr["data"]["products"]:
                    
                    item = {
                        'id':pr['id'],
                        'status':pr['status'],
                        'image':"https://cf.shopee.vn/file/"+pr['cover_image'],
                        'parent_sku':pr['parent_sku'],
                        'name':pr['name'],
                        'statistics_view_count':pr['statistics']['view_count'],
                        'statistics_liked_count':pr['statistics']['liked_count'],
                        'statistics_sold_count':pr['statistics']['sold_count'],
                        'price_detail': f"₫ {int(float(pr['price_detail']['selling_price_min'])):,} - ₫ {int(float(pr['price_detail']['selling_price_max'])):,}",
                        'max_discount':f"{int(pr['price_detail']['max_discount']/100)}%",
                        'today':f"{now.day}/{now.month}/{now.year}"
                    }

                    obj['all'].append(item)
    if request.GET["view"] == "list":
        return render(request, 'show_listsanpham_shopee.html', obj)  
    else:
        return render(request, 'show_luotban_shopee.html', obj)  

YEAR_MONTH_CUSTOMER = {
    2025: {
        7:  751556056,
        8:  751556365,
        9:  751556535,
        10: 751556712,
        11: 751556918,
        12: 751557103,
    },
    # 2026 có thể thêm sau:
    2026: {
        1:  751557300,
        2:  751557450,
        # ...
    }
}
TICKET_CUSTOMER_ID = 750912288

def kho_ticketprocess(request):
    obj = {
        'kho': 'hn'
    }

    kho_q = request.GET.get('kho')
    if kho_q:
        request.session['kho_filter'] = kho_q
    selected_kho = request.session.get('kho_filter')
    obj['kho'] = selected_kho

    return render(request, 'kho_ticketprocess.html', obj) 

def kd_ticketprocess(request):
    obj = {
        'kho': ''
    }

    kho_q = request.GET.get('kho')
    if kho_q:
        request.session['kho_filter'] = kho_q
    selected_kho = request.session.get('kho_filter')
    obj['kho'] = selected_kho

    return render(request, 'kd_ticketprocess.html', obj) 


def api_careorder_listall(request):
    selected_kho = request.session.get('kho_filter', 'all')

    if selected_kho == 'all':
        selected_kho = ''
    # 1. Lấy tất cả note của customer
    notes = js_get_url(f"{MAIN_URL}/customers/750912288.json")["customer"]["notes"]
    tickets = []
    now_local = datetime.datetime.utcnow() + datetime.timedelta(hours=7)

    for note in notes:
        if note.get("status") != "active":
            continue
        try:
            # 2. Parse ticket từ content
            ticket = json.loads(note["content"])
            save_ticket = ticket
            # Nếu trạng thái modify đã quá 1 tiếng thì chuyển về thông báo lại tiếp.
            mod_str = note["modified_on"]  # ex: "2025-07-18T13:13:14Z"
        
            # 3. Loại bỏ timezone suffix ("Z" hoặc "+hh:mm")
            if mod_str.endswith("Z"):
                mod_str = mod_str[:-1]
            elif "+" in mod_str:
                mod_str = mod_str.split("+")[0]

            # 4. Parse thành naive datetime, sau đó +7h
            try:
                dt = datetime.datetime.fromisoformat(mod_str)
            except ValueError:
                continue
            modified_local = dt + datetime.timedelta(hours=7)
            st = ticket.get("status")
            if now_local - modified_local > datetime.timedelta(hours=1) and (st==2 or st==4):
                note["content"] = json.dumps(save_ticket, ensure_ascii=False)
                put_url = f"{MAIN_URL}/customers/{TICKET_CUSTOMER_ID}/notes/{note['id']}.json"
                put_resp = loginss.put(put_url, json=note)
                if put_resp.status_code in (200, 201):
                    print(f"✅ Đã cập nhật note {note['id']} → status {content['status']}")
                else:
                    print(f"❌ Lỗi cập nhật note {note['id']}: HTTP {put_resp.status_code}")

            ticket["note_id"] = note["id"]
            # 3. Lấy chi tiết order (một object) thay vì mảng
            od = js_get_url(f"{MAIN_URL}/orders/{ticket['o_id']}.json")
            # endpoint trả về { "order": { … } }
            ticket["info"] = od.get("order", {})
            ticket["info"] = get_data_packing(ticket["info"])

            if int(ticket['status']) != 9 and selected_kho in ticket['kho']:
                tickets.append(ticket)
        except (KeyError, json.JSONDecodeError):            
            continue

    # 2. Lấy toàn bộ chat notes cho tháng/năm hiện tại
    now = datetime.datetime.now()
    cust_chat = YEAR_MONTH_CUSTOMER.get(now.year, {}).get(now.month)
    chat_notes = []
    if cust_chat:
        resp = js_get_url(f"{MAIN_URL}/customers/{cust_chat}.json")
        chat_notes = resp.get("customer", {}).get("notes", [])

    # 3. Gán log_chat cho mỗi ticket
    for ticket in tickets:
        ti = ticket.get("ti_id")
        log = []
        for cn in chat_notes:
            if cn.get("status") != "active":
                continue
            try:
                msg = json.loads(cn["content"])
                if msg.get("ticket_id") == ti:
                    log.append(msg)
            except:
                continue
        ticket["log_chat"] = log

    # 4. Ưu tiên sort status 3,4 lên trên
    def pri(t):
        return 0 if t.get("status") in (3,4) else 1
    tickets.sort(key=pri)


    return JsonResponse(tickets, safe=False)


def api_careorder_checkorderinfo(request):
    search = request.GET.get('search', '').strip()
    data = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=1&query={search}")
    orders = data.get("orders", [])

    if orders:
        if orders[0]['location_id'] == 241737:
            orders[0]['kho'] = "Kho Hà Nội"
        else:
            orders[0]['kho'] = "Kho HCM"

        # Trả về theo đúng định dạng {"order": { ... }}
        return JsonResponse({"order": orders[0]})
    else:
        # Không tìm thấy thì trả về order = null (hoặc có thể 404 tuỳ ý)
        return JsonResponse({"order": None})


@csrf_exempt
def api_careorder_newticket(request):
    if request.method != "POST":
        return JsonResponse({"status":"error","msg":"Phải POST"}, status=405)

    # 1. Lấy dữ liệu từ form
    o_sn    = request.POST.get("o_sn", "").strip()
    kho     = request.POST.get("kho", "").strip()
    t_type  = request.POST.get("t_type", "").strip()  # nếu bạn có field này
    des     = request.POST.get("des", "").strip()
    # Giả sử client gửi image IDs dưới name="image_ids[]"
    image_urls = request.POST.getlist('ticket_images') 

    if not o_sn:
        return JsonResponse({"status":"error","msg":"Thiếu mã đơn hàng"}, status=400)

    # 2. Tra cứu order để lấy o_id
    data_orders = js_get_url(f"{MAIN_URL}/orders.json?page=1&limit=1&query={o_sn}")
    orders = data_orders.get("orders", [])
    if not orders:
        return JsonResponse({"status":"error","msg":"Không tìm thấy đơn"}, status=404)
    o_id = orders[0].get("id")

    # 3. Sinh ti_id dạng TK-DDMMYYYY-XXX với suffix random
    today = datetime.datetime.now().strftime("%d%m%Y")
    suffix = str(random.randint(0, 999)).zfill(3)
    ti_id = f"TK-{today}-{suffix}"

    # 4. Xây payload của ticket
    ticket = {
        "ti_id": ti_id,
        "kho": kho,
        "o_id": o_id,
        "o_sn": o_sn,
        "status": 1,
        "t_type": t_type,
        "des": des,
        "image": image_urls
    }

    # 5. Gửi lên Sapo tạo note mới
    url_note = f"{MAIN_URL}/customers/750912288/notes.json"

    payload = { "content": json.dumps(ticket, ensure_ascii=False) }

    resp = loginss.post(url_note, json=payload)
    
    if resp.status_code in (200, 201):
        return JsonResponse({"status":"ok"})
    else:
        return JsonResponse({
            "status":"error",
            "msg": f"Lưu note thất bại ({resp.status_code})"
        }, status=500)


@csrf_exempt
def api_careorder_editticket(request):
    if request.method != "POST":
        return JsonResponse({"status":"error","msg":"Phải POST"}, status=405)

    note_id = request.POST.get("note_id")
    if not note_id:
        return JsonResponse({"status":"error","msg":"Thiếu note_id"}, status=400)

    # 1. Lấy note hiện tại từ Sapo
    url_get = f"{MAIN_URL}/customers/{TICKET_CUSTOMER_ID}/notes/{note_id}.json"
    res_get = loginss.get(url_get)
    if res_get.status_code != 200:
        return JsonResponse({"status":"error","msg":"Không tìm thấy note"}, status=404)

    note_obj = res_get.json().get("note")

    try:
        ticket_data = json.loads(note_obj["content"])
    except json.JSONDecodeError:
        return JsonResponse({"status":"error","msg":"Content không phải JSON"}, status=500)

    # 3. Thu thập các field cần update từ request.POST
    updates = {}
    for key, val in request.POST.items():
        if key == "note_id" or key == "ticket_id":
            continue
        if key.endswith("[]"):
            updates[key[:-2]] = request.POST.getlist(key)
        else:
            updates[key] = val

    # 4. Merge vào ticket_data
    ticket_data.update(updates)

    # 5. Gán lại vào note_obj["content"]
    note_obj["content"] = json.dumps(ticket_data, ensure_ascii=False)

    # 3. Gửi PUT lên Sapo để update
    url_put = f"{MAIN_URL}/customers/{TICKET_CUSTOMER_ID}/notes/{note_id}.json"
    resp = loginss.put(url_put, json=note_obj)

    if resp.status_code in (200, 201):
        return JsonResponse({"status":"ok"})
    else:
        return JsonResponse({
            "status":"error",
            "msg": f"Cập nhật thất bại ({resp.status_code})"
        }, status=500)


@csrf_exempt
def api_careorder_newchat(request):
    if request.method != "POST":
        return JsonResponse({"status":"error","msg":"Phải POST"}, status=405)

    ticket_id    = request.POST.get("ticket_id")
    chat_content = request.POST.get("chat_content","").strip()
    sender       = request.POST.get("from","kho").strip()

    if not ticket_id or not chat_content:
        return JsonResponse({"status":"error","msg":"Thiếu ticket_id hoặc chat_content"}, status=400)

    now = datetime.datetime.now()
    year  = now.year
    month = now.month
    customer_id = YEAR_MONTH_CUSTOMER.get(year, {}).get(month)

    # xây payload message
    message = {
        "message_id": f"MSG-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}-{random.randint(0,999):03d}",
        "ticket_id": ticket_id,
        "from": sender,
        "chat_content": chat_content,
        "timestamp": datetime.datetime.utcnow().isoformat() + "Z"
    }

    url = f"{MAIN_URL}/customers/{customer_id}/notes.json"
    payload = {"content": json.dumps(message, ensure_ascii=False)}
    resp = loginss.post(url, json=payload)

    if resp.status_code in (200,201):
        return JsonResponse({"status":"ok"})
    else:
        return JsonResponse({"status":"error","msg":f"Lưu chat thất bại ({resp.status_code})"}, status=500)

# Thay PRODUCT_ID bằng ID sản phẩm bạn dùng để chứa ảnh ticket
PRODUCT_ID = 42673510

@csrf_exempt
def api_uploadimage_ticket(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'msg': 'Phải sử dụng POST'}, status=405)

    # 1. Parse JSON payload
    try:
        payload    = json.loads(request.body.decode('utf-8'))
        b64        = payload.get('base64')
        position   = 1
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'msg': 'JSON không hợp lệ'}, status=400)

    ts = datetime.datetime.now().strftime("%Y%m%d%H%M%S%f")
    file_name = f"{ts}.jpg"

    # 2. Gửi lên Sapo
    url = f"https://sisapsan.mysapogo.com/admin/products/42673510/images.json"
    sapo_payload = {
        'base64':   b64,
        'file_name': file_name,
        'id': 0,
        'isLoading': False,
        'position': 1
    }
    sapo_resp = loginss.post(url, json=sapo_payload)
    if sapo_resp.status_code not in (200, 201):
        return JsonResponse({
            'status': 'error',
            'msg': f'Upload lên Sapo thất bại ({sapo_resp.status_code})'
        }, status=500)

    sapo_data = sapo_resp.json().get('image')

    if not sapo_data:
        return JsonResponse({'status': 'error', 'msg': 'Sapo trả về dữ liệu không hợp lệ'}, status=500)

    # 3. Trả về client đầy đủ thông tin
    return JsonResponse({
        'status':   'ok',
        'image':    sapo_data,
        'position': position
    }, safe=False)


def lovepage(request):
    return render(request, 'love.html')


@csrf_exempt
def sync_nhanphu_from_customer_note(request):
    """
    Đồng bộ thông tin nhãn phụ từ customer note lên product description metagdp.
    
    Customer note ID: 760093681
    Format JSON trong note:
    {
        "product_id": 193446020.0,
        "sku_first": "LT-0437",
        "nsx": "LTENG",
        "brand": "lteng",
        "vi_name": "Giỏ kệ khăn tắm",
        "en_name": "LTENG Wall Towel Basket",
        "descreption": "Giỏ kệ nhôm LTENG dán tường chắc chắn...",
        "material": "Nhôm không gian",
        "sku": ""
    }
    
    Push lên các thông tin: brand, vi_name, en_name, descreption, material
    Lưu vào nhanphu_info của product description metagdp.
    """
    from core.sapo_client import get_sapo_client
    from products.services.sapo_product_service import SapoProductService
    from products.services.dto import NhanPhuInfoDTO
    from products.services.metadata_helper import init_empty_metadata
    
    CUSTOMER_ID = 760093681
    
    result = {
        "status": "ok",
        "processed": 0,
        "success": 0,
        "errors": 0,
        "details": []
    }
    
    try:
        # Khởi tạo Sapo client và product service
        sapo_client = get_sapo_client()
        product_service = SapoProductService(sapo_client)
        
        # Lấy customer note
        print(f"[DEBUG] Đang lấy customer {CUSTOMER_ID}...")
        customer_response = sapo_client.core.get_customer_raw(CUSTOMER_ID)
        customer = customer_response.get("customer", {})
        
        if not customer:
            return JsonResponse({
                "status": "error",
                "message": f"Customer {CUSTOMER_ID} không tồn tại"
            }, safe=False)
        
        # Lấy active notes
        notes = customer.get("notes", [])
        active_notes = [n for n in notes if n.get("status") == "active"]
        print(f"[DEBUG] Tìm thấy {len(active_notes)} active notes")
        
        if not active_notes:
            return JsonResponse({
                "status": "error",
                "message": f"Customer {CUSTOMER_ID} không có active notes"
            }, safe=False)
        
        # Xử lý từng note
        for note in active_notes:
            try:
                content_str = note.get("content", "{}")
                content = json.loads(content_str)
                
                # Kiểm tra có product_id không
                product_id = content.get("product_id")
                if not product_id:
                    print(f"[DEBUG] Skip note {note.get('id')}: Không có product_id")
                    continue
                
                product_id = int(float(product_id))  # Convert to int (có thể là float)
                result["processed"] += 1
                
                # Extract thông tin nhãn phụ
                vi_name = content.get("vi_name", "").strip() or None
                en_name = content.get("en_name", "").strip() or None
                description = content.get("descreption", "").strip() or None  # Note: user viết "descreption"
                material = content.get("material", "").strip() or None
                brand = content.get("brand", "").strip() or None
                
                # Tạo NhanPhuInfoDTO
                nhanphu_info = NhanPhuInfoDTO(
                    vi_name=vi_name,
                    en_name=en_name,
                    description=description,
                    material=material,
                    hdsd=None  # Không có trong customer note
                )
                
                # Lấy product hiện tại
                product = product_service.get_product(product_id)
                if not product:
                    result["errors"] += 1
                    result["details"].append({
                        "product_id": product_id,
                        "status": "error",
                        "reason": "Product không tồn tại"
                    })
                    print(f"[ERROR] Product {product_id} không tồn tại")
                    continue
                
                # Lấy metadata hiện tại
                current_metadata = product.gdp_metadata
                
                if not current_metadata:
                    # Nếu chưa có metadata, init empty metadata trước
                    variant_ids = [v.id for v in product.variants]
                    current_metadata = init_empty_metadata(product_id, variant_ids)
                
                # Update nhanphu_info
                current_metadata.nhanphu_info = nhanphu_info
                
                # Lưu vào product description
                success = product_service.update_product_metadata(
                    product_id,
                    current_metadata,
                    preserve_description=True
                )
                
                if success:
                    result["success"] += 1
                    result["details"].append({
                        "product_id": product_id,
                        "status": "success",
                        "nhanphu_info": {
                            "vi_name": vi_name,
                            "en_name": en_name,
                            "description": description,
                            "material": material,
                            "brand": brand  # Lưu brand vào details (không có trong DTO)
                        }
                    })
                    print(f"[SUCCESS] Đã đồng bộ nhãn phụ cho product {product_id}")
                else:
                    result["errors"] += 1
                    result["details"].append({
                        "product_id": product_id,
                        "status": "error",
                        "reason": "Không thể update product metadata"
                    })
                    print(f"[ERROR] Không thể update product {product_id}")
                    
            except json.JSONDecodeError as e:
                print(f"[ERROR] Lỗi parse JSON note {note.get('id')}: {e}")
                result["errors"] += 1
                result["details"].append({
                    "note_id": note.get("id"),
                    "status": "error",
                    "reason": f"Lỗi parse JSON: {e}"
                })
                continue
            except Exception as e:
                print(f"[ERROR] Lỗi xử lý note {note.get('id')}: {e}")
                result["errors"] += 1
                result["details"].append({
                    "note_id": note.get("id"),
                    "status": "error",
                    "reason": str(e)
                })
                continue
        
        print(f"[DEBUG] Hoàn thành: {result['success']} thành công, {result['errors']} lỗi")
        return JsonResponse(result, safe=False)
        
    except Exception as e:
        print(f"[ERROR] Lỗi tổng quát: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, safe=False) 