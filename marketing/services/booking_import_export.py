# marketing/services/booking_import_export.py
"""
Import/Export service for KOC/KOL Database
Supports CSV and XLSX with idempotent upsert
"""
import csv
import io
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl import Workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

from marketing.models import (
    Creator, CreatorChannel, CreatorContact, CreatorTag, CreatorTagMap
)


# ==================== IMPORT CREATORS ====================

def import_creators_csv(file, dry_run=False):
    """
    Import creators from CSV
    Columns: creator_key, name, alias, gender, dob, location, niche, priority_score, status, note_internal
    """
    results = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
    }
    
    try:
        # Read CSV
        if file.name.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
        else:
            # Assume XLSX
            if not XLSX_AVAILABLE:
                raise Exception('XLSX không được hỗ trợ. Cần cài openpyxl.')
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            reader = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                reader.append(dict(zip(headers, row)))
        
        if not dry_run:
            with transaction.atomic():
                for idx, row in enumerate(reader, start=2):
                    try:
                        creator_key = row.get('creator_key', '').strip()
                        name = row.get('name', '').strip()
                        
                        if not name:
                            results['skipped'] += 1
                            results['errors'].append({
                                'row': idx,
                                'error': 'Thiếu tên creator'
                            })
                            continue
                        
                        # Try to find existing by creator_key or name+location
                        creator = None
                        if creator_key:
                            # Look for existing with same key in note_internal or alias
                            creator = Creator.objects.filter(
                                is_active=True,
                                deleted_at__isnull=True
                            ).filter(
                                Q(alias=creator_key) | Q(note_internal__icontains=f'KEY:{creator_key}')
                            ).first()
                        
                        if not creator:
                            # Try name + location match
                            location = row.get('location', '').strip()
                            if location:
                                creator = Creator.objects.filter(
                                    name__iexact=name,
                                    location__iexact=location,
                                    is_active=True,
                                    deleted_at__isnull=True
                                ).first()
                        
                        # Parse fields
                        gender = row.get('gender', '').strip() or None
                        if gender and gender not in ['male', 'female', 'other']:
                            gender = None
                        
                        dob_str = row.get('dob', '').strip()
                        dob = None
                        if dob_str:
                            try:
                                dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
                            except:
                                pass
                        
                        priority_score = 5
                        try:
                            ps = int(row.get('priority_score', 5))
                            if 1 <= ps <= 10:
                                priority_score = ps
                        except:
                            pass
                        
                        status = row.get('status', 'active').strip()
                        if status not in ['active', 'watchlist', 'blacklist']:
                            status = 'active'
                        
                        if creator:
                            # Update existing
                            creator.name = name
                            creator.alias = row.get('alias', '').strip() or None
                            creator.gender = gender
                            creator.dob = dob
                            creator.location = row.get('location', '').strip() or None
                            creator.niche = row.get('niche', '').strip() or None
                            creator.priority_score = priority_score
                            creator.status = status
                            note_internal = row.get('note_internal', '').strip() or None
                            if creator_key and not creator.note_internal:
                                creator.note_internal = f'KEY:{creator_key}\n{note_internal or ""}'
                            elif note_internal:
                                creator.note_internal = note_internal
                            creator.save()
                            results['updated'] += 1
                        else:
                            # Create new
                            note_internal = row.get('note_internal', '').strip() or None
                            if creator_key:
                                note_internal = f'KEY:{creator_key}\nAUTO-CREATED BY IMPORT\n{note_internal or ""}'
                            else:
                                note_internal = 'AUTO-CREATED BY IMPORT\n' + (note_internal or '')
                            
                            creator = Creator.objects.create(
                                name=name,
                                alias=row.get('alias', '').strip() or None,
                                gender=gender,
                                dob=dob,
                                location=row.get('location', '').strip() or None,
                                niche=row.get('niche', '').strip() or None,
                                note_internal=note_internal,
                                priority_score=priority_score,
                                status=status,
                            )
                            results['created'] += 1
                    except Exception as e:
                        results['skipped'] += 1
                        results['errors'].append({
                            'row': idx,
                            'error': str(e)
                        })
        else:
            # Dry run - just count
            for idx, row in enumerate(reader, start=2):
                try:
                    name = row.get('name', '').strip()
                    if not name:
                        results['skipped'] += 1
                        continue
                    
                    creator_key = row.get('creator_key', '').strip()
                    location = row.get('location', '').strip()
                    
                    # Check if exists
                    exists = False
                    if creator_key:
                        exists = Creator.objects.filter(
                            is_active=True,
                            deleted_at__isnull=True
                        ).filter(
                            Q(alias=creator_key) | Q(note_internal__icontains=f'KEY:{creator_key}')
                        ).exists()
                    
                    if not exists and location:
                        exists = Creator.objects.filter(
                            name__iexact=name,
                            location__iexact=location,
                            is_active=True,
                            deleted_at__isnull=True
                        ).exists()
                    
                    if exists:
                        results['updated'] += 1
                    else:
                        results['created'] += 1
                except Exception as e:
                    results['skipped'] += 1
                    results['errors'].append({
                        'row': idx,
                        'error': str(e)
                    })
        
        results['success'] = True
        results['message'] = f'Import hoàn tất: {results["created"]} tạo mới, {results["updated"]} cập nhật, {results["skipped"]} bỏ qua, {len(results["errors"])} lỗi'
        
    except Exception as e:
        results['success'] = False
        results['message'] = f'Lỗi import: {str(e)}'
        results['errors'].append({
            'row': 0,
            'error': str(e)
        })
    
    return results


# ==================== IMPORT CHANNELS ====================

def import_channels_csv(file, dry_run=False):
    """
    Import channels from CSV
    Columns: creator_key, platform, handle, profile_url, external_id, follower_count, avg_view_10, avg_engagement_rate
    """
    results = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
    }
    
    try:
        # Read CSV
        if file.name.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
        else:
            if not XLSX_AVAILABLE:
                raise Exception('XLSX không được hỗ trợ.')
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            reader = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                reader.append(dict(zip(headers, row)))
        
        if not dry_run:
            with transaction.atomic():
                for idx, row in enumerate(reader, start=2):
                    try:
                        creator_key = row.get('creator_key', '').strip()
                        platform = row.get('platform', '').strip()
                        handle = row.get('handle', '').strip()
                        
                        if not creator_key or not platform or not handle:
                            results['skipped'] += 1
                            results['errors'].append({
                                'row': idx,
                                'error': 'Thiếu creator_key, platform hoặc handle'
                            })
                            continue
                        
                        # Find creator
                        creator = Creator.objects.filter(
                            is_active=True,
                            deleted_at__isnull=True
                        ).filter(
                            Q(alias=creator_key) | Q(note_internal__icontains=f'KEY:{creator_key}')
                        ).first()
                        
                        if not creator:
                            results['skipped'] += 1
                            results['errors'].append({
                                'row': idx,
                                'error': f'Không tìm thấy creator với key: {creator_key}'
                            })
                            continue
                        
                        # Check if channel exists
                        channel = CreatorChannel.objects.filter(
                            platform=platform,
                            handle=handle,
                            is_active=True,
                            deleted_at__isnull=True
                        ).exclude(creator=creator).first()
                        
                        if channel:
                            results['skipped'] += 1
                            results['errors'].append({
                                'row': idx,
                                'error': f'Channel {platform}/{handle} đã tồn tại cho creator khác'
                            })
                            continue
                        
                        # Get or create
                        channel, created = CreatorChannel.objects.get_or_create(
                            creator=creator,
                            platform=platform,
                            handle=handle,
                            defaults={
                                'profile_url': row.get('profile_url', '').strip() or None,
                                'external_id': row.get('external_id', '').strip() or None,
                                'follower_count': int(row.get('follower_count', 0) or 0),
                                'avg_view_10': int(row.get('avg_view_10', 0) or 0),
                                'avg_engagement_rate': Decimal(row.get('avg_engagement_rate', 0) or 0),
                            }
                        )
                        
                        if not created:
                            # Update existing
                            channel.profile_url = row.get('profile_url', '').strip() or None
                            channel.external_id = row.get('external_id', '').strip() or None
                            channel.follower_count = int(row.get('follower_count', 0) or 0)
                            channel.avg_view_10 = int(row.get('avg_view_10', 0) or 0)
                            channel.avg_engagement_rate = Decimal(row.get('avg_engagement_rate', 0) or 0)
                            channel.save()
                            results['updated'] += 1
                        else:
                            results['created'] += 1
                    except Exception as e:
                        results['skipped'] += 1
                        results['errors'].append({
                            'row': idx,
                            'error': str(e)
                        })
        else:
            # Dry run
            for idx, row in enumerate(reader, start=2):
                try:
                    creator_key = row.get('creator_key', '').strip()
                    platform = row.get('platform', '').strip()
                    handle = row.get('handle', '').strip()
                    
                    if not creator_key or not platform or not handle:
                        results['skipped'] += 1
                        continue
                    
                    creator = Creator.objects.filter(
                        is_active=True,
                        deleted_at__isnull=True
                    ).filter(
                        Q(alias=creator_key) | Q(note_internal__icontains=f'KEY:{creator_key}')
                    ).first()
                    
                    if not creator:
                        results['skipped'] += 1
                        continue
                    
                    exists = CreatorChannel.objects.filter(
                        platform=platform,
                        handle=handle,
                        is_active=True,
                        deleted_at__isnull=True
                    ).exists()
                    
                    if exists:
                        results['updated'] += 1
                    else:
                        results['created'] += 1
                except Exception as e:
                    results['skipped'] += 1
                    results['errors'].append({
                        'row': idx,
                        'error': str(e)
                    })
        
        results['success'] = True
        results['message'] = f'Import hoàn tất: {results["created"]} tạo mới, {results["updated"]} cập nhật, {results["skipped"]} bỏ qua, {len(results["errors"])} lỗi'
        
    except Exception as e:
        results['success'] = False
        results['message'] = f'Lỗi import: {str(e)}'
        results['errors'].append({
            'row': 0,
            'error': str(e)
        })
    
    return results


# ==================== IMPORT CONTACTS ====================

def import_contacts_csv(file, dry_run=False):
    """
    Import contacts from CSV
    Columns: creator_key, contact_type, name, phone, zalo, email, wechat, note, is_primary
    """
    results = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
    }
    
    try:
        # Read CSV
        if file.name.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
        else:
            if not XLSX_AVAILABLE:
                raise Exception('XLSX không được hỗ trợ.')
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            reader = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                reader.append(dict(zip(headers, row)))
        
        if not dry_run:
            with transaction.atomic():
                for idx, row in enumerate(reader, start=2):
                    try:
                        creator_key = row.get('creator_key', '').strip()
                        name = row.get('name', '').strip()
                        phone = row.get('phone', '').strip() or None
                        email = row.get('email', '').strip() or None
                        
                        if not creator_key or not name:
                            results['skipped'] += 1
                            results['errors'].append({
                                'row': idx,
                                'error': 'Thiếu creator_key hoặc name'
                            })
                            continue
                        
                        # Find creator
                        creator = Creator.objects.filter(
                            is_active=True,
                            deleted_at__isnull=True
                        ).filter(
                            Q(alias=creator_key) | Q(note_internal__icontains=f'KEY:{creator_key}')
                        ).first()
                        
                        if not creator:
                            results['skipped'] += 1
                            results['errors'].append({
                                'row': idx,
                                'error': f'Không tìm thấy creator với key: {creator_key}'
                            })
                            continue
                        
                        # Check if contact exists (by phone or email)
                        contact = None
                        if phone:
                            contact = CreatorContact.objects.filter(
                                creator=creator,
                                phone=phone,
                                is_active=True,
                                deleted_at__isnull=True
                            ).first()
                        elif email:
                            contact = CreatorContact.objects.filter(
                                creator=creator,
                                email=email,
                                is_active=True,
                                deleted_at__isnull=True
                            ).first()
                        
                        is_primary = str(row.get('is_primary', 'false')).lower() == 'true'
                        
                        # If setting as primary, unset others
                        if is_primary:
                            CreatorContact.objects.filter(
                                creator=creator,
                                is_active=True,
                                deleted_at__isnull=True
                            ).update(is_primary=False)
                        
                        if contact:
                            # Update existing
                            contact.contact_type = row.get('contact_type', 'owner')
                            contact.name = name
                            contact.phone = phone
                            contact.zalo = row.get('zalo', '').strip() or None
                            contact.email = email
                            contact.wechat = row.get('wechat', '').strip() or None
                            contact.note = row.get('note', '').strip() or None
                            contact.is_primary = is_primary
                            contact.save()
                            results['updated'] += 1
                        else:
                            # Create new
                            CreatorContact.objects.create(
                                creator=creator,
                                contact_type=row.get('contact_type', 'owner'),
                                name=name,
                                phone=phone,
                                zalo=row.get('zalo', '').strip() or None,
                                email=email,
                                wechat=row.get('wechat', '').strip() or None,
                                note=row.get('note', '').strip() or None,
                                is_primary=is_primary,
                            )
                            results['created'] += 1
                    except Exception as e:
                        results['skipped'] += 1
                        results['errors'].append({
                            'row': idx,
                            'error': str(e)
                        })
        else:
            # Dry run
            for idx, row in enumerate(reader, start=2):
                try:
                    creator_key = row.get('creator_key', '').strip()
                    name = row.get('name', '').strip()
                    
                    if not creator_key or not name:
                        results['skipped'] += 1
                        continue
                    
                    creator = Creator.objects.filter(
                        is_active=True,
                        deleted_at__isnull=True
                    ).filter(
                        Q(alias=creator_key) | Q(note_internal__icontains=f'KEY:{creator_key}')
                    ).first()
                    
                    if not creator:
                        results['skipped'] += 1
                        continue
                    
                    phone = row.get('phone', '').strip()
                    email = row.get('email', '').strip()
                    
                    exists = False
                    if phone:
                        exists = CreatorContact.objects.filter(
                            creator=creator,
                            phone=phone,
                            is_active=True,
                            deleted_at__isnull=True
                        ).exists()
                    elif email:
                        exists = CreatorContact.objects.filter(
                            creator=creator,
                            email=email,
                            is_active=True,
                            deleted_at__isnull=True
                        ).exists()
                    
                    if exists:
                        results['updated'] += 1
                    else:
                        results['created'] += 1
                except Exception as e:
                    results['skipped'] += 1
                    results['errors'].append({
                        'row': idx,
                        'error': str(e)
                    })
        
        results['success'] = True
        results['message'] = f'Import hoàn tất: {results["created"]} tạo mới, {results["updated"]} cập nhật, {results["skipped"]} bỏ qua, {len(results["errors"])} lỗi'
        
    except Exception as e:
        results['success'] = False
        results['message'] = f'Lỗi import: {str(e)}'
        results['errors'].append({
            'row': 0,
            'error': str(e)
        })
    
    return results


# ==================== EXPORT CREATORS ====================

def export_creators_csv(queryset):
    """Export creators to CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="creators_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Name', 'Alias', 'Gender', 'DOB', 'Location', 'Niche',
        'Priority Score', 'Status', 'Note Internal',
        'Primary Contact Name', 'Primary Contact Phone', 'Primary Contact Email',
        'TikTok Handle', 'TikTok Followers', 'TikTok Avg View',
        'Tags'
    ])
    
    for creator in queryset.select_related().prefetch_related('channels', 'contacts', 'tag_maps__tag'):
        primary_contact = creator.contacts.filter(is_primary=True).first()
        tiktok_channel = creator.channels.filter(platform='tiktok').first()
        tags = ', '.join([tm.tag.name for tm in creator.tag_maps.select_related('tag').filter(tag__is_active=True, tag__deleted_at__isnull=True)])
        
        writer.writerow([
            creator.id,
            creator.name,
            creator.alias or '',
            creator.gender or '',
            creator.dob.strftime('%Y-%m-%d') if creator.dob else '',
            creator.location or '',
            creator.niche or '',
            creator.priority_score,
            creator.status,
            creator.note_internal or '',
            primary_contact.name if primary_contact else '',
            primary_contact.phone if primary_contact else '',
            primary_contact.email if primary_contact else '',
            tiktok_channel.handle if tiktok_channel else '',
            tiktok_channel.follower_count if tiktok_channel else 0,
            tiktok_channel.avg_view_10 if tiktok_channel else 0,
            tags,
        ])
    
    return response


def export_creators_xlsx(queryset):
    """Export creators to XLSX"""
    if not XLSX_AVAILABLE:
        return HttpResponse('XLSX không được hỗ trợ. Cần cài openpyxl.', status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Creators"
    
    # Headers
    headers = [
        'ID', 'Name', 'Alias', 'Gender', 'DOB', 'Location', 'Niche',
        'Priority Score', 'Status', 'Note Internal',
        'Primary Contact Name', 'Primary Contact Phone', 'Primary Contact Email',
        'TikTok Handle', 'TikTok Followers', 'TikTok Avg View',
        'Tags'
    ]
    ws.append(headers)
    
    # Data
    for creator in queryset.select_related().prefetch_related('channels', 'contacts', 'tag_maps__tag'):
        primary_contact = creator.contacts.filter(is_primary=True).first()
        tiktok_channel = creator.channels.filter(platform='tiktok').first()
        tags = ', '.join([tm.tag.name for tm in creator.tag_maps.select_related('tag').filter(tag__is_active=True, tag__deleted_at__isnull=True)])
        
        ws.append([
            creator.id,
            creator.name,
            creator.alias or '',
            creator.gender or '',
            creator.dob.strftime('%Y-%m-%d') if creator.dob else '',
            creator.location or '',
            creator.niche or '',
            creator.priority_score,
            creator.status,
            creator.note_internal or '',
            primary_contact.name if primary_contact else '',
            primary_contact.phone if primary_contact else '',
            primary_contact.email if primary_contact else '',
            tiktok_channel.handle if tiktok_channel else '',
            tiktok_channel.follower_count if tiktok_channel else 0,
            tiktok_channel.avg_view_10 if tiktok_channel else 0,
            tags,
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="creators_export.xlsx"'
    wb.save(response)
    return response

