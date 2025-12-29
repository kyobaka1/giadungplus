# products/services/tkhq_parser.py
"""
Service để parse file Tờ khai hải quan (TKHQ) Excel.
Parse theo khối (block) - mỗi mặt hàng là một block.
"""

import logging
import re
import os
from typing import List, Dict, Any, Optional
from decimal import Decimal
from openpyxl import load_workbook
import pandas as pd

logger = logging.getLogger(__name__)

# DEBUG flag - có thể được set từ bên ngoài
TKHQ_DEBUG = False

def debug_print(*args, **kwargs):
    """Debug print function - chỉ in khi TKHQ_DEBUG = True"""
    if TKHQ_DEBUG:
        print("[TKHQ DEBUG]", *args, **kwargs)

def set_tkhq_debug(enabled: bool):
    """Set debug mode cho TKHQ parser"""
    global TKHQ_DEBUG
    TKHQ_DEBUG = enabled


class TKHQParser:
    """
    Parser cho file Tờ khai hải quan (TKHQ) Excel.
    
    Mỗi mặt hàng là một Block:
    - Bắt đầu bằng dòng có STT ở cột đầu tiên
    - Tên hàng: Cột 1 hoặc 2, cùng dòng hoặc dòng dưới STT
    - Thuế NK: Dòng có "Thuế NK" hoặc mã loại hình thuế
    - Thuế VAT: Dòng có "Thuế GTGT", "VAT" hoặc "10%"
    """
    
    def __init__(self, file_path: str, debug: bool = False):
        """
        Args:
            file_path: Đường dẫn đến file Excel TKHQ
            debug: Bật debug mode (sẽ dùng print thay vì logger)
        """
        self.file_path = file_path
        self.wb = None
        self.ws = None
        self.df = None
        self.all_rows = []
        self.imp_positions = []
        self.debug = debug
        if debug:
            set_tkhq_debug(True)
        # Không load file ngay trong __init__, để parse() tự load
    
    def _unmerge_and_fill_cells(self, sheet):
        """
        Unmerge tất cả merged cells và fill giá trị vào các ô đã unmerge.
        Đây là cách chính xác nhất để xử lý merged cells.
        """
        # Lấy danh sách các vùng bị merge
        merged_cells = list(sheet.merged_cells.ranges)
        
        debug_print(f"Found {len(merged_cells)} merged cell ranges")
        
        for merged_cell in merged_cells:
            # Lấy giá trị của ô đầu tiên trong vùng merge
            min_col = merged_cell.min_col
            min_row = merged_cell.min_row
            max_col = merged_cell.max_col
            max_row = merged_cell.max_row
            
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            
            # Unmerge vùng này
            sheet.unmerge_cells(str(merged_cell))
            
            # Điền giá trị của ô đầu tiên vào tất cả các ô vừa unmerge
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value
        
        return sheet
    
    def load_file(self):
        """Load file Excel, unmerge cells, và tách theo <IMP> - mỗi <IMP> là một trang mới"""
        try:
            debug_print("=== Starting load_file() ===")
            debug_print(f"File path: {self.file_path}")
            debug_print(f"File exists: {os.path.exists(self.file_path) if hasattr(os, 'path') else 'N/A'}")
            
            # Load với openpyxl (dùng data_only để lấy giá trị đã tính toán)
            debug_print("Loading workbook...")
            self.wb = load_workbook(self.file_path, data_only=True)
            self.ws = self.wb.active
            debug_print(f"Workbook loaded. Active sheet: {self.ws.title}")
            
            # Unmerge và fill dữ liệu - đây là bước quan trọng để xử lý merged cells chính xác
            debug_print("Unmerging cells...")
            self.ws = self._unmerge_and_fill_cells(self.ws)
            debug_print("Cells unmerged")
            
            # Đọc toàn bộ nội dung thành từng row
            debug_print("Reading all rows from worksheet...")
            all_rows = []
            for row in self.ws.iter_rows(values_only=True):
                row_values = []
                for cell in row:
                    if cell is None:
                        row_values.append('')
                    else:
                        row_values.append(str(cell).strip())
                all_rows.append(row_values)
            debug_print(f"Read {len(all_rows)} rows from worksheet")
            
            # Tìm tất cả các vị trí <IMP> để phân tách trang
            # Theo yêu cầu: <IMP> ở row C133, tức là cột C (index 2)
            # Tìm cả <IMP>, <IMP >, <IMP>, IMP (case-insensitive)
            imp_positions = []
            imp_variants = ['<IMP>', '<IMP >', '<IMP', 'IMP>', 'IMP']
            
            for row_idx, row in enumerate(all_rows):
                # Ưu tiên kiểm tra cột C (index 2) trước
                if len(row) > 2:
                    cell_c = str(row[2]).strip() if row[2] else ''
                    if cell_c:
                        cell_c_upper = cell_c.upper()
                        for variant in imp_variants:
                            if variant.upper() in cell_c_upper:
                                imp_positions.append(row_idx)
                                debug_print(f"Found <IMP> variant '{variant}' at row {row_idx}, col C (index 2), cell value: '{cell_c}'")
                                break
                
                # Nếu không tìm thấy ở cột C, kiểm tra các cột khác
                if row_idx not in imp_positions:
                    for col_idx, cell_value in enumerate(row):
                        if cell_value:
                            cell_str = str(cell_value).strip().upper()
                            # Kiểm tra các biến thể của IMP
                            for variant in imp_variants:
                                if variant.upper() in cell_str:
                                    imp_positions.append(row_idx)
                                    debug_print(f"Found <IMP> variant '{variant}' at row {row_idx}, col {col_idx}, cell value: '{cell_value}'")
                                    break
                            if row_idx in imp_positions:
                                break
                
                # Nếu vẫn không tìm thấy, thử tìm trong toàn bộ row text
                if row_idx not in imp_positions:
                    row_text = ' '.join([str(c) for c in row]).upper()
                    for variant in imp_variants:
                        if variant.upper() in row_text:
                            imp_positions.append(row_idx)
                            debug_print(f"Found <IMP> variant '{variant}' in row text at row {row_idx}")
                            break
            
            # Loại bỏ duplicate và sort
            imp_positions = sorted(list(set(imp_positions)))
            
            debug_print(f"Loaded file: {self.file_path}")
            debug_print(f"Total rows: {len(all_rows)}")
            debug_print(f"Found {len(imp_positions)} <IMP> markers at rows: {imp_positions}")
            
            # Debug: In một vài rows đầu và rows xung quanh vị trí có thể có <IMP> để kiểm tra
            if len(imp_positions) == 0:
                debug_print("WARNING: No <IMP> markers found!")
                debug_print("First 20 rows sample (checking column C - index 2):")
                for i in range(min(20, len(all_rows))):
                    row = all_rows[i]
                    col_c = row[2] if len(row) > 2 else ''
                    row_sample = row[:10]  # Lấy 10 cột đầu
                    debug_print(f"  Row {i}: col_C='{col_c}', first_10_cols={row_sample}")
                
                # Thử tìm ở các vị trí có thể (row 130-140 nếu file có nhiều rows)
                if len(all_rows) > 140:
                    debug_print("Checking rows 130-140 (around expected IMP position):")
                    for i in range(130, min(141, len(all_rows))):
                        row = all_rows[i]
                        col_c = row[2] if len(row) > 2 else ''
                        if col_c:
                            debug_print(f"  Row {i}: col_C='{col_c}'")
            
            # Lưu toàn bộ rows và imp_positions để parse từng trang
            self.all_rows = all_rows
            self.imp_positions = imp_positions
            
            logger.info(f"[TKHQ Parser] Loaded file: {self.file_path}")
            logger.info(f"[TKHQ Parser] Total rows: {len(all_rows)}, IMP markers: {len(imp_positions)}")
            
            return True
        except Exception as e:
            debug_print(f"ERROR in load_file(): {e}")
            debug_print(f"Exception type: {type(e).__name__}")
            import traceback
            debug_print(f"Traceback: {traceback.format_exc()}")
            logger.error(f"[TKHQ Parser] Error loading file: {e}", exc_info=True)
            return False
    
    def parse(self) -> List[Dict[str, Any]]:
        """
        Parse file TKHQ và trả về danh sách items.
        Tách theo <IMP> - mỗi <IMP> là một trang mới, reset row index.
        
        Returns:
            List of dicts: [
                {
                    'stt': 1,
                    'ma_so_hang_hoa': '70134900',
                    'mo_ta_hang_hoa': 'Bình thuỷ tinh...',
                    'thue_nk_rate': 5.0,
                    'thue_nk_vnd': 4249771.456,
                    'thue_gtgt_rate': 8.0,
                    'thue_gtgt_vnd': 7139616.0461
                },
                ...
            ]
        """
        debug_print("=== Starting parse() ===")
        debug_print(f"File path: {self.file_path}")
        debug_print(f"Has all_rows attr: {hasattr(self, 'all_rows')}")
        
        # Force reload để đảm bảo có dữ liệu mới nhất
        debug_print("Calling load_file() to reload data...")
        if not self.load_file():
            debug_print("ERROR: load_file() returned False")
            return []
        debug_print("load_file() completed successfully")
        
        debug_print(f"Total imp_positions: {len(self.imp_positions) if hasattr(self, 'imp_positions') else 0}")
        debug_print(f"imp_positions: {self.imp_positions if hasattr(self, 'imp_positions') else 'N/A'}")
        debug_print(f"Total all_rows: {len(self.all_rows) if hasattr(self, 'all_rows') else 0}")
        
        # Nếu vẫn không tìm thấy IMP, in thêm thông tin debug
        if len(self.imp_positions) == 0 and len(self.all_rows) > 0:
            debug_print("WARNING: Still no IMP markers found after reload!")
            debug_print("Sample of first 10 rows, column C (index 2):")
            for i in range(min(10, len(self.all_rows))):
                row = self.all_rows[i]
                col_c = row[2] if len(row) > 2 else ''
                debug_print(f"  Row {i}: col_C='{col_c}'")
            
            # Kiểm tra rows 130-140 (vị trí có thể có IMP)
            if len(self.all_rows) > 140:
                debug_print("Checking rows 130-140 (around expected IMP position at row 133):")
                for i in range(130, min(141, len(self.all_rows))):
                    row = self.all_rows[i]
                    col_c = row[2] if len(row) > 2 else ''
                    if col_c:  # Chỉ in nếu có giá trị
                        debug_print(f"  Row {i}: col_C='{col_c}'")
        
        all_items = []
        
        # Parse từng trang (giữa các <IMP>)
        # Trang 3 là bắt đầu có dữ liệu sản phẩm (page_idx 2 vì index bắt đầu từ 0)
        start_page_idx = 2 if len(self.imp_positions) > 2 else 0
        debug_print(f"Starting from page_idx: {start_page_idx}, total pages: {len(self.imp_positions)}")
        
        for page_idx in range(start_page_idx, len(self.imp_positions)):
            start_row = self.imp_positions[page_idx]
            end_row = self.imp_positions[page_idx + 1] if page_idx + 1 < len(self.imp_positions) else len(self.all_rows)
            
            # Lấy dữ liệu trang này
            page_rows = self.all_rows[start_row:end_row]
            
            # Convert sang DataFrame cho trang này
            df_page = pd.DataFrame(page_rows)
            
            debug_print(f"\n=== Parsing Page {page_idx + 1} (rows {start_row} to {end_row}) ===")
            debug_print(f"Page DataFrame shape: {df_page.shape}")
            
            # Parse trang này (row index reset về 0 cho mỗi trang)
            page_items = self._parse_page(df_page, page_idx + 1)
            all_items.extend(page_items)
            
            debug_print(f"Page {page_idx + 1}: Found {len(page_items)} items")
        
        debug_print(f"\n=== Parse completed: Total {len(all_items)} items from {len(self.imp_positions)} pages ===")
        logger.info(f"[TKHQ Parser] Parsed {len(all_items)} items from {len(self.imp_positions)} pages")
        
        return all_items
    
    def _parse_page(self, df_page: pd.DataFrame, page_num: int) -> List[Dict[str, Any]]:
        """
        Parse một trang (giữa 2 <IMP>).
        Row index reset về 0 cho mỗi trang.
        
        Cấu trúc mới dựa trên offset từ <IMP>:
        - <IMP> ở row 0 (trong df_page)
        - Mã số hàng hoá: G143 -> offset 10 rows, cột G (index 6)
        - Mô tả hàng hoá: G144:AG146 -> offset 11 rows, merge từ cột G (6) đến AG (32)
        - Số lượng: V147:AA147 -> offset 14 rows, cột V (21) đến AA (26)
        - Đơn giá tính thuế NK: V152:AA152 -> offset 19 rows, cột V (21) đến AA (26)
        - Thuế suất NK: I153:N153 -> offset 20 rows, cột I (8) đến N (13)
        - Trị giá tính thuế GTGT: I162:O162 -> offset 29 rows, cột I (8) đến O (14)
        - Thuế suất GTGT: I163:O163 -> offset 30 rows, cột I (8) đến O (14)
        - Tổng tiền thuế GTGT: I164:O164 -> offset 31 rows, cột I (8) đến O (14)
        """
        items = []
        
        # Tìm vị trí <IMP> trong page (thường ở row 0 vì đã được tách theo <IMP>)
        # Tìm trong cột C (index 2) vì <IMP> ở row C133
        imp_row_idx = None
        for row_idx in range(min(5, len(df_page))):
            row = df_page.iloc[row_idx]
            # Kiểm tra cột C (index 2) trước, sau đó kiểm tra toàn bộ row
            cell_c = self._get_cell_value(row, 2)  # Cột C = index 2
            if cell_c and '<IMP>' in str(cell_c):
                imp_row_idx = row_idx
                break
            # Nếu không tìm thấy ở cột C, tìm trong toàn bộ row
            row_text = ' '.join([str(self._get_cell_value(row, c) or '') for c in range(min(10, len(row)))])
            if '<IMP>' in row_text:
                imp_row_idx = row_idx
                break
        
        if imp_row_idx is None:
            debug_print(f"  Page {page_num}: No <IMP> marker found, skipping")
            return items
        
        debug_print(f"  Page {page_num}: Found <IMP> at row {imp_row_idx} (in page)")
        
        # Kiểm tra xem có đủ rows không (cần ít nhất offset 31 + imp_row_idx)
        min_required_rows = imp_row_idx + 32
        if len(df_page) < min_required_rows:
            debug_print(f"  Page {page_num}: Not enough rows ({len(df_page)} < {min_required_rows}), skipping")
            return items
        
        # Tạo item mới
        current_item = {
            'stt': page_num - 2,  # Trang 3 là bắt đầu có dữ liệu, STT = page_num - 2
            'ma_so_hang_hoa': '',
            'mo_ta_hang_hoa': '',
            'so_luong': Decimal('0'),
            'don_gia_tinh_thue_nk': Decimal('0'),
            'thue_nk_rate': Decimal('0'),
            'thue_nk_vnd': Decimal('0'),
            'tri_gia_tinh_thue_gtgt': Decimal('0'),
            'thue_gtgt_rate': Decimal('0'),
            'thue_gtgt_vnd': Decimal('0'),
            'page': page_num
        }
        
        # Lấy mã số hàng hoá: G143 -> offset 10 rows từ IMP, cột G (index 6)
        ma_so_row_idx = imp_row_idx + 10
        if ma_so_row_idx < len(df_page):
            row_ma_so = df_page.iloc[ma_so_row_idx]
            ma_so = self._get_merged_cell_value(df_page, ma_so_row_idx, 6)  # Cột G = index 6
            if ma_so:
                current_item['ma_so_hang_hoa'] = str(ma_so).strip()
                debug_print(f"  Row {ma_so_row_idx} (IMP+10): Found ma_so_hang_hoa: {current_item['ma_so_hang_hoa']}")
        
        # Lấy mô tả hàng hoá: G144:AG146 -> offset 11 rows, merge từ G (6) đến AG (32)
        # Vì đã unmerge và fill, tất cả các ô trong vùng merge đều có cùng giá trị
        # Chỉ cần đọc từ ô đầu tiên G144 (row 11, col 6) để tránh trùng lặp
        mo_ta_row_idx = imp_row_idx + 11
        if mo_ta_row_idx < len(df_page):
            # Đọc từ ô G144 (row 11, col 6) - đây là ô đầu tiên của merged cell
            mo_ta = self._get_merged_cell_value(df_page, mo_ta_row_idx, 6)
            if mo_ta:
                current_item['mo_ta_hang_hoa'] = str(mo_ta).strip()
                debug_print(f"  Row {mo_ta_row_idx} (IMP+11): Found mo_ta_hang_hoa: {current_item['mo_ta_hang_hoa'][:100]}")
        
        # Lấy số lượng: V147:AA147 -> offset 14 rows, cột V (21) đến AA (26)
        so_luong_row_idx = imp_row_idx + 14
        if so_luong_row_idx < len(df_page):
            so_luong = self._get_merged_cell_value(df_page, so_luong_row_idx, 21, end_col=26)  # V=21, AA=26
            if so_luong:
                try:
                    cleaned = re.sub(r'[^\d.,]', '', str(so_luong))
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                    current_item['so_luong'] = Decimal(cleaned)
                    debug_print(f"  Row {so_luong_row_idx} (IMP+14): Found so_luong: {current_item['so_luong']}")
                except:
                    pass
        
        # Lấy đơn giá tính thuế NK: V894:AA894 -> offset 19 rows từ IMP, cột V (21) đến AA (26)
        # (Trong file gốc: row 894, nhưng trong page: offset 19 từ IMP)
        don_gia_row_idx = imp_row_idx + 19
        if don_gia_row_idx < len(df_page):
            don_gia = self._get_merged_cell_value(df_page, don_gia_row_idx, 21, end_col=26)
            if don_gia:
                try:
                    cleaned = re.sub(r'[^\d.,]', '', str(don_gia))
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                    current_item['don_gia_tinh_thue_nk'] = Decimal(cleaned)
                    debug_print(f"  Row {don_gia_row_idx} (IMP+19): Found don_gia_tinh_thue_nk: {current_item['don_gia_tinh_thue_nk']}")
                except:
                    pass
        
        # Lấy thuế suất NK: I153:N153 -> offset 20 rows, cột I (8) đến N (13)
        thue_nk_rate_row_idx = imp_row_idx + 20
        if thue_nk_rate_row_idx < len(df_page):
            thue_nk_rate = self._get_merged_cell_value(df_page, thue_nk_rate_row_idx, 8, end_col=13)
            if thue_nk_rate:
                val_str = str(thue_nk_rate).strip()
                rate_match = re.search(r'(\d+(?:\.\d+)?)\s*%', val_str)
                if rate_match:
                    try:
                        current_item['thue_nk_rate'] = Decimal(rate_match.group(1))
                        debug_print(f"  Row {thue_nk_rate_row_idx} (IMP+20): Found thue_nk_rate: {current_item['thue_nk_rate']}%")
                    except:
                        pass
        
        # Lấy số tiền thuế NK từ file: I896:O896 -> offset 21 rows, cột I (8) đến O (14)
        thue_nk_vnd_total_row_idx = imp_row_idx + 21
        if thue_nk_vnd_total_row_idx < len(df_page):
            thue_nk_vnd_total = self._get_merged_cell_value(df_page, thue_nk_vnd_total_row_idx, 8, end_col=14)
            if thue_nk_vnd_total:
                try:
                    cleaned = re.sub(r'[^\d.,]', '', str(thue_nk_vnd_total))
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                    current_item['thue_nk_vnd_total'] = Decimal(cleaned)
                    debug_print(f"  Row {thue_nk_vnd_total_row_idx} (IMP+21): Found thue_nk_vnd_total: {current_item['thue_nk_vnd_total']}")
                except:
                    pass
        
        # Tính các giá trị thuế NK từ số liệu đã đọc
        # Đơn giá tính thuế NK từng chiếc = Đơn giá tính thuế NK (total) / Số lượng
        if current_item['so_luong'] > 0 and current_item['don_gia_tinh_thue_nk'] > 0:
            current_item['don_gia_tinh_thue_nk_per_unit'] = current_item['don_gia_tinh_thue_nk'] / current_item['so_luong']
        else:
            current_item['don_gia_tinh_thue_nk_per_unit'] = Decimal('0')
        
        # Thuế NK VND từng chiếc = Thuế NK VND TOTAL / Số lượng
        if current_item['so_luong'] > 0 and current_item.get('thue_nk_vnd_total', 0) > 0:
            current_item['thue_nk_vnd'] = current_item['thue_nk_vnd_total'] / current_item['so_luong']
        else:
            current_item['thue_nk_vnd'] = Decimal('0')
        
        # Nếu không đọc được từ file, tính từ công thức (fallback)
        if not current_item.get('thue_nk_vnd_total') or current_item['thue_nk_vnd_total'] == 0:
            if current_item['don_gia_tinh_thue_nk'] > 0 and current_item['thue_nk_rate'] > 0:
                current_item['thue_nk_vnd_total'] = (current_item['don_gia_tinh_thue_nk'] * current_item['thue_nk_rate']) / Decimal('100')
                debug_print(f"  Calculated thue_nk_vnd_total (fallback): {current_item['thue_nk_vnd_total']}")
            else:
                current_item['thue_nk_vnd_total'] = Decimal('0')
        
        if current_item['thue_nk_vnd'] == 0 and current_item['don_gia_tinh_thue_nk_per_unit'] > 0 and current_item['thue_nk_rate'] > 0:
            current_item['thue_nk_vnd'] = (current_item['don_gia_tinh_thue_nk_per_unit'] * current_item['thue_nk_rate']) / Decimal('100')
            debug_print(f"  Calculated thue_nk_vnd (fallback): {current_item['thue_nk_vnd']}")
        
        # Lấy trị giá tính thuế GTGT: I162:O162 -> offset 29 rows, cột I (8) đến O (14)
        tri_gia_gtgt_row_idx = imp_row_idx + 29
        if tri_gia_gtgt_row_idx < len(df_page):
            tri_gia = self._get_merged_cell_value(df_page, tri_gia_gtgt_row_idx, 8, end_col=14)
            if tri_gia:
                try:
                    cleaned = re.sub(r'[^\d.,]', '', str(tri_gia))
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                    current_item['tri_gia_tinh_thue_gtgt'] = Decimal(cleaned)
                    debug_print(f"  Row {tri_gia_gtgt_row_idx} (IMP+29): Found tri_gia_tinh_thue_gtgt: {current_item['tri_gia_tinh_thue_gtgt']}")
                except:
                    pass
        
        # Lấy thuế suất GTGT: I163:O163 -> offset 30 rows, cột I (8) đến O (14)
        thue_gtgt_rate_row_idx = imp_row_idx + 30
        if thue_gtgt_rate_row_idx < len(df_page):
            thue_gtgt_rate = self._get_merged_cell_value(df_page, thue_gtgt_rate_row_idx, 8, end_col=14)
            if thue_gtgt_rate:
                val_str = str(thue_gtgt_rate).strip()
                rate_match = re.search(r'(\d+(?:\.\d+)?)\s*%', val_str)
                if rate_match:
                    try:
                        current_item['thue_gtgt_rate'] = Decimal(rate_match.group(1))
                        debug_print(f"  Row {thue_gtgt_rate_row_idx} (IMP+30): Found thue_gtgt_rate: {current_item['thue_gtgt_rate']}%")
                    except:
                        pass
        
        # Lấy tổng tiền thuế GTGT: I164:O164 -> offset 31 rows, cột I (8) đến O (14)
        thue_gtgt_vnd_row_idx = imp_row_idx + 31
        if thue_gtgt_vnd_row_idx < len(df_page):
            thue_gtgt_vnd = self._get_merged_cell_value(df_page, thue_gtgt_vnd_row_idx, 8, end_col=14)
            if thue_gtgt_vnd:
                try:
                    cleaned = re.sub(r'[^\d.,]', '', str(thue_gtgt_vnd))
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                    current_item['thue_gtgt_vnd_total'] = Decimal(cleaned)
                    debug_print(f"  Row {thue_gtgt_vnd_row_idx} (IMP+31): Found thue_gtgt_vnd_total: {current_item['thue_gtgt_vnd_total']}")
                except:
                    pass
        
        # Tính các giá trị thuế GTGT
        # Trị giá tính thuế GTGT từng chiếc = Trị giá tính thuế GTGT (total) / Số lượng
        if current_item['so_luong'] > 0 and current_item['tri_gia_tinh_thue_gtgt'] > 0:
            current_item['tri_gia_tinh_thue_gtgt_per_unit'] = current_item['tri_gia_tinh_thue_gtgt'] / current_item['so_luong']
        else:
            current_item['tri_gia_tinh_thue_gtgt_per_unit'] = Decimal('0')
        
        # Thuế GTGT VND từng chiếc = Trị giá tính thuế GTGT từng chiếc * Thuế suất / 100
        if current_item['tri_gia_tinh_thue_gtgt_per_unit'] > 0 and current_item['thue_gtgt_rate'] > 0:
            current_item['thue_gtgt_vnd'] = (current_item['tri_gia_tinh_thue_gtgt_per_unit'] * current_item['thue_gtgt_rate']) / Decimal('100')
        else:
            current_item['thue_gtgt_vnd'] = Decimal('0')
        
        # Nếu chưa có thue_gtgt_vnd_total, tính từ trị giá total
        if not current_item.get('thue_gtgt_vnd_total') and current_item['tri_gia_tinh_thue_gtgt'] > 0 and current_item['thue_gtgt_rate'] > 0:
            current_item['thue_gtgt_vnd_total'] = (current_item['tri_gia_tinh_thue_gtgt'] * current_item['thue_gtgt_rate']) / Decimal('100')
        
        items.append(current_item)
        debug_print(f"  Completed item STT {current_item['stt']}: ma_so={current_item['ma_so_hang_hoa']}, mo_ta={current_item.get('mo_ta_hang_hoa', '')[:50] if current_item.get('mo_ta_hang_hoa') else 'None'}")
        
        return items
    
    def _get_cell_value(self, row, col_idx: int) -> Any:
        """Lấy giá trị cell từ row và col_idx"""
        try:
            if col_idx < len(row):
                val = row.iloc[col_idx] if hasattr(row, 'iloc') else row[col_idx]
                # Xử lý NaN và empty string
                if pd.isna(val):
                    return None
                val_str = str(val).strip()
                if not val_str or val_str == 'nan' or val_str.lower() == 'none':
                    return None
                return val_str
            return None
        except:
            return None
    
    def _get_merged_cell_value(self, df_page: pd.DataFrame, row_idx: int, start_col: int, end_col: int = None) -> Optional[str]:
        """
        Lấy giá trị từ merged cell.
        Đọc từ start_col đến end_col (nếu có), nếu không thì chỉ đọc start_col.
        Vì đã unmerge và fill, nên chỉ cần đọc ô đầu tiên.
        """
        try:
            if row_idx >= len(df_page):
                return None
            
            row = df_page.iloc[row_idx]
            
            # Nếu có end_col, thử đọc từ start_col đến end_col để tìm giá trị
            if end_col is not None:
                for col in range(start_col, min(end_col + 1, len(row))):
                    val = self._get_cell_value(row, col)
                    if val:
                        return val
            else:
                # Chỉ đọc start_col
                return self._get_cell_value(row, start_col)
            
            return None
        except:
            return None
    
    def _is_stt(self, value: Any) -> bool:
        """Kiểm tra xem giá trị có phải là STT không"""
        if value is None:
            return False
        value_str = str(value).strip()
        # STT thường là số nguyên dương
        try:
            stt_num = int(float(value_str))
            return stt_num > 0
        except:
            return False
    
    def _parse_stt(self, value: Any) -> int:
        """Parse STT từ giá trị"""
        try:
            return int(float(str(value).strip()))
        except:
            return 0
    
    def _find_mo_ta_in_row(self, row, row_idx: int) -> Optional[str]:
        """
        Tìm mô tả hàng hóa trong dòng hiện tại.
        Tìm keyword "Mô tả hàng hóa" và đọc giá trị ở cột G (index 6) - merged cells.
        """
        # Kiểm tra dòng hiện tại xem có keyword "Mô tả hàng hóa" không
        for c in range(min(10, len(row))):
            cell_val = self._get_cell_value(row, c)
            if cell_val and isinstance(cell_val, str) and 'Mô tả hàng hóa' in cell_val:
                # Đọc giá trị ở cột G (index 6)
                mo_ta_val = self._get_cell_value(row, 6)
                if mo_ta_val and isinstance(mo_ta_val, str):
                    mo_ta_text = str(mo_ta_val).strip()
                    if len(mo_ta_text) > 10:  # Bỏ qua text quá ngắn
                        return mo_ta_text
                # Nếu không có ở cùng dòng, thử dòng dưới (trong phạm vi 5 dòng)
                for check_r in range(row_idx + 1, min(len(self.df), row_idx + 6)):
                    check_row = self.df.iloc[check_r]
                    mo_ta_val = self._get_cell_value(check_row, 6)
                    if mo_ta_val and isinstance(mo_ta_val, str):
                        mo_ta_text = str(mo_ta_val).strip()
                        if len(mo_ta_text) > 10:
                            return mo_ta_text
        return None
    
    def _find_thue_nk(self, row, row_idx: int) -> Optional[Dict[str, Any]]:
        """
        Tìm thông tin thuế nhập khẩu.
        Tìm dòng có "Thuế nhập khẩu" hoặc "Thuế NK".
        """
        result = {}
        
        # Kiểm tra dòng hiện tại
        row_text = ' '.join([str(self._get_cell_value(row, c) or '') for c in range(min(15, len(row)))])
        
        if 'Thuế nhập khẩu' in row_text or 'Thuế NK' in row_text:
            # Tìm thuế suất (pattern: "X%" hoặc "C X%")
            rate_match = re.search(r'(\d+(?:\.\d+)?)\s*%', row_text)
            if rate_match:
                try:
                    result['rate'] = Decimal(rate_match.group(1))
                except:
                    pass
            
            # Tìm số tiền thuế (thường ở cột cuối hoặc gần cuối)
            # Đọc từ phải sang trái để tìm số lớn nhất (số tiền thuế)
            for c in range(len(row) - 1, max(0, len(row) - 10), -1):
                val = self._get_cell_value(row, c)
                if val:
                    val_str = str(val).strip()
                    # Bỏ qua nếu là text không phải số
                    if 'VND' in val_str.upper() or any(char.isdigit() for char in val_str):
                        try:
                            # Loại bỏ dấu phẩy, chấm (format số VN)
                            cleaned = re.sub(r'[^\d.,]', '', val_str)
                            cleaned = cleaned.replace('.', '').replace(',', '.')
                            amount = Decimal(cleaned)
                            if amount > 1000:  # Số tiền thuế thường > 1000
                                result['amount'] = amount
                                logger.info(f"[TKHQ Parser] Found thue_nk amount: {amount} at row {row_idx}, col {c}")
                                break
                        except:
                            pass
        
        return result if result else None
    
    def _find_thue_gtgt(self, row, row_idx: int) -> Optional[Dict[str, Any]]:
        """
        Tìm thông tin thuế GTGT.
        Tìm dòng có "Thuế GTGT", "VAT" hoặc "10%".
        """
        result = {}
        
        # Kiểm tra dòng hiện tại
        row_text = ' '.join([str(self._get_cell_value(row, c) or '') for c in range(min(15, len(row)))])
        
        if 'Thuế GTGT' in row_text or 'VAT' in row_text.upper() or 'GTGT' in row_text:
            # Tìm thuế suất (pattern: "X%")
            rate_match = re.search(r'(\d+(?:\.\d+)?)\s*%', row_text)
            if rate_match:
                try:
                    result['rate'] = Decimal(rate_match.group(1))
                except:
                    pass
            
            # Tìm số tiền thuế (thường ở cột cuối)
            for c in range(len(row) - 1, max(0, len(row) - 10), -1):
                val = self._get_cell_value(row, c)
                if val:
                    val_str = str(val).strip()
                    if 'VND' in val_str.upper() or any(char.isdigit() for char in val_str):
                        try:
                            cleaned = re.sub(r'[^\d.,]', '', val_str)
                            cleaned = cleaned.replace('.', '').replace(',', '.')
                            amount = Decimal(cleaned)
                            if amount > 1000:
                                result['amount'] = amount
                                logger.info(f"[TKHQ Parser] Found thue_gtgt amount: {amount} at row {row_idx}, col {c}")
                                break
                        except:
                            pass
        
        return result if result else None

