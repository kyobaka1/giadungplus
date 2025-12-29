from django.shortcuts import render, redirect, get_object_or_404
from kho.utils import admin_only
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from typing import List, Dict, Any
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
from django.utils import timezone
import logging
import io
import os
import requests
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLSXImage
try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None

from core.sapo_client import get_sapo_client
from products.services.sapo_product_service import SapoProductService
from products.services.dto import ProductDTO, ProductVariantDTO
from products.models import (
    ContainerTemplate,
    ContainerTemplateSupplier,
    SumPurchaseOrder,
    SPOPurchaseOrder,
    SPOCost,
    SPODocument,
    SPOPackingListItem,
    PurchaseOrder,
    PurchaseOrderCost,
    PurchaseOrderPayment,
    BalanceTransaction,
    PaymentPeriod,
    PaymentPeriodTransaction,
)
from products.brand_settings import (
    get_disabled_brands,
    is_brand_enabled,
    get_enabled_brands,
    set_brand_enabled,
    reload_settings,
    sync_brands_from_api
)
from products.views_excel import export_variants_excel, import_variants_excel
from products.services.xnk_model_service import XNKModelService
import re

logger = logging.getLogger(__name__)

# Import loginss từ thamkhao để giữ nguyên cách lưu Model XNK
try:
    from thamkhao.apps import loginss
except ImportError:
    # Nếu không import được, sẽ dùng SapoClient.core_session
    loginss = None


@admin_only
def product_list(request: HttpRequest):
    """
    Danh sách sản phẩm:
    - Hiển thị danh sách TẤT CẢ sản phẩm từ Sapo (không phân trang)
    - Có thể tìm kiếm, lọc theo brand, category, status
    - Có thể sửa, xoá sản phẩm
    """
    context = {
        "title": "Danh sách sản phẩm",
        "products": [],
        "total": 0,
    }

    try:
        sapo_client = get_sapo_client()
        product_service = SapoProductService(sapo_client)

        # Lấy TẤT CẢ products (loop qua nhiều pages) - KHÔNG CÓ FILTER GÌ HẾT
        all_products = []
        page = 1
        limit = 250  # Lấy nhiều nhất có thể mỗi page
        
        while True:
            # Chỉ lấy dữ liệu, không có filter gì
            filters = {
                "page": page,
                "limit": limit,
            }
            
            products = product_service.list_products(**filters)
            
            if not products:
                break
            
            all_products.extend(products)
            
            # Nếu số lượng products < limit thì đã hết
            if len(products) < limit:
                break
            
            page += 1
            
            # Giới hạn tối đa 1000 pages để tránh vòng lặp vô hạn
            if page > 1000:
                logger.warning("Reached max pages limit (1000) in product_list")
                break

        # Reload settings để đảm bảo có dữ liệu mới nhất
        reload_settings()
        
        # Convert to dict for template và collect brands
        products_data = []
        brands_set = set()
        statuses_set = set()
        
        for product in all_products:
            brand = product.brand or ""
            product_status = product.status or ""
            
            # Chỉ thêm sản phẩm nếu nhãn hiệu được bật
            if brand and not is_brand_enabled(brand):
                continue  # Bỏ qua sản phẩm có nhãn hiệu bị tắt
            
            if brand:
                brands_set.add(brand)
            if product_status:
                statuses_set.add(product_status)
            
            # Lấy ảnh default
            default_image = None
            if product.images:
                for img in product.images:
                    if img.is_default:
                        default_image = img.full_path
                        break
                # Nếu không có ảnh default, lấy ảnh đầu tiên
                if not default_image and len(product.images) > 0:
                    default_image = product.images[0].full_path
            
            # Parse tags
            tags_list = []
            if product.tags:
                tags_list = [tag.strip() for tag in product.tags.split(',') if tag.strip()]
            
            products_data.append({
                "id": product.id,
                "name": product.name,
                "brand": brand,
                "category": product.category or "",
                "status": product_status,
                "variant_count": product.variant_count,
                "total_inventory": product.total_inventory_all_variants,
                "created_on": product.created_on,
                "modified_on": product.modified_on,
                "gdp_metadata": product.gdp_metadata,
                "image_url": default_image,
                "tags": tags_list,
            })

        context["products"] = products_data
        context["total"] = len(products_data)
        # Chỉ hiển thị nhãn hiệu đã bật trong filter
        context["brands"] = get_enabled_brands(sorted(list(brands_set)))
        context["statuses"] = sorted(list(statuses_set))  # Danh sách statuses để tạo filter buttons

    except Exception as e:
        logger.error(f"Error in product_list: {e}", exc_info=True)
        context["error"] = str(e)

    return render(request, "products/product_list.html", context)


@admin_only
def product_detail(request: HttpRequest, product_id: int):
    """
    Chi tiết sản phẩm:
    - Hiển thị thông tin chi tiết sản phẩm
    - Có thể sửa thông tin
    """
    context = {
        "title": "Chi tiết sản phẩm",
        "product": None,
    }

    try:
        sapo_client = get_sapo_client()
        product_service = SapoProductService(sapo_client)

        product = product_service.get_product(product_id)
        if not product:
            context["error"] = "Không tìm thấy sản phẩm"
            return render(request, "products/product_detail.html", context)

        context["product"] = product

    except Exception as e:
        logger.error(f"Error in product_detail: {e}", exc_info=True)
        context["error"] = str(e)

    return render(request, "products/product_detail.html", context)


def _parse_sku_for_sorting(sku: str) -> tuple:
    """
    Parse SKU để tạo sort key.
    
    Ví dụ: ER-0746-4XM -> (prefix='ER', number=746, suffix_num=4, suffix_letters='XM')
    
    Args:
        sku: SKU string (ví dụ: "ER-0746-4XM")
        
    Returns:
        Tuple (prefix, number, suffix_num, suffix_letters) để dùng cho sorting
    """
    if not sku:
        return ('', 0, 0, '')
    
    # Pattern: PREFIX-NUMBER-SUFFIX
    # Ví dụ: ER-0746-4XM, TG-0201-DEN
    parts = sku.split('-')
    
    if len(parts) < 2:
        # Không match pattern, trả về SKU gốc để sort alphabetically
        return (sku, 0, 0, '')
    
    prefix = parts[0]
    
    # Lấy phần số (có thể có leading zeros)
    try:
        number = int(parts[1])
    except (ValueError, IndexError):
        number = 0
    
    # Lấy suffix (phần cuối)
    if len(parts) >= 3:
        suffix = parts[2]
        # Parse suffix: 4XM -> (4, 'XM')
        suffix_match = re.match(r'^(\d+)([A-Za-z]*)$', suffix)
        if suffix_match:
            suffix_num = int(suffix_match.group(1))
            suffix_letters = suffix_match.group(2) or ''
        else:
            # Không match pattern số+chữ, dùng toàn bộ suffix
            suffix_num = 0
            suffix_letters = suffix
    else:
        suffix_num = 0
        suffix_letters = ''
    
    return (prefix, number, suffix_num, suffix_letters)


@admin_only
def variant_list(request: HttpRequest):
    """
    Danh sách phân loại (variants):
    - Hiển thị variants theo brand_id (mặc định = 833608)
    - Filter server-side để giảm tải
    - Có thể tìm kiếm theo SKU, barcode, tên
    - Có thể sửa, xoá variant
    - Sắp xếp theo SKU: nhóm theo mã số, sắp xếp suffix
    """
    # Brand ID mặc định
    DEFAULT_BRAND_ID = 833608
    
    # Lấy brand_id từ query param
    brand_id = request.GET.get('brand_id', str(DEFAULT_BRAND_ID))
    try:
        brand_id = int(brand_id)
    except (ValueError, TypeError):
        brand_id = DEFAULT_BRAND_ID
    
    context = {
        "title": "Danh sách phân loại",
        "variants": [],
        "total": 0,
        "selected_brand_id": brand_id,
        "brands": [],
    }

    try:
        # Reload settings trước để đảm bảo có dữ liệu mới nhất
        reload_settings()
        
        sapo_client = get_sapo_client()
        core_repo = sapo_client.core
        product_service = SapoProductService(sapo_client)
        
        # Lấy danh sách brands từ API search (đầy đủ hơn)
        brands_response = core_repo.list_brands_search_raw(page=1, limit=220)
        all_brands = brands_response.get("brands", [])
        
        # Đồng bộ brands mới từ API vào settings
        sync_brands_from_api(all_brands)
        
        # Lọc chỉ lấy brands được bật (enabled)
        enabled_brands = [
            brand for brand in all_brands 
            if is_brand_enabled(brand.get("name", ""))
        ]
        context["brands"] = sorted(enabled_brands, key=lambda x: x.get("name", ""))
        
        # Lấy products từ Sapo API (đã bao gồm variants và inventories)
        # Thử filter theo brand_id nếu API hỗ trợ, nếu không thì filter ở client-side
        all_products = []
        page = 1
        limit = 250  # Sapo API limit tối đa là 250
        
        logger.info(f"[variant_list] Starting to fetch products (with variants) for brand_id={brand_id}")
        
        while True:
            # Thử filter theo brand_id nếu API hỗ trợ
            filters = {
                "page": page,
                "limit": limit,
                "status": "active"  # Có thể bỏ nếu muốn lấy cả inactive
            }
            # Thử thêm brand_ids filter (có thể không hỗ trợ, nhưng thử xem)
            try:
                filters["brand_ids"] = brand_id
            except:
                pass
            
            products_response = core_repo.list_products_raw(**filters)
            products_data = products_response.get("products", [])
            
            if not products_data:
                logger.info(f"[variant_list] No more products at page {page}")
                break
            
            all_products.extend(products_data)
            logger.info(f"[variant_list] Fetched page {page}: {len(products_data)} products (total so far: {len(all_products)})")
            
            # Nếu số products trả về ít hơn limit, có nghĩa là đã hết
            if len(products_data) < limit:
                logger.info(f"[variant_list] Received fewer products than limit ({len(products_data)} < {limit}), assuming last page")
                break
            
            page += 1
            
            # Safety limit để tránh vòng lặp vô hạn
            if page > 100:
                logger.warning(f"[variant_list] Reached safety limit of 100 pages, stopping")
                break
        
        # Parse products và extract variants
        # Tạo map product_id -> product để parse metadata
        product_map = {}
        all_variants_from_products = []
        
        for product_data in all_products:
            product_id = product_data.get("id")
            if not product_id:
                continue
            
            # Lấy brand_id từ product_data (có sẵn trong JSON)
            brand_id_from_product = product_data.get("brand_id")
            
            # Parse product để có metadata
            try:
                product = product_service.get_product(product_id)
                if product:
                    product_map[product_id] = product
            except Exception as e:
                logger.warning(f"Failed to parse product {product_id}: {e}")
                # Vẫn dùng product_data nếu parse lỗi
                product = None
            
            # Extract variants từ product (đã có inventories sẵn)
            # CHỈ lấy variants có packsize = false (1 pcs), loại bỏ packsize = true (combo)
            variants = product_data.get("variants", [])
            for variant_data in variants:
                variant_id = variant_data.get("id")
                if not variant_id:
                    continue
                
                # Bỏ qua variant có packsize = true (combo)
                packsize = variant_data.get("packsize", False)
                if packsize is True:
                    continue
                
                # Lưu variant với product_id và brand_id để filter sau
                variant_data["_product_id"] = product_id
                variant_data["_brand_id"] = brand_id_from_product
                variant_data["_product"] = product  # Lưu product object nếu có
                all_variants_from_products.append(variant_data)
        
        logger.info(f"[variant_list] Extracted {len(all_variants_from_products)} variants from {len(all_products)} products")
        
        # Parse variants và lấy metadata, filter theo brand_id
        variants_data = []
        brands_set = set()
        statuses_set = set()
        
        for variant_data in all_variants_from_products:
            variant_id = variant_data.get("id")
            product_id = variant_data.get("_product_id")
            product = variant_data.get("_product")
            
            # Lấy brand_id từ variant_data (đã lưu từ product_data)
            variant_brand_id = variant_data.get("_brand_id")
            
            # Filter theo brand_id (client-side)
            if variant_brand_id != brand_id:
                continue
            
            # Lấy brand name từ product hoặc all_brands
            brand = ""
            if product:
                brand = product.brand or ""
            else:
                # Fallback: tìm brand name từ all_brands
                for b in all_brands:
                    if b.get("id") == variant_brand_id:
                        brand = b.get("name", "")
                        break
            
            # Chỉ thêm variants nếu nhãn hiệu được bật
            if brand and not is_brand_enabled(brand):
                continue
            
            if brand:
                brands_set.add(brand)
            
            variant_status = variant_data.get("status", "")
            if variant_status:
                statuses_set.add(variant_status)
            
            # Lấy metadata từ product nếu có
            variant_meta = None
            if product:
                # Tìm variant trong product để lấy metadata
                for v in product.variants:
                    if v.id == variant_id:
                        variant_meta = v.gdp_metadata
                        break
            
            # Sử dụng inventories từ variant_data (đã có sẵn trong products)
            inventories = variant_data.get("inventories", [])
            total_inventory = sum(inv.get("on_hand", 0) or 0 for inv in inventories)
            total_available = sum(inv.get("available", 0) or 0 for inv in inventories)
            
            # Lấy on_hand theo 3 kho: HN (241737), SG (548744), Showroom (783623)
            inventory_hn = 0
            inventory_sg = 0
            inventory_showroom = 0
            for inv in inventories:
                location_id = inv.get("location_id")
                on_hand = inv.get("on_hand", 0) or 0
                if location_id == 241737:  # Kho HN
                    inventory_hn = on_hand
                elif location_id == 548744:  # Kho SG
                    inventory_sg = on_hand
                elif location_id == 783623:  # Showroom
                    inventory_showroom = on_hand
            
            # Lấy images từ variant_data
            images = variant_data.get("images", [])
            
            # Lấy opt1, opt2, opt3 - kiểm tra cả opt1 và option1
            opt1_raw = variant_data.get("opt1")
            if opt1_raw is None:
                opt1_raw = variant_data.get("option1")
            opt1 = opt1_raw or ""
            
            opt2_raw = variant_data.get("opt2")
            if opt2_raw is None:
                opt2_raw = variant_data.get("option2")
            opt2 = opt2_raw or ""
            
            opt3_raw = variant_data.get("opt3")
            if opt3_raw is None:
                opt3_raw = variant_data.get("option3")
            opt3 = opt3_raw or ""
            
            # Lấy product name
            product_name = ""
            if product:
                product_name = product.name
            else:
                variant_name = variant_data.get("name", "")
                if variant_name:
                    product_name = variant_name.split(" - ")[0]
            
            variants_data.append({
                "id": variant_id,
                "product_id": product_id,
                "product_name": product_name,
                "brand": brand,
                "sku": variant_data.get("sku", ""),
                "barcode": variant_data.get("barcode") or "",
                "name": variant_data.get("name", ""),
                "opt1": opt1,
                "opt2": opt2,
                "opt3": opt3,
                "status": variant_status,
                "variant_retail_price": variant_data.get("variant_retail_price", 0) or 0,
                "variant_whole_price": variant_data.get("variant_whole_price", 0) or 0,
                "total_inventory": total_inventory,
                "total_available": total_available,
                "inventory_hn": inventory_hn,
                "inventory_sg": inventory_sg,
                "inventory_showroom": inventory_showroom,
                "inventories": inventories,  # Giữ lại để dùng sau nếu cần
                "images": images,  # Thêm images
                "weight_value": variant_data.get("weight_value", 0) or 0,
                "weight_unit": variant_data.get("weight_unit", "g"),
                "gdp_metadata": variant_meta,
                # Extract metadata fields
                "price_tq": variant_meta.price_tq if variant_meta else None,
                "sku_tq": variant_meta.sku_tq if variant_meta else None,
                "name_tq": variant_meta.name_tq if variant_meta else None,
                "sku_model_xnk": variant_meta.sku_model_xnk if variant_meta else None,
                "box_info": variant_meta.box_info if variant_meta else None,
                "packed_info": variant_meta.packed_info if variant_meta else None,
                "plan_tags": variant_meta.plan_tags if variant_meta and variant_meta.plan_tags else [],
                "plan_tags_with_color": [],  # Sẽ được populate sau
            })
        
        # Sắp xếp variants theo SKU: nhóm theo mã số, sắp xếp suffix
        # Ví dụ: ER-0746-4XM, ER-0746-5XM, ER-0746-6XM, ER-0746-4XR, ER-0746-5XR
        variants_data.sort(key=lambda v: _parse_sku_for_sorting(v.get("sku", "")))
        
        logger.info(f"[variant_list] Sorted {len(variants_data)} variants by SKU pattern")
        
        # Luôn hiển thị tất cả variants
        # Load variant tags từ database (với màu sắc)
        from settings.models import VariantTag
        all_tags = {}
        all_tags_with_color = {}
        for tag in VariantTag.objects.all():
            all_tags[tag.tags_name] = tag.id
            all_tags_with_color[tag.tags_name] = {
                'id': tag.id,
                'color': tag.color,
                'color_classes': tag.get_color_classes()
            }
        context["all_tags"] = all_tags
        context["all_tags_with_color"] = all_tags_with_color
        
        # Populate plan_tags_with_color cho từng variant
        for variant_data in variants_data:
            plan_tags_with_color = []
            if variant_data.get("plan_tags"):
                for tag_name in variant_data["plan_tags"]:
                    if tag_name in all_tags_with_color:
                        plan_tags_with_color.append({
                            'name': tag_name,
                            'color_classes': all_tags_with_color[tag_name]['color_classes']
                        })
                    else:
                        # Fallback nếu tag không tồn tại trong database
                        plan_tags_with_color.append({
                            'name': tag_name,
                            'color_classes': 'bg-blue-100 text-blue-800 border-blue-200'
                        })
            variant_data["plan_tags_with_color"] = plan_tags_with_color
        
        context["variants"] = variants_data
        context["total"] = len(variants_data)
        context["total_variants"] = len(variants_data)
        context["statuses"] = sorted(list(statuses_set))

    except Exception as e:
        logger.error(f"Error in variant_list: {e}", exc_info=True)
        context["error"] = str(e)

    return render(request, "products/variant_list.html", context)


@admin_only
def variant_detail(request: HttpRequest, variant_id: int):
    """
    Chi tiết phân loại:
    - Hiển thị thông tin chi tiết variant
    - Có thể sửa thông tin
    """
    context = {
        "title": "Chi tiết phân loại",
        "variant": None,
    }

    try:
        sapo_client = get_sapo_client()
        core_repo = sapo_client.core

        # Lấy variant từ Sapo
        variant_response = core_repo.get_variant_raw(variant_id)
        variant_data = variant_response.get("variant")

        if not variant_data:
            context["error"] = "Không tìm thấy phân loại"
            return render(request, "products/variant_detail.html", context)

        # Lấy product để có thông tin đầy đủ
        product_id = variant_data.get("product_id")
        if product_id:
            product_service = SapoProductService(sapo_client)
            product = product_service.get_product(product_id)
            if product:
                # Tìm variant trong product
                for v in product.variants:
                    if v.id == variant_id:
                        context["variant"] = v
                        context["product"] = product
                        break

        if not context.get("variant"):
            context["error"] = "Không tìm thấy phân loại trong product"

    except Exception as e:
        logger.error(f"Error in variant_detail: {e}", exc_info=True)
        context["error"] = str(e)

    return render(request, "products/variant_detail.html", context)


@admin_only
@require_POST
def init_all_products_metadata(request: HttpRequest):
    """
    Init metadata cho sản phẩm.
    
    Endpoint: POST /products/init-all-metadata/
    
    Query parameters:
    - test_mode: true/false - Nếu true: init tất cả sản phẩm (kể cả đã có metadata)
                   Nếu false: chỉ init sản phẩm chưa có metadata
    
    Tự động init metadata cho products, thêm vào cuối description nếu có nội dung.
    """
    try:
        # Lấy test_mode từ request body hoặc query params
        test_mode = False
        if request.content_type == 'application/json':
            try:
                import json
                body = json.loads(request.body)
                test_mode = body.get('test_mode', False)
            except:
                pass
        else:
            test_mode = request.POST.get('test_mode', 'false').lower() == 'true'
        
        sapo_client = get_sapo_client()
        product_service = SapoProductService(sapo_client)
        
        # Lấy tất cả products (active)
        all_products = []
        page = 1
        limit = 250
        max_pages = 100  # Giới hạn để tránh quá tải
        
        while page <= max_pages:
            products = product_service.list_products(page=page, limit=limit, status="active")
            if not products:
                break
            all_products.extend(products)
            if len(products) < limit:
                break
            page += 1
        
        # Xác định products cần init
        products_to_init = []
        products_already_have = []
        products_skipped = []
        
        for product in all_products:
            if test_mode:
                # Test mode: init tất cả sản phẩm
                products_to_init.append(product.id)
            else:
                # Normal mode: chỉ init sản phẩm chưa có metadata
                if product.gdp_metadata:
                    products_already_have.append(product.id)
                else:
                    products_to_init.append(product.id)
        
        # Init metadata cho các products
        success_count = 0
        error_count = 0
        errors = []
        
        for product_id in products_to_init:
            try:
                # Nếu test_mode, force init bằng cách update trực tiếp
                if test_mode:
                    # Lấy product hiện tại
                    product = product_service.get_product(product_id)
                    if product:
                        # Tạo metadata mới với structure đầy đủ
                        from products.services.metadata_helper import init_empty_metadata
                        variant_ids = [v.id for v in product.variants]
                        metadata = init_empty_metadata(product_id, variant_ids)
                        
                        # Update metadata (preserve description nếu có)
                        success = product_service.update_product_metadata(
                            product_id, 
                            metadata, 
                            preserve_description=True
                        )
                        if success:
                            success_count += 1
                        else:
                            error_count += 1
                            errors.append(f"Product {product_id}: Failed to update")
                    else:
                        error_count += 1
                        errors.append(f"Product {product_id}: Not found")
                else:
                    # Normal mode: dùng init_product_metadata (chỉ init nếu chưa có)
                    success = product_service.init_product_metadata(product_id)
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
                        errors.append(f"Product {product_id}: Failed to init")
            except Exception as e:
                error_count += 1
                errors.append(f"Product {product_id}: {str(e)}")
                logger.error(f"Error init metadata for product {product_id}: {e}", exc_info=True)
        
        return JsonResponse({
            "status": "success",
            "test_mode": test_mode,
            "total_products": len(all_products),
            "products_to_init": len(products_to_init),
            "products_already_have": len(products_already_have),
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors[:10] if len(errors) > 10 else errors,  # Chỉ trả về 10 lỗi đầu
        })
        
    except Exception as e:
        logger.error(f"Error in init_all_products_metadata: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
def brand_settings(request: HttpRequest):
    """
    Quản lý cài đặt nhãn hiệu (bật/tắt).
    - Hiển thị danh sách tất cả nhãn hiệu từ sản phẩm
    - Cho phép bật/tắt nhãn hiệu
    """
    context = {
        "title": "Cài đặt nhãn hiệu",
        "brands": [],
        "error": None,
    }
    
    try:
        # Reload settings trước để đảm bảo có dữ liệu mới nhất
        reload_settings()
        
        # Lấy tất cả nhãn hiệu từ API search (đầy đủ hơn)
        sapo_client = get_sapo_client()
        core_repo = sapo_client.core
        
        # Lấy brands từ API search
        brands_response = core_repo.list_brands_search_raw(page=1, limit=220)
        all_brands = brands_response.get("brands", [])
        
        # Đồng bộ brands mới từ API vào settings (tự động thêm brands mới)
        sync_brands_from_api(all_brands)
        
        # Reload lại settings sau khi sync
        reload_settings()
        
        # Lấy danh sách nhãn hiệu bị tắt
        disabled_brands = get_disabled_brands()
        
        # Tạo danh sách nhãn hiệu với trạng thái
        brands_data = []
        for brand in all_brands:
            brand_name = brand.get("name", "")
            if brand_name:
                brands_data.append({
                    "name": brand_name,
                    "id": brand.get("id"),
                    "is_enabled": is_brand_enabled(brand_name),
                })
        
        # Sắp xếp theo tên
        brands_data.sort(key=lambda x: x.get("name", ""))
        
        context["brands"] = brands_data
        context["disabled_count"] = len(disabled_brands)
        context["enabled_count"] = len(brands_data) - len(disabled_brands)
        
    except Exception as e:
        logger.error(f"Error in brand_settings: {e}", exc_info=True)
        context["error"] = str(e)
    
    return render(request, "products/brand_settings.html", context)


@admin_only
@require_POST
def toggle_brand(request: HttpRequest):
    """
    API để bật/tắt nhãn hiệu.
    
    POST data:
    - brand_name: Tên nhãn hiệu
    - enabled: true/false (bật/tắt)
    """
    try:
        # Lấy dữ liệu từ request
        if request.content_type == 'application/json':
            import json
            data = json.loads(request.body)
            brand_name = data.get('brand_name', '').strip()
            enabled = data.get('enabled', True)
        else:
            brand_name = request.POST.get('brand_name', '').strip()
            enabled = request.POST.get('enabled', 'true').lower() == 'true'
        
        if not brand_name:
            return JsonResponse({
                "status": "error",
                "message": "Tên nhãn hiệu không được để trống"
            }, status=400)
        
        # Cập nhật settings
        success = set_brand_enabled(brand_name, enabled)
        
        if success:
            return JsonResponse({
                "status": "success",
                "message": f"Đã {'bật' if enabled else 'tắt'} nhãn hiệu '{brand_name}'"
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": "Không thể lưu cài đặt"
            }, status=500)
            
    except Exception as e:
        logger.error(f"Error in toggle_brand: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
def init_variants_from_old_notes(request: HttpRequest):
    """
    Init variant metadata từ dữ liệu cũ trong customer notes.
    
    Endpoint: POST /products/init-variants-from-old-notes/
    
    Body (JSON):
    - test_mode: true/false - Nếu true: chỉ log không update
    - limit: Số lượng variants tối đa để migrate (optional)
    """
    try:
        # Lấy parameters
        test_mode = False
        limit = None
        
        if request.content_type == 'application/json':
            try:
                import json
                body = json.loads(request.body)
                test_mode = body.get('test_mode', False)
                limit = body.get('limit')
            except:
                pass
        else:
            test_mode = request.POST.get('test_mode', 'false').lower() == 'true'
            limit_str = request.POST.get('limit')
            if limit_str:
                try:
                    limit = int(limit_str)
                except:
                    pass
        
        # Import migration service
        from products.services.variant_migration import init_variants_from_old_data
        
        # Thực hiện migration
        result = init_variants_from_old_data(test_mode=test_mode, limit=limit)
        
        return JsonResponse({
            "status": "success",
            "test_mode": test_mode,
            "result": result
        })
        
    except Exception as e:
        logger.error(f"Error in init_variants_from_old_notes: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
def update_variant_metadata(request: HttpRequest, variant_id: int):
    """
    API endpoint để cập nhật variant metadata.
    Chỉ cập nhật các field có thể edit: price_tq, sku_tq, name_tq, box_info, sku_model_xnk
    """
    try:
        # Lấy dữ liệu từ request body
        import json
        data = json.loads(request.body)
        
        sapo_client = get_sapo_client()
        product_service = SapoProductService(sapo_client)
        
        # Lấy variant để biết product_id
        variant_response = sapo_client.core.get_variant_raw(variant_id)
        variant = variant_response.get("variant", {})
        
        if not variant:
            return JsonResponse({
                "status": "error",
                "message": f"Variant {variant_id} không tồn tại"
            }, status=404)
        
        product_id = variant.get("product_id")
        if not product_id:
            return JsonResponse({
                "status": "error",
                "message": f"Variant {variant_id} không có product_id"
            }, status=400)
        
        # Lấy product hiện tại
        product = product_service.get_product(product_id)
        if not product:
            return JsonResponse({
                "status": "error",
                "message": f"Product {product_id} không tồn tại"
            }, status=404)
        
        # Lấy metadata hiện tại
        from products.services.metadata_helper import init_empty_metadata, update_variant_metadata
        from products.services.dto import VariantMetadataDTO, BoxInfoDTO, PackedInfoDTO, NhanPhuInfoDTO
        
        current_metadata = product.gdp_metadata
        if not current_metadata:
            variant_ids = [v.id for v in product.variants]
            current_metadata = init_empty_metadata(product_id, variant_ids)
        
        # Tìm variant metadata hiện tại
        variant_meta = None
        for vm in current_metadata.variants:
            if vm.id == variant_id:
                variant_meta = vm
                break
        
        if not variant_meta:
            variant_meta = VariantMetadataDTO(id=variant_id)
        
        # Helper functions
        def to_float_safe(val, default=None):
            try:
                if val is None or val == "":
                    return default
                return float(val)
            except:
                return default
        
        def to_int_safe(val, default=None):
            try:
                if val is None or val == "":
                    return default
                return int(float(val))
            except:
                return default
        
        # Cập nhật từ request data
        price_tq = to_float_safe(data.get('price_tq'))
        sku_tq = data.get('sku_tq', '').strip() if data.get('sku_tq') else None
        name_tq = data.get('name_tq', '').strip() if data.get('name_tq') else None
        sku_model_xnk = data.get('sku_model_xnk', '').strip() if data.get('sku_model_xnk') else None
        
        # Plan tags
        plan_tags = data.get('plan_tags', [])
        if plan_tags and isinstance(plan_tags, list):
            # Validate tags exist in database
            from settings.models import VariantTag
            valid_tags = []
            for tag_name in plan_tags:
                if VariantTag.objects.filter(tags_name=tag_name).exists():
                    valid_tags.append(tag_name)
            plan_tags = valid_tags
        else:
            plan_tags = variant_meta.plan_tags if variant_meta and variant_meta.plan_tags else []
        
        # Update box_info
        full_box = to_int_safe(data.get('full_box'))
        box_length = to_float_safe(data.get('box_length_cm'))
        box_width = to_float_safe(data.get('box_width_cm'))
        box_height = to_float_safe(data.get('box_height_cm'))
        
        box_info = None
        if full_box is not None or box_length is not None or box_width is not None or box_height is not None:
            box_info = BoxInfoDTO(
                full_box=full_box if full_box is not None else (variant_meta.box_info.full_box if variant_meta.box_info else None),
                length_cm=box_length if box_length is not None else (variant_meta.box_info.length_cm if variant_meta.box_info else None),
                width_cm=box_width if box_width is not None else (variant_meta.box_info.width_cm if variant_meta.box_info else None),
                height_cm=box_height if box_height is not None else (variant_meta.box_info.height_cm if variant_meta.box_info else None)
            )
        elif variant_meta.box_info:
            box_info = variant_meta.box_info
        
        # Update packed_info
        packed_length = to_float_safe(data.get('packed_length_cm'))
        packed_width = to_float_safe(data.get('packed_width_cm'))
        packed_height = to_float_safe(data.get('packed_height_cm'))
        packed_weight_with_box = to_float_safe(data.get('packed_weight_with_box_g'))
        packed_weight_without_box = to_float_safe(data.get('packed_weight_without_box_g'))
        
        packed_info = None
        if packed_length is not None or packed_width is not None or packed_height is not None or packed_weight_with_box is not None or packed_weight_without_box is not None:
            packed_info = PackedInfoDTO(
                length_cm=packed_length if packed_length is not None else (variant_meta.packed_info.length_cm if variant_meta.packed_info else None),
                width_cm=packed_width if packed_width is not None else (variant_meta.packed_info.width_cm if variant_meta.packed_info else None),
                height_cm=packed_height if packed_height is not None else (variant_meta.packed_info.height_cm if variant_meta.packed_info else None),
                weight_with_box_g=packed_weight_with_box if packed_weight_with_box is not None else (variant_meta.packed_info.weight_with_box_g if variant_meta.packed_info else None),
                weight_without_box_g=packed_weight_without_box if packed_weight_without_box is not None else (variant_meta.packed_info.weight_without_box_g if variant_meta.packed_info else None)
            )
        elif variant_meta.packed_info:
            packed_info = variant_meta.packed_info
        
        # Update nhanphu_info (product level)
        nhanphu_vi_name = data.get('nhanphu_vi_name', '').strip() if data.get('nhanphu_vi_name') else None
        nhanphu_en_name = data.get('nhanphu_en_name', '').strip() if data.get('nhanphu_en_name') else None
        nhanphu_description = data.get('nhanphu_description', '').strip() if data.get('nhanphu_description') else None
        nhanphu_material = data.get('nhanphu_material', '').strip() if data.get('nhanphu_material') else None
        
        nhanphu_info = None
        if nhanphu_vi_name is not None or nhanphu_en_name is not None or nhanphu_description is not None or nhanphu_material is not None:
            nhanphu_info = NhanPhuInfoDTO(
                vi_name=nhanphu_vi_name if nhanphu_vi_name else (current_metadata.nhanphu_info.vi_name if current_metadata.nhanphu_info else None),
                en_name=nhanphu_en_name if nhanphu_en_name else (current_metadata.nhanphu_info.en_name if current_metadata.nhanphu_info else None),
                description=nhanphu_description if nhanphu_description else (current_metadata.nhanphu_info.description if current_metadata.nhanphu_info else None),
                material=nhanphu_material if nhanphu_material else (current_metadata.nhanphu_info.material if current_metadata.nhanphu_info else None),
                hdsd=current_metadata.nhanphu_info.hdsd if current_metadata.nhanphu_info else None
            )
        elif current_metadata.nhanphu_info:
            nhanphu_info = current_metadata.nhanphu_info
        
        # Tạo variant metadata mới
        new_variant_meta = VariantMetadataDTO(
            id=variant_id,
            price_tq=price_tq if price_tq is not None else variant_meta.price_tq,
            sku_tq=sku_tq if sku_tq else variant_meta.sku_tq,
            name_tq=name_tq if name_tq else variant_meta.name_tq,
            box_info=box_info,
            packed_info=packed_info,
            sku_model_xnk=sku_model_xnk if sku_model_xnk else variant_meta.sku_model_xnk,
            web_variant_id=variant_meta.web_variant_id if variant_meta.web_variant_id else [],
            plan_tags=plan_tags
        )
        
        # Update product level nhanphu_info nếu có
        if nhanphu_info:
            current_metadata.nhanphu_info = nhanphu_info
        
        # Update variant metadata
        current_metadata = update_variant_metadata(
            current_metadata,
            variant_id,
            new_variant_meta
        )
        
        # Lưu vào Sapo
        success = product_service.update_product_metadata(
            product_id,
            current_metadata,
            preserve_description=True
        )
        
        if success:
            return JsonResponse({
                "status": "success",
                "message": "Đã cập nhật thành công"
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": "Không thể lưu variant metadata"
            }, status=500)
        
    except Exception as e:
        logger.error(f"Error updating variant metadata {variant_id}: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


# ========================= XNK MODEL MANAGEMENT =========================

def _get_xnk_service():
    """
    Helper để lấy XNKModelService với session phù hợp.
    Sử dụng loginss từ thamkhao nếu có, nếu không thì dùng SapoClient.core_session.
    """
    if loginss is not None:
        return XNKModelService(loginss)
    else:
        # Fallback: dùng SapoClient
        sapo_client = get_sapo_client()
        # Đảm bảo core session đã được init (gọi private method nếu cần)
        try:
            sapo_client._ensure_logged_in()
        except:
            pass  # Nếu không login được, vẫn dùng session hiện tại
        return XNKModelService(sapo_client.core_session)


@admin_only
def xnk_model_list(request: HttpRequest):
    """
    Danh sách Model Xuất Nhập Khẩu:
    - Hiển thị danh sách tất cả model XNK từ customer notes
    - Có thể tìm kiếm, sửa, xóa
    """
    context = {
        "title": "Quản lý Model Xuất Nhập Khẩu",
        "models": [],
        "total": 0,
        "error": None,
    }
    
    try:
        xnk_service = _get_xnk_service()
        all_models = xnk_service.get_all_models()
        
        # Sắp xếp theo SKU
        all_models.sort(key=lambda x: str(x.get("sku", "")).lower())
        
        context["models"] = all_models
        context["total"] = len(all_models)
        
    except Exception as e:
        logger.error(f"Error in xnk_model_list: {e}", exc_info=True)
        context["error"] = str(e)
    
    return render(request, "products/xnk_model_list.html", context)


@admin_only
@require_http_methods(["GET"])
def api_xnk_search(request: HttpRequest):
    """
    API tìm kiếm model XNK theo SKU và tên tiếng anh.
    
    Query params:
    - q: Từ khóa tìm kiếm (SKU hoặc tên tiếng anh)
    
    Returns:
        JSON với danh sách model XNK khớp
    """
    try:
        query = request.GET.get("q", "").strip()
        
        xnk_service = _get_xnk_service()
        results = xnk_service.search_models(query)
        
        return JsonResponse({
            "status": "success",
            "results": results,
            "count": len(results)
        })
        
    except Exception as e:
        logger.error(f"Error in api_xnk_search: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_http_methods(["GET"])
def export_spo_packing_list(request: HttpRequest, spo_id: int):
    """
    Xuất packing_list cho một SPO:
    - Dựa trên các PO trong SPO
    - Dùng GDP Meta để lấy sku_model_xnk cho từng variant
    - Dùng Model XNK để lấy thông tin model (HS code, tên, đơn vị, giá, material, ...)
    - Gom theo SKU nhập khẩu (sku_model_xnk), cộng dồn quantity, box, total_cpm, total_price, total_weight
    - Xuất file Excel để client tải về.
    """
    from products.services.spo_po_service import SPOPOService
    from products.services.sapo_product_service import SapoProductService

    spo = get_object_or_404(SumPurchaseOrder, id=spo_id)

    # Lấy danh sách Sapo PO IDs trong SPO
    spo_po_relations = spo.spo_purchase_orders.select_related("purchase_order").all()
    po_ids = [
        rel.purchase_order.sapo_order_supplier_id
        for rel in spo_po_relations
        if rel.purchase_order and rel.purchase_order.sapo_order_supplier_id
    ]

    if not po_ids:
        return JsonResponse(
            {
                "status": "error",
                "message": "SPO chưa có PO nào để xuất packing_list.",
            },
            status=400,
        )

    sapo_client = get_sapo_client()
    spo_po_service = SPOPOService(sapo_client)
    product_service = SapoProductService(sapo_client)
    xnk_service = _get_xnk_service()

    # BƯỚC 1: Lấy toàn bộ PO data (bao gồm line_items với cpm, full_box...)
    pos_data = spo_po_service.get_pos_for_spo(po_ids)

    # BƯỚC 2: Thu thập product_ids để đọc GDP Meta
    product_ids = set()
    for po_data in pos_data:
        for item in po_data.get("line_items", []):
            product_id = item.get("product_id")
            if product_id:
                product_ids.add(product_id)

    # BƯỚC 3: Xây map variant_id -> sku_model_xnk từ GDP Meta
    variant_sku_xnk_map: Dict[int, str] = {}

    for product_id in product_ids:
        try:
            product_dto = product_service.get_product(product_id)
            if not product_dto or not product_dto.gdp_metadata:
                continue

            for v_meta in product_dto.gdp_metadata.variants:
                if v_meta.sku_model_xnk:
                    variant_sku_xnk_map[v_meta.id] = v_meta.sku_model_xnk.strip()
        except Exception as e:
            logger.warning(f"[export_spo_packing_list] Error loading product {product_id}: {e}")
            continue

    if not variant_sku_xnk_map:
        return JsonResponse(
            {
                "status": "error",
                "message": "Không tìm thấy sku_model_xnk cho bất kỳ variant nào trong SPO.",
            },
            status=400,
        )

    # BƯỚC 4: Load toàn bộ XNK models và tạo map SKU -> model
    try:
        all_xnk_models = xnk_service.get_all_models()
    except Exception as e:
        logger.error(f"[export_spo_packing_list] Error loading XNK models: {e}", exc_info=True)
        return JsonResponse(
            {
                "status": "error",
                "message": "Không thể load dữ liệu Model XNK.",
            },
            status=500,
        )

    xnk_model_map: Dict[str, Dict[str, Any]] = {}
    xnk_model_map_ci: Dict[str, Dict[str, Any]] = {}
    for model in all_xnk_models:
        sku = str(model.get("sku", "")).strip()
        if not sku:
            continue
        xnk_model_map[sku] = model
        xnk_model_map_ci[sku.upper()] = model

    if not xnk_model_map:
        return JsonResponse(
            {
                "status": "error",
                "message": "Không có dữ liệu Model XNK để đối chiếu.",
            },
            status=400,
        )

    # BƯỚC 5: Gom dữ liệu theo SKU nhập khẩu (sku_model_xnk)
    aggregated: Dict[str, Dict[str, Any]] = {}

    for po_data in pos_data:
        for item in po_data.get("line_items", []):
            variant_id = item.get("variant_id")
            quantity = item.get("quantity", 0) or 0

            if not variant_id or quantity <= 0:
                continue

            sku_model_xnk = variant_sku_xnk_map.get(variant_id)
            if not sku_model_xnk:
                # Không có SKU nhập khẩu trong GDP Meta
                continue

            # Tìm model XNK theo SKU nhập khẩu
            model = xnk_model_map.get(sku_model_xnk)
            if not model:
                model = xnk_model_map_ci.get(sku_model_xnk.upper())

            if not model:
                # Không tìm thấy model tương ứng
                continue

            sku_key = str(model.get("sku", "")).strip()
            if not sku_key:
                continue

            # Khởi tạo dòng mới nếu chưa có
            if sku_key not in aggregated:
                price = model.get("usd_price")
                try:
                    price_val = float(price) if price is not None else 0.0
                except (TypeError, ValueError):
                    price_val = 0.0

                aggregated[sku_key] = {
                    "hs_code": model.get("hs_code", "") or "",
                    "sku_nhapkhau": sku_key,
                    "vn_name": model.get("vn_name", "") or "",
                    "en_name": model.get("en_name", "") or "",
                    "cn_name": model.get("china_name", "") or "",
                    "nsx_name": model.get("nsx_name", "") or "",
                    "quantity": 0,
                    "unit": model.get("unit", "") or "",
                    "price": price_val,
                    "box": 0.0,
                    "total_price": 0.0,
                    "total_cpm": 0.0,
                    "material": model.get("material", "") or "",
                    "total_weight": 0.0,
                }

            row_data = aggregated[sku_key]

            # full_box để tính số thùng
            full_box = item.get("full_box") or 1
            try:
                full_box_val = float(full_box) if full_box else 1.0
            except (TypeError, ValueError):
                full_box_val = 1.0

            try:
                qty_val = float(quantity)
            except (TypeError, ValueError):
                qty_val = 0.0

            boxes = qty_val / full_box_val if full_box_val > 0 else qty_val

            # Cộng dồn các chỉ số
            row_data["quantity"] += quantity
            row_data["box"] += boxes
            row_data["total_price"] += qty_val * float(row_data["price"])

            cpm_val = item.get("cpm") or 0
            try:
                cpm_float = float(cpm_val)
            except (TypeError, ValueError):
                cpm_float = 0.0
            row_data["total_cpm"] += cpm_float

    if not aggregated:
        return JsonResponse(
            {
                "status": "error",
                "message": "Không có dữ liệu nào phù hợp để xuất packing_list.",
            },
            status=400,
        )

    # BƯỚC 6: Tính total_weight cho từng dòng (theo công thức tham khảo ALL-HCM mẫu)
    for row in aggregated.values():
        # TOTAL WEIGHT = TOTAL CAPACITY * 1,000,000 / 6,000
        total_cpm = row.get("total_cpm", 0.0) or 0.0
        try:
            total_cpm_val = float(total_cpm)
        except (TypeError, ValueError):
            total_cpm_val = 0.0
        row["total_weight"] = total_cpm_val * 1000000.0 / 6000.0 if total_cpm_val > 0 else 0.0

    # Sắp xếp theo SKU nhập khẩu
    rows = sorted(aggregated.values(), key=lambda r: r.get("sku_nhapkhau", ""))

    # BƯỚC 6.5: Lưu packing list vào database
    # Xóa packing list items cũ (nếu có)
    SPOPackingListItem.objects.filter(sum_purchase_order=spo).delete()
    # Tạo packing list items mới
    for order, item in enumerate(rows):
        SPOPackingListItem.objects.create(
            sum_purchase_order=spo,
            hs_code=item.get("hs_code", "") or "",
            sku_nhapkhau=item.get("sku_nhapkhau", "") or "",
            vn_name=item.get("vn_name", "") or "",
            en_name=item.get("en_name", "") or "",
            cn_name=item.get("cn_name", "") or "",
            nsx_name=item.get("nsx_name", "") or "",
            quantity=Decimal(str(item.get("quantity", 0) or 0)),
            unit=item.get("unit", "") or "",
            box=Decimal(str(item.get("box", 0.0) or 0.0)),
            price=Decimal(str(item.get("price", 0.0) or 0.0)),
            total_price=Decimal(str(item.get("total_price", 0.0) or 0.0)),
            total_cpm=Decimal(str(item.get("total_cpm", 0.0) or 0.0)),
            total_weight=Decimal(str(item.get("total_weight", 0.0) or 0.0)),
            material=item.get("material", "") or "",
            order=order,
            # Các field thuế mặc định = 0 (sẽ được cập nhật sau từ tờ khai hải quan)
            import_tax_rate=Decimal('0'),
            import_tax_amount=Decimal('0'),
            import_tax_total=Decimal('0'),
            vat_rate=Decimal('0'),
            vat_amount=Decimal('0'),
            vat_total=Decimal('0'),
        )

    # BƯỚC 7: Xuất Excel bằng xlsxwriter (giống format ALL-HCM2505)
    if xlsxwriter is None:
        return JsonResponse(
            {
                "status": "error",
                "message": "Thư viện xlsxwriter chưa được cài đặt trên server.",
            },
            status=500,
        )

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet("Packing List")

    # Header
    row_idx = 0
    col = 0
    headers = [
        "MÃ HS",
        "TÊN TIẾNG VIỆT",
        "TÊN TIẾNG ANH",
        "TÊN TIẾNG TRUNG",
        "NSX",
        "QUANTITY",
        "UNIT",
        "BOX",
        "UNIT PRICE",
        "TOTAL PRICE",
        "TOTAL CAPACITY",
        "TOTAL WEIGHT",
        "MATERIAL",
        "SKU-NHAPKHAU",
    ]

    for idx, header in enumerate(headers):
        worksheet.write(row_idx, col + idx, header)

    # Dữ liệu
    row_idx = 1
    for item in rows:
        col = 0
        worksheet.write(row_idx, col, item.get("hs_code", ""))
        worksheet.write(row_idx, col + 1, item.get("vn_name", ""))
        worksheet.write(row_idx, col + 2, item.get("en_name", ""))
        worksheet.write(row_idx, col + 3, item.get("cn_name", ""))
        worksheet.write(row_idx, col + 4, item.get("nsx_name", ""))
        worksheet.write(row_idx, col + 5, item.get("quantity", 0))
        worksheet.write(row_idx, col + 6, item.get("unit", ""))
        worksheet.write(row_idx, col + 7, item.get("box", 0.0))
        worksheet.write(row_idx, col + 8, item.get("price", 0.0))
        worksheet.write(row_idx, col + 9, item.get("total_price", 0.0))
        worksheet.write(row_idx, col + 10, item.get("total_cpm", 0.0))
        worksheet.write(row_idx, col + 11, item.get("total_weight", 0.0))
        worksheet.write(row_idx, col + 12, item.get("material", ""))
        worksheet.write(row_idx, col + 13, item.get("sku_nhapkhau", ""))
        row_idx += 1

    workbook.close()
    output.seek(0)

    filename = f"ALL-{spo.code}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return response


@admin_only
@require_POST
@csrf_exempt
def api_xnk_edit(request: HttpRequest):
    """
    API chỉnh sửa model XNK.
    
    Body (JSON):
    {
        "sku": "SKU123",
        "field1": "value1",
        "field2": "value2",
        ...
    }
    
    Returns:
        JSON với status và message
    """
    try:
        import json
        data = json.loads(request.body)
        
        sku = data.get("sku", "").strip()
        if not sku:
            return JsonResponse({
                "status": "error",
                "message": "Thiếu SKU trong dữ liệu gửi lên"
            }, status=400)
        
        # Loại bỏ SKU khỏi updates (vì SKU là key để tìm)
        updates = {k: v for k, v in data.items() if k != "sku"}
        
        if not updates:
            return JsonResponse({
                "status": "error",
                "message": "Không có dữ liệu cập nhật"
            }, status=400)
        
        xnk_service = _get_xnk_service()
        result = xnk_service.update_model(sku, updates)
        
        if result.get("status") == "success":
            return JsonResponse({
                "status": "success",
                "message": result.get("msg", "Cập nhật thành công")
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": result.get("msg", "Lỗi khi cập nhật")
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({
            "status": "error",
            "message": "Dữ liệu JSON không hợp lệ"
        }, status=400)
    except Exception as e:
        logger.error(f"Error in api_xnk_edit: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
@csrf_exempt
def api_xnk_create(request: HttpRequest):
    """
    API tạo model XNK mới.
    
    Body (JSON):
    {
        "sku": "SKU123",
        "hs_code": "...",
        "en_name": "...",
        ...
    }
    
    Returns:
        JSON với status và message
    """
    try:
        import json
        data = json.loads(request.body)
        
        sku = data.get("sku", "").strip()
        if not sku:
            return JsonResponse({
                "status": "error",
                "message": "Thiếu SKU trong dữ liệu gửi lên"
            }, status=400)
        
        xnk_service = _get_xnk_service()
        result = xnk_service.create_model(data)
        
        if result.get("status") == "success":
            return JsonResponse({
                "status": "success",
                "message": result.get("msg", "Tạo mới thành công"),
                "note_id": result.get("note_id")
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": result.get("msg", "Lỗi khi tạo mới")
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({
            "status": "error",
            "message": "Dữ liệu JSON không hợp lệ"
        }, status=400)
    except Exception as e:
        logger.error(f"Error in api_xnk_create: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
@csrf_exempt
def api_xnk_delete(request: HttpRequest):
    """
    API xóa model XNK (set status = inactive).
    
    Body (JSON):
    {
        "sku": "SKU123"
    }
    
    Returns:
        JSON với status và message
    """
    try:
        import json
        data = json.loads(request.body)
        
        sku = data.get("sku", "").strip()
        if not sku:
            return JsonResponse({
                "status": "error",
                "message": "Thiếu SKU trong dữ liệu gửi lên"
            }, status=400)
        
        xnk_service = _get_xnk_service()
        result = xnk_service.delete_model(sku)
        
        if result.get("status") == "success":
            return JsonResponse({
                "status": "success",
                "message": result.get("msg", "Xóa thành công")
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": result.get("msg", "Lỗi khi xóa")
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({
            "status": "error",
            "message": "Dữ liệu JSON không hợp lệ"
        }, status=400)
    except Exception as e:
        logger.error(f"Error in api_xnk_delete: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


# ========================= SUPPLIER MANAGEMENT =========================

@admin_only
def supplier_list(request: HttpRequest):
    """
    Danh sách nhà cung cấp:
    - Hiển thị danh sách nhà cung cấp đang hoạt động
    - Tên - Logo - Vùng (Tỉnh) - Link web
    - Tags loại phân phối hay độc quyền
    - Mô tả
    - Số sản phẩm thuộc nhà cung cấp đó
    """
    context = {
        "title": "Danh sách nhà cung cấp",
        "suppliers": [],
        "total": 0,
    }

    try:
        from products.services.sapo_supplier_service import SapoSupplierService
        
        sapo_client = get_sapo_client()
        supplier_service = SapoSupplierService(sapo_client)

        # Lấy tất cả suppliers đang hoạt động
        all_suppliers = supplier_service.get_all_suppliers(status="active")
        
        # Bổ sung số lượng sản phẩm
        all_suppliers = supplier_service.enrich_suppliers_with_product_count(all_suppliers)
        
        # Tính gợi ý nhập hàng theo NSX (chỉ từ database, không tính toán lại)
        supplier_purchase_suggestions = {}
        try:
            from products.services.sales_forecast_service import SalesForecastService
            forecast_service = SalesForecastService(sapo_client)
            
            # Chỉ load từ database và tính gợi ý nhập (không tính toán lại forecast)
            supplier_purchase_suggestions = forecast_service.calculate_supplier_purchase_suggestions_from_db(
                days=30
            )
        except Exception as e:
            logger.warning(f"Error calculating supplier purchase suggestions: {e}", exc_info=True)
        
        # Convert to dict for template
        suppliers_data = []
        group_names_set = set()
        
        for supplier in all_suppliers:
            if supplier.group_name:
                group_names_set.add(supplier.group_name)
            
            # Parse websites
            websites_dict = supplier.websites_dict
            
            # Tìm gợi ý nhập hàng cho NSX này (match theo code hoặc name)
            purchase_suggestion = None
            supplier_code_upper = (supplier.code or "").upper()
            supplier_name_upper = (supplier.name or "").upper()
            
            # Tìm match trong supplier_purchase_suggestions
            for brand_name, suggestion_data in supplier_purchase_suggestions.items():
                brand_upper = brand_name.upper()
                if brand_upper == supplier_code_upper or brand_upper == supplier_name_upper:
                    purchase_suggestion = suggestion_data
                    break
            
            suppliers_data.append({
                "id": supplier.id,
                "code": supplier.code,
                "name": supplier.name,
                "description": supplier.description,
                "logo_path": supplier.logo_path,
                "province": supplier.province,
                "company_name": supplier.company_name,
                "group_name": supplier.group_name,
                "tags": supplier.tags,
                "websites": websites_dict,
                "product_count": supplier.product_count,
                "debt": supplier.debt,
                "status": supplier.status,
                "purchase_suggestion": purchase_suggestion,  # Thêm gợi ý nhập hàng
            })

        # Sắp xếp: ưu tiên ĐỘC QUYỀN trước PHÂN PHỐI
        def sort_key(supplier):
            group_name = supplier.get("group_name", "") or ""
            # ĐỘC QUYỀN = 0 (ưu tiên cao nhất), PHÂN PHỐI = 1, khác = 2
            if group_name == "ĐỘC QUYỀN":
                return (0, supplier.get("name", ""))
            elif group_name == "PHÂN PHỐI":
                return (1, supplier.get("name", ""))
            else:
                return (2, supplier.get("name", ""))
        
        suppliers_data.sort(key=sort_key)

        context["suppliers"] = suppliers_data
        context["total"] = len(suppliers_data)
        context["group_names"] = sorted(list(group_names_set))

    except Exception as e:
        logger.error(f"Error in supplier_list: {e}", exc_info=True)
        context["error"] = str(e)

    return render(request, "products/supplier_list.html", context)


@admin_only
@require_POST
@csrf_exempt
def upload_supplier_logo(request: HttpRequest, supplier_id: int):
    """
    API endpoint để upload logo cho nhà cung cấp.
    
    POST data:
    - file: File ảnh (jpg hoặc png)
    
    Returns:
        JSON với status và logo_path
    """
    try:
        from django.conf import settings
        import os
        from pathlib import Path
        
        # Kiểm tra file được upload
        if 'file' not in request.FILES:
            return JsonResponse({
                "status": "error",
                "message": "Không có file được upload"
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Kiểm tra định dạng file
        allowed_extensions = ['.jpg', '.jpeg', '.png']
        file_ext = os.path.splitext(uploaded_file.name)[1].lower()
        if file_ext not in allowed_extensions:
            return JsonResponse({
                "status": "error",
                "message": f"Chỉ chấp nhận file {', '.join(allowed_extensions)}"
            }, status=400)
        
        # Lấy thông tin supplier từ SAPO
        sapo_client = get_sapo_client()
        supplier_response = sapo_client.core.get_supplier_raw(supplier_id)
        supplier_data = supplier_response.get('supplier')
        
        if not supplier_data:
            return JsonResponse({
                "status": "error",
                "message": f"Không tìm thấy nhà cung cấp với ID {supplier_id}"
            }, status=404)
        
        supplier_code = supplier_data.get('code', '')
        
        # Tạo tên file: id_code.jpg hoặc id_code.png
        file_name = f"{supplier_id}_{supplier_code}{file_ext}"
        
        # Tạo đường dẫn thư mục: assets/supplier/logo/
        logo_dir = Path(settings.BASE_DIR) / 'assets' / 'supplier' / 'logo'
        logo_dir.mkdir(parents=True, exist_ok=True)
        
        # Đường dẫn đầy đủ của file
        file_path = logo_dir / file_name
        
        # Lưu file
        with open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        
        # Đường dẫn URL để hiển thị: /static/supplier/logo/id_code.jpg
        logo_url = f"/static/supplier/logo/{file_name}"
        
        # Update supplier address trong SAPO
        # Lấy address đầu tiên (hoặc tạo mới nếu chưa có)
        addresses = supplier_data.get('addresses', [])
        if not addresses:
            return JsonResponse({
                "status": "error",
                "message": "Nhà cung cấp chưa có địa chỉ. Vui lòng thêm địa chỉ trước."
            }, status=400)
        
        address = addresses[0]
        address_id = address.get('id')
        
        if not address_id:
            return JsonResponse({
                "status": "error",
                "message": "Địa chỉ không có ID"
            }, status=400)
        
        # Build address payload với tất cả fields hiện có + update first_name
        address_payload = {
            "id": address_id,
            "country": address.get("country"),
            "city": address.get("city"),
            "district": address.get("district"),
            "ward": address.get("ward"),
            "address1": address.get("address1"),  # Giữ nguyên tỉnh
            "address2": address.get("address2"),
            "zip_code": address.get("zip_code"),
            "email": address.get("email"),
            "first_name": logo_url,  # ⭐ Update logo path vào đây
            "last_name": address.get("last_name"),
            "full_name": address.get("full_name"),  # Tên công ty
            "label": address.get("label"),  # Không dùng đến
            "phone_number": address.get("phone_number"),
            "status": address.get("status", "active"),
        }
        
        # Gọi API update address
        try:
            sapo_client.core.update_supplier_address(
                supplier_id=supplier_id,
                address_id=address_id,
                address_data=address_payload
            )
            logger.info(f"[upload_supplier_logo] ✅ Updated logo for supplier {supplier_id}: {logo_url}")
        except Exception as e:
            logger.error(f"[upload_supplier_logo] Failed to update supplier address: {e}", exc_info=True)
            # Xóa file đã lưu nếu update thất bại
            if file_path.exists():
                file_path.unlink()
            return JsonResponse({
                "status": "error",
                "message": f"Không thể cập nhật địa chỉ trên SAPO: {str(e)}"
            }, status=500)
        
        return JsonResponse({
            "status": "success",
            "message": "Upload logo thành công",
            "logo_path": logo_url
        })
        
    except Exception as e:
        logger.error(f"Error in upload_supplier_logo: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
@csrf_exempt
def add_supplier_website(request: HttpRequest, supplier_id: int):
    """
    API endpoint để thêm website cho nhà cung cấp.
    
    POST data (JSON):
    - website_type: Loại website (1688, tmall, taobao, douyin, other)
    - website_url: URL của website
    
    Returns:
        JSON với status và websites dict
    """
    try:
        import json
        
        # Lấy dữ liệu từ request
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        website_type = data.get('website_type', '').strip()
        website_url = data.get('website_url', '').strip()
        
        if not website_type or not website_url:
            return JsonResponse({
                "status": "error",
                "message": "Thiếu website_type hoặc website_url"
            }, status=400)
        
        # Validate URL
        if not website_url.startswith(('http://', 'https://')):
            website_url = 'https://' + website_url
        
        # Lấy thông tin supplier từ SAPO
        sapo_client = get_sapo_client()
        supplier_response = sapo_client.core.get_supplier_raw(supplier_id)
        supplier_data = supplier_response.get('supplier')
        
        if not supplier_data:
            return JsonResponse({
                "status": "error",
                "message": f"Không tìm thấy nhà cung cấp với ID {supplier_id}"
            }, status=404)
        
        # Parse website hiện tại (có thể là JSON string hoặc string đơn giản)
        current_website = supplier_data.get('website', '')
        websites_dict = {}
        
        if current_website:
            try:
                # Thử parse JSON
                if current_website.startswith('{'):
                    websites_dict = json.loads(current_website)
                else:
                    # Nếu là string đơn giản, lưu vào key "default"
                    websites_dict = {"default": current_website}
            except:
                # Nếu không parse được, coi như string đơn giản
                websites_dict = {"default": current_website}
        
        # Thêm website mới vào dict
        # Nếu là "other", dùng URL làm key (hoặc tên domain)
        if website_type == 'other':
            # Extract domain từ URL để làm key
            from urllib.parse import urlparse
            parsed = urlparse(website_url)
            key = parsed.netloc.replace('www.', '') or 'other'
            websites_dict[key] = website_url
        else:
            websites_dict[website_type] = website_url
        
        # Convert dict thành JSON string
        new_website_json = json.dumps(websites_dict, ensure_ascii=False)
        
        # Build supplier update payload với tất cả fields hiện có
        supplier_payload = {
            "id": supplier_data.get("id"),
            "tenant_id": supplier_data.get("tenant_id"),
            "code": supplier_data.get("code"),
            "name": supplier_data.get("name"),
            "description": supplier_data.get("description"),
            "email": supplier_data.get("email"),
            "fax": supplier_data.get("fax"),
            "phone_number": supplier_data.get("phone_number"),
            "tax_number": supplier_data.get("tax_number"),
            "website": new_website_json,  # ⭐ Update website field
            "supplier_group_id": supplier_data.get("supplier_group_id"),
            "assignee_id": supplier_data.get("assignee_id"),
            "default_payment_term_id": supplier_data.get("default_payment_term_id"),
            "default_payment_method_id": supplier_data.get("default_payment_method_id"),
            "default_tax_type_id": supplier_data.get("default_tax_type_id"),
            "default_discount_rate": supplier_data.get("default_discount_rate"),
            "default_price_list_id": supplier_data.get("default_price_list_id"),
            "tags": supplier_data.get("tags", []),
            "status": supplier_data.get("status", "active"),
            "is_default": supplier_data.get("is_default", False),
        }
        
        # Gọi API update supplier
        try:
            sapo_client.core.update_supplier(
                supplier_id=supplier_id,
                supplier_data=supplier_payload
            )
            logger.info(f"[add_supplier_website] ✅ Added website for supplier {supplier_id}: {website_type}={website_url}")
        except Exception as e:
            logger.error(f"[add_supplier_website] Failed to update supplier: {e}", exc_info=True)
            return JsonResponse({
                "status": "error",
                "message": f"Không thể cập nhật nhà cung cấp trên SAPO: {str(e)}"
            }, status=500)
        
        return JsonResponse({
            "status": "success",
            "message": "Thêm website thành công",
            "websites": websites_dict
        })
        
    except Exception as e:
        logger.error(f"Error in add_supplier_website: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


# ========================= SALES FORECAST =========================

@admin_only
def sales_forecast_list(request: HttpRequest):
    """
    Danh sách dự báo bán hàng & cảnh báo tồn kho:
    - Hiển thị danh sách variants với dự báo bán hàng
    - Có nút refresh data để tính toán lại
    - Cảnh báo màu: đỏ (< 30 ngày), vàng (30-60 ngày), xanh (> 60 ngày)
    - Có thể sắp xếp theo các cột
    """
    context = {
        "title": "Dự báo bán hàng & Cảnh báo tồn kho",
        "variants": [],
        "total": 0,
        "brands_map": {},  # Khởi tạo brands_map để tránh lỗi template
        "brand_filter": None,  # Khởi tạo brand_filter
    }
    
    try:
        from products.services.sales_forecast_service import SalesForecastService
        
        sapo_client = get_sapo_client()
        forecast_service = SalesForecastService(sapo_client)
        
        # Lấy brand filter từ query param (mặc định None = tất cả)
        brand_filter = request.GET.get('brand', '').strip()
        if brand_filter == 'all' or not brand_filter:
            brand_filter = None
        context["brand_filter"] = brand_filter
        
        # KHÔNG force refresh mặc định - chỉ load từ database
        # Chỉ refresh khi bấm nút "Sync Data"
        force_refresh = False
        
        # Load dữ liệu cho cả 30 ngày và 10 ngày
        logger.info(f"[sales_forecast_list] Loading forecast data for 30 days and 10 days")
        forecast_map_30, all_products_30, all_variants_map_30 = forecast_service.calculate_sales_forecast(
            days=30,
            force_refresh=force_refresh
        )
        forecast_map_10, all_products_10, all_variants_map_10 = forecast_service.calculate_sales_forecast(
            days=10,
            force_refresh=force_refresh
        )
        
        # Dùng all_products và all_variants_map từ 30 ngày (giống nhau)
        all_products = all_products_30
        all_variants_map = all_variants_map_30
        
        # Lấy thông tin đầy đủ cho từng variant (dùng variant_data đã có sẵn)
        all_variant_ids = set(forecast_map_30.keys()) | set(forecast_map_10.keys())
        print(f"[DEBUG] [VIEW] Lấy thông tin tồn kho cho {len(all_variant_ids)} variants...")
        import time
        view_start = time.time()
        
        # Tạo map variant_id -> product_data để lấy brand
        variant_to_product: Dict[int, Dict[str, Any]] = {}
        for product in all_products:
            product_id = product.get("id")
            variants = product.get("variants", [])
            for variant in variants:
                variant_id = variant.get("id")
                if variant_id:
                    variant_to_product[variant_id] = product
        
        # Collect brands từ SUPPLIERS ĐANG HOẠT ĐỘNG (nhanh hơn, chỉ lấy active suppliers)
        # Học từ suppliers page: lấy suppliers active, rồi lấy brands từ đó
        from products.services.sapo_supplier_service import SapoSupplierService
        supplier_service = SapoSupplierService(sapo_client)
        
        print(f"[DEBUG] [VIEW] Lấy suppliers đang hoạt động...")
        active_suppliers = supplier_service.get_all_suppliers(status="active")
        
        # Tạo set các brand names từ suppliers (code và name, case-insensitive)
        active_brand_names = set()
        for supplier in active_suppliers:
            if supplier.code:
                active_brand_names.add(supplier.code.upper())
            if supplier.name:
                active_brand_names.add(supplier.name.upper())
        
        print(f"[DEBUG] [VIEW] Có {len(active_suppliers)} suppliers đang hoạt động, {len(active_brand_names)} brand names")
        
        # Collect brands từ variants, nhưng CHỈ giữ lại những brands có trong suppliers active
        all_brands_map = {}
        for variant_id, variant_data in all_variants_map.items():
            # Ưu tiên lấy brand từ product_data, fallback về variant_data
            product_data = variant_to_product.get(variant_id)
            brand = ""
            if product_data:
                brand = product_data.get("brand") or ""
            if not brand:
                brand = variant_data.get("brand") or ""
            brand = brand.strip()
            
            # CHỈ đếm nếu brand có trong suppliers active (case-insensitive)
            if brand and brand.upper() in active_brand_names:
                all_brands_map[brand] = all_brands_map.get(brand, 0) + 1
        
        print(f"[DEBUG] [VIEW] ✅ Đã collect {len(all_brands_map)} brands (chỉ từ suppliers active) từ {len(all_variants_map)} variants")
        if all_brands_map:
            print(f"[DEBUG] [VIEW] Brands: {list(all_brands_map.keys())[:10]}")
        
        variants_data = []
        processed = 0
        
        for variant_id in all_variant_ids:
            # Lấy forecast cho 30 ngày và 10 ngày
            forecast_30 = forecast_map_30.get(variant_id)
            forecast_10 = forecast_map_10.get(variant_id)
            
            # Lấy variant_data từ map (đã có sẵn inventories)
            variant_data = all_variants_map.get(variant_id)
            product_data = variant_to_product.get(variant_id)
            
            # Dùng forecast_30 làm chính (hoặc forecast_10 nếu không có 30)
            main_forecast = forecast_30 or forecast_10
            if not main_forecast:
                continue
                
            variant_info = forecast_service.get_variant_forecast_with_inventory(
                variant_id,
                {variant_id: main_forecast},  # Truyền map với 1 item
                variant_data=variant_data,  # Truyền variant_data đã có sẵn
                product_data=product_data,  # Truyền product_data để lấy brand
                forecast_30=forecast_30  # Truyền forecast_30 để tính gợi ý nhập hàng
            )
            
            # Tính tỉ lệ % cho 30 ngày và 10 ngày
            if forecast_30:
                # Tính ASP (Average Selling Price) = revenue / total_sold
                asp_value = None
                if forecast_30.revenue and forecast_30.revenue > 0 and forecast_30.total_sold and forecast_30.total_sold > 0:
                    asp_value = forecast_30.revenue / forecast_30.total_sold
                
                variant_info["forecast_30"] = {
                    "total_sold": forecast_30.total_sold,
                    "total_sold_previous_period": forecast_30.total_sold_previous_period,
                    "growth_percentage": forecast_30.growth_percentage,
                    "sales_rate": forecast_30.sales_rate,
                    "abc_category": forecast_30.abc_category,
                    "revenue": forecast_30.revenue,
                    "revenue_percentage": forecast_30.revenue_percentage,
                    "cumulative_percentage": forecast_30.cumulative_percentage,
                    "abc_rank": forecast_30.abc_rank,
                    "priority_score": forecast_30.priority_score,
                    "velocity_stability_score": forecast_30.velocity_stability_score,
                    "velocity_score": forecast_30.velocity_score,
                    "stability_bonus": forecast_30.stability_bonus,
                    "asp_score": forecast_30.asp_score,
                    "asp_value": asp_value,  # Giá trị ASP thực tế (VNĐ)
                    "revenue_contribution_score": forecast_30.revenue_contribution_score,
                }
                # Tính tỉ lệ %: (hiện tại - cùng kỳ) / cùng kỳ * 100
                if forecast_30.total_sold_previous_period > 0:
                    variant_info["growth_percentage_30"] = ((forecast_30.total_sold - forecast_30.total_sold_previous_period) / forecast_30.total_sold_previous_period) * 100
                else:
                    variant_info["growth_percentage_30"] = None
            else:
                variant_info["forecast_30"] = None
                variant_info["growth_percentage_30"] = None
                
            if forecast_10:
                # Tính ASP (Average Selling Price) = revenue / total_sold cho 10N
                asp_value_10 = None
                if forecast_10.revenue and forecast_10.revenue > 0 and forecast_10.total_sold and forecast_10.total_sold > 0:
                    asp_value_10 = forecast_10.revenue / forecast_10.total_sold
                
                variant_info["forecast_10"] = {
                    "total_sold": forecast_10.total_sold,
                    "total_sold_previous_period": forecast_10.total_sold_previous_period,
                    "growth_percentage": forecast_10.growth_percentage,
                    "sales_rate": forecast_10.sales_rate,
                    "revenue": forecast_10.revenue,  # Revenue cho 10N (nếu có)
                    "asp_value": asp_value_10,  # Giá trị ASP thực tế (VNĐ) cho 10N
                }
                # Tính tỉ lệ %: (hiện tại - cùng kỳ) / cùng kỳ * 100
                if forecast_10.total_sold_previous_period > 0:
                    variant_info["growth_percentage_10"] = ((forecast_10.total_sold - forecast_10.total_sold_previous_period) / forecast_10.total_sold_previous_period) * 100
                else:
                    variant_info["growth_percentage_10"] = None
            else:
                variant_info["forecast_10"] = None
                variant_info["growth_percentage_10"] = None
            
            variants_data.append(variant_info)
            processed += 1
            
            # Log progress mỗi 100 variants
            if processed % 100 == 0:
                print(f"[DEBUG] [VIEW] Đã xử lý {processed}/{len(all_variant_ids)} variants...")
        
        print(f"[DEBUG] [VIEW] ✅ Đã lấy thông tin cho {len(variants_data)} variants ({time.time() - view_start:.2f}s)")
        
        # Filter theo brand (server-side)
        total_before_filter = len(variants_data)
        if brand_filter:
            print(f"[DEBUG] [VIEW] Filter theo brand: {brand_filter}")
            variants_data = [v for v in variants_data if v.get("brand", "").strip() == brand_filter]
            print(f"[DEBUG] [VIEW] Sau khi filter: {len(variants_data)}/{total_before_filter} variants")
        
        # Sắp xếp mặc định: theo days_remaining (tăng dần - nguy hiểm nhất trước)
        print(f"[DEBUG] [VIEW] Sắp xếp danh sách...")
        sort_start = time.time()
        variants_data.sort(key=lambda x: (
            0 if x["days_remaining"] == float('inf') else x["days_remaining"],
            -x["total_inventory"]
        ))
        print(f"[DEBUG] [VIEW] ✅ Đã sắp xếp ({time.time() - sort_start:.2f}s)")
        
        context["variants"] = variants_data
        context["total"] = len(variants_data)
        context["brands_map"] = all_brands_map  # Tất cả brands để hiển thị filter buttons
        # Convert dict thành sorted list để dễ iterate trong template
        context["brands_list"] = sorted(all_brands_map.items(), key=lambda x: x[0])  # Sort theo tên brand
        
        # Debug: Log brands_map để kiểm tra
        print(f"[DEBUG] [VIEW] brands_map có {len(all_brands_map)} brands: {list(all_brands_map.keys())[:10]}...")
        print(f"[DEBUG] [VIEW] Context brands_map type: {type(context.get('brands_map'))}, size: {len(context.get('brands_map', {}))}")
        print(f"[DEBUG] [VIEW] brands_list có {len(context['brands_list'])} items")
        
    except Exception as e:
        logger.error(f"Error in sales_forecast_list: {e}", exc_info=True)
        context["error"] = str(e)
        # Đảm bảo brands_map luôn có trong context (ngay cả khi lỗi)
        if "brands_map" not in context:
            context["brands_map"] = {}
    
    return render(request, "products/sales_forecast_list.html", context)


# ========================= CONTAINER TEMPLATE =========================

@admin_only
def container_template_list(request: HttpRequest):
    """
    Danh sách INIT CONTAINER.
    """
    context = {
        "title": "Quản lý Container Template",
        "templates": [],
        "total": 0,
    }

    try:
        from products.services.sales_forecast_service import SalesForecastService
        
        sapo_client = get_sapo_client()
        forecast_service = SalesForecastService(sapo_client)
        
        templates = ContainerTemplate.objects.filter(is_active=True).order_by('code')
        # Preload suppliers và format money cho mỗi template
        for template in templates:
            template.suppliers_list = list(template.suppliers.all().order_by('supplier_name'))
            
            # Lấy danh sách SPOs sử dụng container template này
            template.spos_list = list(template.sum_purchase_orders.all().order_by('-created_at'))
            
            # Format avg_total_amount (e.g. 1.200.000.000 -> 1B2, 300.000.000 -> 300M)
            val = template.avg_total_amount or 0
            if val >= 1_000_000_000:
                billions = int(val // 1_000_000_000)
                remainder = int((val % 1_000_000_000) // 100_000_000)
                if remainder > 0:
                    template.formatted_avg_total = f"{billions}B{remainder}"
                else:
                    template.formatted_avg_total = f"{billions}B"
            elif val >= 1_000_000:
                millions = int(val // 1_000_000)
                template.formatted_avg_total = f"{millions}M"
            else:
                template.formatted_avg_total = f"{int(val):,}"
            
            # Tính gợi ý nhập hàng cho container template
            try:
                suppliers_data = [
                    {
                        "supplier_code": s.supplier_code,
                        "supplier_name": s.supplier_name
                    }
                    for s in template.suppliers_list
                ]
                
                template.purchase_suggestion = forecast_service.calculate_container_template_suggestions(
                    suppliers_data,
                    float(template.volume_cbm)
                )
            except Exception as e:
                logger.warning(f"Error calculating purchase suggestion for template {template.code}: {e}", exc_info=True)
                template.purchase_suggestion = {
                    "current_cbm": 0.0,
                    "percentage": 0.0,
                    "daily_cbm_growth": 0.0,
                    "days_to_full": None,
                    "estimated_date": None,
                }

        context["templates"] = templates
        context["total"] = templates.count()
    except Exception as e:
        logger.error(f"Error in container_template_list: {e}", exc_info=True)
        context["error"] = str(e)

    return render(request, "products/container_template_list.html", context)


@admin_only
def container_template_detail(request: HttpRequest, template_id: int):
    """
    Chi tiết INIT CONTAINER.
    """
    context = {
        "title": "Chi tiết Container Template",
        "template": None,
        "suppliers": [],
    }
    
    try:
        from products.services.container_template_service import ContainerTemplateService
        
        sapo_client = get_sapo_client()
        template_service = ContainerTemplateService(sapo_client)
        
        template_data = template_service.get_template_with_suppliers(template_id)
        context["template"] = template_data["template"]
        context["suppliers"] = template_data["suppliers"]
    except ContainerTemplate.DoesNotExist:
        context["error"] = "Container template không tồn tại"
    except Exception as e:
        logger.error(f"Error in container_template_detail: {e}", exc_info=True)
        context["error"] = str(e)
    
    return render(request, "products/container_template_detail.html", context)


@admin_only
@require_POST
def create_container_template(request: HttpRequest):
    """
    API endpoint để tạo container template mới.
    
    POST data:
    - code: str (required)
    - name: str (optional)
    - container_type: str (default: "40ft")
    - volume_cbm: float (default: 65.0)
    - default_supplier_id: int (optional)
    - ship_time_avg_hn: int (default: 0)
    - ship_time_avg_hcm: int (default: 0)
    - departure_port: str (optional)
    
    Returns:
        JSON: {status, message, template_id, template_code}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        from products.services.container_template_service import ContainerTemplateService
        
        sapo_client = get_sapo_client()
        template_service = ContainerTemplateService(sapo_client)
        
        # Thêm created_by
        data['created_by'] = request.user
        
        template = template_service.create_template(data)
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã tạo container template {template.code}",
            "template_id": template.id,
            "template_code": template.code
        })
        
    except Exception as e:
        logger.error(f"Error in create_container_template: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def update_container_template(request: HttpRequest, template_id: int):
    """
    API endpoint để cập nhật container template.
    
    POST data:
    - name, container_type, volume_cbm, etc.
    
    Returns:
        JSON: {status, message}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        from products.services.container_template_service import ContainerTemplateService
        
        sapo_client = get_sapo_client()
        template_service = ContainerTemplateService(sapo_client)
        
        template = template_service.update_template(template_id, data)
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã cập nhật container template {template.code}"
        })
        
    except ContainerTemplate.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Container template không tồn tại"
        }, status=404)
    except Exception as e:
        logger.error(f"Error in update_container_template: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def add_supplier_to_container(request: HttpRequest):
    """
    API endpoint để thêm supplier vào container template.
    
    POST data:
    - container_template_id: int
    - supplier_id: int (Sapo supplier_id)
    
    Returns:
        JSON: {status, message, supplier_data}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        template_id = int(data.get('container_template_id'))
        supplier_id = int(data.get('supplier_id'))
        
        from products.services.container_template_service import ContainerTemplateService
        
        sapo_client = get_sapo_client()
        template_service = ContainerTemplateService(sapo_client)
        
        template_supplier = template_service.add_supplier(template_id, supplier_id)
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã thêm supplier {template_supplier.supplier_name} vào container template",
            "container_template_id": template_id,
            "supplier": {
                "id": template_supplier.supplier_id,
                "code": template_supplier.supplier_code,
                "name": template_supplier.supplier_name,
                "logo_path": template_supplier.supplier_logo_path
            },
            "supplier_data": {
                "id": template_supplier.supplier_id,
                "code": template_supplier.supplier_code,
                "name": template_supplier.supplier_name,
                "logo_path": template_supplier.supplier_logo_path
            }
        })
        
    except Exception as e:
        logger.error(f"Error in add_supplier_to_container: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def set_default_supplier(request: HttpRequest, template_id: int):
    """
    API endpoint để đặt NSX mặc định cho container template.
    
    POST data:
    - supplier_id: int (Sapo supplier_id, có thể null để xóa)
    
    Returns:
        JSON: {status, message}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        supplier_id = data.get('supplier_id')
        if supplier_id:
            supplier_id = int(supplier_id)
        else:
            supplier_id = None
        
        template = ContainerTemplate.objects.get(id=template_id)
        
        # Kiểm tra supplier có trong danh sách suppliers của template không
        if supplier_id:
            supplier_exists = template.suppliers.filter(supplier_id=supplier_id).exists()
            if not supplier_exists:
                return JsonResponse({
                    "status": "error",
                    "message": "Supplier này chưa được thêm vào container template"
                }, status=400)
            
            # Lấy thông tin supplier từ ContainerTemplateSupplier
            template_supplier = template.suppliers.get(supplier_id=supplier_id)
            template.default_supplier_id = supplier_id
            template.default_supplier_code = template_supplier.supplier_code
            template.default_supplier_name = template_supplier.supplier_name
        else:
            # Xóa default supplier
            template.default_supplier_id = None
            template.default_supplier_code = ''
            template.default_supplier_name = ''
        
        template.save()
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã {'đặt' if supplier_id else 'xóa'} NSX mặc định cho container template {template.code}"
        })
        
    except ContainerTemplate.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Container template không tồn tại"
        }, status=404)
    except Exception as e:
        logger.error(f"Error in set_default_supplier: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def resync_container_template_stats(request: HttpRequest, template_id: int):
    """
    API endpoint để tính toán lại avg_total_amount và avg_import_cycle_days từ các SPO completed.
    
    POST /products/container-templates/{template_id}/resync-stats/
    
    Returns:
        JSON: {status, message, avg_total_amount, avg_import_cycle_days, spo_count}
    """
    try:
        from products.services.container_template_service import ContainerTemplateService
        from products.models import ContainerTemplate
        
        template = ContainerTemplate.objects.get(id=template_id)
        
        sapo_client = get_sapo_client()
        template_service = ContainerTemplateService(sapo_client)
        
        result = template_service.resync_template_stats(template_id)
        
        if result['status'] == 'success':
            return JsonResponse({
                "status": "success",
                "message": result['message'],
                "avg_total_amount": result['avg_total_amount'],
                "avg_import_cycle_days": result['avg_import_cycle_days'],
                "spo_count": result['spo_count']
            })
        else:
            return JsonResponse({
                "status": "warning",
                "message": result['message'],
                "spo_count": result['spo_count']
            })
            
    except ContainerTemplate.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Container template không tồn tại"
        }, status=404)
    except Exception as e:
        logger.error(f"Error in resync_container_template_stats: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
def get_suppliers_for_select(request: HttpRequest):
    """
    API endpoint để lấy danh sách suppliers với logo (cho select dropdown).
    
    Returns:
        JSON: {status, suppliers: [{id, code, name, logo_path}, ...]}
    """
    try:
        from products.services.sapo_supplier_service import SapoSupplierService
        
        sapo_client = get_sapo_client()
        supplier_service = SapoSupplierService(sapo_client)
        
        # Lấy tất cả suppliers đang hoạt động
        all_suppliers = supplier_service.get_all_suppliers(status="active")
        
        # Convert to simple dict
        suppliers_data = []
        for supplier in all_suppliers:
            suppliers_data.append({
                "id": supplier.id,
                "code": supplier.code,
                "name": supplier.name,
                "logo_path": supplier.logo_path or "",
            })
        
        # Sort by name
        suppliers_data.sort(key=lambda x: x["name"])
        
        return JsonResponse({
            "status": "success",
            "suppliers": suppliers_data
        })
        
    except Exception as e:
        logger.error(f"Error in get_suppliers_for_select: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def remove_supplier_from_container(request: HttpRequest):
    """
    API endpoint để xóa supplier khỏi container template.
    
    POST data:
    - container_template_id: int
    - supplier_id: int
    
    Returns:
        JSON: {status, message}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        template_id = int(data.get('container_template_id'))
        supplier_id = int(data.get('supplier_id'))
        
        from products.services.container_template_service import ContainerTemplateService
        
        sapo_client = get_sapo_client()
        template_service = ContainerTemplateService(sapo_client)
        
        template_service.remove_supplier(template_id, supplier_id)
        
        return JsonResponse({
            "status": "success",
            "message": "Đã xóa supplier khỏi container template"
        })
        
    except Exception as e:
        logger.error(f"Error in remove_supplier_from_container: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


# ========================= SUM PURCHASE ORDER (SPO) =========================

@admin_only
def sum_purchase_order_list(request: HttpRequest):
    """
    Danh sách SPO.
    """
    context = {
        "title": "Quản lý Đợt Nhập Container (SPO)",
        "spos": [],
        "total": 0,
        "container_templates": [],
        "status_choices": SumPurchaseOrder.STATUS_CHOICES,
    }
    
    try:
        # Lấy container templates để hiển thị trong form tạo SPO
        context["container_templates"] = ContainerTemplate.objects.filter(is_active=True).order_by('code')
        
        # Get all SPOs
        from django.db.models import Count
        from products.services.spo_po_service import SPOPOService
        from products.services.sapo_supplier_service import SapoSupplierService
        
        # Order by created_date (nếu có) hoặc created_at (fallback)
        # Lưu ý: Cần chạy migration 0013_sumpurchaseorder_created_date.py trên server Ubuntu
        try:
            spos = SumPurchaseOrder.objects.select_related('container_template') \
                .prefetch_related('spo_purchase_orders') \
                .annotate(
                    po_count=Count('spo_purchase_orders', distinct=True)
                ).order_by('created_date', 'created_at')
        except Exception as e:
            # Fallback nếu migration chưa chạy (created_date chưa có trong DB)
            logger.warning(f"Ordering by created_date failed, using created_at only: {e}")
            spos = SumPurchaseOrder.objects.select_related('container_template') \
                .prefetch_related('spo_purchase_orders') \
                .annotate(
                    po_count=Count('spo_purchase_orders', distinct=True)
                ).order_by('created_at')
        
        # Tính toán động từ Sapo cho mỗi SPO
        sapo_client = get_sapo_client()
        spo_po_service = SPOPOService(sapo_client)
        supplier_service = SapoSupplierService(sapo_client)

        # Lấy tất cả active suppliers để map tên, logo
        all_suppliers = supplier_service.get_all_suppliers(status="active")
        supplier_map = {s.id: s for s in all_suppliers}
        
        for spo in spos:
            po_ids = [spo_po.purchase_order.sapo_order_supplier_id for spo_po in spo.spo_purchase_orders.all() if spo_po.purchase_order]
            
            # Init list supplier để hiển thị
            spo.suppliers_display_list = []
            
            if po_ids:
                try:
                    # Tính tổng amount và quantity từ Sapo
                    total_amount = Decimal('0')
                    total_quantity = 0
                    
                    # Aggregate theo supplier
                    spo_supplier_map = {} # supplier_id -> {'name': '', 'logo': '', 'cbm': 0, 'count': 0}
                    
                    for po_id in po_ids:
                        po_data = spo_po_service.get_po_from_sapo(po_id)
                        total_amount += po_data.get('total_amount', Decimal('0'))
                        total_quantity += po_data.get('total_quantity', 0)
                        
                        # Supplier info
                        # API PO trả về supplier_id
                        s_id = po_data.get('supplier_id')
                        s_cbm = po_data.get('total_cpm', Decimal('0'))
                        
                        if s_id:
                            if s_id not in spo_supplier_map:
                                # Lookup info từ supplier_map (đã lấy từ trước)
                                s_info = supplier_map.get(s_id)
                                spo_supplier_map[s_id] = {
                                    'id': s_id,
                                    'name': s_info.name if s_info else po_data.get('supplier_name', f'Supplier {s_id}'),
                                    'logo_path': s_info.logo_path if s_info else '',
                                    'cbm': Decimal('0'),
                                    'count': 0
                                }
                            
                            spo_supplier_map[s_id]['cbm'] += s_cbm
                            spo_supplier_map[s_id]['count'] += 1
                    
                    # Convert map to list & sort by CBM desc
                    spo.suppliers_display_list = list(spo_supplier_map.values())
                    spo.suppliers_display_list.sort(key=lambda x: x['cbm'], reverse=True)
                    
                    # Gán vào SPO object (không lưu DB, chỉ để hiển thị)
                    spo.total_po_amount_val = total_amount
                    spo.total_po_quantity_val = total_quantity
                except Exception as e:
                    logger.warning(f"Error calculating totals for SPO {spo.id}: {e}")
                    spo.total_po_amount_val = Decimal('0')
                    spo.total_po_quantity_val = 0
            else:
                spo.total_po_amount_val = Decimal('0')
                spo.total_po_quantity_val = 0
        
        # Tách SPOs thành 2 nhóm: chưa hoàn tất và đã hoàn tất
        incomplete_spos = [spo for spo in spos if spo.status != 'completed']
        completed_spos = [spo for spo in spos if spo.status == 'completed']
        
        # Sắp xếp completed_spos theo updated_at (mới nhất trước)
        completed_spos.sort(key=lambda x: x.updated_at, reverse=True)
        
        context["incomplete_spos"] = incomplete_spos
        context["completed_spos"] = completed_spos
        context["spos"] = spos  # Giữ lại để tương thích nếu có code cũ
        context["total"] = spos.count()
    except Exception as e:
        logger.error(f"Error in sum_purchase_order_list: {e}", exc_info=True)
        context["error"] = str(e)
    
    return render(request, "products/sum_purchase_order_list.html", context)


@admin_only
def sum_purchase_order_detail(request: HttpRequest, spo_id: int):
    """
    Chi tiết SPO.
    """
    context = {
        "title": "Chi tiết Đợt Nhập Container",
        "spo": None,
        "purchase_orders": [],
        "line_items": [],
    }
    
    try:
        from products.services.sum_purchase_order_service import SumPurchaseOrderService
        from products.services.spo_po_service import SPOPOService
        
        spo = SumPurchaseOrder.objects.select_related('container_template').prefetch_related(
            'costs',
            'documents',
            'packing_list_items'
        ).get(id=spo_id)
        context["spo"] = spo
        context["packing_list_items"] = spo.packing_list_items.all()
        
        # Kiểm tra xem có file TKHQ trong documents không
        # Tìm file có tên bắt đầu bằng "TKHQ" hoặc "ToKhaiHQ"
        has_tkhq = False
        tkhq_document = None
        for doc in spo.documents.all():
            # Kiểm tra doc.name
            if doc.name:
                doc_name_upper = doc.name.upper()
                if doc_name_upper.startswith('TKHQ') or doc_name_upper.startswith('TOKHAIHQ'):
                    has_tkhq = True
                    tkhq_document = doc
                    break
            
            # Kiểm tra filename từ file field
            if doc.file:
                file_name = doc.file.name if hasattr(doc.file, 'name') else str(doc.file)
                file_name_upper = file_name.upper()
                if 'TKHQ' in file_name_upper or 'TOKHAIHQ' in file_name_upper:
                    has_tkhq = True
                    tkhq_document = doc
                    break
        context["has_tkhq_document"] = has_tkhq
        context["tkhq_document"] = tkhq_document
        
        # Lấy danh sách PO từ SPOPurchaseOrder với đầy đủ thông tin từ PurchaseOrder model
        spo_po_relations = spo.spo_purchase_orders.select_related('purchase_order').prefetch_related(
            'purchase_order__costs',
            'purchase_order__payments'
        ).all()
        
        # Lấy thông tin PO từ Sapo API và kết hợp với PurchaseOrder model
        sapo_client = get_sapo_client()
        spo_po_service = SPOPOService(sapo_client)
        
        purchase_orders_data = []
        all_line_items = []
        total_quantity = 0
        total_amount = Decimal('0')
        total_packages = 0
        
        for spo_po_rel in spo_po_relations:
            if not spo_po_rel.purchase_order:
                continue
                
            po = spo_po_rel.purchase_order
            po_id = po.sapo_order_supplier_id
            
            try:
                # Lấy thông tin từ Sapo API
                po_data = spo_po_service.get_po_from_sapo(po_id)
                
                # Cập nhật product_amount_cny từ Sapo API (tính từ price_tq trong metadata)
                product_amount_cny_from_api = float(po_data.get('product_amount_cny', 0))
                if product_amount_cny_from_api > 0:
                    # Cập nhật vào DB nếu khác với giá trị hiện tại
                    if po.product_amount_cny != Decimal(str(product_amount_cny_from_api)):
                        po.product_amount_cny = Decimal(str(product_amount_cny_from_api))
                        po.calculate_total_amount()  # Tính lại total_amount_cny
                        po.save()
                
                # Bổ sung thông tin từ PurchaseOrder model
                po_data['po_id'] = po.id
                po_data['delivery_status'] = po.delivery_status
                po_data['delivery_status_display'] = po.get_delivery_status_display()
                po_data['delivery_timeline'] = po.delivery_timeline
                po_data['expected_delivery_date'] = po.expected_delivery_date
                # Sử dụng giá trị từ API nếu có, nếu không thì dùng từ DB
                po_data['product_amount_cny'] = product_amount_cny_from_api if product_amount_cny_from_api > 0 else float(po.product_amount_cny)
                po_data['total_amount_cny'] = float(po.total_amount_cny)
                po_data['paid_amount_cny'] = float(po.paid_amount_cny)
                po_data['remaining_amount_cny'] = float(po.total_amount_cny - po.paid_amount_cny)
                
                # Lấy costs và gộp theo cost_type (mỗi loại chỉ 1 dòng)
                # Lấy total_cbm từ po_data (đã tính từ Sapo API)
                po_total_cbm = float(po_data.get('total_cbm', 0) or 0)
                
                # Gộp costs theo cost_type
                costs_by_type = {}
                for cost in po.costs.all():
                    cost_type = cost.cost_type
                    if cost_type not in costs_by_type:
                        costs_by_type[cost_type] = {
                            'id': cost.id,  # Giữ ID đầu tiên để delete
                            'cost_type': cost_type,
                            'cost_type_display': cost.get_cost_type_display(),
                            'amount_cny': Decimal('0'),
                            'cbm': None,
                            'description': '',  # Gộp descriptions nếu cần
                            'cost_ids': [],  # Lưu tất cả IDs để delete
                        }
                    
                    costs_by_type[cost_type]['amount_cny'] += cost.amount_cny
                    costs_by_type[cost_type]['cost_ids'].append(cost.id)
                    # Lấy CBM từ cost đầu tiên có CBM
                    if not costs_by_type[cost_type]['cbm'] and cost.cbm:
                        costs_by_type[cost_type]['cbm'] = float(cost.cbm)
                    # Gộp description (nếu có)
                    if cost.description:
                        if costs_by_type[cost_type]['description']:
                            costs_by_type[cost_type]['description'] += '; ' + cost.description
                        else:
                            costs_by_type[cost_type]['description'] = cost.description
                
                # Chuyển sang list và tính giá/m³
                import json
                po_data['costs'] = []
                for cost_type, cost_data in costs_by_type.items():
                    amount_cny = float(cost_data['amount_cny'])
                    
                    # Tính giá/m³: ưu tiên dùng CBM của cost, nếu không có thì dùng total_cbm của PO
                    cbm_for_calc = None
                    if cost_data['cbm'] and cost_data['cbm'] > 0:
                        cbm_for_calc = cost_data['cbm']
                    elif po_total_cbm > 0:
                        cbm_for_calc = po_total_cbm
                    
                    final_cost_data = {
                        'id': cost_data['id'],  # ID đầu tiên (để tương thích với delete)
                        'cost_ids': cost_data['cost_ids'],  # Tất cả IDs để delete (list)
                        'cost_ids_json': json.dumps(cost_data['cost_ids']),  # JSON string cho template
                        'cost_type': cost_type,
                        'cost_type_display': cost_data['cost_type_display'],
                        'amount_cny': amount_cny,
                        'cbm': cost_data['cbm'],
                        'description': cost_data['description'],
                    }
                    
                    if cbm_for_calc and cbm_for_calc > 0:
                        final_cost_data['price_per_cbm'] = round(amount_cny / cbm_for_calc, 2)
                    else:
                        final_cost_data['price_per_cbm'] = None
                    
                    po_data['costs'].append(final_cost_data)
                # Tính tổng phí từ costs
                total_costs_cny = sum(float(cost.amount_cny) for cost in po.costs.all())
                po_data['total_costs_cny'] = total_costs_cny
                
                # Lấy payments
                po_data['payments'] = [
                    {
                        'id': payment.id,
                        'payment_type': payment.payment_type,
                        'payment_type_display': payment.get_payment_type_display(),
                        'amount_cny': float(payment.amount_cny),
                        'amount_vnd': float(payment.amount_vnd) if payment.amount_vnd else None,
                        'exchange_rate': float(payment.exchange_rate) if payment.exchange_rate else None,
                        'payment_date': payment.payment_date,
                        'description': payment.description,
                    }
                    for payment in po.payments.all().order_by('-payment_date')
                ]
                
                purchase_orders_data.append(po_data)
                all_line_items.extend(po_data.get('line_items', []))
                total_quantity += po_data.get('total_quantity', 0)
                # Tính total_packages từ box_info.full_box
                po_packages = 0
                for line_item in po_data.get('line_items', []):
                    quantity = line_item.get('quantity', 0)
                    full_box = line_item.get('full_box', 1)  # Số cái/thùng, mặc định 1
                    if full_box > 0:
                        # Số kiện = quantity / full_box (làm tròn lên)
                        import math
                        boxes = math.ceil(quantity / full_box)
                        po_packages += boxes
                    else:
                        # Nếu không có full_box, mặc định 1 item = 1 package
                        po_packages += quantity
                total_packages += po_packages
            except Exception as e:
                logger.warning(f"Error getting PO {po_id}: {e}")
        
        # Tính tổng giá trị hàng từ CNY của tất cả PO
        total_product_amount_cny = Decimal('0')
        for po_data in purchase_orders_data:
            total_product_amount_cny += Decimal(str(po_data.get('product_amount_cny', 0)))
        
        # Tính tỷ giá trung bình từ các payment của các PO
        all_exchange_rates = []
        for po_data in purchase_orders_data:
            for payment in po_data.get('payments', []):
                if payment.get('exchange_rate'):
                    all_exchange_rates.append(float(payment.get('exchange_rate')))
        
        # Nếu chưa có tỷ giá từ payments, lấy từ 3 PO gần nhất
        if not all_exchange_rates:
            # Lấy 3 PO gần nhất có payments
            recent_pos = PurchaseOrder.objects.filter(
                id__in=[po.get('po_id') for po in purchase_orders_data if po.get('po_id')]
            ).order_by('-created_at')[:3]
            
            for po in recent_pos:
                payments = po.payments.all().order_by('-payment_date')[:3]
                for payment in payments:
                    if payment.exchange_rate:
                        all_exchange_rates.append(float(payment.exchange_rate))
        
        # Tính tỷ giá trung bình
        avg_exchange_rate = None
        if all_exchange_rates:
            avg_exchange_rate = sum(all_exchange_rates) / len(all_exchange_rates)
        
        # Tính giá trị VND (làm tròn thành số nguyên)
        total_amount_vnd = None
        if avg_exchange_rate and total_product_amount_cny > 0:
            total_amount_vnd = Decimal(str(round(float(total_product_amount_cny * Decimal(str(avg_exchange_rate))))))
        
        context["purchase_orders"] = purchase_orders_data
        context["line_items"] = all_line_items
        context["total_packages"] = total_packages
        context["total_quantity"] = total_quantity
        context["total_amount"] = total_amount_vnd if total_amount_vnd else total_amount  # Ưu tiên VND, fallback về total_amount cũ
        context["total_product_amount_cny"] = total_product_amount_cny
        context["avg_exchange_rate"] = avg_exchange_rate
        context["total_amount_vnd"] = total_amount_vnd
        
        # Tính lại total_cbm của SPO
        spo_service = SumPurchaseOrderService(sapo_client)
        spo_service._recalculate_spo_cbm(spo)
        spo.refresh_from_db()
        
        # Lấy ngày dự kiến từ warehouse stage trong timeline
        warehouse_planned_date = None
        if spo.timeline:
            warehouse_stage_name = None
            if spo.destination_port == 'hcm':
                warehouse_stage_name = 'arrived_warehouse_hcm'
            elif spo.destination_port == 'haiphong':
                warehouse_stage_name = 'arrived_warehouse_hn'
            
            if warehouse_stage_name:
                for stage in spo.timeline:
                    if stage.get('stage') == warehouse_stage_name and stage.get('planned_date'):
                        from datetime import datetime
                        try:
                            # Parse date string YYYY-MM-DD
                            d_str = stage['planned_date']
                            if 'T' in d_str:
                                warehouse_planned_date = datetime.fromisoformat(d_str.replace('Z', '+00:00'))
                            else:
                                warehouse_planned_date = datetime.strptime(d_str, '%Y-%m-%d')
                        except:
                            pass
                        break
        
        context["warehouse_planned_date"] = warehouse_planned_date
        
        # Analyze timeline stages for display
        from datetime import datetime, timedelta
        today = datetime.now().date()
        stage_analysis = {}
        
        for stage_data in spo.timeline:
            stage_code = stage_data.get('stage')
            planned_date = stage_data.get('planned_date')
            actual_date = stage_data.get('actual_date')
            
            analysis = {
                'has_actual': bool(actual_date),
                'has_planned': bool(planned_date),
                'tag': None,
                'tag_color': None,
                'days_diff': None
            }
            
            # Parse dates
            planned_dt = None
            actual_dt = None
            
            if planned_date:
                try:
                    if 'T' in planned_date:
                        planned_dt = datetime.fromisoformat(planned_date.replace('Z', '+00:00')).date()
                    else:
                        planned_dt = datetime.strptime(planned_date, '%Y-%m-%d').date()
                except: 
                    pass
            
            if actual_date:
                try:
                    if 'T' in actual_date:
                        actual_dt = datetime.fromisoformat(actual_date.replace('Z', '+00:00')).date()
                    else:
                        actual_dt = datetime.strptime(actual_date, '%Y-%m-%d').date()
                except: 
                    pass
            
            # Generate tags
            if actual_dt:
                # Case A: Completed
                if planned_dt:
                    days_diff = (actual_dt - planned_dt).days
                    if days_diff <= 0:
                        analysis['tag'] = 'Đúng hẹn'
                        analysis['tag_color'] = 'emerald'
                    else:
                        analysis['tag'] = f'Trễ {days_diff} ngày'
                        analysis['tag_color'] = 'red'
                    analysis['days_diff'] = days_diff
            elif planned_dt:
                # Case B: Planned but not completed
                days_diff = (planned_dt - today).days
                if days_diff > 0:
                    analysis['tag'] = f'Còn {days_diff} ngày'
                    analysis['tag_color'] = 'blue'
                elif days_diff < 0:
                    analysis['tag'] = f'Trễ deadline {abs(days_diff)} ngày'
                    analysis['tag_color'] = 'red'
                else:
                    analysis['tag'] = 'Hôm nay'
                    analysis['tag_color'] = 'amber'
                analysis['days_diff'] = days_diff
            
            stage_analysis[stage_code] = analysis
        
        context['stage_analysis'] = stage_analysis
        context['today'] = today
        
        # Lấy SPO Costs (dynamic costs) - sắp xếp theo transaction_date
        spo_costs = spo.costs.all().order_by('-transaction_date', '-created_at')
        context["spo_costs"] = [
            {
                'id': cost.id,
                'name': cost.name,
                'cost_side': cost.cost_side,
                'amount_vnd': float(cost.amount_vnd),
                'amount_cny': float(cost.amount_cny) if cost.amount_cny else None,
                'balance_transaction': cost.balance_transaction,
                'note': cost.note,
                'transaction_date': cost.transaction_date,
                'created_at': cost.created_at,
            }
            for cost in spo_costs
        ]
        # Tính tổng chi phí: VND costs (Vietnam side) + VND equivalent của CNY costs (China side)
        # Tạm thời chỉ tính tổng VND costs, có thể cải thiện sau để quy đổi CNY sang VND
        total_spo_costs = sum(float(cost.amount_vnd) for cost in spo_costs if cost.cost_side == 'vietnam')
        # Có thể thêm tổng CNY costs nếu cần
        total_spo_costs_cny = sum(float(cost.amount_cny) for cost in spo_costs if cost.cost_side == 'china' and cost.amount_cny)
        context["total_spo_costs"] = total_spo_costs
        context["total_spo_costs_cny"] = total_spo_costs_cny
        
        # Lấy SPO Documents
        spo_documents = spo.documents.all().order_by('-uploaded_at')
        context["spo_documents"] = [
            {
                'id': doc.id,
                'name': doc.name or (doc.file.name.split('/')[-1] if doc.file else ''),
                'file': doc.file,
                'file_url': doc.get_file_url() if hasattr(doc, 'get_file_url') else (f'/static/{doc.file.name}' if doc.file else None),
                'file_icon': doc.get_file_icon() if hasattr(doc, 'get_file_icon') else None,
                'uploaded_at': doc.uploaded_at,
                'uploaded_by': doc.uploaded_by.username if doc.uploaded_by else None,
            }
            for doc in spo_documents
        ]
    except SumPurchaseOrder.DoesNotExist:
        context["error"] = "SPO không tồn tại"
    except Exception as e:
        logger.error(f"Error in sum_purchase_order_detail: {e}", exc_info=True)
        context["error"] = str(e)
    
    return render(request, "products/sum_purchase_order_detail.html", context)


@admin_only
@require_POST
def create_sum_purchase_order(request: HttpRequest):
    """
    API endpoint để tạo SPO mới.
    
    POST data:
    - container_template_id: int (required)
    - name: str (optional)
    - destination_port: str ('hcm' hoặc 'haiphong', optional)
    - expected_arrival_date: str (YYYY-MM-DD format, optional)
    - created_date: str (YYYY-MM-DD format, optional)
    
    Returns:
        JSON: {status, message, spo_id, spo_code}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        container_template_id = int(data.get('container_template_id'))
        name = data.get('name', '').strip() or None
        destination_port = data.get('destination_port', '').strip() or None
        expected_arrival_date = data.get('expected_arrival_date', '').strip() or None
        created_date = data.get('created_date', '').strip() or None
        
        from products.services.sum_purchase_order_service import SumPurchaseOrderService
        
        sapo_client = get_sapo_client()
        spo_service = SumPurchaseOrderService(sapo_client)
        
        spo = spo_service.create_spo(
            container_template_id, 
            name=name,
            destination_port=destination_port,
            expected_arrival_date=expected_arrival_date,
            created_date=created_date
        )
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã tạo SPO {spo.code}",
            "spo_id": spo.id,
            "spo_code": spo.code
        })
        
    except Exception as e:
        logger.error(f"Error in create_sum_purchase_order: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def add_po_to_spo(request: HttpRequest):
    """
    API endpoint để thêm PO vào SPO.
    
    POST data:
    - spo_id: int (required)
    - po_ids: List[int] (optional) - Sapo order_supplier IDs
    - tag: str (optional) - Tag để tìm PO
    
    Returns:
        JSON: {status, message, added_count}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        spo_id = int(data.get('spo_id'))
        po_ids = data.get('po_ids', [])
        tag = data.get('tag', '').strip() or None
        domestic_shipping_cn = data.get('domestic_shipping_cn')
        expected_production_date = data.get('expected_production_date')
        expected_delivery_date = data.get('expected_delivery_date')
        
        # Convert po_ids to list of ints
        if po_ids and isinstance(po_ids, str):
            import ast
            po_ids = ast.literal_eval(po_ids)
        if po_ids:
            po_ids = [int(pid) for pid in po_ids]
        
        # Parse dates
        prod_date = None
        deliv_date = None
        if expected_production_date:
            try:
                prod_date = date.fromisoformat(expected_production_date)
            except:
                pass
        if expected_delivery_date:
            try:
                deliv_date = date.fromisoformat(expected_delivery_date)
            except:
                pass
        
        # Parse domestic_shipping_cn
        shipping_cn = Decimal('0')
        if domestic_shipping_cn:
            try:
                shipping_cn = Decimal(str(domestic_shipping_cn))
            except:
                pass
        
        from products.services.sum_purchase_order_service import SumPurchaseOrderService
        
        sapo_client = get_sapo_client()
        spo_service = SumPurchaseOrderService(sapo_client)
        
        spo = spo_service.add_po_to_spo(
            spo_id, 
            po_ids=po_ids, 
            tag=tag,
            domestic_shipping_cn=shipping_cn,
            expected_production_date=prod_date,
            expected_delivery_date=deliv_date
        )
        
        added_count = spo.spo_purchase_orders.count()
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã thêm {added_count} PO vào SPO {spo.code}",
            "added_count": added_count
        })
        
    except Exception as e:
        logger.error(f"Error in add_po_to_spo: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def sync_po_from_sapo(request: HttpRequest):
    """
    API endpoint để lấy thông tin PO từ Sapo API (không lưu DB).
    
    POST data:
    - po_id: int (Sapo order_supplier_id) hoặc
    - tag: str (lấy tất cả PO có tag)
    
    Returns:
        JSON: {status, message, po_data hoặc po_list}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        po_id = data.get('po_id')
        tag = data.get('tag', '').strip() or None
        
        from products.services.sum_purchase_order_service import SumPurchaseOrderService
        
        sapo_client = get_sapo_client()
        spo_service = SumPurchaseOrderService(sapo_client)
        
        from products.services.spo_po_service import SPOPOService
        spo_po_service = SPOPOService(sapo_client)
        
        synced_count = 0
        
        if po_id:
            # Lấy 1 PO từ Sapo
            po_id = int(po_id)
            try:
                po_data = spo_po_service.get_po_from_sapo(po_id)
                return JsonResponse({
                    "status": "success",
                    "message": f"Đã lấy thông tin PO {po_data.get('code')}",
                    "po_data": po_data
                })
            except Exception as e:
                return JsonResponse({
                    "status": "error",
                    "message": str(e)
                }, status=400)
        elif tag:
            # Lấy tất cả PO có tag từ Sapo, chỉ chấp nhận status pending, partial, completed
            # Sapo API chỉ hỗ trợ filter 1 status tại một thời điểm, nên lấy tất cả rồi filter
            all_order_suppliers = []
            for status in ['pending', 'partial', 'completed']:
                try:
                    response = sapo_client.core.list_order_suppliers_raw(tags=tag, status=status, limit=250)
                    order_suppliers = response.get('order_suppliers', [])
                    all_order_suppliers.extend(order_suppliers)
                except Exception as e:
                    logger.warning(f"Error getting PO with status {status}: {e}")
                    continue
            
            # Lọc và loại bỏ trùng lặp
            seen_ids = set()
            po_ids = []
            for os in all_order_suppliers:
                os_id = os.get('id')
                if os_id and os_id not in seen_ids:
                    seen_ids.add(os_id)
                    if os.get('status') in ['pending', 'partial', 'completed']:
                        po_ids.append(os_id)
            
            po_list = []
            for po_id in po_ids:
                try:
                    po_data = spo_po_service.get_po_from_sapo(po_id)
                    po_list.append(po_data)
                    synced_count += 1
                except Exception as e:
                    logger.warning(f"Error getting PO {po_id}: {e}")
            
            return JsonResponse({
                "status": "success",
                "message": f"Đã lấy {synced_count} PO từ Sapo",
                "po_list": po_list,
                "synced_count": synced_count
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": "Phải cung cấp po_id hoặc tag"
            }, status=400)
        
    except Exception as e:
        logger.error(f"Error in sync_po_from_sapo: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def update_spo_status(request: HttpRequest):
    """
    API endpoint để cập nhật trạng thái SPO.
    
    POST data:
    - spo_id: int
    - status: str
    - actual_date: str (ISO datetime, optional)
    - note: str (optional)
    
    Returns:
        JSON: {status, message}
    """
    try:
        import json
        from datetime import datetime
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        spo_id = int(data.get('spo_id'))
        new_status = data.get('status')
        actual_date_str = data.get('actual_date', '').strip()
        note = data.get('note', '').strip()
        
        spo = SumPurchaseOrder.objects.get(id=spo_id)
        
        # Validate status
        valid_statuses = [choice[0] for choice in SumPurchaseOrder.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return JsonResponse({
                "status": "error",
                "message": f"Status không hợp lệ. Phải là một trong: {', '.join(valid_statuses)}"
            }, status=400)
        
        # Parse actual_date nếu có
        actual_date = None
        if actual_date_str:
            try:
                actual_date = datetime.fromisoformat(actual_date_str.replace('Z', '+00:00'))
            except ValueError:
                actual_date = None
        
        # Update status
        spo.update_status(new_status, actual_date=actual_date, note=note)
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã cập nhật trạng thái SPO {spo.code} thành {spo.get_status_display()}"
        })
        
    except SumPurchaseOrder.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "SPO không tồn tại"
        }, status=404)
    except Exception as e:
        logger.error(f"Error in update_spo_status: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def update_timeline_planned_date(request: HttpRequest):
    """
    API endpoint để cập nhật planned_date cho một stage trong timeline.
    
    POST data:
    - spo_id: int
    - stage: str (stage name)
    - planned_date: str (YYYY-MM-DD format)
    
    Returns:
        JSON: {status, message}
    """
    try:
        import json
        from datetime import date
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        spo_id = int(data.get('spo_id'))
        stage_name = data.get('stage')
        planned_date_str = data.get('planned_date', '').strip()
        
        if not stage_name:
            return JsonResponse({
                "status": "error",
                "message": "Stage name là bắt buộc"
            }, status=400)
        
        if not planned_date_str:
            return JsonResponse({
                "status": "error",
                "message": "Planned date là bắt buộc"
            }, status=400)
        
        spo = SumPurchaseOrder.objects.get(id=spo_id)
        
        # Parse planned_date
        try:
            planned_date = date.fromisoformat(planned_date_str)
        except ValueError:
            return JsonResponse({
                "status": "error",
                "message": f"Ngày không hợp lệ: {planned_date_str}. Vui lòng dùng format YYYY-MM-DD"
            }, status=400)
        
        # Custom Logic: Sync Expected Arrival Date
        # Nếu stage là 'arrived_vn' hoặc 'completed', cập nhật luôn vào model field
        if stage_name in ['arrived_vn', 'completed']:
             spo.expected_arrival_date = planned_date
             # Không cần lưu vào timeline JSON để tránh duplicate dữ liệu
             # Nhưng để đảm bảo hiển thị đúng ở các chỗ khác, ta vẫn có thể lưu hoặc chỉ lưu model field.
             # Theo yêu cầu "Bỏ 1 trong 2", ta sẽ ưu tiên lưu vào model field.
             # Tuy nhiên, timeline cần record để hiển thị trên UI nếu UI loop qua timeline.
             # Giải pháp: UI sẽ check model field cho stage này. Timeline chỉ lưu record để biết là đã có plan.
             
        # Tìm và cập nhật stage trong timeline
        stage_found = False
        for stage in spo.timeline:
            if stage.get('stage') == stage_name:
                stage['planned_date'] = planned_date.isoformat()
                stage_found = True
                break
        
        # Nếu chưa có trong timeline, thêm mới (để đảm bảo hiển thị trên UI)
        if not stage_found:
             spo.timeline.append({
                'stage': stage_name,
                'planned_date': planned_date.isoformat(),
                'actual_date': None,
                'note': ""
            })
        
        spo.save()
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã cập nhật ngày dự kiến cho stage {stage_name}"
        })
        
    except SumPurchaseOrder.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "SPO không tồn tại"
        }, status=404)
    except Exception as e:
        logger.error(f"Error in update_timeline_planned_date: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def allocate_costs(request: HttpRequest):
    """
    API endpoint để phân bổ chi phí chung của SPO.
    
    POST data:
    - spo_id: int
    
    Returns:
        JSON: {status, message, allocation_details}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        spo_id = int(data.get('spo_id'))
        
        from products.services.sum_purchase_order_service import SumPurchaseOrderService
        
        sapo_client = get_sapo_client()
        spo_service = SumPurchaseOrderService(sapo_client)
        
        result = spo_service.calculate_cost_allocation(spo_id)
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"Error in allocate_costs: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def add_spo_cost(request: HttpRequest):
    """
    API endpoint để thêm chi phí SPO.
    
    POST data:
    - spo_id: int
    - name: str
    - cost_side: str ('china' hoặc 'vietnam')
    - amount: float (CNY nếu cost_side='china', VND nếu cost_side='vietnam')
    - note: str (optional)
    
    Returns:
        JSON: {status, message, cost_id}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        spo_id = int(data.get('spo_id'))
        name = data.get('name', '').strip()
        cost_side = data.get('cost_side', 'china').strip()
        note = data.get('note', '').strip()
        transaction_date_str = data.get('transaction_date', '').strip()
        
        if not name:
            return JsonResponse({
                "status": "error",
                "message": "Tên chi phí không được để trống"
            }, status=400)
        
        if cost_side not in ['china', 'vietnam']:
            return JsonResponse({
                "status": "error",
                "message": "Phía chi phí không hợp lệ. Phải là 'china' hoặc 'vietnam'"
            }, status=400)
        
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        
        amount_cny_decimal = None
        amount_vnd_decimal = Decimal('0')
        balance_txn = None
        
        if cost_side == 'china':
            # Phía Trung Quốc: dùng CNY, trừ từ số dư
            amount_cny_str = data.get('amount', '').strip().replace(',', '')
            if not amount_cny_str:
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền CNY không được để trống cho chi phí phía Trung Quốc"
                }, status=400)
            
            try:
                amount_cny_decimal = Decimal(str(amount_cny_str))
            except (ValueError, InvalidOperation):
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền CNY không hợp lệ"
                }, status=400)
            
            if amount_cny_decimal <= 0:
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền CNY phải lớn hơn 0"
                }, status=400)
            
            # Kiểm tra số dư
            from products.models import BalanceTransaction
            from django.db.models import Sum
            
            current_balance = BalanceTransaction.objects.aggregate(
                total=Sum('amount_cny')
            )['total'] or Decimal('0')
            
            if current_balance < amount_cny_decimal:
                return JsonResponse({
                    "status": "error",
                    "message": f"Số dư không đủ. Số dư hiện tại: {current_balance} CNY, cần: {amount_cny_decimal} CNY"
                }, status=400)
            
            # Parse transaction_date
            transaction_date = timezone.now().date()
            if transaction_date_str:
                try:
                    transaction_date = datetime.strptime(transaction_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({
                        "status": "error",
                        "message": "Định dạng ngày giao dịch không hợp lệ. Phải là YYYY-MM-DD"
                    }, status=400)
            
            # Tạo giao dịch số dư (rút)
            balance_txn = BalanceTransaction.objects.create(
                transaction_type='withdraw_spo_cost',
                amount_cny=-amount_cny_decimal,  # Số âm vì là rút
                transaction_date=transaction_date,
                description=f"Chi phí SPO {spo.code}: {name} (Phía Trung Quốc)",
                created_by=request.user if request.user.is_authenticated else None
            )
            
            # Tự động liên kết với PaymentPeriod dựa trên datetime (tính cả giờ)
            from products.models import PaymentPeriod, PaymentPeriodTransaction
            # Sử dụng created_at của balance_txn để có datetime chính xác
            # Đây là giao dịch rút (withdraw_spo_cost), truyền is_withdraw=True
            period = PaymentPeriod.find_period_for_transaction_datetime(balance_txn.created_at, is_withdraw=True)
            
            if period:
                PaymentPeriodTransaction.objects.get_or_create(
                    payment_period=period,
                    balance_transaction=balance_txn
                )
                # Số dư cuối kỳ tính realtime, không cần lưu
            
            # Không có amount_vnd cho chi phí phía Trung Quốc (chỉ ghi nhận bằng CNY)
            amount_vnd_decimal = Decimal('0')
        else:
            # Phía Việt Nam: dùng VND, không trừ từ số dư
            amount_vnd_str = data.get('amount', '').strip().replace(',', '')
            if not amount_vnd_str:
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền VNĐ không được để trống cho chi phí phía Việt Nam"
                }, status=400)
            
            try:
                amount_vnd_decimal = Decimal(str(round(float(amount_vnd_str))))
            except (ValueError, InvalidOperation):
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền VNĐ không hợp lệ"
                }, status=400)
            
            if amount_vnd_decimal <= 0:
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền VNĐ phải lớn hơn 0"
                }, status=400)
            
            # Không có amount_cny và balance_txn cho chi phí phía Việt Nam
            amount_cny_decimal = None
            balance_txn = None
        
        # Parse transaction_date cho chi phí phía Việt Nam
        if not transaction_date_str:
            transaction_date = timezone.now().date()
        else:
            try:
                transaction_date = datetime.strptime(transaction_date_str, '%Y-%m-%d').date()
            except ValueError:
                transaction_date = timezone.now().date()
        
        cost = SPOCost.objects.create(
            sum_purchase_order=spo,
            name=name,
            cost_side=cost_side,
            amount_vnd=amount_vnd_decimal,
            amount_cny=amount_cny_decimal,  # Lưu số tiền CNY nếu có (chỉ cho phía Trung Quốc)
            balance_transaction=balance_txn,  # Liên kết với giao dịch số dư (chỉ cho phía Trung Quốc)
            note=note,
            transaction_date=transaction_date,
            created_by=request.user if request.user.is_authenticated else None
        )
        
        return JsonResponse({
            "status": "success",
            "message": "Thêm chi phí thành công",
            "cost_id": cost.id
        })
        
    except Exception as e:
        logger.error(f"Error in add_spo_cost: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def edit_spo_cost(request: HttpRequest, cost_id: int):
    """
    API endpoint để sửa chi phí SPO.
    
    POST data:
    - name: str
    - cost_side: str ('china' hoặc 'vietnam')
    - amount: float (CNY nếu cost_side='china', VND nếu cost_side='vietnam')
    - transaction_date: str (YYYY-MM-DD)
    - note: str (optional)
    
    Returns:
        JSON: {status, message}
    """
    try:
        import json
        from datetime import datetime
        
        cost = get_object_or_404(SPOCost, id=cost_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        name = data.get('name', '').strip()
        cost_side = data.get('cost_side', 'china').strip()
        note = data.get('note', '').strip()
        transaction_date_str = data.get('transaction_date', '').strip()
        
        if not name:
            return JsonResponse({
                "status": "error",
                "message": "Tên chi phí không được để trống"
            }, status=400)
        
        if cost_side not in ['china', 'vietnam']:
            return JsonResponse({
                "status": "error",
                "message": "Phía chi phí không hợp lệ. Phải là 'china' hoặc 'vietnam'"
            }, status=400)
        
        # Parse transaction_date
        transaction_date = cost.transaction_date  # Giữ nguyên nếu không có
        if transaction_date_str:
            try:
                transaction_date = datetime.strptime(transaction_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    "status": "error",
                    "message": "Định dạng ngày giao dịch không hợp lệ. Phải là YYYY-MM-DD"
                }, status=400)
        
        # Cập nhật chi phí
        cost.name = name
        cost.cost_side = cost_side
        cost.note = note
        cost.transaction_date = transaction_date
        
        # Cập nhật số tiền
        if cost_side == 'china':
            amount_cny_str = data.get('amount', '').strip().replace(',', '')
            if not amount_cny_str:
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền CNY không được để trống cho chi phí phía Trung Quốc"
                }, status=400)
            
            try:
                amount_cny_decimal = Decimal(str(amount_cny_str))
            except (ValueError, InvalidOperation):
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền CNY không hợp lệ"
                }, status=400)
            
            if amount_cny_decimal <= 0:
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền CNY phải lớn hơn 0"
                }, status=400)
            
            # Cập nhật balance_transaction nếu có
            if cost.balance_transaction:
                # Cập nhật giao dịch số dư
                cost.balance_transaction.amount_cny = -amount_cny_decimal
                cost.balance_transaction.transaction_date = transaction_date
                cost.balance_transaction.description = f"Chi phí SPO {cost.sum_purchase_order.code}: {name} (Phía Trung Quốc)"
                cost.balance_transaction.save()
            
            cost.amount_cny = amount_cny_decimal
            cost.amount_vnd = Decimal('0')
        else:
            amount_vnd_str = data.get('amount', '').strip().replace(',', '')
            if not amount_vnd_str:
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền VNĐ không được để trống cho chi phí phía Việt Nam"
                }, status=400)
            
            try:
                amount_vnd_decimal = Decimal(str(round(float(amount_vnd_str))))
            except (ValueError, InvalidOperation):
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền VNĐ không hợp lệ"
                }, status=400)
            
            if amount_vnd_decimal <= 0:
                return JsonResponse({
                    "status": "error",
                    "message": "Số tiền VNĐ phải lớn hơn 0"
                }, status=400)
            
            cost.amount_vnd = amount_vnd_decimal
            cost.amount_cny = None
            # Xóa balance_transaction nếu chuyển từ china sang vietnam
            if cost.balance_transaction:
                cost.balance_transaction.delete()
                cost.balance_transaction = None
        
        cost.save()
        
        return JsonResponse({
            "status": "success",
            "message": "Cập nhật chi phí thành công"
        })
        
    except Exception as e:
        logger.error(f"Error in edit_spo_cost: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_http_methods(["DELETE"])
def delete_spo_cost(request: HttpRequest, cost_id: int):
    """
    API endpoint để xóa chi phí SPO.
    Tự động xóa giao dịch số dư liên quan (nếu có) và liên kết với kỳ thanh toán.
    
    Returns:
        JSON: {status, message}
    """
    try:
        from products.models import PaymentPeriodTransaction
        
        cost = get_object_or_404(SPOCost, id=cost_id)
        
        # Lấy balance_transaction trước khi xóa cost
        balance_txn = cost.balance_transaction
        
        # Xóa chi phí (sẽ tự động xóa liên kết OneToOne với balance_transaction)
        cost.delete()
        
        # Nếu có giao dịch số dư liên quan, xóa nó và liên kết với kỳ thanh toán
        if balance_txn:
            # Xóa tất cả liên kết với kỳ thanh toán
            PaymentPeriodTransaction.objects.filter(balance_transaction=balance_txn).delete()
            
            # Xóa giao dịch số dư
            balance_txn.delete()
        
        return JsonResponse({
            "status": "success",
            "message": "Xóa chi phí thành công"
        })
        
    except Exception as e:
        logger.error(f"Error in delete_spo_cost: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def upload_spo_document(request: HttpRequest):
    """
    API endpoint để upload chứng từ SPO.
    
    POST data (multipart/form-data):
    - spo_id: int
    - file: File
    - name: str (optional)
    
    Returns:
        JSON: {status, message, document_id}
    """
    try:
        spo_id = int(request.POST.get('spo_id'))
        file = request.FILES.get('file')
        name = request.POST.get('name', '').strip()
        
        if not file:
            return JsonResponse({
                "status": "error",
                "message": "Vui lòng chọn file"
            }, status=400)
        
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        
        # Lưu file vào static/spo_documents
        from django.conf import settings
        from datetime import datetime
        
        # Tạo thư mục nếu chưa có
        year_month = datetime.now().strftime('%Y/%m')
        static_dir = os.path.join(settings.BASE_DIR, 'assets', 'spo_documents', year_month)
        os.makedirs(static_dir, exist_ok=True)
        
        # Lưu file
        filename = file.name
        file_path = os.path.join(static_dir, filename)
        
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        
        # Lưu đường dẫn tương đối trong database (để tương thích với FileField)
        relative_path = f'spo_documents/{year_month}/{filename}'
        
        document = SPODocument.objects.create(
            sum_purchase_order=spo,
            file=relative_path,  # Lưu đường dẫn tương đối
            name=name or file.name,
            uploaded_by=request.user if request.user.is_authenticated else None
        )
        
        return JsonResponse({
            "status": "success",
            "message": "Upload chứng từ thành công",
            "document_id": document.id,
            "document_name": document.name,
            "document_url": document.get_file_url() if hasattr(document, 'get_file_url') else f'/static/spo_documents/{year_month}/{filename}'
        })
        
    except Exception as e:
        logger.error(f"Error in upload_spo_document: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def upload_tkhq(request: HttpRequest, spo_id: int):
    """
    Upload Tờ khai hải quan (TKHQ) và parse dữ liệu để update packing list.
    
    POST data (multipart/form-data):
    - file: File (Excel)
    - name: str (optional)
    
    Returns:
        JSON: {status, message, updated_count}
    """
    try:
        from django.conf import settings
        from datetime import datetime
        
        file = request.FILES.get('file')
        name = request.POST.get('name', '').strip()
        
        if not file:
            return JsonResponse({
                "status": "error",
                "message": "Vui lòng chọn file"
            }, status=400)
        
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        
        # Lưu file vào static/spo_documents
        year_month = datetime.now().strftime('%Y/%m')
        static_dir = os.path.join(settings.BASE_DIR, 'assets', 'spo_documents', year_month)
        os.makedirs(static_dir, exist_ok=True)
        
        filename = file.name
        file_path = os.path.join(static_dir, filename)
        
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        
        # Lưu vào SPODocument
        relative_path = f'spo_documents/{year_month}/{filename}'
        document = SPODocument.objects.create(
            sum_purchase_order=spo,
            file=relative_path,
            name=name or f"TKHQ - {filename}",
            uploaded_by=request.user if request.user.is_authenticated else None
        )
        
        # Parse file và update packing list
        updated_count = _parse_tkhq_and_update_packing_list(file_path, spo)
        
        return JsonResponse({
            "status": "success",
            "message": f"Upload TKHQ thành công. Đã cập nhật {updated_count} dòng packing list.",
            "document_id": document.id,
            "updated_count": updated_count
        })
        
    except Exception as e:
        logger.error(f"Error in upload_tkhq: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
def export_tkhq_info(request: HttpRequest, spo_id: int):
    """
    Xuất file Excel chứa thông tin đã parse từ TKHQ.
    Bao gồm: Mã số hàng hóa, Mô tả hàng hóa, Thuế NK, Thuế GTGT.
    """
    try:
        from django.conf import settings
        from django.http import HttpResponse
        import xlsxwriter
        from io import BytesIO
        
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        
        # Tìm file TKHQ - kiểm tra cả doc.name và filename
        tkhq_document = None
        for doc in spo.documents.all():
            # Kiểm tra doc.name
            if doc.name:
                doc_name_upper = doc.name.upper()
                if doc_name_upper.startswith('TKHQ') or doc_name_upper.startswith('TOKHAIHQ'):
                    tkhq_document = doc
                    break
            
            # Kiểm tra filename từ file field
            if doc.file:
                file_name = doc.file.name if hasattr(doc.file, 'name') else str(doc.file)
                file_name_upper = file_name.upper()
                if 'TKHQ' in file_name_upper or 'TOKHAIHQ' in file_name_upper:
                    tkhq_document = doc
                    break
        
        if not tkhq_document:
            # Debug: In tất cả documents để xem
            all_docs = list(spo.documents.all())
            doc_names = [f"{doc.name} ({doc.file.name if doc.file else 'no file'})" for doc in all_docs]
            logger.warning(f"No TKHQ document found. Available documents: {doc_names}")
            return JsonResponse({
                "status": "error",
                "message": f"Không tìm thấy file TKHQ. Có {len(all_docs)} documents: {', '.join([doc.name or 'No name' for doc in all_docs])}"
            }, status=404)
        
        # Lấy đường dẫn file
        # File được lưu ở assets/spo_documents/, database lưu relative path là spo_documents/...
        # Luôn dùng .name và join với assets/ vì file được lưu thủ công vào assets/
        file_relative_path = tkhq_document.file.name if hasattr(tkhq_document.file, 'name') else str(tkhq_document.file)
        file_path = os.path.join(settings.BASE_DIR, 'assets', file_relative_path)
        
        if not os.path.exists(file_path):
            return JsonResponse({
                "status": "error",
                "message": f"File TKHQ không tồn tại: {file_path}"
            }, status=404)
        
        # Parse file TKHQ
        from products.services.tkhq_parser import TKHQParser, set_tkhq_debug
        # Bật debug mode
        set_tkhq_debug(True)
        parser = TKHQParser(file_path, debug=True)
        items = parser.parse()
        
        # Tạo file Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('TKHQ Info')
        
        # Header
        headers = [
            'STT', 
            'Mã số hàng hóa', 
            'Mô tả hàng hóa', 
            'Số lượng',
            'Đơn giá tính thuế NK (1 chiếc)', 
            'Thuế NK (%)', 
            'Thuế NK VND (1 chiếc)',
            'Thuế NK VND (TOTAL)',
            'Trị giá tính thuế GTGT (1 chiếc)',
            'Thuế GTGT (%)', 
            'Thuế GTGT VND (1 chiếc)',
            'Thuế GTGT VND (TOTAL)',
            'SKU MODEL NHẬP KHẨU'
        ]
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Data
        data_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })
        number_format = workbook.add_format({
            'border': 1,
            'align': 'right',
            'valign': 'vcenter',
            'num_format': '#,##0.00'
        })
        integer_format = workbook.add_format({
            'border': 1,
            'align': 'right',
            'valign': 'vcenter',
            'num_format': '#,##0'
        })
        
        for row_idx, item in enumerate(items, start=1):
            # Xử lý Decimal values an toàn
            def safe_decimal_to_float(val, default=0):
                try:
                    if isinstance(val, Decimal):
                        return float(val)
                    elif val is None or val == '':
                        return default
                    return float(val)
                except:
                    return default
            
            col = 0
            worksheet.write(row_idx, col, item.get('stt', ''), data_format); col += 1
            worksheet.write(row_idx, col, item.get('ma_so_hang_hoa', ''), data_format); col += 1
            worksheet.write(row_idx, col, item.get('mo_ta_hang_hoa', ''), data_format); col += 1
            worksheet.write(row_idx, col, int(safe_decimal_to_float(item.get('so_luong', 0))), integer_format); col += 1
            worksheet.write(row_idx, col, safe_decimal_to_float(item.get('don_gia_tinh_thue_nk_per_unit', 0)), number_format); col += 1
            worksheet.write(row_idx, col, safe_decimal_to_float(item.get('thue_nk_rate', 0)), number_format); col += 1
            worksheet.write(row_idx, col, safe_decimal_to_float(item.get('thue_nk_vnd', 0)), number_format); col += 1
            worksheet.write(row_idx, col, safe_decimal_to_float(item.get('thue_nk_vnd_total', 0)), number_format); col += 1
            worksheet.write(row_idx, col, safe_decimal_to_float(item.get('tri_gia_tinh_thue_gtgt_per_unit', 0)), number_format); col += 1
            worksheet.write(row_idx, col, safe_decimal_to_float(item.get('thue_gtgt_rate', 0)), number_format); col += 1
            worksheet.write(row_idx, col, safe_decimal_to_float(item.get('thue_gtgt_vnd', 0)), number_format); col += 1
            worksheet.write(row_idx, col, safe_decimal_to_float(item.get('thue_gtgt_vnd_total', 0)), number_format); col += 1
            worksheet.write(row_idx, col, '', data_format)  # SKU MODEL NHẬP KHẨU - để trống để user điền
        
        # Auto-fit columns
        worksheet.set_column(0, 0, 8)  # STT
        worksheet.set_column(1, 1, 20)  # Mã số hàng hóa
        worksheet.set_column(2, 2, 60)  # Mô tả hàng hóa
        worksheet.set_column(3, 3, 12)  # Số lượng
        worksheet.set_column(4, 11, 20)  # Các cột thuế và giá
        worksheet.set_column(12, 12, 25)  # SKU MODEL NHẬP KHẨU
        
        workbook.close()
        output.seek(0)
        
        # Trả về file
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="TKHQ_Info_SPO_{spo.id}.xlsx"'
        return response
        
    except Exception as e:
        logger.error(f"Error in export_tkhq_info: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def update_packing_list_from_tkhq(request: HttpRequest, spo_id: int):
    """
    Update packing list từ file TKHQ đã upload (có SKU MODEL NHẬP KHẨU đã điền).
    """
    try:
        from django.conf import settings
        import tempfile
        
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        
        # Kiểm tra file upload
        if 'file' not in request.FILES:
            return JsonResponse({
                "status": "error",
                "message": "Vui lòng upload file TKHQ đã điền SKU MODEL NHẬP KHẨU"
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Lưu file tạm để xử lý
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in uploaded_file.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        try:
            # Parse và update packing list
            from products.services.tkhq_parser import set_tkhq_debug
            set_tkhq_debug(True)  # Bật debug mode
            updated_count = _parse_tkhq_and_update_packing_list(tmp_file_path, spo)
            
            return JsonResponse({
                "status": "success",
                "message": f"Đã cập nhật {updated_count} dòng packing list từ TKHQ.",
                "updated_count": updated_count
            })
        finally:
            # Xóa file tạm
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
        
    except Exception as e:
        logger.error(f"Error in update_packing_list_from_tkhq: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


def _parse_tkhq_and_update_packing_list(file_path: str, spo: SumPurchaseOrder) -> int:
    """
    Parse file Excel đã xuất từ TKHQ (có cột SKU MODEL NHẬP KHẨU đã điền) và update packing list items.
    
    Cấu trúc file Excel:
    - Row 1: Header
    - Row 2+: Data
    - Cột 1: STT
    - Cột 2: Mã số hàng hóa
    - Cột 3: Mô tả hàng hóa
    - Cột 4: Số lượng
    - Cột 5: Đơn giá tính thuế NK (1 chiếc)
    - Cột 6: Thuế NK (%)
    - Cột 7: Thuế NK VND (1 chiếc)
    - Cột 8: Thuế NK VND (TOTAL)
    - Cột 9: Trị giá tính thuế GTGT (1 chiếc)
    - Cột 10: Thuế GTGT (%)
    - Cột 11: Thuế GTGT VND (1 chiếc)
    - Cột 12: Thuế GTGT VND (TOTAL)
    - Cột 13: SKU MODEL NHẬP KHẨU
    
    Args:
        file_path: Đường dẫn đến file Excel đã xuất (có SKU đã điền)
        spo: SumPurchaseOrder instance
        
    Returns:
        Số lượng items đã được update
    """
    try:
        import openpyxl
        from decimal import Decimal
        
        # Đọc file Excel
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Lấy packing list items
        packing_items = list(SPOPackingListItem.objects.filter(sum_purchase_order=spo).order_by('order'))
        if not packing_items:
            logger.warning(f"No packing list items found for SPO {spo.id}")
            return 0
        
        # Tạo map SKU -> packing item
        packing_sku_map = {}
        for item in packing_items:
            if item.sku_nhapkhau:
                sku = str(item.sku_nhapkhau).strip()
                if sku:
                    packing_sku_map[sku] = item
        
        updated_count = 0
        
        # Đọc từ row 2 (row 1 là header)
        for row_idx in range(2, ws.max_row + 1):
            # Đọc SKU MODEL NHẬP KHẨU (cột 13)
            sku_model = ws.cell(row_idx, 13).value
            if not sku_model:
                continue
            
            sku_model = str(sku_model).strip()
            if not sku_model:
                continue
            
            # Tìm packing item tương ứng
            packing_item = packing_sku_map.get(sku_model)
            if not packing_item:
                logger.warning(f"No packing item found for SKU: {sku_model}")
                continue
            
            # Đọc các giá trị từ file Excel
            try:
                # Mã số hàng hóa (cột 2)
                ma_so = ws.cell(row_idx, 2).value
                if ma_so:
                    packing_item.hs_code = str(ma_so).strip()
                
                # Mô tả hàng hóa (cột 3)
                mo_ta = ws.cell(row_idx, 3).value
                if mo_ta:
                    packing_item.vn_name = str(mo_ta).strip()
                
                # Số lượng (cột 4)
                so_luong = ws.cell(row_idx, 4).value
                if so_luong is not None:
                    try:
                        packing_item.quantity = Decimal(str(so_luong))
                    except:
                        pass
                
                # Thuế NK (%) (cột 6)
                thue_nk_rate = ws.cell(row_idx, 6).value
                if thue_nk_rate is not None:
                    try:
                        packing_item.import_tax_rate = Decimal(str(thue_nk_rate))
                    except:
                        pass
                
                # Thuế NK VND (TOTAL) (cột 8)
                thue_nk_total = ws.cell(row_idx, 8).value
                if thue_nk_total is not None:
                    try:
                        packing_item.import_tax_total = Decimal(str(thue_nk_total))
                    except:
                        pass
                
                # Thuế NK VND (1 chiếc) (cột 7)
                thue_nk_per_unit = ws.cell(row_idx, 7).value
                if thue_nk_per_unit is not None:
                    try:
                        packing_item.import_tax_amount = Decimal(str(thue_nk_per_unit))
                    except:
                        pass
                
                # Thuế GTGT (%) (cột 10)
                thue_gtgt_rate = ws.cell(row_idx, 10).value
                if thue_gtgt_rate is not None:
                    try:
                        packing_item.vat_rate = Decimal(str(thue_gtgt_rate))
                    except:
                        pass
                
                # Thuế GTGT VND (TOTAL) (cột 12)
                thue_gtgt_total = ws.cell(row_idx, 12).value
                if thue_gtgt_total is not None:
                    try:
                        packing_item.vat_total = Decimal(str(thue_gtgt_total))
                    except:
                        pass
                
                # Thuế GTGT VND (1 chiếc) (cột 11)
                thue_gtgt_per_unit = ws.cell(row_idx, 11).value
                if thue_gtgt_per_unit is not None:
                    try:
                        packing_item.vat_amount = Decimal(str(thue_gtgt_per_unit))
                    except:
                        pass
                
                packing_item.save()
                updated_count += 1
                logger.info(f"[TKHQ Update] Updated packing item {packing_item.sku_nhapkhau} from Excel file")
                
            except Exception as e:
                logger.error(f"Error updating row {row_idx} for SKU {sku_model}: {e}", exc_info=True)
                continue
        
        logger.info(f"[TKHQ Update] Updated {updated_count} items out of {len(packing_items)} total")
        return updated_count
        
    except Exception as e:
        logger.error(f"Error parsing Excel file {file_path}: {e}", exc_info=True)
        raise
        
        # Tìm các keyword và parse dữ liệu
        items_data = []
        
        # Tìm "Mô tả hàng hóa"
        mo_ta_rows = []
        for row_idx in range(1, min(ws.max_row + 1, 500)):
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and isinstance(cell_value, str) and 'Mô tả hàng hóa' in cell_value:
                    mo_ta_rows.append((row_idx, col_idx))
        
        # Tìm "Thuế nhập khẩu"
        thue_nk_rows = []
        for row_idx in range(1, min(ws.max_row + 1, 500)):
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and isinstance(cell_value, str) and 'Thuế nhập khẩu' in cell_value:
                    thue_nk_rows.append((row_idx, col_idx))
        
        # Tìm "Thuế và thu khác" hoặc "Thuế GTGT"
        thue_gtgt_rows = []
        for row_idx in range(1, min(ws.max_row + 1, 500)):
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and isinstance(cell_value, str):
                    if 'Thuế và thu khác' in cell_value or 'Thuế GTGT' in cell_value or ('GTGT' in cell_value and 'Thuế' in cell_value):
                        # Kiểm tra xem có phải là dòng GTGT không (tránh match nhầm)
                        # Đọc các ô tiếp theo để xác nhận
                        next_cells = []
                        for c in range(col_idx, min(ws.max_column + 1, col_idx + 5)):
                            next_val = ws.cell(row_idx + 1, c).value
                            if next_val:
                                next_cells.append(str(next_val))
                        if any('GTGT' in str(cell).upper() for cell in next_cells):
                            thue_gtgt_rows.append((row_idx, col_idx))
        
        # Parse từng hàng dữ liệu
        # Tìm pattern: mỗi hàng hàng hóa sẽ có mô tả, thuế nhập khẩu, thuế GTGT ở các hàng tương ứng
        
        # Strategy: Tìm các hàng có dữ liệu mô tả hàng hóa (ở cột G sau keyword "Mô tả hàng hóa")
        # Sau đó tìm các hàng thuế tương ứng
        
        processed_rows = set()
        
        # Duyệt qua các hàng mô tả hàng hóa
        for mo_ta_row, mo_ta_col in mo_ta_rows:
            # Tìm dữ liệu ở cột G (col 7) sau keyword "Mô tả hàng hóa"
            # Thường là ở cùng hàng hoặc hàng tiếp theo
            for check_row in range(mo_ta_row, min(ws.max_row + 1, mo_ta_row + 20)):
                if check_row in processed_rows:
                    continue
                
                # Đọc cột G (col 7) - mô tả hàng hóa
                mo_ta_value = ws.cell(check_row, 7).value  # Cột G
                if not mo_ta_value or not isinstance(mo_ta_value, str):
                    continue
                
                mo_ta_text = str(mo_ta_value).strip()
                if not mo_ta_text or len(mo_ta_text) < 10:  # Bỏ qua text quá ngắn
                    continue
                
                # Tìm thuế nhập khẩu tương ứng (gần hàng này)
                import_tax_rate = Decimal('0')
                import_tax_amount = Decimal('0')
                
                for nk_row, nk_col in thue_nk_rows:
                    # Tìm hàng gần nhất với check_row
                    if abs(nk_row - check_row) < 30:
                        # Đọc thuế suất (thường ở cột sau keyword)
                        # Tìm pattern "X%" trong các ô xung quanh
                        for r in range(nk_row, min(ws.max_row + 1, nk_row + 15)):
                            for c in range(nk_col, min(ws.max_column + 1, nk_col + 10)):
                                val = ws.cell(r, c).value
                                if val:
                                    val_str = str(val).strip()
                                    # Tìm thuế suất (ví dụ: "5%", "C 5%")
                                    import_tax_match = re.search(r'(\d+(?:\.\d+)?)\s*%', val_str)
                                    if import_tax_match:
                                        try:
                                            import_tax_rate = Decimal(import_tax_match.group(1))
                                        except:
                                            pass
                                    
                                    # Tìm số tiền thuế (số lớn, có thể có VND)
                                    # Bỏ qua các số quá nhỏ (có thể là index)
                                    try:
                                        # Loại bỏ dấu phẩy, chấm (format số VN)
                                        cleaned = re.sub(r'[^\d.,]', '', val_str)
                                        cleaned = cleaned.replace('.', '').replace(',', '.')
                                        tax_amount_val = Decimal(cleaned)
                                        if tax_amount_val > 1000:  # Số tiền thuế thường > 1000
                                            import_tax_amount = tax_amount_val
                                    except:
                                        pass
                        
                        break
                
                # Tìm thuế GTGT tương ứng
                vat_rate = Decimal('0')
                vat_amount = Decimal('0')
                
                for gtgt_row, gtgt_col in thue_gtgt_rows:
                    if abs(gtgt_row - check_row) < 30:
                        # Tìm pattern tương tự
                        for r in range(gtgt_row, min(ws.max_row + 1, gtgt_row + 20)):
                            for c in range(gtgt_col, min(ws.max_column + 1, gtgt_col + 10)):
                                val = ws.cell(r, c).value
                                if val:
                                    val_str = str(val).strip()
                                    # Tìm thuế suất GTGT
                                    vat_match = re.search(r'(\d+(?:\.\d+)?)\s*%', val_str)
                                    if vat_match:
                                        try:
                                            vat_rate = Decimal(vat_match.group(1))
                                        except:
                                            pass
                                    
                                    # Tìm số tiền thuế GTGT
                                    try:
                                        cleaned = re.sub(r'[^\d.,]', '', val_str)
                                        cleaned = cleaned.replace('.', '').replace(',', '.')
                                        vat_amount_val = Decimal(cleaned)
                                        if vat_amount_val > 1000:
                                            vat_amount = vat_amount_val
                                    except:
                                        pass
                        
                        break
                
                # Match với packing list item (theo mô tả hàng hóa)
                # Tìm item có vn_name gần giống nhất
                best_match = None
                best_score = 0
                
                for item in packing_items:
                    if item.vn_name:
                        # Tính độ tương đồng đơn giản (có thể cải thiện sau)
                        item_desc_lower = item.vn_name.lower()
                        mo_ta_lower = mo_ta_text.lower()
                        
                        # Kiểm tra từ khóa chung (ví dụ: SKU, tên sản phẩm)
                        # Đếm số từ khóa trùng
                        item_words = set(item_desc_lower.split())
                        mo_ta_words = set(mo_ta_lower.split())
                        common_words = item_words.intersection(mo_ta_words)
                        
                        if len(common_words) >= 2:  # Có ít nhất 2 từ trùng
                            score = len(common_words)
                            if score > best_score:
                                best_score = score
                                best_match = item
                
                # Nếu không match được, thử match theo thứ tự (nếu số lượng items ít)
                if not best_match and len(items_data) < len(packing_items):
                    # Match theo thứ tự
                    idx = len(items_data)
                    if idx < len(packing_items):
                        best_match = packing_items[idx]
                
                if best_match:
                    # Update packing list item
                    best_match.vn_name = mo_ta_text  # Update mô tả
                    best_match.import_tax_rate = import_tax_rate
                    best_match.import_tax_amount = import_tax_amount
                    best_match.import_tax_total = import_tax_amount  # Tổng = số tiền thuế
                    best_match.vat_rate = vat_rate
                    best_match.vat_amount = vat_amount
                    best_match.vat_total = vat_amount  # Tổng = số tiền thuế
                    best_match.save()
                    
                    items_data.append({
                        'item': best_match,
                        'mo_ta': mo_ta_text,
                        'import_tax_rate': import_tax_rate,
                        'import_tax_amount': import_tax_amount,
                        'vat_rate': vat_rate,
                        'vat_amount': vat_amount
                    })
                    
                    updated_count += 1
                    processed_rows.add(check_row)
                
                break  # Chỉ xử lý 1 item mỗi lần tìm thấy keyword
        
        return updated_count
        
    except Exception as e:
        logger.error(f"Error parsing TKHQ file {file_path}: {e}", exc_info=True)
        raise


@admin_only
@require_http_methods(["DELETE"])
def delete_spo_document(request: HttpRequest, document_id: int):
    """
    API endpoint để xóa chứng từ SPO.
    
    Returns:
        JSON: {status, message}
    """
    try:
        document = get_object_or_404(SPODocument, id=document_id)
        document.delete()
        
        return JsonResponse({
            "status": "success",
            "message": "Xóa chứng từ thành công"
        })
        
    except Exception as e:
        logger.error(f"Error in delete_spo_document: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def update_po_delivery_status(request: HttpRequest, po_id: int):
    """
    API endpoint để cập nhật trạng thái giao hàng của PO.
    
    POST data:
    - delivery_status: str (ordered, sent_label, production, delivered)
    - date: str (YYYY-MM-DD format, optional)
    - note: str (optional)
    
    Returns:
        JSON: {status, message}
    """
    try:
        import json
        from datetime import datetime
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        delivery_status = data.get('delivery_status', '').strip()
        date_str = data.get('date', '').strip()
        note = data.get('note', '').strip()
        
        if not delivery_status:
            return JsonResponse({
                "status": "error",
                "message": "Trạng thái không được để trống"
            }, status=400)
        
        # Validate delivery_status
        valid_statuses = ['ordered', 'sent_label', 'production', 'delivered']
        if delivery_status not in valid_statuses:
            return JsonResponse({
                "status": "error",
                "message": f"Trạng thái không hợp lệ. Phải là một trong: {', '.join(valid_statuses)}"
            }, status=400)
        
        # Parse date
        date_obj = None
        if date_str:
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    "status": "error",
                    "message": "Định dạng ngày không hợp lệ. Phải là YYYY-MM-DD"
                }, status=400)
        
        # Get PO
        po = get_object_or_404(PurchaseOrder, id=po_id)
        
        # Update delivery status
        po.update_delivery_status(delivery_status, date_obj, note)
        
        return JsonResponse({
            "status": "success",
            "message": "Cập nhật trạng thái thành công"
        })
        
    except Exception as e:
        logger.error(f"Error in update_po_delivery_status: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def add_po_cost(request: HttpRequest, po_id: int):
    """
    API endpoint để thêm chi phí cho PO.
    Chi phí tự động phân bổ theo CBM của PO (không cần nhập CBM thủ công).
    
    POST data:
    - cost_type: str
    - amount_cny: float
    - description: str (optional)
    
    Returns:
        JSON: {status, message, cost_id, price_per_cbm}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        cost_type = data.get('cost_type', '').strip()
        amount_cny = Decimal(str(data.get('amount_cny', 0)))
        description = data.get('description', '').strip()
        
        if not cost_type:
            return JsonResponse({
                "status": "error",
                "message": "Loại chi phí không được để trống"
            }, status=400)
        
        if amount_cny <= 0:
            return JsonResponse({
                "status": "error",
                "message": "Số tiền phải lớn hơn 0"
            }, status=400)
        
        # Get PO
        po = get_object_or_404(PurchaseOrder, id=po_id)
        
        # Lấy total_cbm của PO từ Sapo API
        from products.services.spo_po_service import SPOPOService
        sapo_client = get_sapo_client()
        spo_po_service = SPOPOService(sapo_client)
        
        try:
            po_data = spo_po_service.get_po_from_sapo(po.sapo_order_supplier_id)
            total_cbm = float(po_data.get('total_cbm', 0))
        except Exception as e:
            logger.warning(f"Error getting PO CBM: {e}")
            total_cbm = 0
        
        # Tính CBM tự động (toàn bộ CBM của PO)
        cbm_value = Decimal(str(total_cbm)) if total_cbm > 0 else None
        
        # Tính giá/m³
        price_per_cbm = None
        if cbm_value and cbm_value > 0:
            price_per_cbm = float(amount_cny / cbm_value)
        
        # Create cost
        cost = PurchaseOrderCost.objects.create(
            purchase_order=po,
            cost_type=cost_type,
            amount_cny=amount_cny,
            cbm=cbm_value,
            description=description,
            created_by=request.user if request.user.is_authenticated else None
        )
        
        # Update PO total_amount
        po.calculate_total_amount()
        
        return JsonResponse({
            "status": "success",
            "message": "Thêm chi phí thành công",
            "cost_id": cost.id,
            "price_per_cbm": round(price_per_cbm, 2) if price_per_cbm else None,
            "cbm": float(cbm_value) if cbm_value else None
        })
        
    except Exception as e:
        logger.error(f"Error in add_po_cost: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_http_methods(["DELETE"])
def delete_po_cost(request: HttpRequest, po_id: int, cost_id: int):
    """
    API endpoint để xóa chi phí PO.
    
    Returns:
        JSON: {status, message}
    """
    try:
        cost = get_object_or_404(PurchaseOrderCost, id=cost_id, purchase_order_id=po_id)
        po = cost.purchase_order
        cost.delete()
        
        # Update PO total_amount
        po.calculate_total_amount()
        
        return JsonResponse({
            "status": "success",
            "message": "Xóa chi phí thành công"
        })
        
    except Exception as e:
        logger.error(f"Error in delete_po_cost: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def add_po_payment(request: HttpRequest, po_id: int):
    """
    API endpoint để thêm thanh toán cho PO.
    
    POST data:
    - payment_type: str
    - amount_cny: float
    - amount_vnd: float (optional)
    - payment_date: str (YYYY-MM-DD, optional)
    - description: str (optional)
    
    Returns:
        JSON: {status, message, payment_id}
    """
    try:
        import json
        from datetime import datetime
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        payment_type = data.get('payment_type', '').strip()
        amount_cny = data.get('amount_cny')
        payment_date_str = data.get('payment_date', '').strip()
        description = data.get('description', '').strip()
        
        if not payment_type:
            return JsonResponse({
                "status": "error",
                "message": "Loại thanh toán không được để trống"
            }, status=400)
        
        # Parse số tiền CNY
        amount_cny_decimal = Decimal(str(amount_cny)) if amount_cny else Decimal('0')
        
        # Validate: phải có số tiền CNY
        if amount_cny_decimal <= 0:
            return JsonResponse({
                "status": "error",
                "message": "Số tiền nhân dân tệ (CNY) phải lớn hơn 0"
            }, status=400)
        
        # Parse date
        payment_date = None
        if payment_date_str:
            try:
                payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    "status": "error",
                    "message": "Định dạng ngày không hợp lệ. Phải là YYYY-MM-DD"
                }, status=400)
        
        # Get PO
        po = get_object_or_404(PurchaseOrder, id=po_id)
        
        # Kiểm tra số dư hiện tại
        from products.models import BalanceTransaction
        from django.db.models import Sum
        
        current_balance = BalanceTransaction.objects.aggregate(
            total=Sum('amount_cny')
        )['total'] or Decimal('0')
        
        if current_balance < amount_cny_decimal:
            return JsonResponse({
                "status": "error",
                "message": f"Số dư không đủ. Số dư hiện tại: {current_balance} CNY, cần: {amount_cny_decimal} CNY"
            }, status=400)
        
        # Tự động liên kết với PaymentPeriod dựa trên ngày thanh toán (theo logic mới: khoảng trống thuộc kỳ cũ)
        payment_date_for_period = payment_date or timezone.now().date()
        from products.models import PaymentPeriod, PaymentPeriodTransaction
        
        # Create payment trước (chỉ lưu amount_cny, không lưu amount_vnd và exchange_rate)
        payment = PurchaseOrderPayment.objects.create(
            purchase_order=po,
            payment_type=payment_type,
            amount_cny=amount_cny_decimal,
            amount_vnd=None,  # Không lưu VNĐ nữa
            exchange_rate=None,  # Không lưu tỷ giá nữa
            payment_date=payment_date or timezone.now().date(),
            description=description,
            balance_transaction=None,  # Sẽ cập nhật sau
            created_by=request.user if request.user.is_authenticated else None
        )
        
        # Tạo giao dịch số dư (rút) - tự động tạo khi thanh toán
        balance_txn = BalanceTransaction.objects.create(
            transaction_type='withdraw_po',
            amount_cny=-amount_cny_decimal,  # Số âm vì là rút
            transaction_date=payment_date or timezone.now().date(),
            description=f"Thanh toán PO {po.sapo_order_supplier_id}: {description}" if description else f"Thanh toán PO {po.sapo_order_supplier_id}",
            purchase_order_payment=payment,  # Liên kết với thanh toán PO
            created_by=request.user if request.user.is_authenticated else None
        )
        
        # Cập nhật payment với balance_transaction
        payment.balance_transaction = balance_txn
        payment.save()
        
        # Liên kết với PaymentPeriod dựa trên datetime (tính cả giờ)
        # Sử dụng created_at của balance_txn để có datetime chính xác
        # Đây là giao dịch rút (withdraw_po), truyền is_withdraw=True
        period = PaymentPeriod.find_period_for_transaction_datetime(balance_txn.created_at, is_withdraw=True)
        
        if period:
            PaymentPeriodTransaction.objects.get_or_create(
                payment_period=period,
                balance_transaction=balance_txn
            )
            # Số dư cuối kỳ tính realtime, không cần lưu
        
        # Update PO paid_amount
        po.calculate_paid_amount()
        
        return JsonResponse({
            "status": "success",
            "message": "Thêm thanh toán thành công",
            "payment_id": payment.id
        })
        
    except Exception as e:
        logger.error(f"Error in add_po_payment: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_http_methods(["DELETE"])
def delete_po_payment(request: HttpRequest, po_id: int, payment_id: int):
    """
    API endpoint để xóa thanh toán PO.
    Sẽ xóa cả BalanceTransaction tương ứng.
    
    Returns:
        JSON: {status, message}
    """
    try:
        payment = get_object_or_404(PurchaseOrderPayment, id=payment_id, purchase_order_id=po_id)
        po = payment.purchase_order
        
        # Xóa BalanceTransaction tương ứng nếu có
        if payment.balance_transaction:
            balance_txn = payment.balance_transaction
            # Xóa liên kết với PaymentPeriod trước
            from products.models import PaymentPeriodTransaction
            PaymentPeriodTransaction.objects.filter(balance_transaction=balance_txn).delete()
            # Xóa BalanceTransaction
            balance_txn.delete()
        
        # Xóa payment
        payment.delete()
        
        # Update PO paid_amount
        po.calculate_paid_amount()
        
        return JsonResponse({
            "status": "success",
            "message": "Xóa thanh toán thành công"
        })
        
    except Exception as e:
        logger.error(f"Error in delete_po_payment: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_http_methods(["GET"])
def export_po_excel(request: HttpRequest, po_id: int):
    """
    Xuất Excel PO để gửi cho NSX.
    Bắt đầu từ file template PO_TEMPLATE.xlsx và ghi dữ liệu từ hàng 11.
    Cấu trúc cột: Image (B), SKU (C), NSX SKU (D), Description (E), Quantity (F), 
    Unit Price (G), Total (RMB) (H), Total (Box) (I), Total (CPM) (J)
    """
    try:
        # Get PO
        po = get_object_or_404(PurchaseOrder, id=po_id)
        
        # Lấy thông tin PO từ Sapo API
        from products.services.spo_po_service import SPOPOService
        sapo_client = get_sapo_client()
        spo_po_service = SPOPOService(sapo_client)
        
        po_data = spo_po_service.get_po_from_sapo(po.sapo_order_supplier_id)
        line_items = po_data.get('line_items', [])
        
        if not line_items:
            return JsonResponse({
                "status": "error",
                "message": "PO không có line items"
            }, status=400)
        
        # Tạo thư mục nếu chưa có
        excel_dir = "assets/excel_po"
        os.makedirs(excel_dir, exist_ok=True)
        
        # Load template file
        template_path = os.path.join("products", "PO_TEMPLATE.xlsx")
        if not os.path.exists(template_path):
            return JsonResponse({
                "status": "error",
                "message": f"Template file không tồn tại: {template_path}"
            }, status=404)
        
        # Load workbook từ template
        workbook = load_workbook(template_path)
        worksheet = workbook.active
        
        # Thiết lập độ rộng cột (tùy chọn, có thể điều chỉnh)
        worksheet.column_dimensions['B'].width = 15  # Image
        worksheet.column_dimensions['C'].width = 20  # SKU
        worksheet.column_dimensions['D'].width = 25  # NSX SKU
        worksheet.column_dimensions['E'].width = 30  # Description
        worksheet.column_dimensions['F'].width = 12  # Quantity
        worksheet.column_dimensions['G'].width = 15  # Unit Price
        worksheet.column_dimensions['H'].width = 15  # Total (RMB)
        worksheet.column_dimensions['I'].width = 12  # Total (Box)
        worksheet.column_dimensions['J'].width = 15  # Total (CPM)
        
        # Alignment format
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # ========== LOAD TOÀN BỘ SẢN PHẨM TRƯỚC ==========
        logger.info("Loading all products to create variant map...")
        product_service = SapoProductService(sapo_client)
        
        # Tạo map variant_id -> ProductVariantDTO
        variant_map: Dict[int, ProductVariantDTO] = {}
        
        # Lấy tất cả products với pagination
        all_products = []
        page = 1
        limit = 250  # Max limit theo Sapo API
        max_pages = 100  # Giới hạn an toàn
        
        while page <= max_pages:
            try:
                products = product_service.list_products(page=page, limit=limit, status="active")
                if not products:
                    break
                
                all_products.extend(products)
                logger.info(f"Loaded page {page}: {len(products)} products (total: {len(all_products)})")
                
                # Nếu số products < limit thì đã hết
                if len(products) < limit:
                    break
                
                page += 1
            except Exception as e:
                logger.error(f"Error loading products page {page}: {e}", exc_info=True)
                break
        
        # Tạo map variant_id -> ProductVariantDTO từ tất cả products
        for product in all_products:
            for variant in product.variants:
                variant_map[variant.id] = variant
        
        logger.info(f"Created variant map with {len(variant_map)} variants")
        
        # Bắt đầu từ hàng 11 (index 10 trong 0-based)
        start_row = 11
        current_row = start_row
        
        # Tạo border style với màu #A6A6A6
        border_style = Border(
            left=Side(style='thin', color='A6A6A6'),
            right=Side(style='thin', color='A6A6A6'),
            top=Side(style='thin', color='A6A6A6'),
            bottom=Side(style='thin', color='A6A6A6')
        )
        
        # Process line items
        for item in line_items:
            variant_id = item.get('variant_id')
            quantity = item.get('quantity', 0)
            sku = item.get('sku', '')
            
            if quantity <= 0:
                continue
            
            # Lấy thông tin variant từ map đã load trước
            try:
                # Tra cứu variant từ map
                variant_dto = variant_map.get(variant_id)
                if not variant_dto:
                    logger.warning(f"Variant {variant_id} not found in variant map")
                    continue
                
                # Lấy metadata từ variant_dto
                variant_meta = variant_dto.gdp_metadata
                
                # Lấy ảnh từ variant_dto.images (ProductVariantDTO có images)
                image_url = None
                if variant_dto.images and len(variant_dto.images) > 0:
                    image_url = variant_dto.images[0].full_path
                    logger.info(f"Variant {variant_id} has image: {image_url}")
                else:
                    logger.warning(f"Variant {variant_id} has no images")
                
                # Lấy thông tin từ metadata
                box_info = variant_meta.box_info if variant_meta else None
                tq_price = variant_meta.price_tq if variant_meta and variant_meta.price_tq else 0
                tq_name = variant_meta.name_tq if variant_meta and variant_meta.name_tq else ""
                tq_sku = variant_meta.sku_tq if variant_meta and variant_meta.sku_tq else ""
                
                # Tính số thùng và CBM
                full_box = box_info.full_box if box_info and box_info.full_box else 1
                num_boxes = round(float(quantity) / full_box, 2) if full_box > 0 else quantity
                
                # Tính Total CPM (mét khối)
                total_cpm = 0.0
                description = ""
                if box_info and box_info.length_cm and box_info.width_cm and box_info.height_cm:
                    # Description: "Box: 60 pcs (56 x 64 x 54cm)"
                    description = f"Box: {full_box} pcs ({int(box_info.length_cm)} x {int(box_info.width_cm)} x {int(box_info.height_cm)}cm)"
                    # CPM = (dài * rộng * cao * số thùng) / 1,000,000
                    total_cpm = (box_info.length_cm * box_info.width_cm * box_info.height_cm * num_boxes) / 1000000
                else:
                    description = f"Box: {full_box} pcs" if full_box > 0 else ""
                
                # NSX SKU: tq_name + tq_sku
                nsx_sku = f"{tq_name} {tq_sku}".strip() if tq_name or tq_sku else ""
                
                # Total (RMB) = quantity * tq_price
                total_rmb = quantity * tq_price
                
                # Set row height để hình vuông
                # Column width 15 characters ≈ 112 pixels ≈ 84 points
                # Set row height tương ứng để tạo hình vuông
                row_height_points = 84  # points (Excel uses points for row height)
                worksheet.row_dimensions[current_row].height = row_height_points
                
                # Cột B: Image - Xử lý trong memory, không lưu file
                if image_url:
                    try:
                        logger.info(f"Downloading and inserting image for variant {variant_id} from {image_url}")
                        # Download ảnh vào memory
                        r = requests.get(image_url, allow_redirects=True, timeout=10)
                        if r.status_code == 200:
                            # Xử lý ảnh trong memory
                            img = Image.open(io.BytesIO(r.content))
                            
                            # Resize để fit vào cell vuông (giữ tỷ lệ, max 100px)
                            max_size = 100  # pixels
                            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
                            
                            # Lưu ảnh đã resize vào BytesIO
                            img_bytes = io.BytesIO()
                            img.save(img_bytes, format='JPEG')
                            img_bytes.seek(0)  # Reset pointer về đầu
                            
                            # Insert image vào cell B từ memory
                            img_obj = XLSXImage(img_bytes)
                            # Set kích thước ảnh (openpyxl uses pixels)
                            img_size = 100  # pixels
                            img_obj.width = img_size
                            img_obj.height = img_size
                            
                            # Anchor image vào cell B
                            cell_ref = f'B{current_row}'
                            worksheet.add_image(img_obj, cell_ref)
                            logger.info(f"Image inserted into {cell_ref}")
                        else:
                            logger.warning(f"Failed to download image: HTTP {r.status_code}")
                    except Exception as e:
                        logger.error(f"Error processing image for variant {variant_id}: {e}", exc_info=True)
                else:
                    logger.warning(f"No image URL for variant {variant_id}")
                
                # Cột C: SKU
                cell = worksheet.cell(row=current_row, column=3, value=sku)
                cell.alignment = center_align
                cell.border = border_style
                
                # Cột D: NSX SKU
                cell = worksheet.cell(row=current_row, column=4, value=nsx_sku)
                cell.alignment = left_align
                cell.border = border_style
                
                # Cột E: Description
                cell = worksheet.cell(row=current_row, column=5, value=description)
                cell.alignment = left_align
                cell.border = border_style
                
                # Cột F: Quantity
                cell = worksheet.cell(row=current_row, column=6, value=quantity)
                cell.alignment = center_align
                cell.border = border_style
                
                # Cột G: Unit Price (tq_price)
                cell = worksheet.cell(row=current_row, column=7, value=tq_price)
                cell.alignment = center_align
                cell.border = border_style
                
                # Cột H: Total (RMB)
                cell = worksheet.cell(row=current_row, column=8, value=total_rmb)
                cell.alignment = center_align
                cell.border = border_style
                
                # Cột I: Total (Box)
                cell = worksheet.cell(row=current_row, column=9, value=num_boxes)
                cell.alignment = center_align
                cell.border = border_style
                
                # Cột J: Total (CPM)
                cell = worksheet.cell(row=current_row, column=10, value=round(total_cpm, 4))
                cell.alignment = center_align
                cell.border = border_style
                
                # Cột B: Image - cũng cần border
                cell_b = worksheet.cell(row=current_row, column=2)
                cell_b.border = border_style
                
                current_row += 1
                
            except Exception as e:
                logger.error(f"Error processing line item {variant_id}: {e}", exc_info=True)
                continue
        
        # Thêm border cho tất cả các hàng từ 8 đến hàng cuối cùng (current_row - 1)
        # Áp dụng border cho các cột B đến J
        last_row = current_row - 1  # Hàng cuối cùng có dữ liệu
        for row_num in range(8, last_row + 1):
            for col_num in range(2, 11):  # Cột B (2) đến J (10)
                cell = worksheet.cell(row=row_num, column=col_num)
                if cell.border is None or not any([cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]):
                    cell.border = border_style
        
        # Sửa công thức SUM nếu có trong template (từ hàng 12 thành 11)
        # Tìm các cell có công thức SUM và sửa nếu cần
        for row in worksheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=15):
            for cell in row:
                if cell.data_type == 'f' and cell.value:  # Công thức
                    formula = str(cell.value)
                    # Sửa SUM từ hàng 12 thành 11
                    if 'SUM' in formula and ':12:' in formula:
                        new_formula = formula.replace(':12:', ':11:')
                        cell.value = new_formula
                        logger.info(f"Updated SUM formula in {cell.coordinate}: {formula} -> {new_formula}")
        
        # Tên file output
        supplier_name = po_data.get('supplier_name', 'Supplier').replace('/', '_').replace('\\', '_')
        po_code = po_data.get('code', f'PO-{po_id}').replace('/', '_').replace('\\', '_')
        filename = f"{supplier_name}-{po_code}.xlsx"
        filepath = os.path.join(excel_dir, filename)
        
        # Save workbook
        workbook.save(filepath)
        
        # Return file
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
    except Exception as e:
        logger.error(f"Error in export_po_excel: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_http_methods(["GET"])
def export_po_labels(request: HttpRequest, po_id: int):
    """
    Xuất nhãn phụ (HTML) cho PO - nhãn xuất nhập khẩu 10x15 (A6).
    Mỗi SKU có 1 nhãn, kết hợp thông tin sản phẩm và model XNK.
    """
    try:
        # Get PO
        po = get_object_or_404(PurchaseOrder, id=po_id)
        
        # Lấy thông tin PO từ Sapo API
        from products.services.spo_po_service import SPOPOService
        from products.services.xnk_model_service import XNKModelService
        
        sapo_client = get_sapo_client()
        spo_po_service = SPOPOService(sapo_client)
        xnk_service = XNKModelService(sapo_client.core_session)
        
        po_data = spo_po_service.get_po_from_sapo(po.sapo_order_supplier_id)
        line_items = po_data.get('line_items', [])
        
        if not line_items:
            return JsonResponse({
                "status": "error",
                "message": "PO không có line items"
            }, status=400)
        
        # ===== BƯỚC 1: Load toàn bộ XNK models =====
        logger.info(f"[export_po_labels] [BƯỚC 1] Loading all XNK models...")
        all_xnk_models = xnk_service.get_all_models()
        # Tạo map với key là SKU (trimmed, normalized)
        xnk_model_map = {}
        for model in all_xnk_models:
            sku = model.get('sku', '').strip()
            if sku:
                # Lưu với key là SKU đã trim (case-sensitive để chính xác)
                xnk_model_map[sku] = model
        
        logger.info(f"[export_po_labels] [BƯỚC 1] ✅ Loaded {len(all_xnk_models)} XNK models. Map has {len(xnk_model_map)} unique SKUs")
        
        # ===== BƯỚC 2: Lấy thông tin supplier và brand =====
        supplier_name = po_data.get('supplier_name', '')
        supplier_code = po_data.get('supplier_code', '')
        supplier_id = po_data.get('supplier_id')
        
        # Lấy supplier address và brand_id từ Sapo
        supplier_address = ""
        brand_id = None
        
        try:
            if supplier_id:
                supplier_data = sapo_client.core.get_supplier_raw(supplier_id)
                if supplier_data:
                    # Lấy address
                    if supplier_data.get('addresses'):
                        addr = supplier_data['addresses'][0]
                        supplier_address = addr.get('address1', '')
                    
                    # Lấy brand_id từ supplier code
                    supplier_code_from_api = supplier_data.get('code', '') or supplier_code
                    if supplier_code_from_api:
                        logger.info(f"[export_po_labels] [BƯỚC 2] Looking for brand with code/name: {supplier_code_from_api}")
                        # Tìm brand theo code hoặc name
                        brands_response = sapo_client.core.list_brands_search_raw(page=1, limit=250, query=supplier_code_from_api)
                        brands = brands_response.get("brands", [])
                        for brand in brands:
                            if brand.get('code', '').upper() == supplier_code_from_api.upper() or brand.get('name', '').upper() == supplier_code_from_api.upper():
                                brand_id = brand.get('id')
                                logger.info(f"[export_po_labels] [BƯỚC 2] ✅ Found brand_id: {brand_id} for supplier {supplier_code_from_api}")
                                break
        except Exception as e:
            logger.warning(f"[export_po_labels] Error getting supplier/brand info: {e}")
        
        # ===== BƯỚC 3: Load toàn bộ variants của NSX (theo brand_id) =====
        variant_sku_xnk_map = {}  # Map: variant_id -> sku_model_xnk
        variant_metadata_map = {}  # Map: variant_id -> variant_metadata (để lấy tq_name, tq_sku)
        
        if brand_id:
            logger.info(f"[export_po_labels] [BƯỚC 3] Loading all variants for brand_id={brand_id}...")
            product_service = SapoProductService(sapo_client)
            core_repo = sapo_client.core
            
            # Load tất cả products của brand này (giống variant_list)
            all_products = []
            page = 1
            limit = 250
            
            while True:
                filters = {
                    "page": page,
                    "limit": limit,
                    "status": "active"
                }
                
                products_response = core_repo.list_products_raw(**filters)
                products_data = products_response.get("products", [])
                
                if not products_data:
                    break
                
                # Filter theo brand_id (client-side)
                for product_data in products_data:
                    if product_data.get("brand_id") == brand_id:
                        all_products.append(product_data)
                
                if len(products_data) < limit:
                    break
                
                page += 1
                if page > 100:  # Safety limit
                    break
            
            logger.info(f"[export_po_labels] [BƯỚC 3] Found {len(all_products)} products for brand_id={brand_id}")
            
            # Parse variants và lấy sku_model_xnk từ metadata
            for product_data in all_products:
                product_id = product_data.get("id")
                if not product_id:
                    continue
                
                try:
                    product_dto = product_service.get_product(product_id)
                    if product_dto and product_dto.gdp_metadata:
                        # Lấy sku_model_xnk và variant metadata từ tất cả variants
                        for v_meta in product_dto.gdp_metadata.variants:
                            variant_id = v_meta.id
                            sku_model_xnk = v_meta.sku_model_xnk
                            if sku_model_xnk:
                                variant_sku_xnk_map[variant_id] = sku_model_xnk.strip()
                            # Lưu variant metadata để lấy tq_name và tq_sku
                            variant_metadata_map[variant_id] = v_meta
                except Exception as e:
                    logger.warning(f"[export_po_labels] Error parsing product {product_id}: {e}")
                    continue
            
            logger.info(f"[export_po_labels] [BƯỚC 3] ✅ Loaded {len(variant_sku_xnk_map)} variants with sku_model_xnk, {len(variant_metadata_map)} variants with metadata")
        else:
            logger.warning(f"[export_po_labels] [BƯỚC 3] ⚠️ No brand_id found, skipping variant loading")
        
        # Process line items để tạo labels
        label_items = []
        
        for item in line_items:
            variant_id = item.get('variant_id')
            quantity = item.get('quantity', 0)
            sku = item.get('sku', '')
            
            if quantity <= 0:
                continue
            
            try:
                # Lấy variant từ Sapo (chỉ cần basic info)
                variant_data = sapo_client.core.get_variant_raw(variant_id)
                if not variant_data:
                    continue
                
                # ===== BƯỚC 4: So trùng SKU nhập khẩu với XNK models =====
                xnk_model = None
                nsx_address = supplier_address  # Mặc định dùng supplier address
                sku_model_xnk_value = variant_sku_xnk_map.get(variant_id)  # Lấy từ map đã load
                
                if sku_model_xnk_value:
                    logger.info(f"[export_po_labels] [BƯỚC 4] Variant {variant_id} (SKU: {sku}) - sku_model_xnk: {sku_model_xnk_value}")
                    
                    # Tìm trong XNK model map (exact match)
                    xnk_model = xnk_model_map.get(sku_model_xnk_value)
                    
                    # Nếu không tìm thấy, thử case-insensitive search
                    if not xnk_model:
                        for map_sku, map_model in xnk_model_map.items():
                            if map_sku.upper() == sku_model_xnk_value.upper():
                                xnk_model = map_model
                                break
                    
                    if xnk_model:
                        logger.info(f"[export_po_labels] [BƯỚC 4] ✅ Found XNK model for SKU: '{sku_model_xnk_value}' -> Model SKU: '{xnk_model.get('sku', '')}'")
                        
                        # Lấy NSX address từ XNK model nếu có
                        nsx_address = (
                            xnk_model.get('nsx_address', '') or 
                            xnk_model.get('supplier_address', '') or 
                            xnk_model.get('address', '') or 
                            supplier_address
                        )
                    else:
                        # Log để debug - chỉ lấy 10 SKUs đầu tiên
                        available_skus = list(xnk_model_map.keys())[:10]
                        logger.warning(f"[export_po_labels] [BƯỚC 4] ⚠️ XNK model NOT found for SKU: '{sku_model_xnk_value}'. Available SKUs sample: {available_skus}")
                else:
                    logger.warning(f"[export_po_labels] [BƯỚC 4] ⚠️ Variant {variant_id} (SKU: {sku}) - No sku_model_xnk in variant map")
                
                # Lấy các field từ XNK model (xử lý cả en_name và name_en)
                hs_code = ''
                en_name = ''
                vn_name = ''
                nsx_name = supplier_name
                unit = ''
                
                if xnk_model:
                    hs_code = xnk_model.get('hs_code', '') or ''
                    # Xử lý cả en_name và name_en
                    en_name = xnk_model.get('en_name', '') or xnk_model.get('name_en', '') or ''
                    vn_name = xnk_model.get('vn_name', '') or xnk_model.get('name_vn', '') or ''
                    nsx_name = xnk_model.get('nsx_name', '') or supplier_name
                    unit = xnk_model.get('unit', '') or ''
                
                # Lấy opt1 từ nhiều nguồn (variant_data, item, hoặc parse từ variant_name)
                opt1 = (
                    variant_data.get('opt1', '') or 
                    item.get('opt1', '') or 
                    item.get('variant_options', '') or
                    ''
                )
                
                # Nếu vẫn không có, thử parse từ variant_name (format: "Product Name - opt1")
                if not opt1 and item.get('variant_name'):
                    variant_name = item.get('variant_name', '')
                    if ' - ' in variant_name:
                        parts = variant_name.split(' - ', 1)
                        if len(parts) > 1:
                            opt1 = parts[1].strip()
                
                # Lấy tq_name và tq_sku từ variant metadata
                tq_name = ''
                tq_sku = ''
                variant_meta = variant_metadata_map.get(variant_id)
                if variant_meta:
                    tq_name = variant_meta.name_tq or ''
                    tq_sku = variant_meta.sku_tq or ''
                
                # Tạo label data cho mỗi sản phẩm (1 label = 1 SKU)
                label_data = {
                    'sku': sku,
                    'barcode': variant_data.get('barcode', ''),
                    'opt1': opt1,
                    'product_name': item.get('product_name', ''),
                    'variant_name': item.get('variant_name', ''),
                    'quantity': quantity,
                    # XNK Model info (từ SKU nhập khẩu)
                    'hs_code': hs_code,
                    'en_name_xnk': en_name,
                    'vn_name_xnk': vn_name,
                    'nsx_name': nsx_name,
                    'nsx_address': nsx_address,
                    'unit': unit,
                    # Thông tin nhà sản xuất (tq_name / tq_sku)
                    'tq_name': tq_name,
                    'tq_sku': tq_sku,
                    # Debug info
                    'sku_model_xnk': sku_model_xnk_value or '',
                }
                
                # Tạo 1 label cho mỗi SKU (không phải mỗi quantity)
                label_items.append(label_data)
                
            except Exception as e:
                logger.error(f"Error processing label for variant {variant_id}: {e}", exc_info=True)
                continue
        
        # Render template
        context = {
            'po': po,
            'po_data': po_data,
            'label_items': label_items,
        }
        
        return render(request, 'products/po_labels.html', context)
        
    except Exception as e:
        logger.error(f"Error in export_po_labels: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
def delete_container_template(request: HttpRequest, template_id: int):
    """
    API endpoint để xóa container template.
    
    Returns:
        JSON: {status, message}
    """
    try:
        template = get_object_or_404(ContainerTemplate, id=template_id)
        
        # Kiểm tra xem template có đang được sử dụng bởi SPO nào không
        spo_count = template.sum_purchase_orders.count()
        if spo_count > 0:
            return JsonResponse({
                "status": "error",
                "message": f"Không thể xóa template này vì đang được sử dụng bởi {spo_count} SPO. Vui lòng xóa các SPO liên quan trước."
            }, status=400)
        
        # Xóa template (sẽ tự động xóa các suppliers liên quan do CASCADE)
        template_code = template.code
        template.delete()
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã xóa container template {template_code}"
        })
        
    except ContainerTemplate.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Container template không tồn tại"
        }, status=404)
    except Exception as e:
        logger.error(f"Error in delete_container_template: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_POST
def delete_sum_purchase_order(request: HttpRequest, spo_id: int):
    """
    API endpoint để xóa Sum Purchase Order (SPO).
    
    Returns:
        JSON: {status, message}
    """
    try:
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        
        # Kiểm tra trạng thái - chỉ cho phép xóa nếu ở trạng thái draft hoặc cancelled
        if spo.status not in ['draft', 'cancelled']:
            return JsonResponse({
                "status": "error",
                "message": f"Không thể xóa SPO ở trạng thái '{spo.get_status_display()}'. Chỉ có thể xóa SPO ở trạng thái 'Nháp' hoặc 'Đã hủy'."
            }, status=400)
        
        # Xóa SPO (sẽ tự động xóa các SPOPurchaseOrder liên quan do CASCADE)
        spo_code = spo.code
        spo.delete()
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã xóa SPO {spo_code}"
        })
        
    except SumPurchaseOrder.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "SPO không tồn tại"
        }, status=404)
    except Exception as e:
        logger.error(f"Error in delete_sum_purchase_order: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)


@admin_only
@require_http_methods(["GET"])
def get_valid_pos_for_spo(request: HttpRequest, spo_id: int):
    """
    API endpoint để lấy danh sách PO hợp lệ có thể thêm vào SPO.
    
    Điều kiện:
    - PO chưa thuộc SPO nào
    - statuses=pending
    - Có nhà sản xuất (NSX) thuộc SPO Template của SPO này
    
    Returns:
        JSON: {status, valid_pos: List[Dict]}
    """
    try:
        from products.models import SumPurchaseOrder, SPOPurchaseOrder, PurchaseOrder, ContainerTemplateSupplier
        
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        container_template = spo.container_template
        
        # Lấy danh sách supplier_ids thuộc container template
        template_suppliers = ContainerTemplateSupplier.objects.filter(
            container_template=container_template
        ).values_list('supplier_id', flat=True)
        
        if not template_suppliers:
            return JsonResponse({
                "status": "success",
                "valid_pos": [],
                "message": "Container template chưa có NSX nào"
            })
        
        # Lấy tất cả PO đã thuộc SPO nào đó
        po_ids_in_spo = SPOPurchaseOrder.objects.values_list('purchase_order_id', flat=True).distinct()
        
        # Lấy PO chưa thuộc SPO nào, có status pending, và có supplier thuộc template
        valid_pos_db = PurchaseOrder.objects.filter(
            supplier_id__in=template_suppliers,
            delivery_status='ordered'  # pending status
        ).exclude(
            id__in=po_ids_in_spo
        ).order_by('-created_at')[:50]  # Giới hạn 50 PO gần nhất
        
        # Lấy thông tin từ Sapo API
        sapo_client = get_sapo_client()
        from products.services.spo_po_service import SPOPOService
        spo_po_service = SPOPOService(sapo_client)
        
        valid_pos = []
        for po_db in valid_pos_db:
            try:
                # Lấy thông tin từ Sapo để kiểm tra status
                po_data = spo_po_service.get_po_from_sapo(po_db.sapo_order_supplier_id)
                sapo_status = po_data.get('status', '')
                
                # Chỉ lấy PO có status pending trong Sapo
                if sapo_status == 'pending':
                    valid_pos.append({
                        'sapo_order_supplier_id': po_db.sapo_order_supplier_id,
                        'code': po_data.get('code', ''),
                        'supplier_name': po_data.get('supplier_name', ''),
                        'total_amount_cny': float(po_data.get('product_amount_cny', 0)),
                        'total_cbm': float(po_data.get('total_cbm', 0)),
                        'total_quantity': po_data.get('total_quantity', 0),
                    })
            except Exception as e:
                logger.warning(f"Error getting PO {po_db.sapo_order_supplier_id} from Sapo: {e}")
                continue
        
        return JsonResponse({
            "status": "success",
            "valid_pos": valid_pos,
            "count": len(valid_pos)
        })
        
    except Exception as e:
        logger.error(f"Error in get_valid_pos_for_spo: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
def update_ship_info(request: HttpRequest, spo_id: int):
    """
    API endpoint để cập nhật thông tin tàu và tracking link.
    
    POST data:
    - ship_name: str (optional)
    - ship_tracking_link: str (optional)
    
    Returns:
        JSON: {status, message}
    """
    try:
        import json
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
        
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        
        # Cho phép cập nhật hoặc xóa (set về None nếu rỗng)
        if 'ship_name' in data:
            ship_name = data.get('ship_name', '').strip()
            spo.ship_name = ship_name if ship_name else None
        
        if 'ship_tracking_link' in data:
            ship_tracking_link = data.get('ship_tracking_link', '').strip()
            spo.ship_tracking_link = ship_tracking_link if ship_tracking_link else None
        
        spo.save()
        
        return JsonResponse({
            "status": "success",
            "message": "Đã cập nhật thông tin tàu"
        })
        
    except Exception as e:
        logger.error(f"Error in update_ship_info: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


def remove_po_from_spo(request: HttpRequest, spo_id: int, po_id: int):
    """
    API endpoint để xóa PO khỏi SPO (không xóa PO, chỉ xóa quan hệ).
    
    Args:
        spo_id: SumPurchaseOrder ID
        po_id: PurchaseOrder.id (database ID, không phải sapo_order_supplier_id)
    
    Returns:
        JSON: {status, message}
    """
    try:
        from products.models import SumPurchaseOrder, SPOPurchaseOrder, PurchaseOrder
        
        spo = get_object_or_404(SumPurchaseOrder, id=spo_id)
        
        # Tìm PurchaseOrder theo id (database ID, không phải sapo_order_supplier_id)
        try:
            purchase_order = PurchaseOrder.objects.get(id=po_id)
        except PurchaseOrder.DoesNotExist:
            return JsonResponse({
                "status": "error",
                "message": f"Không tìm thấy PO với ID {po_id}"
            }, status=404)
        
        # Tìm và xóa SPOPurchaseOrder
        try:
            spo_po = SPOPurchaseOrder.objects.get(
                sum_purchase_order=spo,
                purchase_order=purchase_order
            )
            spo_po.delete()
            
            # Tính lại total_cbm của SPO
            from products.services.sum_purchase_order_service import SumPurchaseOrderService
            sapo_client = get_sapo_client()
            spo_service = SumPurchaseOrderService(sapo_client)
            spo_service._recalculate_spo_cbm(spo)
            
            logger.info(f"[remove_po_from_spo] Removed PO {po_id} (sapo_id: {purchase_order.sapo_order_supplier_id}) from SPO {spo.code}")
            
            return JsonResponse({
                "status": "success",
                "message": f"Đã xóa PO khỏi SPO {spo.code}"
            })
        except SPOPurchaseOrder.DoesNotExist:
            return JsonResponse({
                "status": "error",
                "message": "PO không có trong SPO này"
            }, status=404)
        
    except Exception as e:
        logger.error(f"Error in remove_po_from_spo: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
def refresh_sales_forecast(request: HttpRequest):
    """
    API endpoint để refresh dữ liệu dự báo bán hàng.
    Tự động tính cả 30 ngày và 10 ngày.
    
    Returns:
        JSON với status và message
    """
    try:
        from products.services.sales_forecast_service import SalesForecastService
        
        sapo_client = get_sapo_client()
        forecast_service = SalesForecastService(sapo_client)
        
        # Tính toán lại với force_refresh=True cho cả 30 và 10 ngày
        logger.info(f"[refresh_sales_forecast] Refreshing forecast for 30 days and 10 days")
        
        # Tính 30 ngày
        forecast_map_30, all_products_30, all_variants_map_30 = forecast_service.calculate_sales_forecast(
            days=30,
            force_refresh=True
        )
        
        # Tính 10 ngày
        forecast_map_10, all_products_10, all_variants_map_10 = forecast_service.calculate_sales_forecast(
            days=10,
            force_refresh=True
        )
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã tính toán lại dự báo cho {len(forecast_map_30)} variants (30 ngày) và {len(forecast_map_10)} variants (10 ngày)",
            "count_30": len(forecast_map_30),
            "count_10": len(forecast_map_10)
        })
        
    except Exception as e:
        error_msg = str(e)
        # Xử lý lỗi Bad Gateway hoặc JSON parsing
        if "Bad Gateway" in error_msg or "502" in error_msg:
            error_msg = "Lỗi kết nối tới Sapo (Bad Gateway). Vui lòng thử lại sau vài giây."
        elif "JSON" in error_msg or "json" in error_msg.lower() or "Unexpected token" in error_msg:
            error_msg = "Lỗi xử lý dữ liệu từ Sapo. Vui lòng thử lại."
        
        logger.error(f"Error in refresh_sales_forecast: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": error_msg
        }, status=500)


@admin_only
def create_purchase_order_from_supplier(request: HttpRequest, supplier_id: int):
    """
    Màn hình lên đơn nhập hàng từ nhà cung cấp.
    Hiển thị danh sách variants với thông tin đầy đủ và form nhập số lượng.
    """
    context = {
        "title": "Lên đơn nhập hàng",
        "supplier": None,
        "variants": [],
        "error": None,
        "po_mode": "create",
        "current_po": None,
        "current_po_id": None,
    }
    
    try:
        from products.services.sapo_supplier_service import SapoSupplierService
        from products.services.sales_forecast_service import SalesForecastService
        from products.services.metadata_helper import extract_gdp_metadata
        
        sapo_client = get_sapo_client()
        supplier_service = SapoSupplierService(sapo_client)
        forecast_service = SalesForecastService(sapo_client)
        
        # Lấy thông tin supplier
        all_suppliers = supplier_service.get_all_suppliers(status="active")
        supplier = None
        for s in all_suppliers:
            if s.id == supplier_id:
                supplier = s
                break
        
        if not supplier:
            context["error"] = f"Không tìm thấy nhà cung cấp với ID {supplier_id}"
            return render(request, "products/create_purchase_order.html", context)
        
        context["supplier"] = {
            "id": supplier.id,
            "code": supplier.code,
            "name": supplier.name,
            "logo_path": supplier.logo_path,
        }
        
        # Lấy brand từ supplier (code hoặc name)
        brand_name = supplier.code or supplier.name
        
        # Lấy brand_id từ Sapo
        from products.services.sapo_product_service import SapoProductService
        product_service = SapoProductService(sapo_client)
        
        # Tìm brand_id
        brand_id = None
        try:
            # Lấy brands từ Sapo
            brands_response = sapo_client.core.list_brands_raw(query=brand_name)
            brands = brands_response.get("brands", [])
            for brand in brands:
                if (brand.get("code", "").upper() == brand_name.upper() or 
                    brand.get("name", "").upper() == brand_name.upper()):
                    brand_id = brand.get("id")
                    break
        except Exception as e:
            logger.warning(f"Error getting brand_id for {brand_name}: {e}")
        
        # Lấy products của brand này
        all_products = []
        all_variants_map = {}
        
        if brand_id:
            # Lấy products theo brand_id
            page = 1
            limit = 250
            while True:
                response = sapo_client.core.list_products_raw(
                    page=page,
                    limit=limit,
                    status="active",
                    product_types="normal",
                    brand_ids=str(brand_id)
                )
                
                products_data = response.get("products", [])
                if not products_data:
                    break
                
                all_products.extend(products_data)
                
                # Extract variants (chỉ lấy packsize=false)
                for product in products_data:
                    variants = product.get("variants", [])
                    for variant in variants:
                        variant_id = variant.get("id")
                        if variant_id:
                            packsize = variant.get("packsize", False)
                            if packsize is not True:
                                all_variants_map[variant_id] = variant
                
                if len(products_data) < limit:
                    break
                
                page += 1
                if page > 100:
                    break
        
        # Lấy forecast data từ database (cả 30 ngày và 10 ngày)
        forecast_map_30, all_products_30, all_variants_map_30 = forecast_service.calculate_sales_forecast(
            days=30,
            force_refresh=False
        )
        forecast_map_10, all_products_10, all_variants_map_10 = forecast_service.calculate_sales_forecast(
            days=10,
            force_refresh=False
        )
        
        # Dùng all_products và all_variants_map từ 30 ngày
        all_products_forecast = all_products_30
        all_variants_map_forecast = all_variants_map_30
        
        # Import LOCATION constants
        from products.services.sales_forecast_service import LOCATION_HN, LOCATION_SG
        
        # ========== LOAD PO DRAFT TỪ SAPO (NẾU CÓ) ==========
        draft_po = None
        draft_po_line_items_map = {}  # variant_id -> quantity (pcs)
        
        try:
            # Tìm PO DRAFT đang active cho supplier này
            po_response = sapo_client.core.list_order_suppliers_raw(
                page=1,
                limit=20,
                statuses="pending",
                supplier_ids=str(supplier_id),
                tags="DRAFT"
            )
            
            order_suppliers = po_response.get("order_suppliers", [])
            if order_suppliers:
                # Lấy PO đầu tiên (mới nhất)
                draft_po = order_suppliers[0]
                logger.info(f"[create_purchase_order] Found DRAFT PO: {draft_po.get('code')} (ID: {draft_po.get('id')})")
                
                # Lấy chi tiết PO để có line_items
                po_id = draft_po.get("id")
                if po_id:
                    po_detail = sapo_client.core.get_order_supplier_raw(po_id)
                    po_data = po_detail.get("order_supplier", {})
                    line_items = po_data.get("line_items", [])
                    
                    # Tạo map variant_id -> quantity
                    for item in line_items:
                        variant_id = item.get("variant_id")
                        quantity = item.get("quantity", 0)
                        if variant_id and quantity > 0:
                            draft_po_line_items_map[variant_id] = quantity
                    
                    logger.info(f"[create_purchase_order] Loaded {len(draft_po_line_items_map)} line items from DRAFT PO")
        except Exception as e:
            logger.warning(f"[create_purchase_order] Error loading DRAFT PO: {e}")
            # Tiếp tục với gợi ý nếu không load được PO

        # Nếu có draft_po thì chuyển sang chế độ edit PO hiện có
        if draft_po:
            context["po_mode"] = "edit"
            context["current_po"] = {
                "id": draft_po.get("id"),
                "code": draft_po.get("code"),
            }
            context["current_po_id"] = draft_po.get("id")
        
        # Tạo map variant_id -> product_data
        variant_to_product: Dict[int, Dict[str, Any]] = {}
        for product in all_products:
            variants = product.get("variants", [])
            for variant in variants:
                variant_id = variant.get("id")
                if variant_id:
                    variant_to_product[variant_id] = product
        
        # Lấy variants và thông tin đầy đủ (giống sales_forecast_list)
        variants_data = []
        for variant_id, variant in all_variants_map.items():
            # Lấy forecast cho 30 ngày và 10 ngày
            forecast_30 = forecast_map_30.get(variant_id)
            forecast_10 = forecast_map_10.get(variant_id)
            
            # Lấy variant_data từ map (đã có sẵn inventories)
            variant_data = all_variants_map_forecast.get(variant_id)
            product_data = variant_to_product.get(variant_id)
            
            # Dùng forecast_30 làm chính (hoặc forecast_10 nếu không có 30)
            main_forecast = forecast_30 or forecast_10
            if not main_forecast:
                continue
            
            # Sử dụng get_variant_forecast_with_inventory để lấy đầy đủ thông tin
            variant_info = forecast_service.get_variant_forecast_with_inventory(
                variant_id,
                {variant_id: main_forecast},  # Truyền map với 1 item
                variant_data=variant_data,  # Truyền variant_data đã có sẵn
                product_data=product_data,  # Truyền product_data để lấy brand
                forecast_30=forecast_30  # Truyền forecast_30 để tính gợi ý nhập hàng
            )
            
            # Tính tỉ lệ % cho 30 ngày và 10 ngày
            if forecast_30:
                # Tính ASP (Average Selling Price) = revenue / total_sold
                asp_value = None
                if forecast_30.revenue and forecast_30.revenue > 0 and forecast_30.total_sold and forecast_30.total_sold > 0:
                    asp_value = forecast_30.revenue / forecast_30.total_sold
                
                variant_info["forecast_30"] = {
                    "total_sold": forecast_30.total_sold,
                    "total_sold_previous_period": forecast_30.total_sold_previous_period,
                    "growth_percentage": forecast_30.growth_percentage,
                    "sales_rate": forecast_30.sales_rate,
                    "abc_category": forecast_30.abc_category,
                    "revenue": forecast_30.revenue,
                    "revenue_percentage": forecast_30.revenue_percentage,
                    "cumulative_percentage": forecast_30.cumulative_percentage,
                    "abc_rank": forecast_30.abc_rank,
                    "priority_score": forecast_30.priority_score,
                    "velocity_stability_score": forecast_30.velocity_stability_score,
                    "velocity_score": forecast_30.velocity_score,
                    "stability_bonus": forecast_30.stability_bonus,
                    "asp_score": forecast_30.asp_score,
                    "asp_value": asp_value,  # Giá trị ASP thực tế (VNĐ)
                    "revenue_contribution_score": forecast_30.revenue_contribution_score,
                }
                # Tính tỉ lệ %: (hiện tại - cùng kỳ) / cùng kỳ * 100
                if forecast_30.total_sold_previous_period > 0:
                    variant_info["growth_percentage_30"] = ((forecast_30.total_sold - forecast_30.total_sold_previous_period) / forecast_30.total_sold_previous_period) * 100
                else:
                    variant_info["growth_percentage_30"] = None
            else:
                variant_info["forecast_30"] = None
                variant_info["growth_percentage_30"] = None
                
            if forecast_10:
                # Tính ASP (Average Selling Price) = revenue / total_sold cho 10N
                asp_value_10 = None
                if forecast_10.revenue and forecast_10.revenue > 0 and forecast_10.total_sold and forecast_10.total_sold > 0:
                    asp_value_10 = forecast_10.revenue / forecast_10.total_sold
                
                variant_info["forecast_10"] = {
                    "total_sold": forecast_10.total_sold,
                    "total_sold_previous_period": forecast_10.total_sold_previous_period,
                    "growth_percentage": forecast_10.growth_percentage,
                    "sales_rate": forecast_10.sales_rate,
                    "revenue": forecast_10.revenue,  # Revenue cho 10N (nếu có)
                    "asp_value": asp_value_10,  # Giá trị ASP thực tế (VNĐ) cho 10N
                }
                # Tính tỉ lệ %: (hiện tại - cùng kỳ) / cùng kỳ * 100
                if forecast_10.total_sold_previous_period > 0:
                    variant_info["growth_percentage_10"] = ((forecast_10.total_sold - forecast_10.total_sold_previous_period) / forecast_10.total_sold_previous_period) * 100
                else:
                    variant_info["growth_percentage_10"] = None
            else:
                variant_info["forecast_10"] = None
                variant_info["growth_percentage_10"] = None
            
            # Lấy box_info và full_box từ metadata
            box_info = None
            full_box = None
            price_tq = None
            if product_data:
                description = product_data.get("description") or ""
                if description:
                    metadata, _ = extract_gdp_metadata(description)
                    if metadata:
                        for v_meta in metadata.variants:
                            if v_meta.id == variant_id:
                                if v_meta.box_info:
                                    box_info = v_meta.box_info
                                    full_box = box_info.full_box
                                if v_meta.price_tq:
                                    price_tq = v_meta.price_tq
                                break
            
            # Tính số thùng gợi ý (nếu có)
            suggested_boxes = None
            if variant_info.get("suggested_purchase_qty") and full_box and full_box > 0:
                boxes_float = variant_info["suggested_purchase_qty"] / full_box
                if boxes_float >= 0.5:
                    suggested_boxes = int(boxes_float + 0.5)  # Làm tròn 0.5 lên
                else:
                    suggested_boxes = 0
            
            # Nếu có PO DRAFT, ưu tiên dùng số lượng từ PO
            draft_quantity_pcs = draft_po_line_items_map.get(variant_id)
            draft_quantity_boxes = None
            if draft_quantity_pcs and full_box and full_box > 0:
                draft_quantity_boxes = int(draft_quantity_pcs / full_box + 0.5)  # Làm tròn
            
            # Thêm các thông tin bổ sung
            variant_info["draft_quantity_pcs"] = draft_quantity_pcs
            variant_info["draft_quantity_boxes"] = draft_quantity_boxes
            variant_info["suggested_boxes"] = suggested_boxes
            variant_info["full_box"] = full_box or 1
            variant_info["price_tq"] = price_tq
            variant_info["box_info"] = box_info
            
            variants_data.append(variant_info)
        
        context["variants"] = variants_data
        
        # ========== TÍNH TỔNG THÔNG TIN ĐƠN HÀNG NHÁP ==========
        total_quantity_pcs = 0
        total_quantity_boxes = 0
        total_volume_cbm = 0.0  # Tổng khối (m³)
        
        for variant in variants_data:
            # Dùng số lượng từ PO DRAFT nếu có, nếu không thì dùng gợi ý
            qty_pcs = variant.get("draft_quantity_pcs") or 0
            qty_boxes = variant.get("draft_quantity_boxes") or 0
            
            if qty_pcs > 0:
                total_quantity_pcs += qty_pcs
                total_quantity_boxes += qty_boxes
                
                # Tính khối (nếu có box_info)
                box_info = variant.get("box_info")
                if box_info:
                    length = getattr(box_info, 'length_cm', None)
                    width = getattr(box_info, 'width_cm', None)
                    height = getattr(box_info, 'height_cm', None)
                    if length and width and height:
                        # Tính khối 1 thùng (cm³ -> m³)
                        box_volume_cm3 = length * width * height
                        box_volume_m3 = box_volume_cm3 / 1000000  # cm³ -> m³
                        total_volume_cbm += box_volume_m3 * qty_boxes
        
        context["draft_po"] = draft_po
        context["draft_po_summary"] = {
            "total_quantity_pcs": total_quantity_pcs,
            "total_quantity_boxes": total_quantity_boxes,
            "total_volume_cbm": total_volume_cbm,
        }
        
    except Exception as e:
        logger.error(f"Error in create_purchase_order_from_supplier: {e}", exc_info=True)
        context["error"] = str(e)
    
    return render(request, "products/create_purchase_order.html", context)


@admin_only
def edit_purchase_order(request: HttpRequest, po_id: int):
    """
    Màn hình sửa một Purchase Order có sẵn (theo Sapo order_supplier.id).
    Giao diện giống create_purchase_order_from_supplier nhưng:
    - Luôn load từ PO hiện tại (không dùng tag DRAFT)
    - Nút hành động sẽ cập nhật vào chính PO này.
    """
    context = {
        "title": "Sửa đơn nhập hàng",
        "supplier": None,
        "variants": [],
        "error": None,
        "po_mode": "edit",
        "current_po": None,
        "current_po_id": po_id,
    }

    try:
        from products.services.sapo_supplier_service import SapoSupplierService
        from products.services.sales_forecast_service import SalesForecastService
        from products.services.metadata_helper import extract_gdp_metadata
        from products.services.sapo_product_service import SapoProductService

        sapo_client = get_sapo_client()
        supplier_service = SapoSupplierService(sapo_client)
        forecast_service = SalesForecastService(sapo_client)
        product_service = SapoProductService(sapo_client)

        # Lấy chi tiết PO từ Sapo
        po_detail = sapo_client.core.get_order_supplier_raw(po_id)
        po_data = po_detail.get("order_supplier", {})
        if not po_data:
            context["error"] = f"Không tìm thấy đơn hàng với ID {po_id}"
            return render(request, "products/create_purchase_order.html", context)

        context["current_po"] = {
            "id": po_data.get("id"),
            "code": po_data.get("code"),
            "status": po_data.get("status"),
        }

        supplier_id = po_data.get("supplier_id")
        if not supplier_id:
            context["error"] = "Đơn hàng không có thông tin nhà cung cấp"
            return render(request, "products/create_purchase_order.html", context)

        # Lấy thông tin supplier
        all_suppliers = supplier_service.get_all_suppliers(status="active")
        supplier = None
        for s in all_suppliers:
            if s.id == supplier_id:
                supplier = s
                break

        if not supplier:
            context["error"] = f"Không tìm thấy nhà cung cấp với ID {supplier_id}"
            return render(request, "products/create_purchase_order.html", context)

        context["supplier"] = {
            "id": supplier.id,
            "code": supplier.code,
            "name": supplier.name,
            "logo_path": supplier.logo_path,
        }

        # Lấy brand từ supplier (code hoặc name)
        brand_name = supplier.code or supplier.name

        # Lấy brand_id từ Sapo
        brand_id = None
        try:
            brands_response = sapo_client.core.list_brands_raw(query=brand_name)
            brands = brands_response.get("brands", [])
            for brand in brands:
                if (
                    brand.get("code", "").upper() == brand_name.upper()
                    or brand.get("name", "").upper() == brand_name.upper()
                ):
                    brand_id = brand.get("id")
                    break
        except Exception as e:
            logger.warning(f"Error getting brand_id for {brand_name}: {e}")

        # Lấy products của brand này
        all_products: List[Dict[str, Any]] = []
        all_variants_map: Dict[int, Dict[str, Any]] = {}

        if brand_id:
            page = 1
            limit = 250
            while True:
                response = sapo_client.core.list_products_raw(
                    page=page,
                    limit=limit,
                    status="active",
                    product_types="normal",
                    brand_ids=str(brand_id),
                )
                products_data = response.get("products", [])
                if not products_data:
                    break

                all_products.extend(products_data)

                # Extract variants (chỉ lấy packsize=false)
                for product in products_data:
                    variants = product.get("variants", [])
                    for variant in variants:
                        variant_id = variant.get("id")
                        if variant_id:
                            packsize = variant.get("packsize", False)
                            if packsize is not True:
                                all_variants_map[variant_id] = variant

                if len(products_data) < limit:
                    break

                page += 1
                if page > 100:
                    break

        # Forecast data
        forecast_map_30, all_products_30, all_variants_map_30 = forecast_service.calculate_sales_forecast(
            days=30,
            force_refresh=False,
        )
        forecast_map_10, all_products_10, all_variants_map_10 = forecast_service.calculate_sales_forecast(
            days=10,
            force_refresh=False,
        )

        all_variants_map_forecast = all_variants_map_30

        # Tạo map variant_id -> product_data
        variant_to_product: Dict[int, Dict[str, Any]] = {}
        for product in all_products:
            variants = product.get("variants", [])
            for variant in variants:
                variant_id = variant.get("id")
                if variant_id:
                    variant_to_product[variant_id] = product

        # Map variant_id -> quantity từ chính PO này
        draft_po_line_items_map: Dict[int, int] = {}
        for item in po_data.get("line_items", []):
            variant_id = item.get("variant_id")
            quantity = item.get("quantity", 0)
            if variant_id and quantity > 0:
                draft_po_line_items_map[variant_id] = quantity

        # Lấy variants và thông tin đầy đủ (giống create_purchase_order_from_supplier)
        variants_data: List[Dict[str, Any]] = []
        for variant_id, variant in all_variants_map.items():
            forecast_30 = forecast_map_30.get(variant_id)
            forecast_10 = forecast_map_10.get(variant_id)

            variant_data = all_variants_map_forecast.get(variant_id)
            product_data = variant_to_product.get(variant_id)

            main_forecast = forecast_30 or forecast_10
            if not main_forecast:
                continue

            variant_info = forecast_service.get_variant_forecast_with_inventory(
                variant_id,
                {variant_id: main_forecast},
                variant_data=variant_data,
                product_data=product_data,
                forecast_30=forecast_30,
            )

            # Tương tự create_purchase_order_from_supplier: gán forecast_30 / forecast_10
            if forecast_30:
                asp_value = None
                if (
                    forecast_30.revenue
                    and forecast_30.revenue > 0
                    and forecast_30.total_sold
                    and forecast_30.total_sold > 0
                ):
                    asp_value = forecast_30.revenue / forecast_30.total_sold

                variant_info["forecast_30"] = {
                    "total_sold": forecast_30.total_sold,
                    "total_sold_previous_period": forecast_30.total_sold_previous_period,
                    "growth_percentage": forecast_30.growth_percentage,
                    "sales_rate": forecast_30.sales_rate,
                    "abc_category": forecast_30.abc_category,
                    "revenue": forecast_30.revenue,
                    "revenue_percentage": forecast_30.revenue_percentage,
                    "cumulative_percentage": forecast_30.cumulative_percentage,
                    "abc_rank": forecast_30.abc_rank,
                    "priority_score": forecast_30.priority_score,
                    "velocity_stability_score": forecast_30.velocity_stability_score,
                    "velocity_score": forecast_30.velocity_score,
                    "stability_bonus": forecast_30.stability_bonus,
                    "asp_score": forecast_30.asp_score,
                    "asp_value": asp_value,
                    "revenue_contribution_score": forecast_30.revenue_contribution_score,
                }
                if forecast_30.total_sold_previous_period > 0:
                    variant_info["growth_percentage_30"] = (
                        (forecast_30.total_sold - forecast_30.total_sold_previous_period)
                        / forecast_30.total_sold_previous_period
                    ) * 100
                else:
                    variant_info["growth_percentage_30"] = None
            else:
                variant_info["forecast_30"] = None
                variant_info["growth_percentage_30"] = None

            if forecast_10:
                asp_value_10 = None
                if (
                    forecast_10.revenue
                    and forecast_10.revenue > 0
                    and forecast_10.total_sold
                    and forecast_10.total_sold > 0
                ):
                    asp_value_10 = forecast_10.revenue / forecast_10.total_sold

                variant_info["forecast_10"] = {
                    "total_sold": forecast_10.total_sold,
                    "total_sold_previous_period": forecast_10.total_sold_previous_period,
                    "growth_percentage": forecast_10.growth_percentage,
                    "sales_rate": forecast_10.sales_rate,
                    "revenue": forecast_10.revenue,
                    "asp_value": asp_value_10,
                }
                if forecast_10.total_sold_previous_period > 0:
                    variant_info["growth_percentage_10"] = (
                        (forecast_10.total_sold - forecast_10.total_sold_previous_period)
                        / forecast_10.total_sold_previous_period
                    ) * 100
                else:
                    variant_info["growth_percentage_10"] = None
            else:
                variant_info["forecast_10"] = None
                variant_info["growth_percentage_10"] = None

            # Lấy box_info, full_box, price_tq từ metadata
            box_info = None
            full_box = None
            price_tq = None
            if product_data:
                description = product_data.get("description") or ""
                if description:
                    metadata, _ = extract_gdp_metadata(description)
                    if metadata:
                        for v_meta in metadata.variants:
                            if v_meta.id == variant_id:
                                if v_meta.box_info:
                                    box_info = v_meta.box_info
                                    full_box = box_info.full_box
                                if v_meta.price_tq:
                                    price_tq = v_meta.price_tq
                                break

            # Số lượng từ PO hiện tại (pcs -> boxes)
            draft_quantity_pcs = draft_po_line_items_map.get(variant_id)
            draft_quantity_boxes = None
            if draft_quantity_pcs and full_box and full_box > 0:
                draft_quantity_boxes = int(draft_quantity_pcs / full_box + 0.5)

            variant_info["draft_quantity_pcs"] = draft_quantity_pcs
            variant_info["draft_quantity_boxes"] = draft_quantity_boxes
            variant_info["suggested_boxes"] = None
            variant_info["full_box"] = full_box or 1
            variant_info["price_tq"] = price_tq
            variant_info["box_info"] = box_info

            variants_data.append(variant_info)

        context["variants"] = variants_data

        # Tính tổng summary từ chính PO này
        total_quantity_pcs = 0
        total_quantity_boxes = 0
        total_volume_cbm = 0.0

        for variant in variants_data:
            qty_pcs = variant.get("draft_quantity_pcs") or 0
            qty_boxes = variant.get("draft_quantity_boxes") or 0

            if qty_pcs > 0:
                total_quantity_pcs += qty_pcs
                total_quantity_boxes += qty_boxes

                box_info = variant.get("box_info")
                if box_info:
                    length = getattr(box_info, "length_cm", None)
                    width = getattr(box_info, "width_cm", None)
                    height = getattr(box_info, "height_cm", None)
                    if length and width and height:
                        box_volume_cm3 = length * width * height
                        box_volume_m3 = box_volume_cm3 / 1000000
                        total_volume_cbm += box_volume_m3 * qty_boxes

        context["draft_po"] = po_data
        context["draft_po_summary"] = {
            "total_quantity_pcs": total_quantity_pcs,
            "total_quantity_boxes": total_quantity_boxes,
            "total_volume_cbm": total_volume_cbm,
        }

    except Exception as e:
        logger.error(f"Error in edit_purchase_order ({po_id}): {e}", exc_info=True)
        context["error"] = str(e)

    return render(request, "products/create_purchase_order.html", context)


@admin_only
@require_POST
@csrf_exempt
def submit_purchase_order_draft(request: HttpRequest):
    """
    API endpoint để tạo đơn nháp (Purchase Order) trong Sapo.
    
    Request body:
        {
            "supplier_id": int,
            "location_id": int (241737 = Gele/HN, 548744 = Toky/SG),
            "line_items": [
                {
                    "variant_id": int,
                    "quantity": int (số pcs),
                    "boxes": int (số thùng)
                },
                ...
            ]
        }
    
    Returns:
        JSON: {status, message, order_code, order_id}
    """
    try:
        import json
        from products.services.sales_forecast_service import LOCATION_HN, LOCATION_SG
        from products.services.metadata_helper import extract_gdp_metadata
        
        data = json.loads(request.body)
        supplier_id = data.get("supplier_id")
        location_id = data.get("location_id")
        line_items = data.get("line_items", [])
        po_id = data.get("po_id")  # Sapo order_supplier.id cần update (nếu có)
        
        if not supplier_id or not location_id:
            return JsonResponse({
                "status": "error",
                "message": "Thiếu supplier_id hoặc location_id"
            }, status=400)
        
        if not line_items:
            return JsonResponse({
                "status": "error",
                "message": "Vui lòng nhập số lượng cho ít nhất một phân loại"
            }, status=400)
        
        # Validate location_id
        if location_id not in [LOCATION_HN, LOCATION_SG]:
            return JsonResponse({
                "status": "error",
                "message": f"Location ID không hợp lệ: {location_id}"
            }, status=400)
        
        sapo_client = get_sapo_client()
        
        # Lấy thông tin supplier
        from products.services.sapo_supplier_service import SapoSupplierService
        supplier_service = SapoSupplierService(sapo_client)
        all_suppliers = supplier_service.get_all_suppliers(status="active")
        supplier = None
        for s in all_suppliers:
            if s.id == supplier_id:
                supplier = s
                break
        
        if not supplier:
            return JsonResponse({
                "status": "error",
                "message": f"Không tìm thấy nhà cung cấp với ID {supplier_id}"
            }, status=404)
        
        # Tạo order_supplier payload
        order_supplier_data = {
            "location_id": location_id,
            "supplier_id": supplier_id,
            "assignee_id": 319911,  # Default assignee
            "line_items": []
        }
        
        # Xử lý line_items: lấy price_tq từ metadata
        for item in line_items:
            variant_id = item.get("variant_id")
            quantity = item.get("quantity", 0)  # Số pcs
            
            if not variant_id or quantity <= 0:
                continue
            
            # Lấy variant để lấy product_id và price_tq
            try:
                variant_response = sapo_client.core.get_variant_raw(variant_id)
                variant = variant_response.get("variant", {})
                product_id = variant.get("product_id")
                
                # Lấy price_tq từ metadata
                price_tq = 0.0
                product_response = sapo_client.core.get_product_raw(product_id)
                product = product_response.get("product", {})
                description = product.get("description", "")
                
                if description:
                    metadata, _ = extract_gdp_metadata(description)
                    if metadata:
                        for v_meta in metadata.variants:
                            if v_meta.id == variant_id and v_meta.price_tq:
                                price_tq = v_meta.price_tq
                                break
                
                # Tạo line_item
                line_item = {
                    "variant_id": variant_id,
                    "product_id": product_id,
                    "quantity": quantity,
                    "units": [
                        {
                            "variant_id": variant_id,
                            "name": "chiếc"
                        }
                    ],
                    "tax": {
                        "tax_value": 0,
                        "tax_type": "exclude"
                    },
                    "excluded_tax_begin_amount": 0,
                    "excluded_tax_line_amount": 0
                }
                
                # Thêm price nếu có
                if price_tq > 0:
                    # Convert CNY sang VNĐ (tỷ giá tạm thời, có thể cần config)
                    price_vnd = int(price_tq * 3600)  # Tỷ giá tạm thời
                    line_item["price"] = price_vnd
                
                order_supplier_data["line_items"].append(line_item)
            except Exception as e:
                logger.warning(f"Error processing variant {variant_id}: {e}")
                continue
        
        if not order_supplier_data["line_items"]:
            return JsonResponse({
                "status": "error",
                "message": "Không có line items hợp lệ để tạo đơn"
            }, status=400)

        # ========== TRƯỜNG HỢP UPDATE VÀO PO CỤ THỂ (CÓ po_id) ==========
        if po_id:
            try:
                po_detail = sapo_client.core.get_order_supplier_raw(po_id)
                old_po = po_detail.get("order_supplier", {})
                if not old_po:
                    return JsonResponse({
                        "status": "error",
                        "message": f"Không tìm thấy PO với ID {po_id}"
                    }, status=404)

                order_supplier_data["id"] = old_po.get("id")
                order_supplier_data["code"] = old_po.get("code")
                order_supplier_data["supplier_id"] = old_po.get("supplier_id")
                order_supplier_data["tags"] = old_po.get("tags", ["CN"])
                # Giữ location_id theo yêu cầu người dùng
                order_supplier_data["location_id"] = location_id

                response = sapo_client.core.update_order_supplier_raw(
                    old_po.get("id"),
                    order_supplier_data,
                )
                order_supplier = response.get("order_supplier", {})
                if not order_supplier:
                    return JsonResponse({
                        "status": "error",
                        "message": "Không thể cập nhật đơn trong Sapo"
                    }, status=500)

                logger.info(f"[submit_purchase_order_draft] Updated specific PO: {old_po.get('code')} (ID: {order_supplier.get('id')})")

                return JsonResponse({
                    "status": "success",
                    "message": "Đã cập nhật đơn hàng thành công",
                    "order_code": old_po.get("code"),
                    "order_id": order_supplier.get("id"),
                    "is_update": True,
                    "redirect_url": f"/products/purchase-orders/{order_supplier.get('id')}/view-draft/"
                })
            except Exception as e:
                logger.error(f"[submit_purchase_order_draft] Error updating specific PO {po_id}: {e}", exc_info=True)
                return JsonResponse({
                    "status": "error",
                    "message": f"Lỗi khi cập nhật PO: {e}"
                }, status=500)

        # ========== CHECK XEM ĐÃ CÓ PO DRAFT CHO NSX NÀY CHƯA (FLOW CŨ) ==========
        # Tạo mã đơn tự động (CN-2025-SXX)
        try:
            # Lấy đơn mới nhất để tạo mã tiếp theo
            last_po_response = sapo_client.core.list_order_suppliers_raw(limit=1, page=1)
            last_pos = last_po_response.get("order_suppliers", [])
            
            if last_pos:
                last_code = last_pos[0].get("code", "")
                if "CN-2025-S" in last_code:
                    next_num = int(last_code.replace("CN-2025-S", "")) + 1
                    order_code = f"CN-2025-S{next_num:02d}"
                else:
                    order_code = "CN-2025-S01"
            else:
                order_code = "CN-2025-S01"
        except Exception as e:
            logger.warning(f"Error generating order code: {e}")
            order_code = f"CN-{datetime.now().strftime('%Y-%m-%d')}"
        
        # ========== CHECK XEM ĐÃ CÓ PO DRAFT CHƯA ==========
        existing_draft_po = None
        try:
            draft_po_response = sapo_client.core.list_order_suppliers_raw(
                page=1,
                limit=1,
                statuses="pending",
                supplier_ids=str(supplier_id),
                tags="DRAFT"
            )
            
            draft_pos = draft_po_response.get("order_suppliers", [])
            if draft_pos:
                # Đã có PO DRAFT -> lấy chi tiết để update
                existing_draft_po = draft_pos[0]
                logger.info(f"[submit_purchase_order_draft] Found existing DRAFT PO: {existing_draft_po.get('code')} (ID: {existing_draft_po.get('id')})")
                
                # Lấy chi tiết PO để giữ nguyên các thông tin khác
                po_detail = sapo_client.core.get_order_supplier_raw(existing_draft_po.get("id"))
                old_po = po_detail.get("order_supplier", {})
                
                # Cập nhật order_supplier_data với thông tin từ PO cũ
                order_supplier_data["id"] = old_po.get("id")
                order_supplier_data["code"] = old_po.get("code")
                order_supplier_data["supplier_id"] = old_po.get("supplier_id")
                order_supplier_data["tags"] = old_po.get("tags", ["CN", "DRAFT"])
                # Giữ nguyên location_id từ request (có thể thay đổi)
                order_supplier_data["location_id"] = location_id
                
                # Update đơn trong Sapo
                response = sapo_client.core.update_order_supplier_raw(
                    existing_draft_po.get("id"),
                    order_supplier_data
                )
                order_supplier = response.get("order_supplier", {})
                
                if not order_supplier:
                    return JsonResponse({
                        "status": "error",
                        "message": "Không thể cập nhật đơn trong Sapo"
                    }, status=500)
                
                logger.info(f"[submit_purchase_order_draft] Updated PO: {old_po.get('code')} (ID: {order_supplier.get('id')}) for supplier {supplier_id}")
                
                return JsonResponse({
                    "status": "success",
                    "message": f"Đã cập nhật đơn nháp thành công",
                    "order_code": old_po.get("code"),
                    "order_id": order_supplier.get("id"),
                    "is_update": True,
                    "redirect_url": f"/products/purchase-orders/{order_supplier.get('id')}/view-draft/"
                })
        except Exception as e:
            logger.warning(f"[submit_purchase_order_draft] Error checking existing DRAFT PO: {e}, will create new")
            # Tiếp tục tạo mới nếu có lỗi khi check
        
        # ========== TẠO MỚI NẾU CHƯA CÓ PO DRAFT ==========
        # Tạo mã đơn tự động (CN-2025-SXX)
        try:
            # Lấy đơn mới nhất để tạo mã tiếp theo
            last_po_response = sapo_client.core.list_order_suppliers_raw(limit=1, page=1)
            last_pos = last_po_response.get("order_suppliers", [])
            
            if last_pos:
                last_code = last_pos[0].get("code", "")
                if "CN-2025-S" in last_code:
                    next_num = int(last_code.replace("CN-2025-S", "")) + 1
                    order_code = f"CN-2025-S{next_num:02d}"
                else:
                    order_code = "CN-2025-S01"
            else:
                order_code = "CN-2025-S01"
        except Exception as e:
            logger.warning(f"Error generating order code: {e}")
            order_code = f"CN-{datetime.now().strftime('%Y-%m-%d')}"
        
        order_supplier_data["code"] = order_code
        order_supplier_data["tags"] = ["CN", "DRAFT"]
        
        # Tạo đơn trong Sapo
        response = sapo_client.core.create_order_supplier_raw(order_supplier_data)
        order_supplier = response.get("order_supplier", {})
        
        if not order_supplier:
            return JsonResponse({
                "status": "error",
                "message": "Không thể tạo đơn trong Sapo"
            }, status=500)
        
        logger.info(f"[submit_purchase_order_draft] Created PO: {order_code} (ID: {order_supplier.get('id')}) for supplier {supplier_id}")
        
        return JsonResponse({
            "status": "success",
            "message": f"Đã tạo đơn nháp thành công",
            "order_code": order_code,
            "order_id": order_supplier.get("id"),
            "is_update": False,
            "redirect_url": f"/products/purchase-orders/{order_supplier.get('id')}/view-draft/"
        })
        
    except Exception as e:
        logger.error(f"Error in submit_purchase_order_draft: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
def view_purchase_order_draft(request: HttpRequest, po_id: int):
    """
    Màn hình xem lại đơn hàng nháp đã tạo.
    Hiển thị table các sản phẩm trong đơn với thông tin đầy đủ.
    """
    context = {
        "title": "Chi tiết đơn hàng nháp",
        "po": None,
        "line_items": [],
        "supplier": None,
        "error": None,
    }
    
    try:
        sapo_client = get_sapo_client()
        
        # Lấy chi tiết đơn hàng từ Sapo
        po_detail = sapo_client.core.get_order_supplier_raw(po_id)
        po_data = po_detail.get("order_supplier", {})
        
        if not po_data:
            context["error"] = f"Không tìm thấy đơn hàng với ID {po_id}"
            return render(request, "products/view_purchase_order_draft.html", context)
        
        context["po"] = {
            "id": po_data.get("id"),
            "code": po_data.get("code"),
            "status": po_data.get("status"),
            "created_at": po_data.get("created_at"),
            "location_id": po_data.get("location_id"),
            "location_name": "Gele (Hà Nội)" if po_data.get("location_id") == 241737 else "Toky (Hồ Chí Minh)",
            "tags": po_data.get("tags", []),
        }
        
        # Lấy thông tin supplier
        supplier_id = po_data.get("supplier_id")
        if supplier_id:
            from products.services.sapo_supplier_service import SapoSupplierService
            supplier_service = SapoSupplierService(sapo_client)
            all_suppliers = supplier_service.get_all_suppliers(status="active")
            for s in all_suppliers:
                if s.id == supplier_id:
                    context["supplier"] = {
                        "id": s.id,
                        "code": s.code,
                        "name": s.name,
                        "logo_path": s.logo_path,
                    }
                    break
        
        # Lấy line_items và thông tin đầy đủ cho từng sản phẩm
        line_items_raw = po_data.get("line_items", [])
        line_items_data = []
        
        from products.services.metadata_helper import extract_gdp_metadata
        
        for item in line_items_raw:
            variant_id = item.get("variant_id")
            product_id = item.get("product_id")
            quantity = item.get("quantity", 0)
            price = item.get("price", 0)
            
            if not variant_id:
                continue
            
            # Lấy thông tin variant
            try:
                variant_response = sapo_client.core.get_variant_raw(variant_id)
                variant = variant_response.get("variant", {})
                
                # Lấy thông tin product để lấy metadata
                product_response = sapo_client.core.get_product_raw(product_id)
                product = product_response.get("product", {})
                
                # Lấy box_info từ metadata
                box_info = None
                full_box = None
                price_tq = None
                description = product.get("description", "")
                
                if description:
                    metadata, _ = extract_gdp_metadata(description)
                    if metadata:
                        for v_meta in metadata.variants:
                            if v_meta.id == variant_id:
                                if v_meta.box_info:
                                    box_info = v_meta.box_info
                                    full_box = box_info.full_box
                                if v_meta.price_tq:
                                    price_tq = v_meta.price_tq
                                break
                
                # Tính số thùng
                boxes = None
                if quantity > 0 and full_box and full_box > 0:
                    boxes = int(quantity / full_box + 0.5)
                
                # Tính khối (nếu có box_info)
                volume_cbm = None
                if boxes and box_info and box_info.length_cm and box_info.width_cm and box_info.height_cm:
                    box_volume_cm3 = box_info.length_cm * box_info.width_cm * box_info.height_cm
                    box_volume_m3 = box_volume_cm3 / 1000000  # cm³ -> m³
                    volume_cbm = box_volume_m3 * boxes
                
                # Lấy ảnh
                image_url = None
                images = variant.get("images", [])
                if images and len(images) > 0:
                    image_url = images[0].get("full_path") or images[0].get("path")
                
                line_items_data.append({
                    "variant_id": variant_id,
                    "product_id": product_id,
                    "sku": variant.get("sku", ""),
                    "name": variant.get("name", ""),
                    "opt1": variant.get("opt1", ""),
                    "image_url": image_url,
                    "quantity": quantity,  # Số pcs
                    "boxes": boxes,  # Số thùng
                    "volume_cbm": volume_cbm,  # Số khối (m³)
                    "price": price,  # Giá VNĐ
                    "price_tq": price_tq,  # Giá CNY
                    "full_box": full_box or 1,
                    "total_price": price * quantity if price and quantity else 0,  # Tổng giá VNĐ
                })
            except Exception as e:
                logger.warning(f"Error loading variant {variant_id}: {e}")
                # Vẫn thêm item nhưng thiếu thông tin
                line_items_data.append({
                    "variant_id": variant_id,
                    "product_id": product_id,
                    "sku": f"Variant {variant_id}",
                    "name": "N/A",
                    "opt1": "",
                    "image_url": None,
                    "quantity": quantity,
                    "boxes": None,
                    "volume_cbm": None,
                    "price": price,
                    "price_tq": None,
                    "full_box": 1,
                    "total_price": price * quantity if price and quantity else 0,
                })
        
        context["line_items"] = line_items_data
        
        # Tính tổng
        total_quantity = sum(item["quantity"] for item in line_items_data)
        total_boxes = sum(item["boxes"] or 0 for item in line_items_data)
        total_volume = sum(item["volume_cbm"] or 0 for item in line_items_data)
        total_price = sum(item["total_price"] for item in line_items_data)
        
        context["summary"] = {
            "total_quantity": total_quantity,
            "total_boxes": total_boxes,
            "total_volume": total_volume,
            "total_price": total_price,
        }
        
    except Exception as e:
        logger.error(f"Error in view_purchase_order_draft: {e}", exc_info=True)
        context["error"] = str(e)
    
    return render(request, "products/view_purchase_order_draft.html", context)
