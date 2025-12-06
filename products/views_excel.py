# products/views_excel.py
"""
Excel export/import views cho variants.
Tách riêng để dễ quản lý.
"""

from kho.utils import admin_only
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.views.decorators.http import require_POST
import logging
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from core.sapo_client import get_sapo_client
from products.services.sapo_product_service import SapoProductService
from products.brand_settings import (
    is_brand_enabled,
    reload_settings
)
from products.services.metadata_helper import init_empty_metadata, update_variant_metadata
from products.services.dto import VariantMetadataDTO, BoxInfoDTO

logger = logging.getLogger(__name__)


@admin_only
def export_variants_excel(request: HttpRequest):
    """
    Export danh sách variants ra file Excel.
    Chỉ export những variants trong brand_id đang lọc (theo query param brand_id).
    """
    try:
        # Brand ID mặc định
        DEFAULT_BRAND_ID = 833608
        
        # Lấy brand_id từ query param
        brand_id = request.GET.get('brand_id', str(DEFAULT_BRAND_ID))
        try:
            brand_id = int(brand_id)
        except (ValueError, TypeError):
            brand_id = DEFAULT_BRAND_ID
        
        sapo_client = get_sapo_client()
        core_repo = sapo_client.core
        product_service = SapoProductService(sapo_client)
        
        # Reload settings
        reload_settings()
        
        # Lấy products từ Sapo API (đã bao gồm variants và inventories)
        all_products = []
        page = 1
        limit = 250
        
        logger.info(f"[export_variants_excel] Starting to fetch products (with variants) for brand_id={brand_id}")
        
        while True:
            # Thử filter theo brand_id nếu API hỗ trợ
            filters = {
                "page": page,
                "limit": limit,
                "status": "active"
            }
            try:
                filters["brand_ids"] = brand_id
            except:
                pass
            
            products_response = core_repo.list_products_raw(**filters)
            products_data = products_response.get("products", [])
            
            if not products_data:
                logger.info(f"[export_variants_excel] No more products at page {page}")
                break
            
            all_products.extend(products_data)
            logger.info(f"[export_variants_excel] Fetched page {page}: {len(products_data)} products (total so far: {len(all_products)})")
            
            if len(products_data) < limit:
                break
            
            page += 1
            
            if page > 100:
                break
        
        # Parse products và extract variants
        product_map = {}
        all_variants_from_products = []
        
        for product_data in all_products:
            product_id = product_data.get("id")
            if not product_id:
                continue
            
            brand_id_from_product = product_data.get("brand_id")
            
            # Parse product để có metadata
            try:
                product = product_service.get_product(product_id)
                if product:
                    product_map[product_id] = product
            except Exception as e:
                logger.warning(f"Failed to parse product {product_id}: {e}")
                product = None
            
            # Extract variants từ product (đã có inventories sẵn)
            variants = product_data.get("variants", [])
            for variant_data in variants:
                variant_id = variant_data.get("id")
                if not variant_id:
                    continue
                
                variant_data["_product_id"] = product_id
                variant_data["_brand_id"] = brand_id_from_product
                variant_data["_product"] = product
                all_variants_from_products.append(variant_data)
        
        logger.info(f"[export_variants_excel] Extracted {len(all_variants_from_products)} variants from {len(all_products)} products")
        
        # Parse variants và lấy metadata, filter theo brand_id
        variants_data = []
        
        for variant_data in all_variants_from_products:
            variant_id = variant_data.get("id")
            product_id = variant_data.get("_product_id")
            product = variant_data.get("_product")
            
            # Filter theo brand_id (client-side)
            variant_brand_id = variant_data.get("_brand_id")
            if variant_brand_id != brand_id:
                continue
            
            # Lấy brand name
            brand = ""
            if product:
                brand = product.brand or ""
            
            # Chỉ thêm variants nếu nhãn hiệu được bật
            if brand and not is_brand_enabled(brand):
                continue
            
            # Lấy metadata từ product nếu có
            variant_meta = None
            if product:
                for v in product.variants:
                    if v.id == variant_id:
                        variant_meta = v.gdp_metadata
                        break
            
            # Lấy opt1 (tên phân loại) - kiểm tra cả opt1 và option1
            opt1_raw = variant_data.get("opt1")
            if opt1_raw is None:
                opt1_raw = variant_data.get("option1")
            opt1 = opt1_raw or ""
            
            # Chuẩn bị dữ liệu cho Excel
            row_data = {
                "sku": variant_data.get("sku", ""),
                "opt1": opt1,  # Tên phân loại - thêm sau SKU
                "vari_id": variant_id,
                "price_tq": variant_meta.price_tq if variant_meta and variant_meta.price_tq else "",
                "sku_tq": variant_meta.sku_tq if variant_meta and variant_meta.sku_tq else "",
                "name_tq": variant_meta.name_tq if variant_meta and variant_meta.name_tq else "",
                "full_box": variant_meta.box_info.full_box if variant_meta and variant_meta.box_info and variant_meta.box_info.full_box else "",
                "box_length_cm": variant_meta.box_info.length_cm if variant_meta and variant_meta.box_info and variant_meta.box_info.length_cm else "",
                "box_width_cm": variant_meta.box_info.width_cm if variant_meta and variant_meta.box_info and variant_meta.box_info.width_cm else "",
                "box_height_cm": variant_meta.box_info.height_cm if variant_meta and variant_meta.box_info and variant_meta.box_info.height_cm else "",
                "sku_model_xnk": variant_meta.sku_model_xnk if variant_meta and variant_meta.sku_model_xnk else "",
                "update": ""  # Cột update để user đánh dấu
            }
            variants_data.append(row_data)
        
        # Tạo Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Variants"
        
        # Headers - opt1 (tên phân loại) sau SKU
        headers = [
            "sku", "opt1", "vari_id", "price_tq", "sku_tq", "name_tq",
            "full_box", "box_length_cm", "box_width_cm", "box_height_cm",
            "sku_model_xnk", "update"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row_data in enumerate(variants_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="variants_export.xlsx"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting variants Excel: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@admin_only
@require_POST
def import_variants_excel(request: HttpRequest):
    """
    Import và cập nhật variants từ file Excel.
    Chỉ cập nhật những dòng có giá trị trong cột "update".
    """
    try:
        if 'file' not in request.FILES:
            return JsonResponse({
                "status": "error",
                "message": "Không có file được upload"
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Kiểm tra file extension
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                "status": "error",
                "message": "File phải là định dạng Excel (.xlsx hoặc .xls)"
            }, status=400)
        
        # Đọc file Excel
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        
        # Đọc headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Tìm index của các cột
        header_map = {}
        for idx, header in enumerate(headers):
            if header:
                header_map[str(header).strip()] = idx
        
        # Kiểm tra các cột bắt buộc
        required_headers = ['vari_id', 'update']
        missing_headers = [h for h in required_headers if h not in header_map]
        if missing_headers:
            return JsonResponse({
                "status": "error",
                "message": f"Thiếu các cột bắt buộc: {', '.join(missing_headers)}"
            }, status=400)
        
        # Đọc dữ liệu
        sapo_client = get_sapo_client()
        product_service = SapoProductService(sapo_client)
        
        update_idx = header_map['update']
        vari_id_idx = header_map.get('vari_id')
        
        results = {
            "total_rows": 0,
            "processed": 0,
            "success": 0,
            "skipped": 0,
            "errors": 0,
            "error_details": []
        }
        
        # Process từng dòng
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            results["total_rows"] += 1
            
            # Kiểm tra cột update
            update_cell = row[update_idx] if update_idx < len(row) else None
            update_value = update_cell.value if update_cell else None
            
            # Chỉ xử lý nếu có giá trị trong cột update
            if not update_value or str(update_value).strip() == "":
                results["skipped"] += 1
                continue
            
            try:
                # Lấy vari_id
                vari_id = None
                if vari_id_idx is not None and vari_id_idx < len(row):
                    vari_id_value = row[vari_id_idx].value
                    if vari_id_value:
                        try:
                            vari_id = int(float(vari_id_value))
                        except:
                            pass
                
                if not vari_id:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Không tìm thấy variant_id")
                    continue
                
                # Lấy product_id từ variant
                variant_response = sapo_client.core.get_variant_raw(vari_id)
                variant = variant_response.get("variant", {})
                if not variant:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Variant {vari_id} không tồn tại")
                    continue
                
                product_id = variant.get("product_id")
                if not product_id:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Variant {vari_id} không có product_id")
                    continue
                
                # Lấy product hiện tại
                product = product_service.get_product(product_id)
                if not product:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Product {product_id} không tồn tại")
                    continue
                
                # Lấy metadata hiện tại
                current_metadata = product.gdp_metadata
                if not current_metadata:
                    variant_ids = [v.id for v in product.variants]
                    current_metadata = init_empty_metadata(product_id, variant_ids)
                
                # Tìm variant metadata hiện tại
                variant_meta = None
                for vm in current_metadata.variants:
                    if vm.id == vari_id:
                        variant_meta = vm
                        break
                
                if not variant_meta:
                    variant_meta = VariantMetadataDTO(id=vari_id)
                
                # Cập nhật từ Excel
                def get_cell_value(header_name, default=None):
                    if header_name in header_map:
                        idx = header_map[header_name]
                        if idx < len(row):
                            value = row[idx].value
                            return value if value is not None else default
                    return default
                
                def to_float_safe(val, default=None):
                    try:
                        if val is None or val == "":
                            return default
                        return float(val)
                    except:
                        return default
                
                def to_int_safe(val, default=None):
                    try:
                        if val is None or val == "":
                            return default
                        return int(float(val))
                    except:
                        return default
                
                # Update các field
                price_tq = to_float_safe(get_cell_value('price_tq'))
                sku_tq = get_cell_value('sku_tq', '').strip() if get_cell_value('sku_tq') else None
                name_tq = get_cell_value('name_tq', '').strip() if get_cell_value('name_tq') else None
                sku_model_xnk = get_cell_value('sku_model_xnk', '').strip() if get_cell_value('sku_model_xnk') else None
                
                # Update box_info
                full_box = to_int_safe(get_cell_value('full_box'))
                box_length = to_float_safe(get_cell_value('box_length_cm'))
                box_width = to_float_safe(get_cell_value('box_width_cm'))
                box_height = to_float_safe(get_cell_value('box_height_cm'))
                
                box_info = None
                if full_box or box_length or box_width or box_height:
                    box_info = BoxInfoDTO(
                        full_box=full_box if full_box else (variant_meta.box_info.full_box if variant_meta.box_info else None),
                        length_cm=box_length if box_length else (variant_meta.box_info.length_cm if variant_meta.box_info else None),
                        width_cm=box_width if box_width else (variant_meta.box_info.width_cm if variant_meta.box_info else None),
                        height_cm=box_height if box_height else (variant_meta.box_info.height_cm if variant_meta.box_info else None)
                    )
                elif variant_meta.box_info:
                    box_info = variant_meta.box_info
                
                # Tạo variant metadata mới
                new_variant_meta = VariantMetadataDTO(
                    id=vari_id,
                    price_tq=price_tq if price_tq is not None else variant_meta.price_tq,
                    sku_tq=sku_tq if sku_tq else variant_meta.sku_tq,
                    name_tq=name_tq if name_tq else variant_meta.name_tq,
                    box_info=box_info,
                    packed_info=variant_meta.packed_info,
                    sku_model_xnk=sku_model_xnk if sku_model_xnk else variant_meta.sku_model_xnk,
                    web_variant_id=variant_meta.web_variant_id if variant_meta.web_variant_id else []
                )
                
                # Update variant metadata
                current_metadata = update_variant_metadata(
                    current_metadata,
                    vari_id,
                    new_variant_meta
                )
                
                # Lưu vào Sapo
                success = product_service.update_product_metadata(
                    product_id,
                    current_metadata,
                    preserve_description=True
                )
                
                if success:
                    results["success"] += 1
                else:
                    results["errors"] += 1
                    results["error_details"].append(f"Dòng {row_idx}: Không thể lưu variant {vari_id}")
                
                results["processed"] += 1
                
            except Exception as e:
                results["errors"] += 1
                error_msg = f"Dòng {row_idx}: {str(e)}"
                results["error_details"].append(error_msg)
                logger.error(f"Error processing row {row_idx}: {e}", exc_info=True)
        
        return JsonResponse({
            "status": "success",
            "result": results
        })
        
    except Exception as e:
        logger.error(f"Error importing variants Excel: {e}", exc_info=True)
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)

